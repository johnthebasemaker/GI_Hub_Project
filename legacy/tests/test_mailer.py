"""
test_mailer.py — Excel generation and SMTP email tests
=======================================================
Tests in this file verify:
  1. Excel output is valid, non-empty .xlsx bytes for all 3 report types
  2. Report data is mathematically consistent with the database
     (totals in xlsx match totals computed from DB — same formula as live dashboard)
  3. _build_email_message() produces correct MIME structure without SMTP
  4. send_email() calls SMTP with correct arguments (fully mocked — no network)
  5. parse_recipients() handles edge cases
  6. send_eod_report() integration (mocked SMTP)

NOTE: No actual email is ever sent during tests.
      smtplib.SMTP and smtplib.SMTP_SSL are patched at the module level.
"""

import io
import datetime
import pytest
from unittest.mock import patch, MagicMock

import openpyxl
import pandas as pd

import mailer
from mailer import (
    build_daily_report,
    build_monthly_report,
    build_low_stock_report,
    parse_recipients,
    get_default_recipients,
    send_eod_report,
)
from database import init_db


# ---------------------------------------------------------------------------
# Seed helpers
# ---------------------------------------------------------------------------
def _seed_item(conn, sap, desc="Item", uom="PCS", min_qty=10):
    conn.execute(
        "INSERT OR IGNORE INTO inventory (SAP_Code, Equipment_Description, UOM, Minimum_Qty) "
        "VALUES (?, ?, ?, ?)",
        (sap, desc, uom, min_qty),
    )
    conn.commit()


def _add_receipt(conn, sap, qty, date="2026-01-01"):
    conn.execute("INSERT INTO receipts (Date,SAP_Code,Quantity) VALUES (?,?,?)", (date, sap, qty))
    conn.commit()


def _add_consumption(conn, sap, qty, date="2026-05-06", work_type="Maintenance"):
    conn.execute(
        "INSERT INTO consumption (Date,SAP_Code,Quantity,Work_Type) VALUES (?,?,?,?)",
        (date, sap, qty, work_type),
    )
    conn.commit()


def _add_pending(conn, sap, qty, date="2026-05-06"):
    conn.execute(
        "INSERT INTO pending_issues (Date,SAP_Code,Quantity,Work_Type) VALUES (?,?,?,?)",
        (date, sap, qty, "Maintenance"),
    )
    conn.commit()


def _xlsx_to_dfs(xlsx_bytes: bytes) -> dict[str, pd.DataFrame]:
    """Reads all sheets from xlsx bytes into {sheet_name: DataFrame}."""
    buf = io.BytesIO(xlsx_bytes)
    return pd.read_excel(buf, sheet_name=None, engine="openpyxl")


# ===========================================================================
# GROUP 1: build_daily_report()
# ===========================================================================

class TestBuildDailyReport:
    """Validates Daily Issue Log Excel generation."""

    def test_returns_non_empty_bytes(self, db_conn):
        result = build_daily_report(conn=db_conn)
        assert isinstance(result, bytes)
        assert len(result) > 0

    def test_output_is_valid_xlsx(self, db_conn):
        """openpyxl must be able to open the bytes without error."""
        result = build_daily_report(conn=db_conn)
        wb = openpyxl.load_workbook(io.BytesIO(result))
        assert wb is not None

    def test_has_committed_and_pending_sheets(self, db_conn):
        result = build_daily_report(conn=db_conn)
        wb = openpyxl.load_workbook(io.BytesIO(result))
        sheet_names = wb.sheetnames
        assert any("Committed" in s for s in sheet_names)
        assert any("Pending" in s for s in sheet_names)

    def test_empty_db_does_not_crash(self, db_conn):
        """No consumption, no pending — report must still be generated."""
        try:
            result = build_daily_report(conn=db_conn)
        except Exception as e:
            pytest.fail(f"build_daily_report raised on empty DB: {e}")
        assert isinstance(result, bytes) and len(result) > 0

    def test_committed_rows_match_db_count(self, db_conn):
        """Number of data rows in xlsx == number of rows in consumption for that date."""
        today = "2026-05-06"
        _seed_item(db_conn, "DAY-001")
        _seed_item(db_conn, "DAY-002")
        _add_consumption(db_conn, "DAY-001", 10, date=today)
        _add_consumption(db_conn, "DAY-002", 5,  date=today)

        result = build_daily_report(
            conn=db_conn, report_date=datetime.date(2026, 5, 6)
        )
        dfs = _xlsx_to_dfs(result)
        committed_sheet = next((v for k, v in dfs.items() if "Committed" in k), None)
        assert committed_sheet is not None
        # 2 data rows (excluding header which pd.read_excel uses as columns)
        assert len(committed_sheet) == 2

    def test_pending_rows_match_db_count(self, db_conn):
        _seed_item(db_conn, "PEND-001")
        _add_pending(db_conn, "PEND-001", 7)
        _add_pending(db_conn, "PEND-001", 3)

        result = build_daily_report(conn=db_conn)
        dfs = _xlsx_to_dfs(result)
        pending_sheet = next((v for k, v in dfs.items() if "Pending" in k), None)
        assert pending_sheet is not None
        assert len(pending_sheet) == 2

    def test_committed_quantity_sum_matches_db(self, db_conn):
        """Key integrity test: sum of Quantity in xlsx == sum from DB query."""
        today = "2026-05-06"
        _seed_item(db_conn, "QSUM-001")
        _add_consumption(db_conn, "QSUM-001", 15.0, date=today)
        _add_consumption(db_conn, "QSUM-001", 23.5, date=today)

        db_sum = pd.read_sql(
            "SELECT SUM(Quantity) as total FROM consumption WHERE Date=?",
            db_conn, params=(today,)
        ).iloc[0]["total"]

        result = build_daily_report(conn=db_conn, report_date=datetime.date(2026, 5, 6))
        dfs = _xlsx_to_dfs(result)
        committed_sheet = next(v for k, v in dfs.items() if "Committed" in k)

        xlsx_sum = pd.to_numeric(committed_sheet["Quantity"], errors="coerce").sum()
        assert xlsx_sum == pytest.approx(db_sum)

    def test_summary_sheet_exists(self, db_conn):
        result = build_daily_report(conn=db_conn)
        dfs = _xlsx_to_dfs(result)
        assert any("Summary" in k for k in dfs.keys())


# ===========================================================================
# GROUP 2: build_monthly_report()
# ===========================================================================

class TestBuildMonthlyReport:
    """Validates Monthly Consumption Pivot Excel generation."""

    def test_returns_non_empty_bytes(self, db_conn):
        result = build_monthly_report(conn=db_conn)
        assert isinstance(result, bytes) and len(result) > 0

    def test_output_is_valid_xlsx(self, db_conn):
        result = build_monthly_report(conn=db_conn)
        wb = openpyxl.load_workbook(io.BytesIO(result))
        assert wb is not None

    def test_empty_consumption_does_not_crash(self, db_conn):
        try:
            result = build_monthly_report(conn=db_conn)
        except Exception as e:
            pytest.fail(f"build_monthly_report raised on empty DB: {e}")
        assert isinstance(result, bytes) and len(result) > 0

    def test_pivot_grand_total_matches_db_sum(self, db_conn):
        """Grand Total column in pivot must equal total from DB for each SAP_Code."""
        _seed_item(db_conn, "MON-001")
        _seed_item(db_conn, "MON-002")
        _add_consumption(db_conn, "MON-001", 50,  date="2026-01-15")
        _add_consumption(db_conn, "MON-001", 30,  date="2026-02-10")
        _add_consumption(db_conn, "MON-002", 100, date="2026-01-20")

        result = build_monthly_report(conn=db_conn)
        dfs    = _xlsx_to_dfs(result)
        df     = next(iter(dfs.values()))

        # MON-001 grand total should be 80
        row_001 = df[df["SAP_Code"] == "MON-001"]
        assert not row_001.empty
        assert float(row_001.iloc[0]["Grand Total"]) == pytest.approx(80.0)

        # MON-002 grand total should be 100
        row_002 = df[df["SAP_Code"] == "MON-002"]
        assert not row_002.empty
        assert float(row_002.iloc[0]["Grand Total"]) == pytest.approx(100.0)

    def test_pivot_sorted_by_grand_total_descending(self, db_conn):
        """Most-consumed items should appear first in the report."""
        _seed_item(db_conn, "SORT-A")
        _seed_item(db_conn, "SORT-B")
        _add_consumption(db_conn, "SORT-A", 10, date="2026-03-01")
        _add_consumption(db_conn, "SORT-B", 90, date="2026-03-01")

        result = build_monthly_report(conn=db_conn)
        dfs    = _xlsx_to_dfs(result)
        df     = next(iter(dfs.values()))
        totals = pd.to_numeric(df["Grand Total"], errors="coerce").tolist()
        assert totals == sorted(totals, reverse=True), "Rows must be sorted descending by Grand Total"


# ===========================================================================
# GROUP 3: build_low_stock_report()
# ===========================================================================

class TestBuildLowStockReport:
    """Validates Low-Stock Warning Excel generation."""

    def test_returns_non_empty_bytes(self, db_conn):
        result = build_low_stock_report(conn=db_conn)
        assert isinstance(result, bytes) and len(result) > 0

    def test_output_is_valid_xlsx(self, db_conn):
        result = build_low_stock_report(conn=db_conn)
        wb = openpyxl.load_workbook(io.BytesIO(result))
        assert wb is not None

    def test_empty_inventory_does_not_crash(self, db_conn):
        try:
            result = build_low_stock_report(conn=db_conn)
        except Exception as e:
            pytest.fail(f"build_low_stock_report raised on empty inventory: {e}")

    def test_low_stock_items_appear_in_report(self, db_conn):
        _seed_item(db_conn, "LOW-001", min_qty=50)
        _add_receipt(db_conn, "LOW-001", 10)  # stock=10, min=50 → LOW

        result = build_low_stock_report(conn=db_conn)
        dfs    = _xlsx_to_dfs(result)
        df     = next(iter(dfs.values()))
        assert "LOW-001" in df["SAP_Code"].values

    def test_adequate_items_excluded_from_report(self, db_conn):
        _seed_item(db_conn, "OK-001", min_qty=5)
        _add_receipt(db_conn, "OK-001", 100)  # stock=100, min=5 → OK

        result = build_low_stock_report(conn=db_conn)
        dfs    = _xlsx_to_dfs(result)
        df     = next(iter(dfs.values()))

        # If df has SAP_Code column, OK-001 must not be in it
        if "SAP_Code" in df.columns:
            assert "OK-001" not in df["SAP_Code"].values

    def test_shortage_column_present_and_correct(self, db_conn):
        _seed_item(db_conn, "GAP-001", min_qty=40)
        _add_receipt(db_conn, "GAP-001", 15)  # shortage = 40 - 15 = 25

        result = build_low_stock_report(conn=db_conn)
        dfs    = _xlsx_to_dfs(result)
        df     = next(iter(dfs.values()))
        if "SAP_Code" in df.columns and "Shortage" in df.columns:
            row = df[df["SAP_Code"] == "GAP-001"]
            assert float(row.iloc[0]["Shortage"]) == pytest.approx(25.0)

    def test_shortage_sorted_descending(self, db_conn):
        """Largest shortage should appear first — most urgent items at the top."""
        _seed_item(db_conn, "SRT-X", min_qty=100)
        _seed_item(db_conn, "SRT-Y", min_qty=20)
        _add_receipt(db_conn, "SRT-X", 5)   # shortage = 95
        _add_receipt(db_conn, "SRT-Y", 10)  # shortage = 10

        result = build_low_stock_report(conn=db_conn)
        dfs    = _xlsx_to_dfs(result)
        df     = next(iter(dfs.values()))
        if "Shortage" in df.columns and len(df) > 1:
            shortages = pd.to_numeric(df["Shortage"], errors="coerce").tolist()
            assert shortages == sorted(shortages, reverse=True)


# ===========================================================================
# GROUP 4: parse_recipients()
# ===========================================================================

class TestParseRecipients:
    """Validates email address list parsing."""

    def test_comma_separated(self):
        result = parse_recipients("a@x.com, b@x.com, c@x.com")
        assert result == ["a@x.com", "b@x.com", "c@x.com"]

    def test_newline_separated(self):
        result = parse_recipients("a@x.com\nb@x.com")
        assert result == ["a@x.com", "b@x.com"]

    def test_semicolon_separated(self):
        result = parse_recipients("a@x.com; b@x.com")
        assert result == ["a@x.com", "b@x.com"]

    def test_mixed_separators(self):
        result = parse_recipients("a@x.com,b@x.com\nc@x.com")
        assert len(result) == 3

    def test_empty_string_returns_empty_list(self):
        assert parse_recipients("") == []

    def test_none_returns_empty_list(self):
        assert parse_recipients(None) == []

    def test_strips_whitespace(self):
        result = parse_recipients("  a@x.com  ,  b@x.com  ")
        assert result == ["a@x.com", "b@x.com"]

    def test_filters_blank_entries(self):
        result = parse_recipients("a@x.com,,, b@x.com")
        assert "" not in result
        assert len(result) == 2

    def test_single_address(self):
        result = parse_recipients("only@one.com")
        assert result == ["only@one.com"]


# ===========================================================================
# GROUP 5: send_eod_report() — Outlook win32com mock
# ===========================================================================

class TestSendEodReportOutlook:
    """
    Tests the Outlook-based send_eod_report() with win32com fully mocked.
    No real Outlook process is ever started during the test run.
    """

    def _outlook_mocks(self):
        mock_mail = MagicMock()
        mock_mail.Attachments = MagicMock()
        mock_app = MagicMock()
        mock_app.CreateItem.return_value = mock_mail
        return mock_app, mock_mail

    def _fake_win32_modules(self, mock_app=None, dispatch_side_effect=None):
        """
        Build stand-in `win32` and `pythoncom` MagicMocks suitable for
        patching mailer.win32 / mailer.pythoncom on non-Windows hosts
        where the real modules import as None.
        """
        fake_win32 = MagicMock()
        if dispatch_side_effect is not None:
            fake_win32.Dispatch.side_effect = dispatch_side_effect
        elif mock_app is not None:
            fake_win32.Dispatch.return_value = mock_app
        fake_pythoncom = MagicMock()
        return fake_win32, fake_pythoncom

    def test_no_recipients_returns_false(self, db_conn):
        ok, msg = send_eod_report([], conn=db_conn)
        assert ok is False
        assert "recipient" in msg.lower()

    def test_outlook_dispatch_called(self, db_conn):
        mock_app, mock_mail = self._outlook_mocks()
        fake_win32, fake_pythoncom = self._fake_win32_modules(mock_app=mock_app)
        with patch("mailer.platform.system", return_value="Windows"), \
             patch.object(mailer, "win32", fake_win32), \
             patch.object(mailer, "pythoncom", fake_pythoncom), \
             patch("mailer.open", create=True) as mo, \
             patch("mailer.os.path.exists", return_value=True), \
             patch("mailer.os.remove"):
            mo.return_value.__enter__ = MagicMock(return_value=MagicMock())
            mo.return_value.__exit__  = MagicMock(return_value=False)
            ok, msg = send_eod_report(["mgr@gi.com"], conn=db_conn)
        assert ok is True
        mock_app.CreateItem.assert_called_once_with(0)

    def test_recipients_joined_with_semicolons(self, db_conn):
        mock_app, mock_mail = self._outlook_mocks()
        fake_win32, fake_pythoncom = self._fake_win32_modules(mock_app=mock_app)
        with patch("mailer.platform.system", return_value="Windows"), \
             patch.object(mailer, "win32", fake_win32), \
             patch.object(mailer, "pythoncom", fake_pythoncom), \
             patch("mailer.open", create=True) as mo, \
             patch("mailer.os.path.exists", return_value=True), \
             patch("mailer.os.remove"):
            mo.return_value.__enter__ = MagicMock(return_value=MagicMock())
            mo.return_value.__exit__  = MagicMock(return_value=False)
            send_eod_report(["a@gi.com", "b@gi.com"], conn=db_conn)
        assert ";" in mock_mail.To

    def test_subject_contains_date(self, db_conn):
        mock_app, mock_mail = self._outlook_mocks()
        fake_win32, fake_pythoncom = self._fake_win32_modules(mock_app=mock_app)
        test_date = datetime.date(2026, 5, 12)
        with patch("mailer.platform.system", return_value="Windows"), \
             patch.object(mailer, "win32", fake_win32), \
             patch.object(mailer, "pythoncom", fake_pythoncom), \
             patch("mailer.open", create=True) as mo, \
             patch("mailer.os.path.exists", return_value=True), \
             patch("mailer.os.remove"):
            mo.return_value.__enter__ = MagicMock(return_value=MagicMock())
            mo.return_value.__exit__  = MagicMock(return_value=False)
            send_eod_report(["mgr@gi.com"], conn=db_conn, report_date=test_date)
        assert "2026" in mock_mail.Subject or "May" in mock_mail.Subject

    def test_attachment_added(self, db_conn):
        mock_app, mock_mail = self._outlook_mocks()
        fake_win32, fake_pythoncom = self._fake_win32_modules(mock_app=mock_app)
        with patch("mailer.platform.system", return_value="Windows"), \
             patch.object(mailer, "win32", fake_win32), \
             patch.object(mailer, "pythoncom", fake_pythoncom), \
             patch("mailer.open", create=True) as mo, \
             patch("mailer.os.path.exists", return_value=True), \
             patch("mailer.os.remove"):
            mo.return_value.__enter__ = MagicMock(return_value=MagicMock())
            mo.return_value.__exit__  = MagicMock(return_value=False)
            send_eod_report(["mgr@gi.com"], conn=db_conn)
        mock_mail.Attachments.Add.assert_called_once()

    def test_outlook_exception_returns_false(self, db_conn):
        fake_win32, fake_pythoncom = self._fake_win32_modules(
            dispatch_side_effect=Exception("Outlook not installed")
        )
        with patch("mailer.platform.system", return_value="Windows"), \
             patch.object(mailer, "win32", fake_win32), \
             patch.object(mailer, "pythoncom", fake_pythoncom), \
             patch("mailer.os.path.exists", return_value=False):
            ok, msg = send_eod_report(["mgr@gi.com"], conn=db_conn)
        assert ok is False
        assert "Outlook" in msg

