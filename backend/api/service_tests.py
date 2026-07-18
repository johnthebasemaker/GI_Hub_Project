"""
backend/api/service_tests.py — service-level + auth/role guard tests (CI gate).

Two suites, both run against a **populated** Postgres (the same one dual_ci.py
loads from gi_database.db):

  A. Service invariants — call the write services inside a transaction, assert
     their effects (rows, audit, notifications) via count-deltas, then ROLL BACK.
     Nothing persists, so there is no cleanup and no divergence from SQLite.

  B. Auth/role guards — drive the real ASGI app with httpx and assert the
     endpoint guards: 401 without a token, 403 for the wrong role (including the
     master-data write gate), 200 on an open read.

Run:  DATABASE_URL=postgresql+psycopg2://…  python backend/api/service_tests.py
Exit code is non-zero if any check fails (so CI fails the build).
"""
from __future__ import annotations

import asyncio
import os
import sys

# CI determinism: never let backend.api.config pull a developer's local
# .env/deploy/.env into the test process — a real WHATSAPP_TOKEN there would
# flip wa.enabled() True and un-mocked suites could hit Meta live. Must be set
# BEFORE the .db/.main imports below trigger the config loader.
os.environ.setdefault("GI_DOTENV", "0")

from httpx import ASGITransport, AsyncClient
from sqlalchemy import func, select

from .db import SessionLocal, engine
from .main import app
from .services import ledger, notifications, procurement, supervisor

_MD = ledger._MD
pr_master_t = _MD.tables["pr_master"]
audit_t = _MD.tables["system_audit_log"]
notif_t = _MD.tables["app_notifications"]
smr_t = _MD.tables["supervisor_material_requests"]
smr_items_t = _MD.tables["supervisor_material_request_items"]
pending_issues_t = _MD.tables["pending_issues"]
receipts_t = _MD.tables["receipts"]
lots_t = _MD.tables["lots"]

PASSED: list[str] = []
FAILED: list[str] = []


def check(name: str, cond: bool, detail: str = "") -> None:
    (PASSED if cond else FAILED).append(name)
    mark = "✅" if cond else "❌"
    line = f"  {mark} {name}"
    if not cond and detail:
        line += f"  — {detail}"
    print(line)


async def _count(session, table, *where) -> int:
    stmt = select(func.count()).select_from(table)
    for w in where:
        stmt = stmt.where(w)
    return (await session.execute(stmt)).scalar_one()


# --- Suite A: service invariants (rolled back) -------------------------------
async def test_create_and_submit_pr():
    async with SessionLocal() as s:
        # Delta-counted: PR numbers restart per day, so a leftover audit row
        # from a cleaned-up test PR with the same number must not fail this.
        audit_before = await _count(s, audit_t, audit_t.c["action_type"] == "CREATE_PR")
        res = await procurement.create_pr(
            s, username="svc_hod", site_id="CNCEC",
            lines=[{"SAP_Code": "1001", "Requested_Qty": 3},
                   {"SAP_Code": "1002", "Requested_Qty": 2}])
        pr = res.get("pr_number")
        check("create_pr returns created", res.get("created") is True, str(res))
        n_lines = await _count(s, pr_master_t, pr_master_t.c["PR_Number"] == pr)
        check("create_pr writes one row per line", n_lines == 2, f"got {n_lines}")
        audit_after = await _count(s, audit_t, audit_t.c["action_type"] == "CREATE_PR")
        check("create_pr writes a CREATE_PR audit", audit_after == audit_before + 1,
              f"{audit_before} → {audit_after}")

        sub = await procurement.submit_pr(s, username="svc_hod", pr_number=pr, site_id="CNCEC")
        check("submit_pr succeeds", sub.get("submitted") is True, str(sub))
        n_notif = await _count(
            s, notif_t, notif_t.c["event_key"] == "pr_submitted_to_logistics",
            notif_t.c["related_ref"] == pr, notif_t.c["recipient_role"] == "logistics")
        check("submit_pr notifies logistics", n_notif == 1, f"got {n_notif}")
        await s.rollback()


async def test_smr_create_and_approve():
    async with SessionLocal() as s:
        created = await supervisor.create_smr(
            s, supervisor="svc_sup", site_id="CNCEC", worker_id="30001",
            job_tank_place="svc test", old_ppe_returned=1, no_return_reason=None,
            items=[{"SAP_Code": "1001", "Requested_Qty": 2}])
        rid = created.get("request_id")
        no = created.get("request_no")
        check("create_smr succeeds", created.get("created") is True, str(created))
        n_items = await _count(s, smr_items_t, smr_items_t.c["request_id"] == rid)
        check("create_smr writes items", n_items == 1, f"got {n_items}")
        n_notif = await _count(
            s, notif_t, notif_t.c["event_key"] == "smr_created",
            notif_t.c["related_ref"] == no, notif_t.c["recipient_role"] == "store_keeper",
            notif_t.c["recipient_site"] == "CNCEC")
        check("create_smr notifies store-keeper@site", n_notif == 1, f"got {n_notif}")

        item_id = (await supervisor.smr_items(s, rid))[0]["id"]
        appr = await supervisor.approve_smr(s, sk_username="svc_sk", request_id=rid,
                                            qty_overrides={item_id: 1.5})
        check("approve_smr succeeds", appr.get("approved") is True, str(appr))
        n_pending = await _count(
            s, pending_issues_t, pending_issues_t.c["Source_Ref"].like(f"SMR:{no}:%"),
            pending_issues_t.c["status"] == "pending_hod")
        check("approve_smr stages pending_issues", n_pending == 1, f"got {n_pending}")
        staged_qty = (await s.execute(select(pending_issues_t.c["Quantity"]).where(
            pending_issues_t.c["Source_Ref"].like(f"SMR:{no}:%")))).scalar_one()
        check("SK qty-override lands on the staged issue (2 → 1.5)",
              abs(float(staged_qty) - 1.5) < 1e-9, f"got {staged_qty}")
        n_fb = await _count(
            s, notif_t, notif_t.c["event_key"] == "smr_approved",
            notif_t.c["related_ref"] == no, notif_t.c["recipient_user"] == "svc_sup")
        check("approve_smr notifies the requester", n_fb == 1, f"got {n_fb}")
        await s.rollback()


async def test_receipt_ledger():
    async with SessionLocal() as s:
        r_before = await _count(s, receipts_t, receipts_t.c["SAP_Code"] == "1001",
                                receipts_t.c["Site_ID"] == "CNCEC")
        a_before = await _count(s, audit_t, audit_t.c["action_type"] == "POST_RECEIPT")
        res = await ledger.post_receipt(s, username="svc", data={
            "Date": "2026-07-04", "SAP_Code": "1001", "Quantity": 5, "Site_ID": "CNCEC",
            "Supplier": "svctest", "Remarks": "svctest", "Expiry_Date": "2027-06-01"})
        check("post_receipt returns a receipt_id", bool(res.get("receipt_id")), str(res))
        r_after = await _count(s, receipts_t, receipts_t.c["SAP_Code"] == "1001",
                               receipts_t.c["Site_ID"] == "CNCEC")
        check("post_receipt inserts one receipt", r_after == r_before + 1)
        lot = res.get("lot_number")
        n_lot = await _count(s, lots_t, lots_t.c["Lot_Number"] == lot)
        check("post_receipt auto-creates the lot", bool(lot) and n_lot >= 1, f"lot={lot}")
        a_after = await _count(s, audit_t, audit_t.c["action_type"] == "POST_RECEIPT")
        check("post_receipt writes an audit row", a_after == a_before + 1)
        await s.rollback()


async def test_submitter_resolution():
    """hod._submitter maps each pending kind to its submitter column (receipts=None)."""
    from . import hod
    async with SessionLocal() as s:
        r = await ledger.stage_return(s, username="svc_sub", data={
            "Date": "2026-07-04", "SAP_Code": "1001", "Quantity": 1, "Site_ID": "CNCEC"})
        sub = await hod._submitter(s, "returns", r["pending_id"])
        check("_submitter resolves the return's submitter", sub == "svc_sub", f"got {sub}")
        rr = await ledger.stage_receipt(s, username="svc_sub", data={
            "Date": "2026-07-04", "SAP_Code": "1001", "Quantity": 1, "Site_ID": "CNCEC"})
        none_sub = await hod._submitter(s, "receipts", rr["pending_id"])
        check("_submitter is None for receipts (no submitter column)", none_sub is None, f"got {none_sub}")
        await s.rollback()


async def test_notification_visibility():
    async with SessionLocal() as s:
        await notifications.notify(s, event_key="svc_role_ev", title="t", recipient_role="logistics")
        await notifications.notify(s, event_key="svc_user_ev", title="t", recipient_user="svc_alice")

        sk = await notifications.list_for(s, username="svc_bob", role="store_keeper",
                                          site_id="CNCEC", warehouse_id=None, limit=200)
        check("isolation: store-keeper can't see a logistics broadcast",
              not any(n["event_key"] == "svc_role_ev" for n in sk))
        lg = await notifications.list_for(s, username="svc_carol", role="logistics",
                                          site_id=None, warehouse_id=None, limit=200)
        check("logistics sees the role broadcast",
              any(n["event_key"] == "svc_role_ev" for n in lg))
        al = await notifications.list_for(s, username="svc_alice", role="store_keeper",
                                          site_id="ZZ", warehouse_id=None, limit=200)
        check("user-targeted notification is visible to its recipient",
              any(n["event_key"] == "svc_user_ev" for n in al))

        nid = (await s.execute(select(notif_t.c["id"])
               .where(notif_t.c["event_key"] == "svc_user_ev"))).scalars().first()
        ok = await notifications.mark_read(s, notif_id=nid, username="svc_bob",
                                           role="store_keeper", site_id="CNCEC", warehouse_id=None)
        check("mark_read guard blocks a non-recipient", ok is False)
        ok2 = await notifications.mark_read(s, notif_id=nid, username="svc_alice",
                                            role="store_keeper", site_id="ZZ", warehouse_id=None)
        check("mark_read succeeds for the recipient", ok2 is True)
        await s.rollback()


# --- Suite B: auth/role guards (live ASGI app) -------------------------------
async def test_auth_guards():
    transport = ASGITransport(app=app)
    async with AsyncClient(transport=transport, base_url="http://svc") as ac:
        async def token(u, p):
            r = await ac.post("/auth/login", json={"username": u, "password": p})
            return r.json().get("access_token")

        def H(t):
            return {"Authorization": f"Bearer {t}"}

        admin_t = await token("admin", "admin2026")
        worker_t = await token("worker", "floor2026")
        check("admin + worker can log in", bool(admin_t) and bool(worker_t))

        r = await ac.get("/inventory")
        check("no token → 401 on a protected read", r.status_code == 401, f"got {r.status_code}")
        r = await ac.get("/inventory", headers=H(worker_t))
        check("worker → 200 on an open read", r.status_code == 200, f"got {r.status_code}")

        for path in ("/admin/users", "/hod/pending", "/logistics/prs"):
            r = await ac.get(path, headers=H(worker_t))
            check(f"worker (lvl 0) → 403 on {path}", r.status_code == 403, f"got {r.status_code}")

        # The hardening fix: master-data writes are role-gated (level ≥ 3).
        r = await ac.post("/vendors", headers=H(worker_t), json={"Vendor_Name": "svc_x"})
        check("worker → 403 on POST /vendors (master-data write gate)",
              r.status_code == 403, f"got {r.status_code}")
        # Admin passes the role gate (bad column → 422, proving the guard let it through
        # without persisting anything).
        r = await ac.post("/vendors", headers=H(admin_t), json={"__not_a_column__": 1})
        check("admin passes the write gate (422 on bad body, not 403)",
              r.status_code == 422, f"got {r.status_code}")
        check("admin → 200 on /admin/users",
              (await ac.get("/admin/users", headers=H(admin_t))).status_code == 200)

        # Inventory editor guards (non-persisting: duplicate SAP + delete-with-movements).
        r = await ac.post("/admin/inventory", headers=H(worker_t), json={"SAP_Code": "svc_x"})
        check("worker → 403 on POST /admin/inventory", r.status_code == 403, f"got {r.status_code}")
        r = await ac.post("/admin/inventory", headers=H(admin_t), json={"SAP_Code": "1001"})
        check("admin → 409 creating a duplicate SAP (no persist)", r.status_code == 409, f"got {r.status_code}")
        r = await ac.request("DELETE", "/admin/inventory/1001", headers=H(admin_t))
        check("admin → 409 deleting an item with movements", r.status_code == 409, f"got {r.status_code}")

        # 2FA self-enrollment guards (non-persisting).
        r = await ac.get("/auth/2fa/status", headers=H(worker_t))
        check("2fa status → 200 for any authed user",
              r.status_code == 200 and r.json().get("enabled") is False, f"got {r.status_code}")
        r = await ac.post("/auth/2fa/verify", headers=H(worker_t), json={"code": "000000"})
        check("2fa verify without enrollment → 409", r.status_code == 409, f"got {r.status_code}")
        r = await ac.post("/auth/2fa/disable", headers=H(worker_t), json={"code": "000000"})
        check("2fa disable when not enabled → 409", r.status_code == 409, f"got {r.status_code}")

        # Reports (read-only, level ≥ 2): role gate + format validation.
        r = await ac.get("/reports/stock", params={"format": "xlsx"}, headers=H(worker_t))
        check("worker (lvl 0) → 403 on a report", r.status_code == 403, f"got {r.status_code}")
        r = await ac.get("/reports", headers=H(admin_t))
        check("admin → 200 on /reports list",
              r.status_code == 200 and len(r.json().get("reports", [])) >= 1, f"got {r.status_code}")
        r = await ac.get("/reports/stock", params={"format": "xlsx"}, headers=H(admin_t))
        check("stock report xlsx → 200 + spreadsheet content-type",
              r.status_code == 200 and "spreadsheetml" in r.headers.get("content-type", ""),
              f"got {r.status_code} / {r.headers.get('content-type')}")
        r = await ac.get("/reports/nope", params={"format": "xlsx"}, headers=H(admin_t))
        check("unknown report → 404", r.status_code == 404, f"got {r.status_code}")
        r = await ac.get("/reports/stock", params={"format": "docx"}, headers=H(admin_t))
        check("bad report format → 400", r.status_code == 400, f"got {r.status_code}")

        # Registration + access requests (non-persisting: only failing paths + reads).
        r = await ac.post("/auth/register", json={"username": "svc_admin_wannabe",
                          "password": "secret123", "role": "admin"})
        check("register requesting admin role → 422 (no self-elevation)",
              r.status_code == 422, f"got {r.status_code}")
        r = await ac.post("/auth/register", json={"username": "admin",
                          "password": "secret123", "role": "store_keeper"})
        check("register an existing username → 409", r.status_code == 409, f"got {r.status_code}")
        r = await ac.post("/auth/register", json={"username": "svc_x", "password": "no",
                          "role": "store_keeper"})
        check("register short password → 422", r.status_code == 422, f"got {r.status_code}")
        r = await ac.get("/admin/pending-users", headers=H(worker_t))
        check("worker → 403 on /admin/pending-users", r.status_code == 403, f"got {r.status_code}")
        r = await ac.get("/admin/pending-users", headers=H(admin_t))
        check("admin → 200 on /admin/pending-users", r.status_code == 200, f"got {r.status_code}")
        r = await ac.post("/admin/pending-users/999999/approve", headers=H(admin_t), json={})
        check("approve a non-existent request → 404", r.status_code == 404, f"got {r.status_code}")

        # T4 — role-conditional site rules. Public site list first (no auth).
        r = await ac.get("/auth/register/sites")
        check("public /auth/register/sites → 200 with a list",
              r.status_code == 200 and isinstance(r.json().get("sites"), list),
              f"got {r.status_code}")
        _sites = r.json().get("sites", [])
        _made_site = False
        if not _sites:  # fresh env: create a throwaway admin site to test against
            r = await ac.post("/admin/sites", headers=H(admin_t), json={"name": "SVC-SITE"})
            _made_site = r.status_code == 201
            _sites = ["SVC-SITE"]
        _site = _sites[0]

        # Distinct X-Real-IPs isolate these from the 5/min register cap (same
        # pattern as the login rate-limit test below).
        _t4a, _t4b = {"X-Real-IP": "203.0.113.41"}, {"X-Real-IP": "203.0.113.42"}
        r = await ac.post("/auth/register", headers=_t4a, json={"username": "svc_t4_hod",
                          "password": "secret123", "role": "hod"})
        check("scoped role without a site → 422", r.status_code == 422, f"got {r.status_code}")
        r = await ac.post("/auth/register", headers=_t4a, json={"username": "svc_t4_hod",
                          "password": "secret123", "role": "hod", "site_id": "NOT-A-SITE"})
        check("scoped role with an unknown site → 422", r.status_code == 422, f"got {r.status_code}")
        r = await ac.post("/auth/register", headers=_t4a, json={"username": "svc_t4_log",
                          "password": "secret123", "role": "logistics", "site_id": _site})
        check("unscoped role WITH a site → 422 (global roles carry no site)",
              r.status_code == 422, f"got {r.status_code}")

        # Happy paths — register, verify surfaced fields, then reject (cleanup;
        # re-runs revive the rejected row instead of colliding).
        r = await ac.post("/auth/register", headers=_t4b, json={"username": "svc_t4_hod",
                          "password": "secret123", "role": "hod", "site_id": _site})
        check("scoped role with an admin-created site → 201",
              r.status_code == 201, f"got {r.status_code}")
        r = await ac.post("/auth/register", headers=_t4b, json={"username": "svc_t4_log",
                          "password": "secret123", "role": "logistics",
                          "location": "Central Warehouse, Dammam"})
        check("unscoped role with a free-text location → 201",
              r.status_code == 201, f"got {r.status_code}")
        rows = (await ac.get("/admin/pending-users", headers=H(admin_t))).json()["items"]
        _hod_row = next((x for x in rows if x["username"] == "svc_t4_hod"), None)
        _log_row = next((x for x in rows if x["username"] == "svc_t4_log"), None)
        check("pending hod row carries the picked Site_ID",
              _hod_row is not None and _hod_row["Site_ID"] == _site, f"row={_hod_row}")
        check("pending logistics row carries Location + empty Site_ID",
              _log_row is not None and _log_row["Site_ID"] == ""
              and _log_row.get("Location") == "Central Warehouse, Dammam", f"row={_log_row}")
        for _row in (_hod_row, _log_row):  # cleanup → rejected (revivable)
            if _row is not None:
                await ac.post(f"/admin/pending-users/{_row['id']}/reject", headers=H(admin_t))
        if _made_site:  # drop the throwaway site so re-runs stay clean
            sid = next((s["id"] for s in (await ac.get("/admin/sites", headers=H(admin_t))
                        ).json()["items"] if s["name"] == "SVC-SITE"), None)
            if sid is not None:
                await ac.delete(f"/admin/sites/{sid}", headers=H(admin_t))

        # Rate limiting on public auth (isolated by a unique X-Real-IP so it does
        # not affect the other logins in this suite; login cap is 10/min).
        rl = {"X-Real-IP": "203.0.113.7"}
        codes = [(await ac.post("/auth/login", json={"username": "nobody", "password": "x"}, headers=rl)).status_code
                 for _ in range(12)]
        check("public /auth/login is rate-limited (429 past the cap)", 429 in codes, f"codes={codes}")
        check("attempts under the cap are 401, not 429", codes[0] == 401, f"first={codes[0]}")


async def test_token_refresh():
    """Access/refresh split: cookie issuance, rotation, reuse detection
    (family revocation), and logout revocation."""
    transport = ASGITransport(app=app)
    async with AsyncClient(transport=transport, base_url="http://svc") as ac:
        r = await ac.post("/auth/login", json={"username": "worker", "password": "floor2026"})
        old_refresh = r.cookies.get("gi_refresh")
        check("login sets the httpOnly refresh cookie", bool(old_refresh),
              f"status={r.status_code}")

        r2 = await ac.post("/auth/refresh")
        check("refresh → 200 + a new access token",
              r2.status_code == 200 and bool(r2.json().get("access_token")),
              f"got {r2.status_code}")
        new_refresh = r2.cookies.get("gi_refresh")
        check("refresh rotates the cookie", bool(new_refresh) and new_refresh != old_refresh)

        # NB: manual cookie sets use a different jar key (domain) than
        # response-set cookies — clear the jar first or requests carry BOTH
        # gi_refresh cookies and the server reads the stale one.
        def use_cookie(tok):
            ac.cookies.clear()
            ac.cookies.set("gi_refresh", tok, domain="svc")

        # Replaying the OLD (rotated) token must trip reuse detection…
        use_cookie(old_refresh)
        r3 = await ac.post("/auth/refresh")
        check("replaying a rotated token → 401 (reuse detection)",
              r3.status_code == 401, f"got {r3.status_code}")
        # …which revokes the whole family, including the successor.
        use_cookie(new_refresh)
        r4 = await ac.post("/auth/refresh")
        check("reuse detection also revoked the successor token",
              r4.status_code == 401, f"got {r4.status_code}")

        # Fresh session → logout revokes it server-side.
        ac.cookies.clear()
        r5 = await ac.post("/auth/login", json={"username": "worker", "password": "floor2026"})
        tok5 = r5.cookies.get("gi_refresh")
        r6 = await ac.post("/auth/logout")
        check("logout → 200", r6.status_code == 200, f"got {r6.status_code}")
        use_cookie(tok5)
        r7 = await ac.post("/auth/refresh")
        check("refresh after logout → 401 (session revoked)",
              r7.status_code == 401, f"got {r7.status_code}")


async def test_site_scoping():
    """Multi-site isolation: below logistics (level 3), every read is pinned to
    the caller's own Site_ID; admin/logistics stay global."""
    transport = ASGITransport(app=app)
    async with AsyncClient(transport=transport, base_url="http://svc") as ac:
        async def token(u, p):
            r = await ac.post("/auth/login", json={"username": u, "password": p})
            return r.json().get("access_token")

        def H(t):
            return {"Authorization": f"Bearer {t}"}

        worker_t = await token("worker", "floor2026")   # store_keeper @ CNCEC
        hod_t = await token("hod", "hod2026")           # hod @ CNCEC (level 2 → scoped)
        admin_t = await token("admin", "admin2026")     # level 4 → global

        r = await ac.get("/receipts", params={"limit": 500}, headers=H(worker_t))
        items = r.json().get("items", [])
        check("scoped list returns only own-site rows",
              r.status_code == 200 and len(items) > 0
              and all((i.get("Site_ID") or "").strip() == "CNCEC" for i in items),
              f"status={r.status_code} n={len(items)}")

        r = await ac.get("/receipts", params={"site_id": "HQ"}, headers=H(worker_t))
        check("scoped user asking for another site → 403", r.status_code == 403,
              f"got {r.status_code}")

        ra = (await ac.get("/receipts", params={"limit": 1}, headers=H(admin_t))).json()
        rw = (await ac.get("/receipts", params={"limit": 1}, headers=H(worker_t))).json()
        check("admin sees at least as many rows as a scoped user",
              ra.get("total", 0) >= rw.get("total", 0), f"{ra.get('total')} vs {rw.get('total')}")

        # Cross-site get-one must 404 (not leak existence). Only checkable when a
        # foreign-site row exists in the data.
        foreign = (await ac.get("/receipts", params={"limit": 1, "site_id": "HQ"},
                                headers=H(admin_t))).json().get("items", [])
        if foreign:
            r = await ac.get(f"/receipts/{foreign[0]['id']}", headers=H(worker_t))
            check("scoped get-one of another site's row → 404", r.status_code == 404,
                  f"got {r.status_code}")
        else:
            check("scoped get-one of another site's row → 404 (skipped: no HQ receipts)", True)

        check("scoped user → 403 on /stock/live (cross-site aggregate)",
              (await ac.get("/stock/live", headers=H(worker_t))).status_code == 403)
        check("admin → 200 on /stock/live",
              (await ac.get("/stock/live", headers=H(admin_t))).status_code == 200)

        r = await ac.get("/stock/by-site", params={"limit": 500}, headers=H(worker_t))
        rows = r.json().get("items", [])
        check("stock/by-site forced to the user's own site",
              r.status_code == 200 and all(i.get("Site_ID") == "CNCEC" for i in rows),
              f"status={r.status_code}")

        r = await ac.get("/meta/sites", headers=H(worker_t))
        check("meta/sites returns only the user's site", r.json().get("sites") == ["CNCEC"],
              str(r.json()))
        r = await ac.get("/meta/sites", headers=H(admin_t))
        check("meta/sites unrestricted for admin", len(r.json().get("sites", [])) >= 1)

        r = await ac.get("/meta/inventory-summary", headers=H(worker_t))
        bs = r.json().get("by_site", [])
        check("inventory-summary by_site scoped to one site",
              len(bs) <= 1 and all(x.get("Site_ID") == "CNCEC" for x in bs), str(bs))

        r = await ac.get("/hod/pending", params={"site_id": "HQ"}, headers=H(hod_t))
        check("scoped hod asking for a foreign approvals queue → 403",
              r.status_code == 403, f"got {r.status_code}")
        r = await ac.get("/hod/pending", headers=H(hod_t))
        check("scoped hod pending counts → 200 (own site)", r.status_code == 200)

        r = await ac.get("/reports/stock", params={"format": "csv"}, headers=H(hod_t))
        check("scoped hod report → 200 (forced to own site)", r.status_code == 200,
              f"got {r.status_code}")
        r = await ac.get("/reports/stock", params={"format": "csv", "site_id": "HQ"},
                         headers=H(hod_t))
        check("scoped hod report for a foreign site → 403", r.status_code == 403,
              f"got {r.status_code}")

        # Work-queue badge counts are role- and site-aware.
        j = (await ac.get("/meta/work-queues", headers=H(worker_t))).json()
        check("work-queues: store keeper gets site queues but no approvals",
              "approvals" not in j and "incoming_dns" in j and "sk_requests" in j, str(j))
        j = (await ac.get("/meta/work-queues", headers=H(hod_t))).json()
        check("work-queues: hod gets the approvals count",
              isinstance(j.get("approvals"), int), str(j))
        j = (await ac.get("/meta/work-queues", headers=H(admin_t))).json()
        check("work-queues: admin gets the warehouse workload too",
              isinstance(j.get("warehouse"), int), str(j))

        # R2 lock: entry staging is exact-locked to store_keeper (+admin).
        r = await ac.post("/entry/receipts", headers=H(hod_t), json={})
        check("hod → 403 staging an entry (R2 exact lock)",
              r.status_code == 403, f"got {r.status_code}")
        r = await ac.post("/entry/receipts", headers=H(worker_t), json={})
        check("store keeper passes the entry gate (422 on empty body, not 403)",
              r.status_code == 422, f"got {r.status_code}")

        # Warehouse binding: policy unit-checks on the resolver.
        from .auth import resolve_warehouse_param, warehouse_scope
        wu = {"role": "warehouse_user", "warehouse_id": "WH-01"}
        check("warehouse_user pinned to own warehouse",
              resolve_warehouse_param(wu, None) == "WH-01")
        try:
            resolve_warehouse_param(wu, "WH-02")
            blocked = False
        except Exception:  # noqa: BLE001 — HTTPException(403)
            blocked = True
        check("warehouse_user asking for another warehouse → 403", blocked)
        check("logistics passes warehouse params through",
              resolve_warehouse_param({"role": "logistics", "warehouse_id": ""}, "WH-02") == "WH-02")
        check("unbound warehouse_user fails closed (scope='')",
              warehouse_scope({"role": "warehouse_user", "warehouse_id": ""}) == "")

        # HOD operations pack (non-persisting guard checks; commit machinery is
        # covered by the rolled-back suite-A service tests).
        r = await ac.patch("/hod/pending/returns/999999", headers=H(hod_t),
                           json={"fields": {"Quantity": 5}})
        check("edit of a non-existent staged row → 404", r.status_code == 404,
              f"got {r.status_code}")
        r = await ac.patch("/hod/pending/returns/1", headers=H(hod_t),
                           json={"fields": {"status": "approved"}})
        check("editing a non-whitelisted field → 422", r.status_code == 422,
              f"got {r.status_code}")
        r = await ac.get("/hod/preflight", headers=H(hod_t))
        check("negative-stock pre-flight → 200 + items list",
              r.status_code == 200 and isinstance(r.json().get("items"), list),
              f"got {r.status_code}")
        r = await ac.post("/hod/pending/issues/bulk-approve", headers=H(hod_t),
                          json={"ids": []})
        check("bulk-approve with no ids → 422", r.status_code == 422, f"got {r.status_code}")
        r = await ac.get("/hod/low-stock", headers=H(hod_t))
        check("low-stock view → 200 for a scoped hod", r.status_code == 200,
              f"got {r.status_code}")
        r = await ac.post("/hod/prs/auto-draft", headers=H(hod_t), json={"site_id": "HQ"})
        check("scoped hod auto-drafting a foreign-site PR → 403", r.status_code == 403,
              f"got {r.status_code}")
        r = await ac.get("/hod/prs/PR-NOPE-0000/pdf", headers=H(hod_t))
        check("PDF of a non-existent PR → 404", r.status_code == 404, f"got {r.status_code}")
        r = await ac.get("/hod/preflight", headers=H(worker_t))
        check("worker (lvl 0) → 403 on the HOD ops pack", r.status_code == 403,
              f"got {r.status_code}")

        # Warehouse completion pack (non-persisting guard checks).
        r = await ac.get("/warehouse/returns", headers=H(admin_t))
        check("returns-from-site queue → 200 for admin",
              r.status_code == 200 and isinstance(r.json().get("items"), list),
              f"got {r.status_code}")
        r = await ac.get("/warehouse/returns", headers=H(worker_t))
        check("worker → 403 on the warehouse returns queue", r.status_code == 403,
              f"got {r.status_code}")
        r = await ac.post("/warehouse/returns", headers=H(admin_t), json={})
        check("recording a return with an empty body → 422", r.status_code == 422,
              f"got {r.status_code}")
        r = await ac.post("/warehouse/returns/999999/disposition", headers=H(admin_t),
                          json={"status": "hold"})
        check("disposition of a non-existent return → 404", r.status_code == 404,
              f"got {r.status_code}")
        r = await ac.post("/warehouse/returns/1/disposition", headers=H(admin_t),
                          json={"status": "yeet"})
        check("invalid disposition value → 422", r.status_code == 422, f"got {r.status_code}")
        r = await ac.get("/warehouse/history", headers=H(admin_t))
        j = r.json() if r.status_code == 200 else {}
        check("warehouse history → 200 with dns/assignments/throughput",
              r.status_code == 200 and {"dns", "assignments", "throughput"} <= set(j),
              f"got {r.status_code}")

        # Store-keeper toolbox (non-persisting guard checks).
        r = await ac.get("/entry/count-sheet", headers=H(worker_t))
        check("count sheet → 200 for a store keeper (own site)",
              r.status_code == 200 and isinstance(r.json().get("items"), list),
              f"got {r.status_code}")
        r = await ac.post("/entry/count-sheet", headers=H(worker_t),
                          json={"site_id": "CNCEC", "rows": []})
        check("count submit with no rows → 422", r.status_code == 422, f"got {r.status_code}")
        r = await ac.post("/entry/count-sheet", headers=H(worker_t),
                          json={"site_id": "CNCEC", "reason_code": "yeet",
                                "rows": [{"SAP_Code": "1001", "counted_qty": 1}]})
        check("count submit with a bad reason → 422", r.status_code == 422,
              f"got {r.status_code}")
        r = await ac.get("/entry/bins/1001", headers=H(worker_t))
        check("bin locations → 200 + bins list",
              r.status_code == 200 and isinstance(r.json().get("bins"), list),
              f"got {r.status_code}")
        r = await ac.get("/entry/returnables", headers=H(worker_t))
        check("returnables list → 200 for a store keeper", r.status_code == 200,
              f"got {r.status_code}")
        r = await ac.get("/entry/returnables", headers=H(hod_t))
        check("hod → 403 on returnables (SK exact lock)", r.status_code == 403,
              f"got {r.status_code}")
        r = await ac.post("/entry/returnables", headers=H(worker_t),
                          json={"material_name": "svc wrench", "borrower_name": "svc",
                                "expected_return_time": "not-a-date"})
        check("loan with a bad datetime → 422", r.status_code == 422, f"got {r.status_code}")
        r = await ac.post("/entry/returnables/999999/return", headers=H(worker_t))
        check("returning a non-existent loan → 404", r.status_code == 404,
              f"got {r.status_code}")
        j = (await ac.get("/meta/work-queues", headers=H(worker_t))).json()
        check("work-queues: store keeper gets the returnables_overdue count",
              isinstance(j.get("returnables_overdue"), int), str(j))

        # Phase-5 reports: every new key renders (csv), scoped gates hold.
        new_keys = ("daily-consumption", "monthly-summary", "wbs", "low-stock",
                    "burn-rate", "valuation", "fefo", "audit",
                    "warehouse-throughput", "force-closures", "intent-vs-actual")
        bad = []
        for k in new_keys:
            rr = await ac.get(f"/reports/{k}", params={"format": "csv"}, headers=H(admin_t))
            if rr.status_code != 200:
                bad.append(f"{k}={rr.status_code}")
        check("all 11 Phase-5 reports render as CSV for admin", not bad, ", ".join(bad))
        r = await ac.get("/reports/audit", params={"format": "csv"}, headers=H(hod_t))
        check("scoped hod → 403 on the global-only audit report",
              r.status_code == 403, f"got {r.status_code}")

        # Archive lifecycle (created → listed → downloaded → deleted).
        r = await ac.post("/reports/archive", headers=H(admin_t),
                          json={"key": "stock", "format": "csv"})
        aid = r.json().get("id") if r.status_code == 201 else None
        check("archive a report → 201 + id", bool(aid), f"got {r.status_code}")
        r = await ac.get("/reports/archive", headers=H(admin_t))
        check("archive list contains the new entry",
              any(x["id"] == aid for x in r.json().get("items", [])))
        r = await ac.get(f"/reports/archive/{aid}/download", headers=H(admin_t))
        check("archived file re-downloads", r.status_code == 200, f"got {r.status_code}")
        r = await ac.request("DELETE", f"/reports/archive/{aid}", headers=H(admin_t))
        check("archive delete → 200 (cleanup)", r.status_code == 200, f"got {r.status_code}")

        # Scheduler: validation + run-now + the daemon's atomic claim.
        r = await ac.post("/reports/schedules", headers=H(admin_t),
                          json={"label": "svc bad", "report_type": "stock",
                                "frequency": "whenever"})
        check("bad schedule frequency → 422", r.status_code == 422, f"got {r.status_code}")
        r = await ac.post("/reports/schedules", headers=H(admin_t),
                          json={"label": "svc daily", "report_type": "stock",
                                "frequency": "daily 00:00", "format": "csv"})
        sid = r.json().get("id")
        check("create schedule → 201 + id", bool(sid), f"got {r.status_code}")
        r = await ac.post(f"/reports/schedules/{sid}/run", headers=H(admin_t))
        ran_aid = (r.json().get("archive") or {}).get("id")
        check("run-now → archives + returns the archive id",
              r.status_code == 200 and bool(ran_aid), f"got {r.status_code}")

        from .report_center import run_due_schedules
        from .db import SessionLocal as _SL
        from sqlalchemy import update as _upd
        from .services.ledger import _MD as _md
        async with _SL() as s2:
            await s2.execute(_upd(_md.tables["report_schedules"])
                             .where(_md.tables["report_schedules"].c["id"] == sid)
                             .values(last_run=None))
            await s2.commit()
        n1 = await run_due_schedules()
        check("daemon tick runs the due schedule", n1 >= 1, f"ran {n1}")
        n2 = await run_due_schedules()
        check("second tick does NOT rerun (atomic claim holds)", n2 == 0, f"ran {n2}")

        # Cleanup: schedule + every archive row this test created.
        r = await ac.request("DELETE", f"/reports/schedules/{sid}", headers=H(admin_t))
        check("schedule delete → 200 (cleanup)", r.status_code == 200, f"got {r.status_code}")
        arch = (await ac.get("/reports/archive", headers=H(admin_t))).json().get("items", [])
        for x in arch:
            if x["id"] == ran_aid or str(x.get("generated_by", "")).startswith("scheduler:"):
                await ac.request("DELETE", f"/reports/archive/{x['id']}", headers=H(admin_t))

        # Phase-6 documents: label/badge PDFs, reference docs, master exports.
        r = await ac.get("/documents/qr-labels", headers=H(admin_t))
        check("QR bin labels → 200 + PDF",
              r.status_code == 200 and "application/pdf" in r.headers.get("content-type", ""),
              f"got {r.status_code}/{r.headers.get('content-type')}")
        r = await ac.get("/documents/employee-badges", headers=H(admin_t))
        check("employee badges → 200 + PDF",
              r.status_code == 200 and "application/pdf" in r.headers.get("content-type", "")
              and len(r.content) > 800, f"got {r.status_code} len={len(r.content)}")
        r = await ac.get("/documents/qr-labels", headers=H(worker_t))
        check("store keeper (lvl 0) → 403 on QR labels", r.status_code == 403, f"got {r.status_code}")
        r = await ac.get("/documents/reference/manual", headers=H(worker_t))
        check("any authed user can download the manual → 200 PDF",
              r.status_code == 200 and "application/pdf" in r.headers.get("content-type", ""),
              f"got {r.status_code}")
        r = await ac.get("/documents/reference/nope", headers=H(worker_t))
        check("unknown reference doc → 404", r.status_code == 404, f"got {r.status_code}")
        r = await ac.get("/documents/master/vendors", params={"format": "xlsx"}, headers=H(admin_t))
        check("vendor master export → 200 + spreadsheet",
              r.status_code == 200 and "spreadsheet" in r.headers.get("content-type", ""),
              f"got {r.status_code}")
        r = await ac.get("/documents/master/nope", params={"format": "xlsx"}, headers=H(admin_t))
        check("unknown master entity → 404", r.status_code == 404, f"got {r.status_code}")
        r = await ac.get("/documents/master/vendors", params={"format": "docx"}, headers=H(admin_t))
        check("bad export format → 400", r.status_code == 400, f"got {r.status_code}")
        r = await ac.get("/documents/master/employees", params={"format": "csv"}, headers=H(hod_t))
        check("scoped hod employee export → 200 (forced to own site)",
              r.status_code == 200, f"got {r.status_code}")

        # Phase-8 SME read-parity (pure reads; SME Canon — no write endpoints exist).
        r = await ac.get("/sme/equipment-report", headers=H(hod_t))
        check("SME equipment report → 200 + items",
              r.status_code == 200 and isinstance(r.json().get("items"), list),
              f"got {r.status_code}")
        r = await ac.get("/sme/consumption-comparison", headers=H(hod_t))
        check("SME consumption comparison → 200 + items",
              r.status_code == 200 and isinstance(r.json().get("items"), list),
              f"got {r.status_code}")
        r = await ac.get("/sme/demand-matrix", headers=H(hod_t))
        dm = r.json() if r.status_code == 200 else {}
        check("SME demand matrix → 200 with lines + totals",
              r.status_code == 200 and {"lines", "totals"} <= set(dm),
              f"got {r.status_code}")
        check("demand lines hold allocated + shortfall == demand",
              all(abs(l["Allocated_Qty"] + l["Shortfall_Qty"] - l["Demand_Qty"]) < 1e-6
                  for l in dm.get("lines", [])))
        line_sum = sum(l["Demand_Qty"] for l in dm.get("lines", []))
        tot_sum = sum(t["Demand_Qty"] for t in dm.get("totals", []))
        check("demand totals reconcile with the lines",
              abs(line_sum - tot_sum) < 1e-3, f"{line_sum} vs {tot_sum}")
        r = await ac.get("/sme/export/demand-totals", params={"format": "xlsx"}, headers=H(hod_t))
        check("SME export → 200 + spreadsheet",
              r.status_code == 200 and "spreadsheet" in r.headers.get("content-type", ""),
              f"got {r.status_code}")
        r = await ac.get("/sme/export/nope", headers=H(hod_t))
        check("unknown SME export → 404", r.status_code == 404, f"got {r.status_code}")
        r = await ac.get("/sme/demand-matrix", headers=H(worker_t))
        check("worker (lvl 0) → 403 on SME views", r.status_code == 403, f"got {r.status_code}")

        # ---- Phase-9 admin console -------------------------------------------
        import os as _os

        # Global sites CRUD lifecycle.
        r = await ac.post("/admin/sites", headers=H(admin_t), json={"name": "SVC-SITE"})
        new_site = r.json().get("id")
        check("add site → 201", r.status_code == 201 and bool(new_site), f"got {r.status_code}")
        r = await ac.post("/admin/sites", headers=H(admin_t), json={"name": "SVC-SITE"})
        check("duplicate site → 409", r.status_code == 409, f"got {r.status_code}")
        r = await ac.get("/admin/sites", headers=H(admin_t))
        check("sites list contains the new site",
              any(s["name"] == "SVC-SITE" for s in r.json().get("items", [])))
        r = await ac.request("DELETE", f"/admin/sites/{new_site}", headers=H(admin_t))
        check("delete site → 200 (cleanup)", r.status_code == 200, f"got {r.status_code}")
        sites_all = (await ac.get("/admin/sites", headers=H(admin_t))).json()["items"]
        cncec = next((s for s in sites_all if s["name"] == "CNCEC"), None)
        if cncec:
            r = await ac.request("DELETE", f"/admin/sites/{cncec['id']}", headers=H(admin_t))
            check("deleting a site with bound users → 409", r.status_code == 409,
                  f"got {r.status_code}")
        else:
            check("deleting a site with bound users → 409 (skipped: no CNCEC row)", True)

        # Settings + the maintenance-mode login gate (isolated rate-limit bucket;
        # ALWAYS restored in the finally so later suites can log in).
        r = await ac.put("/admin/settings", headers=H(admin_t),
                         json={"key": "nope", "value": "1"})
        check("non-whitelisted setting key → 422", r.status_code == 422, f"got {r.status_code}")
        mip = {"X-Real-IP": "203.0.113.9"}
        r = await ac.put("/admin/settings", headers=H(admin_t),
                         json={"key": "maintenance_mode", "value": "1"})
        check("maintenance mode ON → 200", r.status_code == 200, f"got {r.status_code}")
        try:
            r = await ac.post("/auth/login", headers=mip,
                              json={"username": "supervisor", "password": "super2026"})
            check("non-admin login during maintenance → 503", r.status_code == 503,
                  f"got {r.status_code}")
            r = await ac.post("/auth/login", headers=mip,
                              json={"username": "admin", "password": "admin2026"})
            check("admin login during maintenance → 200", r.status_code == 200,
                  f"got {r.status_code}")
        finally:
            rr = await ac.put("/admin/settings", headers=H(admin_t),
                              json={"key": "maintenance_mode", "value": "0"})
            check("maintenance mode OFF restored", rr.status_code == 200,
                  f"got {rr.status_code}")

        # Manual backup trigger (200 where pg_dump exists, else a clear 501).
        r = await ac.post("/admin/backup", headers=H(admin_t))
        if r.status_code == 200:
            p = r.json().get("file", "")
            ok = _os.path.exists(p) and r.json().get("size_bytes", 0) > 0
            if ok:
                _os.remove(p)
            check("manual backup → dump file written (cleaned up)", ok, p)
        else:
            check("manual backup → 501 when pg_dump unavailable",
                  r.status_code == 501, f"got {r.status_code}")

        # Sessions viewer + admin revocation ends a live session.
        r = await ac.get("/admin/sessions", headers=H(admin_t), params={"username": "worker"})
        check("sessions list → 200, no token material",
              r.status_code == 200 and all("refresh_hash" not in s
                                           for s in r.json().get("items", [])),
              f"got {r.status_code}")
        async with AsyncClient(transport=transport, base_url="http://svc") as ac2:
            lr = await ac2.post("/auth/login", headers=mip,
                                json={"username": "worker", "password": "floor2026"})
            check("victim login for revocation test → 200", lr.status_code == 200)
            r = await ac.post("/admin/sessions/revoke-user/worker", headers=H(admin_t))
            check("revoke-user → 200 + revoked ≥ 1",
                  r.status_code == 200 and r.json().get("revoked", 0) >= 1, str(r.json()))
            r = await ac2.post("/auth/refresh")
            check("revoked session's refresh → 401", r.status_code == 401,
                  f"got {r.status_code}")

        # Oversight KPIs: admin 200 with the expected blocks; hod 403.
        r = await ac.get("/admin/oversight", headers=H(admin_t))
        j = r.json() if r.status_code == 200 else {}
        check("logistics oversight → 200 with KPI blocks",
              r.status_code == 200 and {"prs_by_state", "pos_by_status", "dns_by_status",
                                        "warehouse_load"} <= set(j), f"got {r.status_code}")
        r = await ac.get("/admin/oversight", headers=H(hod_t))
        check("hod (lvl 2) → 403 on oversight", r.status_code == 403, f"got {r.status_code}")

        # Cross-site requests: hod raises → admin decides → cleanup.
        r = await ac.post("/xsite", headers=H(worker_t),
                          json={"target_site": "HQ", "SAP_Code": "1001", "requested_qty": 1})
        check("worker (lvl 0) → 403 raising a cross-site request",
              r.status_code == 403, f"got {r.status_code}")
        r = await ac.post("/xsite", headers=H(hod_t),
                          json={"target_site": "HQ", "SAP_Code": "1001", "requested_qty": 2})
        xid = r.json().get("id")
        check("hod raises a cross-site request → 201 + availability snapshot",
              r.status_code == 201 and "available_at_target" in r.json(),
              f"got {r.status_code}")
        r = await ac.get("/xsite", headers=H(hod_t), params={"mine": "true"})
        check("hod 'my requests' lists it",
              any(x["id"] == xid for x in r.json().get("items", [])))
        r = await ac.post(f"/xsite/{xid}/decide", headers=H(hod_t), json={"action": "approve"})
        check("hod cannot decide (admin only) → 403", r.status_code == 403,
              f"got {r.status_code}")
        r = await ac.post(f"/xsite/{xid}/decide", headers=H(admin_t),
                          json={"action": "approve", "suggested_qty": 1.5})
        check("admin decides → approved", r.status_code == 200
              and r.json().get("status") == "approved", f"got {r.status_code}")
        r = await ac.post(f"/xsite/{xid}/decide", headers=H(admin_t), json={"action": "reject"})
        check("double-decide → 409", r.status_code == 409, f"got {r.status_code}")
        r = await ac.request("DELETE", f"/xsite/{xid}", headers=H(admin_t))
        check("admin deletes the test request (cleanup)", r.status_code == 200,
              f"got {r.status_code}")

        # Feedback: worker submits → admin responds → cleanup.
        r = await ac.post("/feedback", headers=H(worker_t),
                          json={"type": "nope", "description": "x"})
        check("bad feedback type → 422", r.status_code == 422, f"got {r.status_code}")
        r = await ac.post("/feedback", headers=H(worker_t),
                          json={"type": "bug", "description": "svc test report", "page": "/stock"})
        fid = r.json().get("id")
        check("submit feedback → 201", r.status_code == 201 and bool(fid), f"got {r.status_code}")
        r = await ac.get("/feedback/mine", headers=H(worker_t))
        check("'my feedback' lists it", any(x["id"] == fid for x in r.json().get("items", [])))
        r = await ac.patch(f"/admin/feedback/{fid}", headers=H(admin_t),
                           json={"status": "resolved", "admin_response": "done"})
        check("admin resolves feedback → 200", r.status_code == 200, f"got {r.status_code}")
        r = await ac.request("DELETE", f"/admin/feedback/{fid}", headers=H(admin_t))
        check("admin deletes the test report (cleanup)", r.status_code == 200,
              f"got {r.status_code}")


async def test_manhours():
    """Phase-10 Man-Hours portal: exact {hod, admin} lock, roster upserts, the
    ported hour math (8h normal + OT, overnight wrap), SQM distribution,
    estimate-vs-actual variance, attendance-xlsx import, exports. Uses a unique
    future work-date + SVC- codes, and cleans every mh_* row up afterwards."""
    transport = ASGITransport(app=app)
    ip = {"X-Real-IP": "203.0.113.10"}  # own rate-limit bucket
    D = "2031-01-15"                    # far-future date: never collides with real data
    async with AsyncClient(transport=transport, base_url="http://svc") as ac:
        async def token(u, p):
            r = await ac.post("/auth/login", json={"username": u, "password": p}, headers=ip)
            return r.json().get("access_token")

        def H(t):
            return {"Authorization": f"Bearer {t}"}

        admin_t = await token("admin", "admin2026")
        hod_t = await token("hod", "hod2026")        # hod @ CNCEC
        worker_t = await token("worker", "floor2026")

        try:
            # Exact role lock: {hod, admin} only.
            r = await ac.get("/mh/employees", headers=H(worker_t))
            check("worker (lvl 0) → 403 on the MH portal", r.status_code == 403,
                  f"got {r.status_code}")
            r = await ac.get("/mh/meta", headers=H(hod_t))
            check("hod → 200 on /mh/meta with SME dropdowns",
                  r.status_code == 200 and len(r.json().get("equipment_tags", [])) > 0
                  and len(r.json().get("system_codes", [])) > 0, f"got {r.status_code}")

            # Roster: upsert + re-upsert updates in place (no duplicate).
            r = await ac.post("/mh/employees", headers=H(hod_t), json={
                "employee_code": "SVC-EMP-1", "name": "Svc One", "worker_type": "nope"})
            check("bad worker_type → 422", r.status_code == 422, f"got {r.status_code}")
            for code, name in (("SVC-EMP-1", "Svc One"), ("SVC-EMP-2", "Svc Two"),
                               ("SVC-EMP-3", "Svc Three")):
                await ac.post("/mh/employees", headers=H(hod_t), json={
                    "employee_code": code, "name": name, "worker_type": "OWN"})
            r = await ac.post("/mh/employees", headers=H(hod_t), json={
                "employee_code": "SVC-EMP-1", "name": "Svc One Renamed",
                "worker_type": "Supply", "company": "ACME"})
            check("roster upsert → 200", r.status_code == 200, f"got {r.status_code}")
            emps = (await ac.get("/mh/employees", headers=H(hod_t))).json()["items"]
            mine = [e for e in emps if e["Employee_Code"] == "SVC-EMP-1"]
            check("re-upsert updates in place (1 row, new name/type)",
                  len(mine) == 1 and mine[0]["Name"] == "Svc One Renamed"
                  and mine[0]["Worker_Type"] == "Supply", str(mine))
            r = await ac.patch(f"/mh/employees/{mine[0]['id']}/status",
                               headers=H(hod_t), params={"status": "inactive"})
            check("status flip → inactive", r.status_code == 200, f"got {r.status_code}")
            r = await ac.patch(f"/mh/employees/{mine[0]['id']}/status",
                               headers=H(hod_t), params={"status": "active"})
            check("status flip back → active", r.status_code == 200, f"got {r.status_code}")

            # Site scoping: the hod (CNCEC) may not read/write another site.
            r = await ac.get("/mh/employees", headers=H(hod_t), params={"site_id": "HQ"})
            check("hod requesting another site → 403", r.status_code == 403,
                  f"got {r.status_code}")
            r = await ac.post("/mh/employees", headers=H(admin_t), json={
                "employee_code": "SVC-X", "name": "x"})
            check("admin write without site_id → 422", r.status_code == 422,
                  f"got {r.status_code}")

            # Timesheet batch: ported hour math (8h normal + unpaid break + OT,
            # overnight wraps +24h). 07:30–16:30→8.0 · 07:00–18:30→10.5 (2.5 OT)
            # · 22:00–06:00→7.0.
            r = await ac.post("/mh/timesheets", headers=H(hod_t), json={
                "work_date": D, "equipment_tag": "SVC-TAG", "system_code": "99",
                "location": "SVC-LOC", "break_mins": 60, "rows": [
                    {"employee_code": "SVC-EMP-1", "in_time": "07:30", "out_time": "16:30"},
                    {"employee_code": "SVC-EMP-2", "in_time": "07:00", "out_time": "18:30"},
                    {"employee_code": "SVC-EMP-3", "in_time": "22:00", "out_time": "06:00"},
                ]})
            check("timesheet batch → 3 saved", r.status_code == 200
                  and r.json().get("saved") == 3, f"got {r.status_code} {r.text[:120]}")
            ts = (await ac.get("/mh/timesheets", headers=H(hod_t),
                               params={"work_date": D})).json()["items"]
            hours = {t["Employee_Code"]: (t["Total_Hours"], t["Normal_Hours"], t["OT_Hours"])
                     for t in ts}
            check("hour math: 07:30–16:30 − 60min → 8.0 / 8.0 / 0",
                  hours.get("SVC-EMP-1") == (8.0, 8.0, 0.0), str(hours.get("SVC-EMP-1")))
            check("hour math: 07:00–18:30 → 10.5 total with 2.5 OT",
                  hours.get("SVC-EMP-2") == (10.5, 8.0, 2.5), str(hours.get("SVC-EMP-2")))
            check("hour math: overnight 22:00–06:00 wraps → 7.0",
                  hours.get("SVC-EMP-3") == (7.0, 7.0, 0.0), str(hours.get("SVC-EMP-3")))
            # Re-posting the same day/tag/system upserts (no duplicate rows).
            await ac.post("/mh/timesheets", headers=H(hod_t), json={
                "work_date": D, "equipment_tag": "SVC-TAG", "system_code": "99",
                "break_mins": 60, "rows": [
                    {"employee_code": "SVC-EMP-1", "in_time": "07:30", "out_time": "16:30"}]})
            ts2 = (await ac.get("/mh/timesheets", headers=H(hod_t),
                                params={"work_date": D})).json()["items"]
            check("batch re-post upserts in place (still 3 rows)", len(ts2) == 3,
                  f"got {len(ts2)}")

            # Team SQM distribution: even, then pro-rata by hours.
            r = await ac.post("/mh/production", headers=H(hod_t), json={
                "work_date": D, "equipment_tag": "SVC-TAG", "system_code": "99",
                "sqm_done": 30, "distribution_method": "even"})
            check("production even-distribute hits 3 rows", r.status_code == 200
                  and r.json().get("distributed_rows") == 3, r.text[:120])
            ts3 = (await ac.get("/mh/timesheets", headers=H(hod_t),
                                params={"work_date": D})).json()["items"]
            check("even split → 10 SQM each",
                  all(abs(float(t["Allocated_SQM"]) - 10.0) < 1e-6 for t in ts3),
                  str([t["Allocated_SQM"] for t in ts3]))
            await ac.post("/mh/production", headers=H(hod_t), json={
                "work_date": D, "equipment_tag": "SVC-TAG", "system_code": "99",
                "sqm_done": 30, "distribution_method": "by_hours"})
            ts4 = (await ac.get("/mh/timesheets", headers=H(hod_t),
                                params={"work_date": D})).json()["items"]
            sqm_by = {t["Employee_Code"]: float(t["Allocated_SQM"]) for t in ts4}
            # total hours 25.5 → EMP-2's pro-rata share = 30 × 10.5 / 25.5
            check("by-hours split is pro-rata on Total_Hours",
                  abs(sqm_by.get("SVC-EMP-2", 0) - round(30 * 10.5 / 25.5, 3)) < 1e-6,
                  str(sqm_by))

            # Estimator + variance (the inlined v_mh_estimate_vs_actual port):
            # estimate 20 vs actual 25.5 → +5.5 / +27.5%.
            r = await ac.post("/mh/estimates", headers=H(hod_t), json={
                "equipment_tag": "SVC-TAG", "system_code": "99",
                "estimated_manhours": 20, "estimated_sqm": 60, "basis": "svc test"})
            check("estimate upsert → 200", r.status_code == 200, f"got {r.status_code}")
            v = (await ac.get("/mh/variance", headers=H(hod_t))).json()
            row = next((x for x in v["items"] if x["Equipment_Tag"] == "SVC-TAG"), None)
            check("variance row: actual 25.5 vs estimated 20 → +5.5",
                  row is not None and abs(float(row["Actual_Manhours"]) - 25.5) < 1e-6
                  and abs(float(row["Variance_Manhours"]) - 5.5) < 1e-6, str(row))
            check("variance pct 27.5 + SQM rollup 30",
                  row is not None and abs(float(row["Variance_Pct"]) - 27.5) < 1e-6
                  and abs(float(row["SQM_Done"]) - 30.0) < 1e-6, str(row))
            check("variance KPIs count the over-consumer",
                  v["kpis"]["scopes"] >= 1 and v["kpis"]["over_consuming"] >= 1,
                  str(v["kpis"]))
            r = await ac.post("/mh/variance/reason", headers=H(hod_t), json={
                "equipment_tag": "SVC-TAG", "system_code": "99",
                "reason": "svc: rework after hydrotest"})
            check("variance reason saved", r.status_code == 200, f"got {r.status_code}")
            v2 = (await ac.get("/mh/variance", headers=H(hod_t))).json()["items"]
            row2 = next((x for x in v2 if x["Equipment_Tag"] == "SVC-TAG"), None)
            check("reason lands on the variance row",
                  row2 is not None and row2["Variance_Reason"] == "svc: rework after hydrotest",
                  str(row2 and row2["Variance_Reason"]))

            # Employee-wise timeline (roster-name join + windowing).
            tl = (await ac.get("/mh/employee-timeline", headers=H(hod_t), params={
                "employee_code": "SVC-EMP-2", "date_from": D, "date_to": D})).json()
            check("employee timeline: 1 row, joined name, 10.5h total",
                  len(tl["items"]) == 1 and tl["items"][0]["Name"] == "Svc Two"
                  and abs(tl["total_hours"] - 10.5) < 1e-6, str(tl)[:160])

            # Attendance-xlsx import: dry-run preview, replace import, idempotent
            # re-import (replace deletes the file's dates first).
            import io as _io

            from openpyxl import Workbook
            wb = Workbook()
            ws = wb.active
            ws.title = "ADD EMPLOYEE"
            ws.append(["Code", "Name", "Designation", "Type", "Company"])
            ws.append(["SVC-IMP-1", "Svc Import One", "Blaster", "Supply", "ACME"])
            sar = wb.create_sheet("SAR")
            sar.append(["Location", "Equipment Tag #", "Code", "Name", "Work Date",
                        "In Time", "Out Time", "Status", "Remarks"])
            sar.append(["YARD", "SVC-TAG", "SVC-IMP-1", "Svc Import One", "2031-02-01",
                        "07:30", "16:30", "PR", ""])
            sar.append(["YARD", "SVC-TAG", "SVC-IMP-2", "Svc Import Two", "2031-02-01",
                        "07:00", "18:30", "PR", ""])
            buf = _io.BytesIO()
            wb.save(buf)
            xlsx = ("att.xlsx", buf.getvalue(),
                    "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
            r = await ac.post("/mh/import", headers=H(hod_t), files={"file": xlsx},
                              params={"dry_run": "true"})
            check("import dry-run parses 2 employees / 2 rows / 1 date",
                  r.status_code == 200 and r.json().get("employees") == 2
                  and r.json().get("timesheets") == 2 and len(r.json().get("dates", [])) == 1,
                  r.text[:160])
            r = await ac.post("/mh/import", headers=H(hod_t), files={"file": xlsx},
                              params={"replace": "true"})
            check("import (replace) → 2 employees + 2 timesheets", r.status_code == 200
                  and r.json().get("timesheets") == 2, r.text[:160])
            imp = (await ac.get("/mh/timesheets", headers=H(hod_t),
                                params={"work_date": "2031-02-01"})).json()["items"]
            check("imported rows carry recomputed hours (SAR worker merged to roster)",
                  len(imp) == 2 and {float(t["Total_Hours"]) for t in imp} == {8.0, 10.5},
                  str([(t['Employee_Code'], t['Total_Hours']) for t in imp]))
            r = await ac.post("/mh/import", headers=H(hod_t), files={"file": xlsx},
                              params={"replace": "true"})
            imp2 = (await ac.get("/mh/timesheets", headers=H(hod_t),
                                 params={"work_date": "2031-02-01"})).json()["items"]
            check("replace re-import is idempotent (still 2 rows)", len(imp2) == 2,
                  f"got {len(imp2)}")
            emp_row = next((e for e in (await ac.get("/mh/employees", headers=H(hod_t)))
                            .json()["items"] if e["Employee_Code"] == "SVC-IMP-1"), None)
            check("ADD EMPLOYEE attributes land on the roster",
                  emp_row is not None and emp_row["Worker_Type"] == "Supply"
                  and emp_row["Company"] == "ACME", str(emp_row))
            r = await ac.post("/mh/import", headers=H(hod_t),
                              files={"file": ("junk.xlsx", b"not an xlsx", "application/octet-stream")})
            check("unparseable workbook → 422", r.status_code == 422, f"got {r.status_code}")

            # ---- Phase-11A: import fit + bulk-assign -----------------------
            # Legend defaults: a SAR-only worker gets OWN→GI.
            emp2 = next((e for e in (await ac.get("/mh/employees", headers=H(hod_t)))
                         .json()["items"] if e["Employee_Code"] == "SVC-IMP-2"), None)
            check("legend default: SAR-only worker → OWN/GI",
                  emp2 is not None and emp2["Worker_Type"] == "OWN"
                  and emp2["Company"] == "GI", str(emp2))

            # Workbook 2: literal 'nan' junk cells, a duplicated (code,date)
            # row, and a Supply employee with a blank Company cell.
            wb2 = Workbook()
            ws2 = wb2.active
            ws2.title = "ADD EMPLOYEE"
            ws2.append(["Code", "Name", "Designation", "Type", "Company"])
            ws2.append(["SVC-IMP-3", "Svc Import Three", "nan", "Supply", None])
            sar2 = wb2.create_sheet("SAR")
            sar2.append(["Location", "Equipment Tag #", "Code", "Name", "Work Date",
                         "In Time", "Out Time", "Status", "Remarks"])
            sar2.append(["nan", "nan", "SVC-IMP-1", "Svc Import One", "2031-02-02",
                         "07:30", "16:30", "PR", "nan"])
            sar2.append(["nan", "nan", "SVC-IMP-1", "Svc Import One", "2031-02-02",
                         "08:00", "17:00", "PR", ""])   # dup (code,date,tag) — last wins
            sar2.append([None, None, "SVC-IMP-3", "Svc Import Three", "2031-02-02",
                         "07:30", "16:30", "PR", ""])
            buf2 = _io.BytesIO()
            wb2.save(buf2)
            xlsx2 = ("att2.xlsx", buf2.getvalue(),
                     "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
            r = await ac.post("/mh/import", headers=H(hod_t), files={"file": xlsx2},
                              params={"replace": "true"})
            check("in-file dedupe: 3 SAR rows → 2 imported (last dup wins)",
                  r.status_code == 200 and r.json().get("timesheets") == 2, r.text[:160])
            d2 = (await ac.get("/mh/timesheets", headers=H(hod_t),
                               params={"work_date": "2031-02-02"})).json()["items"]
            one = next((x for x in d2 if x["Employee_Code"] == "SVC-IMP-1"), None)
            check("'nan' cells → NULL (import guard) + dup row's later times won",
                  one is not None and one["Equipment_Tag"] is None
                  and one["Location"] is None and one["In_Time"] == "08:00",
                  str(one))
            emp3 = next((e for e in (await ac.get("/mh/employees", headers=H(hod_t)))
                         .json()["items"] if e["Employee_Code"] == "SVC-IMP-3"), None)
            check("legend default: Supply + blank company → DMC ('nan' designation cleaned)",
                  emp3 is not None and emp3["Company"] == "DMC"
                  and emp3["Designation"] == "", str(emp3))

            # Unassigned filter: only wb2's NULL-tag rows qualify (the Phase-10
            # workbook filled Equipment Tag # = SVC-TAG, so its rows are assigned).
            r = await ac.get("/mh/timesheets", headers=H(hod_t), params={
                "unassigned": "true", "date_from": "2031-02-01", "date_to": "2031-02-02"})
            un = r.json()
            check("unassigned filter: wb2's 2 NULL-tag rows with a total_hours rollup",
                  len(un["items"]) == 2 and un["total_hours"] == 16.0,
                  f"{len(un.get('items', []))} rows, {un.get('total_hours')}")

            # Bulk-assign: unassigned rows → a real SME scope; Location auto-fills.
            meta_j = (await ac.get("/mh/meta", headers=H(hod_t))).json()
            atag = meta_j["equipment_tags"][0]
            aloc = meta_j["tag_locations"].get(atag)
            ids1 = [x["id"] for x in un["items"]]
            r = await ac.patch("/mh/timesheets/assign", headers=H(hod_t), json={
                "ids": ids1, "equipment_tag": atag, "system_code": "99"})
            check("bulk-assign → all rows assigned, SME location auto-filled",
                  r.status_code == 200 and r.json().get("assigned") == len(ids1)
                  and r.json().get("location") == aloc, r.text[:200])
            left = (await ac.get("/mh/timesheets", headers=H(hod_t), params={
                "unassigned": "true", "work_date": "2031-02-02"})).json()["items"]
            check("assigned rows leave the unassigned queue", len(left) == 0,
                  f"got {len(left)}")

            # Append-overlap warning: dry-run flags dates that already have rows.
            r = await ac.post("/mh/import", headers=H(hod_t), files={"file": xlsx2},
                              params={"dry_run": "true"})
            check("dry-run reports overlap_dates for append preview",
                  r.json().get("overlap_dates") == ["2031-02-02"], r.text[:160])
            r = await ac.post("/mh/import", headers=H(hod_t), files={"file": xlsx2},
                              params={"replace": "false"})
            check("append into an existing date → overlap_dates in the response",
                  r.status_code == 200 and r.json().get("overlap_dates") == ["2031-02-02"],
                  r.text[:160])
            d2b = (await ac.get("/mh/timesheets", headers=H(hod_t),
                                params={"work_date": "2031-02-02"})).json()["items"]
            check("append duplicates NULL-tag rows (the documented reason for the warning)",
                  len(d2b) == 4, f"got {len(d2b)}")

            # Conflict skip: the appended twins target a scope where each
            # worker/date already has an assigned row — all skipped, none merged.
            ids2 = [x["id"] for x in d2b if x["Equipment_Tag"] is None]
            r = await ac.patch("/mh/timesheets/assign", headers=H(hod_t), json={
                "ids": ids2, "equipment_tag": atag, "system_code": "99"})
            check("assign skips unique-key twins and reports them (0 assigned + 2 conflicts)",
                  r.status_code == 200 and r.json().get("assigned") == 0
                  and len(r.json().get("conflicts", [])) == 2, r.text[:200])
            r = await ac.patch("/mh/timesheets/assign", headers=H(hod_t), json={
                "ids": [], "equipment_tag": atag, "system_code": "99"})
            check("assign with no ids → 422", r.status_code == 422, f"got {r.status_code}")
            r = await ac.patch("/mh/timesheets/assign", headers=H(hod_t), json={
                "ids": ids2, "equipment_tag": "nan", "system_code": "99"})
            check("assign with a blank-ish tag → 422", r.status_code == 422,
                  f"got {r.status_code}")

            # ---- Phase-11B: SME link layer (read-only joins) ----------------
            # Pick a real SME scope with planned SQM and no existing estimate,
            # book SVC labor + production + an estimate on it, and assert every
            # joined column on the scorecard.
            sc0 = (await ac.get("/mh/scorecard", headers=H(hod_t))).json()
            check("scorecard unions SME scopes with MH-only scopes",
                  any(x["In_SME"] for x in sc0["items"])
                  and any(not x["In_SME"] and x["Equipment_Tag"] == "SVC-TAG"
                          for x in sc0["items"]),
                  f"{sc0['kpis']}")
            real = next(x for x in sc0["items"]
                        if x["In_SME"] and (x["Planned_SQM"] or 0) > 0
                        and x["Estimated_Manhours"] is None
                        and x["Actual_Manhours"] == 0)
            rtag, rsys = real["Equipment_Tag"], real["System_Code"]
            await ac.post("/mh/timesheets", headers=H(hod_t), json={
                "work_date": "2031-03-01", "equipment_tag": rtag, "system_code": rsys,
                "break_mins": 60, "rows": [
                    {"employee_code": "SVC-EMP-1", "in_time": "07:30", "out_time": "16:30"}]})
            await ac.post("/mh/production", headers=H(hod_t), json={
                "work_date": "2031-03-01", "equipment_tag": rtag, "system_code": rsys,
                "sqm_done": 40, "distribution_method": "even"})
            await ac.post("/mh/estimates", headers=H(hod_t), json={
                "equipment_tag": rtag, "system_code": rsys,
                "estimated_manhours": 10, "estimated_sqm": 50, "basis": "svc scorecard"})
            sc = (await ac.get("/mh/scorecard", headers=H(hod_t))).json()
            row_sc = next(x for x in sc["items"]
                          if x["Equipment_Tag"] == rtag and x["System_Code"] == rsys)
            check("scorecard row: 8h labor + est 10 → labor variance −20%",
                  row_sc["Actual_Manhours"] == 8.0 and row_sc["Estimated_Manhours"] == 10.0
                  and row_sc["Labor_Variance_Pct"] == -20.0, str(row_sc))
            check("scorecard row: labor-reported 40 SQM → MH/SQM 0.2",
                  row_sc["Done_SQM_Labor"] == 40.0 and row_sc["MH_per_SQM"] == 0.2,
                  str(row_sc))
            check("reconciliation flags drift (labor says 40, SME says 0)",
                  row_sc["Reconciliation"] == "drift" and row_sc["Done_SQM_SME"] == 0.0,
                  str(row_sc))
            check("scorecard KPIs count labor + drift",
                  sc["kpis"]["with_labor"] >= 1 and sc["kpis"]["drift"] >= 1,
                  str(sc["kpis"]))

            prod = (await ac.get("/mh/productivity", headers=H(hod_t))).json()
            row_p = next(x for x in prod["items"]
                         if x["Equipment_Tag"] == rtag and x["System_Code"] == rsys)
            check("productivity norms: 0.2 MH/SQM · 5 SQM/MH · est-norm 0.2",
                  row_p["MH_per_SQM"] == 0.2 and row_p["SQM_per_MH"] == 5.0
                  and row_p["Est_MH_per_SQM"] == 0.2, str(row_p))
            check("site norm aggregates scopes with both hours and SQM",
                  (prod["site_norm"]["mh_per_sqm"] or 0) > 0, str(prod["site_norm"]))
            r = await ac.get("/mh/scorecard", headers=H(worker_t))
            check("worker (lvl 0) → 403 on the scorecard", r.status_code == 403,
                  f"got {r.status_code}")
            r = await ac.get("/mh/export/scorecard", headers=H(hod_t),
                             params={"format": "pdf"})
            check("scorecard export → 200 + pdf",
                  r.status_code == 200 and "pdf" in r.headers.get("content-type", ""),
                  f"got {r.status_code}")
            r = await ac.get("/mh/export/productivity", headers=H(hod_t),
                             params={"format": "xlsx"})
            check("productivity export → 200 + spreadsheet",
                  r.status_code == 200 and "spreadsheet" in r.headers.get("content-type", ""),
                  f"got {r.status_code}")

            # ---- Phase-11C: planning automation ------------------------------
            # Auto-draft preview: only unestimated SME scopes with remaining SQM;
            # the seeded history (33.5 h over 70 SQM) yields a real site norm.
            ad = (await ac.get("/mh/estimates/auto-draft", headers=H(hod_t))).json()
            check("auto-draft preview: unestimated scopes only, site norm learned",
                  len(ad["items"]) > 0 and ad["site_norm"] is not None
                  and not any(x["Equipment_Tag"] == rtag and x["System_Code"] == rsys
                              for x in ad["items"]), f"{len(ad['items'])} rows, norm={ad['site_norm']}")
            check("draft math: Draft_MH == Remaining_SQM × Norm_Used on every row",
                  all(x["Draft_Manhours"] is not None
                      and abs(x["Draft_Manhours"] - round(x["Remaining_SQM"] * x["Norm_Used"], 1)) < 0.11
                      for x in ad["items"]), str(ad["items"][:2]))
            ad5 = (await ac.get("/mh/estimates/auto-draft", headers=H(hod_t),
                                params={"norm": 0.5})).json()
            check("norm override: every draft = remaining × 0.5, source 'override'",
                  all(x["Norm_Source"] == "override"
                      and abs(x["Draft_Manhours"] - round(x["Remaining_SQM"] * 0.5, 1)) < 0.11
                      for x in ad5["items"]), str(ad5["items"][:1]))

            # Save two reviewed drafts (edited MH) → they appear as estimates and
            # leave the draftable pool.
            pick = ad["items"][:2]
            r = await ac.post("/mh/estimates/auto-draft", headers=H(hod_t), json={
                "rows": [{"equipment_tag": x["Equipment_Tag"],
                          "system_code": x["System_Code"],
                          "estimated_manhours": 123.0,
                          "estimated_sqm": x["Remaining_SQM"],
                          "location": x["Location"], "basis": "svc auto-draft"}
                         for x in pick]})
            check("auto-draft save → 2 estimates", r.status_code == 200
                  and r.json().get("saved") == 2, r.text[:120])
            ests = (await ac.get("/mh/estimates", headers=H(hod_t))).json()["items"]
            check("saved drafts land in mh_manhour_estimates with the reviewed MH",
                  sum(1 for e in ests if e["Basis"] == "svc auto-draft"
                      and float(e["Estimated_Manhours"]) == 123.0) == 2, "")
            ad2 = (await ac.get("/mh/estimates/auto-draft", headers=H(hod_t))).json()
            check("saved scopes leave the draftable pool",
                  len(ad2["items"]) == len(ad["items"]) - 2,
                  f"{len(ad['items'])} → {len(ad2['items'])}")
            r = await ac.post("/mh/estimates/auto-draft", headers=H(hod_t), json={"rows": []})
            check("auto-draft save with no rows → 422", r.status_code == 422,
                  f"got {r.status_code}")

            # Manpower forecast: estimate-based remaining (10 est − 8 actual = 2)
            # + norm-based scopes; fully-consumed SVC-TAG/99 drops out.
            fc = (await ac.get("/mh/forecast", headers=H(hod_t),
                               params={"crew_size": 10, "hours_per_day": 8})).json()
            fr = next((x for x in fc["items"] if x["Equipment_Tag"] == rtag
                       and x["System_Code"] == rsys), None)
            check("forecast: estimate-based scope has 2 MH remaining",
                  fr is not None and fr["Basis"] == "estimate"
                  and fr["Remaining_Manhours"] == 2.0 and fr["Days_To_Complete"] > 0,
                  str(fr))
            check("forecast: fully-consumed scope drops out (SVC-TAG 25.5h > 20 est)",
                  not any(x["Equipment_Tag"] == "SVC-TAG" for x in fc["items"]), "")
            check("forecast: norm-based scopes included + rollup sums",
                  any(x["Basis"] == "norm" for x in fc["items"])
                  and fc["rollup"]["total_remaining_manhours"] > 0
                  and fc["rollup"]["days_to_complete"] > 0, str(fc["rollup"]))
            r = await ac.get("/mh/forecast", headers=H(hod_t), params={"crew_size": 0})
            check("forecast crew_size=0 → 422", r.status_code == 422, f"got {r.status_code}")
            r = await ac.get("/mh/forecast", headers=H(worker_t))
            check("worker (lvl 0) → 403 on the forecast", r.status_code == 403,
                  f"got {r.status_code}")

            # Exports reuse the shared report renderers.
            r = await ac.get("/mh/export/variance", headers=H(hod_t),
                             params={"format": "xlsx"})
            check("MH export → 200 + spreadsheet", r.status_code == 200
                  and "spreadsheet" in r.headers.get("content-type", ""),
                  f"got {r.status_code}")
            r = await ac.get("/mh/export/nope", headers=H(hod_t))
            check("unknown MH export → 404", r.status_code == 404, f"got {r.status_code}")
        finally:
            # Cleanup: remove every SVC- artifact this suite created.
            from sqlalchemy import text as _text
            async with SessionLocal() as s:
                await s.execute(_text(
                    "DELETE FROM mh_timesheets WHERE \"Employee_Code\" LIKE 'SVC-%'"))
                await s.execute(_text(
                    "DELETE FROM mh_employees WHERE \"Employee_Code\" LIKE 'SVC-%'"))
                await s.execute(_text(
                    "DELETE FROM mh_production WHERE \"Equipment_Tag\" = 'SVC-TAG' "
                    "OR \"Work_Date\" LIKE '2031-%'"))
                await s.execute(_text(
                    "DELETE FROM mh_manhour_estimates WHERE \"Equipment_Tag\" = 'SVC-TAG' "
                    "OR \"Basis\" LIKE 'svc%'"))
                await s.execute(_text(
                    "DELETE FROM mh_variance_notes WHERE \"Equipment_Tag\" = 'SVC-TAG'"))
                await s.commit()


async def test_ai_layer():
    """Phase AI-0/AI-1: safety-gate + fuzzy ports (pure functions), role-gated
    manual retrieval, and the SSE assistant endpoint with a MOCKED Ollama
    client (tests never require a live model server)."""
    import backend.api.ai.client as aic
    from backend.api.ai import fuzzy, manual_qa
    from backend.api.ai.safety import is_safe_select, scrub_sql

    # --- safety gate (PG-hardened port) ----------------------------------
    ok, _ = is_safe_select("SELECT * FROM receipts -- Replace with real Site_ID")
    check("safety: forbidden keyword in a comment does NOT trip", ok)
    ok, _ = is_safe_select("SELECT * FROM receipts WHERE note = 'please DELETE later'")
    check("safety: forbidden keyword in a string literal does NOT trip", ok)
    check("safety: multi-statement blocked",
          not is_safe_select("SELECT 1; DROP TABLE receipts")[0])
    check("safety: UPDATE blocked", not is_safe_select("UPDATE receipts SET x=1")[0])
    check("safety: users table blocked",
          not is_safe_select("SELECT * FROM users")[0])
    check("safety: auth_sessions blocked (new-stack addition)",
          not is_safe_select("SELECT * FROM auth_sessions")[0])
    check("safety: pg_catalog blocked (PG addition)",
          not is_safe_select("SELECT * FROM pg_catalog.pg_tables")[0])
    check("safety: COPY blocked (PG addition)",
          not is_safe_select("SELECT 1 UNION COPY x TO '/tmp/f'")[0])
    check("safety: WITH...SELECT CTE allowed",
          is_safe_select("WITH t AS (SELECT 1 AS n) SELECT n FROM t")[0])
    check("safety: LIMIT injected on a new line",
          scrub_sql("SELECT * FROM receipts").endswith("\nLIMIT 500"))
    check("safety: existing LIMIT kept",
          scrub_sql("SELECT 1 LIMIT 7") == "SELECT 1 LIMIT 7")
    check("safety: 'limit' inside a comment still gets a real LIMIT",
          "LIMIT 500" in scrub_sql("SELECT 1 -- no limit here"))

    # --- fuzzy matcher (pandas-free port) ---------------------------------
    inv = [{"SAP_Code": "1001", "Equipment_Description": "Pipe 6m DN50", "UOM": "PCS"},
           {"SAP_Code": "1002", "Equipment_Description": "Double Clamp 2in", "UOM": "PCS"},
           {"SAP_Code": "1003", "Equipment_Description": "Axial Fan 500mm", "UOM": "EA"}]
    check("fuzzy: normalise drops punctuation + UOM noise",
          fuzzy.normalise("Pipe, 6m (PCS)") == "pipe 6m")
    bm = fuzzy.best_match("pipe 6m dn50", inv)
    check("fuzzy: exact-ish query auto-fills", bm is not None and bm["SAP_Code"] == "1001",
          str(bm))
    rows = fuzzy.resolve_rows([{"material_text": "6m pipe"},
                               {"material_text": "totally unknown thing"}], inv)
    check("fuzzy: reordered tokens → auto/pick (never unknown), junk → unknown",
          rows[0]["match_state"] in ("auto", "pick") and rows[1]["match_state"] == "unknown",
          str([(r['match_state'], r['score']) for r in rows]))

    # --- role-gated manual retrieval --------------------------------------
    sections = manual_qa._load_sections()
    check("manual: sections parsed (v3.0 has 19)", len(sections) >= 17, f"got {len(sections)}")
    sk_ctx = manual_qa._context_for_role("store_keeper")
    adm_ctx = manual_qa._context_for_role("admin")
    hod_ctx = manual_qa._context_for_role("hod")
    check("manual: store keeper sees §4, physically NOT §7 (admin chapter)",
          "=== Section 4:" in sk_ctx and "=== Section 7:" not in sk_ctx)
    check("manual: admin sees §7 untruncated (longer context than SK)",
          "=== Section 7:" in adm_ctx and len(adm_ctx) > len(sk_ctx))
    check("manual: hod allowlist grew §18 SME + §19 Man-Hours",
          "=== Section 18:" in hod_ctx and "=== Section 19:" in hod_ctx)
    check("manual: greeting fast-path (no LLM)",
          manual_qa.greeting_reply("hi") is not None
          and manual_qa.greeting_reply("how do I stage a return?") is None)

    # --- /ai endpoints over the live ASGI app (mocked Ollama) --------------
    transport = ASGITransport(app=app)
    ip = {"X-Real-IP": "203.0.113.15"}
    async with AsyncClient(transport=transport, base_url="http://svc") as ac:
        async def token(u, p):
            r = await ac.post("/auth/login", json={"username": u, "password": p}, headers=ip)
            return r.json().get("access_token")

        def H(t):
            return {"Authorization": f"Bearer {t}"}

        admin_t = await token("admin", "admin2026")
        worker_t = await token("worker", "floor2026")

        r = await ac.get("/ai/health")
        check("ai health without a token → 401", r.status_code == 401, f"got {r.status_code}")
        r = await ac.get("/ai/health", headers=H(worker_t))
        check("ai health → 200 with ok/enabled/message",
              r.status_code == 200 and {"ok", "enabled", "message"} <= set(r.json()),
              r.text[:120])

        # Greeting streams without any model (works even with Ollama down).
        r = await ac.post("/ai/assistant", headers=H(worker_t), json={"question": "hi"})
        check("assistant greeting → SSE tokens + done (no LLM involved)",
              r.status_code == 200 and '"token"' in r.text and '"done": true' in r.text,
              r.text[:160])

        # Mock the Ollama client and prove the stream + the role gate.
        saved = (aic.health, aic.list_models, aic.stream)
        captured: dict = {}

        async def fake_health():
            return True

        async def fake_models():
            return [aic.MODEL_CHAT]

        async def fake_stream(model, prompt, *, system=None, **kw):
            captured["system"] = system or ""
            for t in ("Go to ", "Entry Log."):
                yield t
        try:
            aic.health, aic.list_models, aic.stream = fake_health, fake_models, fake_stream
            r = await ac.post("/ai/assistant", headers=H(worker_t),
                              json={"question": "how do I stage a return?"})
            check("assistant streams model chunks as SSE events in order",
                  r.text.index('"Go to "') < r.text.index('"Entry Log."')
                  and '"done": true' in r.text, r.text[:200])
            check("role gate: the store keeper's PROMPT physically lacks the admin chapter",
                  "=== Section 4:" in captured["system"]
                  and "=== Section 7:" not in captured["system"], "")

            # Feature flag: switch the assistant off → error event; restore.
            r = await ac.put("/admin/settings", headers=H(admin_t),
                             json={"key": "ai_assistant_enabled", "value": "0"})
            check("ai flag accepted by the settings whitelist", r.status_code == 200,
                  f"got {r.status_code}")
            try:
                r = await ac.post("/ai/assistant", headers=H(worker_t),
                                  json={"question": "how do I stage a return?"})
                check("assistant while flagged off → SSE error event",
                      '"error"' in r.text and '"done": true' in r.text, r.text[:160])
                r = await ac.get("/ai/health", headers=H(worker_t))
                check("health reports enabled:false while flagged off",
                      r.json().get("enabled") is False, r.text[:120])
            finally:
                rr = await ac.put("/admin/settings", headers=H(admin_t),
                                  json={"key": "ai_assistant_enabled", "value": "1"})
                check("ai flag restored", rr.status_code == 200, f"got {rr.status_code}")
        finally:
            aic.health, aic.list_models, aic.stream = saved

        # ---- Phase AI-2: document intelligence (PR/PO PDF extraction) -------
        # Synthetic PDFs built with fpdf2 (already a dep) — one PR + the three
        # legacy PO layouts. Extraction is preview-only: nothing may persist.
        from fpdf import FPDF

        def make_pdf(lines):
            p = FPDF()
            p.add_page()
            p.set_font("Helvetica", size=10)
            for ln in lines:
                p.cell(0, 6, ln, new_x="LMARGIN", new_y="NEXT")
            return bytes(p.output())

        mt = "application/pdf"
        hod_t2 = await token("hod", "hod2026")
        pr_pdf = make_pdf([
            "GENERAL INDUSTRIES - PURCHASE REQUEST",
            "Purch. Req. No. : 3001234567",
            "001 GI-7003055 SOME MATERIAL DESC 25.00 KG",
            "002 GI-7002999 OTHER MATERIAL 10 PCS",
            "003 GI-9999999 UNKNOWN THING 5 EA",
        ])
        r = await ac.post("/ai/extract/pr", headers=H(worker_t),
                          files={"file": ("pr.pdf", pr_pdf, mt)})
        check("extract/pr: worker (lvl 0) → 403", r.status_code == 403,
              f"got {r.status_code}")
        n_pr_before = None
        async with SessionLocal() as s:
            n_pr_before = await _count(s, pr_master_t)
        r = await ac.post("/ai/extract/pr", headers=H(hod_t2),
                          files={"file": ("pr.pdf", pr_pdf, mt)})
        j = r.json()
        check("extract/pr: PR number + strict matching (2 matched / 1 unmatched)",
              r.status_code == 200 and j.get("pr_number") == "3001234567"
              and len(j.get("matched", [])) == 2 and len(j.get("unmatched", [])) == 1,
              r.text[:200])
        check("extract/pr: matched rows are pre-shaped create-PR lines w/ SAP + qty",
              j["matched"][0]["SAP_Code"] and j["matched"][0]["Requested_Qty"] == 25.0
              and j["unmatched"][0]["material_code"] == "GI-9999999", str(j["matched"][0]))
        async with SessionLocal() as s:
            n_pr_after = await _count(s, pr_master_t)
        check("extract/pr is preview-only (silent-insert flaw fixed — no rows written)",
              n_pr_after == n_pr_before, f"{n_pr_before} → {n_pr_after}")

        # Confirm path = the EXISTING audited service (create_pr). Audit proof
        # is DELTA-counted: PR numbers restart per day, so audit rows from
        # earlier (cleaned-up) test PRs can share the number.
        async with SessionLocal() as s:
            audit_before = await _count(s, audit_t,
                                        audit_t.c["action_type"] == "CREATE_PR")
        r = await ac.post("/hod/prs", headers=H(hod_t2), json={
            "site_id": "CNCEC", "notes": "Imported from PR PDF 3001234567",
            "lines": [{"SAP_Code": m["SAP_Code"], "Requested_Qty": m["Requested_Qty"]}
                      for m in j["matched"]]})
        conf = r.json()
        check("confirm → PR created through the audited service", r.status_code == 201
              and conf.get("created") is True and conf.get("lines") == 2, r.text[:160])
        async with SessionLocal() as s:
            audit_after = await _count(s, audit_t,
                                       audit_t.c["action_type"] == "CREATE_PR")
            check("confirm wrote the CREATE_PR audit row (legacy never did)",
                  audit_after == audit_before + 1, f"{audit_before} → {audit_after}")
            # Cleanup the PR rows (the audit row stays — it's a true record of
            # a write that really happened; delta counting makes that safe).
            await s.execute(pr_master_t.delete().where(
                pr_master_t.c["PR_Number"] == conf["pr_number"]))
            await s.commit()

        po_a = make_pdf([
            "Purch. Order. No. : 4710003114",
            "Purch. Order. Date : 15.06.2026",
            "Vendor : 0000123456",
            "ACME TRADING EST",
            "Payment Terms : NET 30",
            "GI-7002522",
            "001 SS 316L FILLER WIRE DIA 2.4 MM 20.00 KG 85.00 255.00 1,955.00",
            "SHIPMENT 01 BRICK MATERIALS 05.02.2026",
            "Total Amount 1,955.00",
        ])
        po_b = make_pdf([
            "Purch. Order No. : 4710003115",
            "Vendor : 0000654321",
            "GULF SUPPLIES CO",
            "001 GI-8003100 CUMIFURAN SYRUP GRADE A 5,025.00 KG 10.00 50,250.00",
        ])
        po_c = make_pdf([
            "Purch. Order No. : 4710003116",
            "Vendor : 0000111222",
            "DESERT MATERIALS LLC",
            "001 GI-7002522",
            "CUMIFURAN SYRUP SPECIAL 5,025.00 KG 10.00 50,250.00",
        ])
        r = await ac.post("/ai/extract/po", headers=H(hod_t2),
                          files={"file": ("po.pdf", po_a, mt)})
        check("extract/po: hod (lvl 2) → 403 (logistics gate)", r.status_code == 403,
              f"got {r.status_code}")
        r = await ac.post("/ai/extract/po", headers=H(admin_t),
                          files={"file": ("po.pdf", po_a, mt)})
        ja = r.json()
        it = ja["items"][0] if ja.get("items") else {}
        check("PO layout A (code-line + 7-col w/ VAT): header + item + prices",
              r.status_code == 200 and ja["header"].get("PO_Number") == "4710003114"
              and ja["header"].get("Vendor_Code") == "123456"
              and ja["header"].get("Vendor_Name") == "ACME TRADING EST"
              and it.get("Qty") == 20.0 and it.get("Unit_Price") == 85.0
              and it.get("Total_Price") == 1955.0, r.text[:240])
        check("PO layout A: annexure schedule parsed to ISO date",
              ja.get("shipment_schedule")
              and ja["shipment_schedule"][0]["target_date"] == "2026-02-05",
              str(ja.get("shipment_schedule")))
        r = await ac.post("/ai/extract/po", headers=H(admin_t),
                          files={"file": ("po.pdf", po_b, mt)})
        jb = r.json()
        check("PO layout B (inline 6-col): comma-qty + prices",
              jb["items"] and jb["items"][0]["Material_Code"] == "GI-8003100"
              and jb["items"][0]["Qty"] == 5025.0
              and jb["items"][0]["Total_Price"] == 50250.0, r.text[:200])
        r = await ac.post("/ai/extract/po", headers=H(admin_t),
                          files={"file": ("po.pdf", po_c, mt)})
        jc = r.json()
        check("PO layout C (split-line pair): desc + numbers recovered",
              jc["items"] and jc["items"][0]["Material_Code"] == "GI-7002522"
              and jc["items"][0]["Qty"] == 5025.0
              and "CUMIFURAN" in jc["items"][0]["Description"], r.text[:200])
        r = await ac.post("/ai/extract/po", headers=H(admin_t),
                          files={"file": ("junk.pdf", b"not a pdf", mt)})
        check("unparseable PDF → 422", r.status_code == 422, f"got {r.status_code}")

        # Feature flag: doc-intel off → 503; restored in a finally.
        r = await ac.put("/admin/settings", headers=H(admin_t),
                         json={"key": "ai_doc_intel_enabled", "value": "0"})
        check("doc-intel flag accepted by the settings whitelist",
              r.status_code == 200, f"got {r.status_code}")
        try:
            r = await ac.post("/ai/extract/pr", headers=H(hod_t2),
                              files={"file": ("pr.pdf", pr_pdf, mt)})
            check("extract while flagged off → 503", r.status_code == 503,
                  f"got {r.status_code}")
        finally:
            rr = await ac.put("/admin/settings", headers=H(admin_t),
                              json={"key": "ai_doc_intel_enabled", "value": "1"})
            check("doc-intel flag restored", rr.status_code == 200,
                  f"got {rr.status_code}")

        # ---- Phase AI-3: handwriting OCR (async jobs, mocked vision) ---------
        import io as _io2

        from PIL import Image as _Image

        import backend.api.ai.jobs as ai_jobs_mod
        ai_jobs_t = _MD.tables["ai_jobs"]

        def tiny_jpeg() -> bytes:
            buf = _io2.BytesIO()
            _Image.new("RGB", (40, 40), (200, 180, 40)).save(buf, format="JPEG")
            return buf.getvalue()

        async def poll_until_final(jid: int, tok: str) -> dict:
            for _ in range(80):
                await asyncio.sleep(0.05)
                r = await ac.get(f"/ai/jobs/{jid}", headers=H(tok))
                if r.json().get("status") in ("done", "error"):
                    return r.json()
            return r.json()

        try:
            # Exact lock: {store_keeper, admin} — the legacy Daily Issue Log gate.
            r = await ac.post("/ai/jobs", headers=H(hod_t2),
                              params={"kind": "ocr_consumption"},
                              files={"file": ("l.jpg", tiny_jpeg(), "image/jpeg")})
            check("ocr job: hod → 403 (exact store_keeper lock)", r.status_code == 403,
                  f"got {r.status_code}")
            r = await ac.post("/ai/jobs", headers=H(worker_t),
                              params={"kind": "nope"},
                              files={"file": ("l.jpg", tiny_jpeg(), "image/jpeg")})
            check("ocr job: bad kind → 422", r.status_code == 422, f"got {r.status_code}")
            r = await ac.post("/ai/jobs", headers=H(worker_t),
                              params={"kind": "ocr_consumption"},
                              files={"file": ("l.jpg", b"not an image", "image/jpeg")})
            check("ocr job: corrupt image fails FAST at upload (422, no dead job)",
                  r.status_code == 422, f"got {r.status_code}")

            # Full lifecycle with a mocked vision model.
            saved2 = (aic.health, aic.list_models, aic.generate)
            seen: dict = {}

            async def ok_health():
                return True

            async def ok_models():
                return [aic.MODEL_VISION, aic.MODEL_CHAT]

            async def fake_vision(model, prompt, **kw):
                seen["model"] = model
                seen["images"] = bool(kw.get("images"))
                seen["system"] = kw.get("system") or ""
                import json as _json
                return _json.dumps({"rows": [
                    {"issued_to": "Imran", "material_text": "water storage tank 10000",
                     "uom": "Each", "quantity": 2, "work_type": "site"},
                    {"issued_to": "Ali", "material_text": "zzz nonexistent widget",
                     "uom": "PCS", "quantity": 5, "work_type": ""}]})

            aic.health, aic.list_models, aic.generate = ok_health, ok_models, fake_vision
            try:
                r = await ac.post("/ai/jobs", headers=H(worker_t),
                                  params={"kind": "ocr_consumption"},
                                  files={"file": ("l.jpg", tiny_jpeg(), "image/jpeg")})
                check("ocr job accepted → 202 + id", r.status_code == 202
                      and r.json().get("job_id"), r.text[:120])
                jid = r.json()["job_id"]
                j = await poll_until_final(jid, worker_t)
                check("job lifecycle: queued → done via the atomic-claim worker",
                      j["status"] == "done", str(j)[:200])
                check("vision call: right model + image attached + strict JSON prompt",
                      seen.get("model") == aic.MODEL_VISION and seen.get("images")
                      and "STRICT JSON" in seen.get("system", ""), str(seen)[:120])
                rr_ = {x["material_text"]: x for x in j["result"]["rows"]}
                check("fuzzy resolution: legible text → auto w/ SAP, junk → unknown",
                      rr_["water storage tank 10000"]["match_state"] == "auto"
                      and rr_["water storage tank 10000"]["SAP_Code"] == "1001"
                      and rr_["zzz nonexistent widget"]["match_state"] == "unknown",
                      str(j["result"]["rows"])[:200])
                r = await ac.get(f"/ai/jobs/{jid}", headers=H(admin_t))
                check("admin may inspect any job", r.status_code == 200,
                      f"got {r.status_code}")

                # DN kind: header + items shape survives the round trip.
                async def fake_vision_dn(model, prompt, **kw):
                    import json as _json
                    return _json.dumps({
                        "header": {"DN_No": "15668", "Date": "2026-06-02",
                                   "Mob_From": "GI - ABU HADRIYAH", "Driver_Name": "Imran",
                                   "Vehicle_No": "3909", "Prepared_by": "H", "Mob_To": "CNCEC"},
                        "items": [{"material_text": "air compressor 750",
                                   "uom": "Each", "quantity": 1}]})
                aic.generate = fake_vision_dn
                r = await ac.post("/ai/jobs", headers=H(worker_t),
                                  params={"kind": "ocr_delivery_note"},
                                  files={"file": ("dn.jpg", tiny_jpeg(), "image/jpeg")})
                j = await poll_until_final(r.json()["job_id"], worker_t)
                check("DN job: header preserved + items fuzzy-resolved",
                      j["status"] == "done" and j["result"]["header"]["DN_No"] == "15668"
                      and j["result"]["items"][0]["match_state"] == "auto"
                      and j["result"]["items"][0]["SAP_Code"] == "1003", str(j)[:240])

                # Unparseable model reply → clean job error, not a crash.
                async def garbage_vision(model, prompt, **kw):
                    return "I cannot read this image, sorry!"
                aic.generate = garbage_vision
                r = await ac.post("/ai/jobs", headers=H(worker_t),
                                  params={"kind": "ocr_consumption"},
                                  files={"file": ("l.jpg", tiny_jpeg(), "image/jpeg")})
                j = await poll_until_final(r.json()["job_id"], worker_t)
                check("unparseable model reply → job error w/ paste-tab hint",
                      j["status"] == "error" and "Paste" in (j.get("error") or ""),
                      str(j)[:160])

                # Ollama offline → job error with the friendly preflight message.
                async def down_health():
                    return False
                aic.health = down_health
                r = await ac.post("/ai/jobs", headers=H(worker_t),
                                  params={"kind": "ocr_consumption"},
                                  files={"file": ("l.jpg", tiny_jpeg(), "image/jpeg")})
                j = await poll_until_final(r.json()["job_id"], worker_t)
                check("Ollama offline → job error names the Paste fallback",
                      j["status"] == "error" and "offline" in (j.get("error") or "").lower(),
                      str(j)[:160])
            finally:
                aic.health, aic.list_models, aic.generate = saved2

            # Paste lane: pure-Python, works with NO mock (Ollama-independent).
            r = await ac.post("/ai/paste/ocr_consumption", headers=H(worker_t),
                              json={"text": "Imran\tair compressor 750\tEach\t3\n"
                                            "Ali, water storage tank 10000, Each, 1"})
            j = r.json()
            check("paste lane (offline): both delimiter styles resolve to SAP codes",
                  r.status_code == 200 and len(j["rows"]) == 2
                  and {x["SAP_Code"] for x in j["rows"]} == {"1003", "1001"},
                  r.text[:200])
            r = await ac.post("/ai/paste/ocr_delivery_note", headers=H(worker_t),
                              json={"text": "Customer: GI - HQ\nDriver: Imran\n"
                                            "air compressor 750, Each, 2"})
            j = r.json()
            check("DN paste: header synonyms map (Customer→Mob_From) + items resolve",
                  j["header"]["Mob_From"] == "GI - HQ"
                  and j["items"][0]["match_state"] == "auto", r.text[:200])
            r = await ac.post("/ai/paste/ocr_consumption", headers=H(worker_t),
                              json={"text": "   "})
            check("empty paste → 422", r.status_code == 422, f"got {r.status_code}")

            # Orphan sweep: a queued row from a 'dead process' gets failed.
            async with SessionLocal() as s:
                from sqlalchemy import insert as _ins
                orphan_id = (await s.execute(_ins(ai_jobs_t).values(
                    kind="ocr_consumption", status="running", actor="worker",
                    payload_json="{}").returning(ai_jobs_t.c["id"]))).scalar_one()
                await s.commit()
            n = await ai_jobs_mod.fail_orphans()
            check("startup orphan sweep fails stranded jobs with a clear message",
                  n >= 1, f"swept {n}")
            r = await ac.get(f"/ai/jobs/{orphan_id}", headers=H(worker_t))
            check("orphaned job reads back as error → 'resubmit the photo'",
                  r.json()["status"] == "error"
                  and "resubmit" in (r.json().get("error") or ""), r.text[:160])

            # Flag off → both lanes 503; restored in the finally below.
            r = await ac.put("/admin/settings", headers=H(admin_t),
                             json={"key": "ai_ocr_enabled", "value": "0"})
            check("ocr flag accepted by the settings whitelist", r.status_code == 200,
                  f"got {r.status_code}")
            try:
                r = await ac.post("/ai/jobs", headers=H(worker_t),
                                  params={"kind": "ocr_consumption"},
                                  files={"file": ("l.jpg", tiny_jpeg(), "image/jpeg")})
                check("ocr job while flagged off → 503", r.status_code == 503,
                      f"got {r.status_code}")
                r = await ac.post("/ai/paste/ocr_consumption", headers=H(worker_t),
                                  json={"text": "a\tb\tc\t1"})
                check("paste while flagged off → 503", r.status_code == 503,
                      f"got {r.status_code}")
            finally:
                rr = await ac.put("/admin/settings", headers=H(admin_t),
                                  json={"key": "ai_ocr_enabled", "value": "1"})
                check("ocr flag restored", rr.status_code == 200, f"got {rr.status_code}")
            # ---- Phase AI-4: Smart Scan (badge verify + tool identify) -------
            # Tier 1: QR decode is client-side; the server only verifies the
            # decoded ID string against employees (active check).
            emp_row = None
            async with SessionLocal() as s:
                emp_t = _MD.tables["employees"]
                from sqlalchemy import select as _sel
                emp_row = (await s.execute(_sel(
                    emp_t.c["ID_Number"], emp_t.c["Name"])
                    .where(emp_t.c["status"] == "active").limit(1))).first()
            r = await ac.get(f"/ai/badge/{emp_row.ID_Number}", headers=H(worker_t))
            check("badge verify: active employee found + prefill fields",
                  r.status_code == 200 and r.json()["found"] is True
                  and r.json()["active"] is True
                  and r.json()["name"] == emp_row.Name, r.text[:160])
            r = await ac.get("/ai/badge/no-such-badge-999", headers=H(worker_t))
            check("badge verify: unknown id → found:false + friendly message",
                  r.json().get("found") is False and "No employee" in r.json()["message"],
                  r.text[:120])
            r = await ac.get(f"/ai/badge/{emp_row.ID_Number}", headers=H(hod_t2))
            check("badge verify: hod → 403 (exact store_keeper lock)",
                  r.status_code == 403, f"got {r.status_code}")

            # Tier 2: tool_identify vision job — catalogue-constrained when
            # tool_catalogue has rows (seed two, clean up after).
            async with SessionLocal() as s:
                from sqlalchemy import insert as _ins2
                cat_t = _MD.tables["tool_catalogue"]
                await s.execute(_ins2(cat_t).values(
                    class_name="svc_angle_grinder", display_name="Angle Grinder 9in"))
                await s.execute(_ins2(cat_t).values(
                    class_name="svc_torque_wrench", display_name="Torque Wrench 1/2in"))
                await s.commit()
            saved3 = (aic.health, aic.list_models, aic.generate)
            seen_tool: dict = {}

            async def tool_vision(model, prompt, **kw):
                seen_tool["system"] = kw.get("system") or ""
                import json as _json
                return _json.dumps({"name": "svc_angle_grinder",
                                    "alternatives": ["svc_torque_wrench", "Crowbar"],
                                    "description": "A 9-inch angle grinder."})
            try:
                aic.health, aic.list_models, aic.generate = ok_health, ok_models, tool_vision
                r = await ac.post("/ai/jobs", headers=H(worker_t),
                                  params={"kind": "tool_identify"},
                                  files={"file": ("t.jpg", tiny_jpeg(), "image/jpeg")})
                check("tool_identify accepted as a job kind", r.status_code == 202,
                      f"got {r.status_code}")
                j = await poll_until_final(r.json()["job_id"], worker_t)
                tool = (j.get("result") or {}).get("tool") or {}
                check("tool job: catalogue classes in the PROMPT",
                      "svc_angle_grinder" in seen_tool.get("system", "")
                      and "Angle Grinder 9in" in seen_tool.get("system", ""), "")
                check("tool job: class names map to display names",
                      j["status"] == "done" and tool.get("name") == "Angle Grinder 9in"
                      and tool.get("class_name") == "svc_angle_grinder", str(tool))
                check("tool job: catalogue alt mapped + freeform alt passes through",
                      tool.get("alternatives", [{}])[0].get("name") == "Torque Wrench 1/2in"
                      and tool.get("alternatives", [{}, {}])[1].get("name") == "Crowbar"
                      and tool["alternatives"][1]["class_name"] is None,
                      str(tool.get("alternatives")))

                # Empty-catalogue path: freeform naming still works.
                async with SessionLocal() as s:
                    from sqlalchemy import delete as _del2
                    await s.execute(_del2(cat_t).where(
                        cat_t.c["class_name"].like("svc_%")))
                    await s.commit()

                async def tool_vision_free(model, prompt, **kw):
                    import json as _json
                    return _json.dumps({"name": "Pipe Wrench 24in",
                                        "alternatives": [], "description": "x"})
                aic.generate = tool_vision_free
                r = await ac.post("/ai/jobs", headers=H(worker_t),
                                  params={"kind": "tool_identify"},
                                  files={"file": ("t.jpg", tiny_jpeg(), "image/jpeg")})
                j = await poll_until_final(r.json()["job_id"], worker_t)
                check("tool job: empty catalogue → freeform name (class_name null)",
                      j["status"] == "done"
                      and j["result"]["tool"]["name"] == "Pipe Wrench 24in"
                      and j["result"]["tool"]["class_name"] is None, str(j)[:160])
            finally:
                aic.health, aic.list_models, aic.generate = saved3
                async with SessionLocal() as s:
                    from sqlalchemy import delete as _del3
                    await s.execute(_del3(_MD.tables["tool_catalogue"]).where(
                        _MD.tables["tool_catalogue"].c["class_name"].like("svc_%")))
                    await s.commit()
        finally:
            # Cleanup: OCR jobs are test artifacts — remove every row we made.
            from sqlalchemy import text as _text2
            async with SessionLocal() as s:
                await s.execute(_text2("DELETE FROM ai_jobs"))
                await s.commit()

        # ---- Phase AI-5: analytics AI ------------------------------------------
        import json

        from backend.api.ai import analytics

        # NL→SQL is gated to UNSCOPED roles only (V1 site-scoping ruling).
        r = await ac.post("/ai/nl-search", headers=H(worker_t), json={"question": "x"})
        check("nl-search: store keeper (lvl 0) → 403", r.status_code == 403,
              f"got {r.status_code}")
        r = await ac.post("/ai/nl-search", headers=H(hod_t2), json={"question": "x"})
        check("nl-search: hod (SCOPED role) → 403 by design", r.status_code == 403,
              f"got {r.status_code}")

        saved5 = (aic.health, aic.list_models, aic.generate, aic.stream)

        async def coder_ok(model, prompt, **kw):
            check("nl-search uses the CODER model", model == aic.MODEL_CODER, model)
            return ('```sql\nSELECT "Supplier", COUNT(*) AS orders FROM receipts '
                    'WHERE "Supplier" IS NOT NULL GROUP BY "Supplier" '
                    'ORDER BY orders DESC\n```')

        try:
            aic.health, aic.list_models, aic.generate = ok_health, ok_models, coder_ok
            r = await ac.post("/ai/nl-search", headers=H(admin_t),
                              json={"question": "top suppliers by orders"})
            j = r.json()
            check("nl-search: fenced SQL extracted, executed on the RO engine, "
                  "LIMIT injected",
                  r.status_code == 200 and j["ok"] is True and len(j["rows"]) > 0
                  and j["columns"] == ["Supplier", "orders"]
                  and "LIMIT 500" in j["sql"], r.text[:200])

            async def coder_evil(model, prompt, **kw):
                return "UPDATE receipts SET \"Quantity\" = 0"
            aic.generate = coder_evil
            r = await ac.post("/ai/nl-search", headers=H(admin_t),
                              json={"question": "zero everything"})
            check("nl-search: model-emitted UPDATE rejected by the safety gate",
                  r.json()["ok"] is False and "safety gate" in r.json()["message"],
                  r.text[:160])

            async def coder_snoop(model, prompt, **kw):
                return "SELECT * FROM users"
            aic.generate = coder_snoop
            r = await ac.post("/ai/nl-search", headers=H(admin_t),
                              json={"question": "show users"})
            check("nl-search: users table blocked by the gate",
                  r.json()["ok"] is False, r.text[:160])

            # Wall #2 — the TRUE read-only role. Bypass the text gate entirely
            # and hit the RO engine directly: writes and users are physically
            # impossible even if the gate ever failed.
            from sqlalchemy import text as _t5
            ro_write_blocked = ro_users_blocked = False
            try:
                async with analytics.ro_engine().connect() as conn:
                    await conn.execute(_t5(
                        'INSERT INTO vendors ("Vendor_Name") VALUES (\'svc-ro-test\')'))
            except Exception as e:
                ro_write_blocked = "read-only" in str(e).lower()
            try:
                async with analytics.ro_engine().connect() as conn:
                    await conn.execute(_t5("SELECT COUNT(*) FROM users"))
            except Exception as e:
                ro_users_blocked = "permission denied" in str(e).lower()
            check("RO role: INSERT physically impossible (default_transaction_read_only)",
                  ro_write_blocked, "")
            check("RO role: users unreadable even bypassing the gate (REVOKE)",
                  ro_users_blocked, "")

            # Insights SSE: probe events first (deterministic), then commentary.
            async def commentary_ok(model, prompt, **kw):
                import json as _json
                return _json.dumps({"title": "Svc headline", "body": "Svc body.",
                                    "recs": ["r1", "r2", "r3"]})
            aic.generate = commentary_ok
            r = await ac.post("/ai/insights", headers=H(hod_t2))
            evs = [json.loads(x[6:]) for x in r.text.splitlines()
                   if x.startswith("data: ")]
            probe_ids = [e["probe"]["id"] for e in evs if "probe" in e]
            comm_ids = [e["commentary"]["id"] for e in evs if "commentary" in e]
            check("insights: health-score probe always fires w/ real numbers",
                  "inventory_health_score" in probe_ids
                  and any("probe" in e and e["probe"]["id"] == "inventory_health_score"
                          and e["probe"]["data"]["n_total"] > 0 for e in evs),
                  str(probe_ids))
            check("insights: every fired probe gets commentary + done event",
                  set(probe_ids) == set(comm_ids)
                  and any(e.get("done") for e in evs)
                  and all(e["commentary"]["title"] == "Svc headline"
                          for e in evs if "commentary" in e),
                  f"probes={probe_ids} comms={comm_ids}")
            probe_idx = next(i for i, e in enumerate(evs) if "probe" in e)
            comm_idx = next(i for i, e in enumerate(evs) if "commentary" in e)
            check("insights: probes stream BEFORE commentary (numbers never wait)",
                  probe_idx < comm_idx, f"{probe_idx} vs {comm_idx}")

            # Ollama down → deterministic fallback commentary, stream survives.
            async def down_health5():
                return False
            aic.health = down_health5
            r = await ac.post("/ai/insights", headers=H(hod_t2))
            evs = [json.loads(x[6:]) for x in r.text.splitlines()
                   if x.startswith("data: ")]
            check("insights: Ollama down → deterministic fallback commentary",
                  any("commentary" in e and "unavailable" in e["commentary"]["body"]
                      for e in evs) and any(e.get("done") for e in evs),
                  r.text[:200])

            # EOD summary: streams tokens; hod is site-pinned (403 on foreign site).
            aic.health = ok_health

            async def eod_stream(model, prompt, **kw):
                check("eod: context carries real DB numbers",
                      "Consumption rows today" in prompt, prompt[:80])
                for t in ("Calm ", "day."):
                    yield t
            aic.stream = eod_stream
            r = await ac.post("/ai/eod-summary", headers=H(hod_t2),
                              json={"date": "2026-06-15"})
            check("eod-summary: SSE tokens + done",
                  '"Calm "' in r.text and '"done": true' in r.text, r.text[:160])
            r = await ac.post("/ai/eod-summary", headers=H(hod_t2),
                              json={"date": "2026-06-15", "site_id": "HQ"})
            check("eod-summary: hod requesting a foreign site → 403 (scoping held)",
                  r.status_code == 403, f"got {r.status_code}")
            r = await ac.post("/ai/insights", headers=H(worker_t))
            check("insights: store keeper (lvl 0) → 403", r.status_code == 403,
                  f"got {r.status_code}")

            # Flag off → nl-search 503 (restored in the finally).
            r = await ac.put("/admin/settings", headers=H(admin_t),
                             json={"key": "ai_nl_search_enabled", "value": "0"})
            check("nl-search flag accepted by the settings whitelist",
                  r.status_code == 200, f"got {r.status_code}")
            try:
                r = await ac.post("/ai/nl-search", headers=H(admin_t),
                                  json={"question": "x"})
                check("nl-search while flagged off → 503", r.status_code == 503,
                      f"got {r.status_code}")
            finally:
                rr = await ac.put("/admin/settings", headers=H(admin_t),
                                  json={"key": "ai_nl_search_enabled", "value": "1"})
                check("nl-search flag restored", rr.status_code == 200,
                      f"got {rr.status_code}")
        finally:
            aic.health, aic.list_models, aic.generate, aic.stream = saved5
            await analytics.ro_engine().dispose()


def test_config_jwt():
    """JWT_SECRET hardening: dev is lenient, production fails fast on a weak key."""
    import os
    from .config import _DEV_JWT_SECRET, jwt_secret
    saved = {k: os.environ.get(k) for k in ("GI_ENV", "JWT_SECRET")}
    try:
        os.environ.pop("GI_ENV", None)
        os.environ.pop("JWT_SECRET", None)
        check("dev jwt_secret ≥ 32 chars (no HMAC warning)", len(jwt_secret()) >= 32)
        os.environ["GI_ENV"] = "production"
        try:
            jwt_secret()
            raised = False
        except RuntimeError:
            raised = True
        check("production without JWT_SECRET fails fast", raised)
        os.environ["JWT_SECRET"] = "x" * 40
        check("production accepts a strong secret", jwt_secret() == "x" * 40)
        os.environ["JWT_SECRET"] = _DEV_JWT_SECRET
        try:
            jwt_secret()
            rejected = False
        except RuntimeError:
            rejected = True
        check("production rejects the dev-default secret", rejected)
    finally:
        for k, v in saved.items():
            if v is None:
                os.environ.pop(k, None)
            else:
                os.environ[k] = v


# --- Suite G: SME plan layer (Phase S1 — engine port + parity oracle) --------
def _sme_deep_diff(a, b, path="", tol=1e-9) -> str:
    """First mismatch between two JSON-ish structures ('' if equal).
    Mirrors the comparator in frontend/scripts/sme_parity.mjs."""
    if isinstance(a, (int, float)) and isinstance(b, (int, float)) \
            and not isinstance(a, bool) and not isinstance(b, bool):
        return "" if abs(a - b) <= tol else f"{path}: {a} != {b}"
    if isinstance(a, list) and isinstance(b, list):
        if len(a) != len(b):
            return f"{path}: length {len(a)} != {len(b)}"
        for i, (x, y) in enumerate(zip(a, b)):
            d = _sme_deep_diff(x, y, f"{path}[{i}]", tol)
            if d:
                return d
        return ""
    if isinstance(a, dict) and isinstance(b, dict):
        if sorted(a) != sorted(b):
            return f"{path}: keys {sorted(a)} != {sorted(b)}"
        for k in a:
            d = _sme_deep_diff(a[k], b[k], f"{path}.{k}", tol)
            if d:
                return d
        return ""
    return "" if a == b else f"{path}: {a!r} != {b!r}"


async def test_sme_plan_layer():
    """Phase S1: the Python cascade engine against the shared golden fixture,
    then the snapshot/cascade endpoints (role locks, site pinning, and
    endpoint ≡ pure-engine self-consistency). The SAME golden is asserted
    against the TypeScript engine by frontend/scripts/sme_parity.mjs —
    golden equality on both sides proves the TS engine ≡ this oracle."""
    import json
    from pathlib import Path

    from . import sme_engine as E

    here = Path(__file__).parent
    fx = json.loads((here / "sme_parity_fixture.json").read_text())
    golden = json.loads((here / "sme_parity_golden.json").read_text())
    m = fx["model"]
    model = E.build_model(m["equipment"], m["recipes"], m["materials"], m["progress"])

    check("engine: default order matches golden",
          model["default_order"] == golden["_default_order"],
          str(model["default_order"]))
    for case in fx["cases"]:
        got = {**E.run_plan(model, case["order"]),
               **E.run_suggestion_engine(model, case["order"])}
        d = _sme_deep_diff(got, golden[case["name"]])
        check(f"engine matches golden: {case['name']}", d == "", d)

    # Semantic pins the golden encodes (guard against regenerating it wrong).
    feas = {f["Equipment_Tag_No"]: f for f in golden["priority-order"]["feasibility"]}
    check("engine: staged SQM folds into done (TK-A remaining 15 → demand 30)",
          any(ln["Equipment_Tag_No"] == "TK-A" and ln["Material_Code"] == "M1"
              and abs(ln["Demand_Qty"] - 30.0) < 1e-9
              for ln in golden["priority-order"]["lines"]))
    check("engine: statuses span FULL/PARTIAL/BLOCKED + zero-demand tag is FULL",
          feas["TK-A"]["Status"] == E.STATUS_FULL
          and feas["TK-B"]["Status"].startswith("🟡")
          and feas["TK-C"]["Status"] == E.STATUS_BLOCKED
          and feas["TK-D"]["Status"] == E.STATUS_FULL, str(feas))
    check("engine: priority inversion flips TK-A FULL → BLOCKED",
          golden["reordered-subset"]["feasibility"][1]["Equipment_Tag_No"] == "TK-A"
          and golden["reordered-subset"]["feasibility"][1]["Status"] == E.STATUS_BLOCKED)
    check("engine: suggestion sim finds pausing TK-B completes TK-A (+33.33%)",
          golden["reordered-subset"]["suggestions"][0]["Pause_Tag"] == "TK-B"
          and golden["reordered-subset"]["suggestions"][0]["Newly_Completable_Count"] == 1
          and golden["reordered-subset"]["suggestions"][0]["Recommended"] is True)

    transport = ASGITransport(app=app)
    async with AsyncClient(transport=transport, base_url="http://svc") as ac:
        async def token(u, p):
            r = await ac.post("/auth/login", json={"username": u, "password": p})
            return r.json().get("access_token")

        def H(t):
            return {"Authorization": f"Bearer {t}"}

        worker_t = await token("worker", "floor2026")   # store_keeper (level 0)
        hod_t = await token("hod", "hod2026")           # hod @ CNCEC (level 2)
        admin_t = await token("admin", "admin2026")     # level 4 → global

        r = await ac.get("/sme/model-snapshot", headers=H(worker_t))
        check("worker (lvl 0) → 403 on model snapshot", r.status_code == 403,
              f"got {r.status_code}")
        r = await ac.post("/sme/plan/cascade", headers=H(worker_t),
                          json={"priority_order": []})
        check("worker (lvl 0) → 403 on plan cascade", r.status_code == 403,
              f"got {r.status_code}")

        r = await ac.get("/sme/model-snapshot", params={"site_id": "HQ"},
                         headers=H(hod_t))
        check("hod requesting a foreign-site snapshot → 403", r.status_code == 403,
              f"got {r.status_code}")
        r = await ac.get("/sme/model-snapshot", headers=H(hod_t))
        check("hod snapshot (no param) is pinned to own site",
              r.status_code == 200 and r.json().get("site_id") == "CNCEC",
              f"got {r.status_code} site={r.json().get('site_id') if r.status_code == 200 else '—'}")

        r = await ac.get("/sme/model-snapshot", headers=H(admin_t))
        snap = r.json() if r.status_code == 200 else {}
        check("admin snapshot → 200 with all model sections",
              r.status_code == 200 and
              {"equipment", "recipes", "materials", "progress", "default_order"} <= set(snap),
              f"got {r.status_code}")
        check("snapshot default_order is sorted + deduped",
              snap.get("default_order") == sorted(set(snap.get("default_order", []))))

        # Endpoint ≡ pure engine on the SAME snapshot (self-consistency): the
        # server built its model from DB rows, we rebuild from the JSON round-
        # trip — results must agree within float tolerance.
        live_model = E.build_model(snap["equipment"], snap["recipes"],
                                   snap["materials"], snap["progress"])
        order = snap["default_order"]
        expected = E.run_plan(live_model, order)
        r = await ac.post("/sme/plan/cascade", headers=H(admin_t),
                          json={"priority_order": order})
        body = r.json() if r.status_code == 200 else {}
        d = _sme_deep_diff({k: body.get(k) for k in
                            ("order_used", "lines", "feasibility", "totals", "procurement")},
                           expected)
        check("cascade endpoint ≡ pure engine on the live snapshot",
              r.status_code == 200 and d == "", d or f"got {r.status_code}")

        r = await ac.post("/sme/plan/cascade", headers=H(admin_t),
                          json={"priority_order": [], "include_suggestions": True})
        j = r.json() if r.status_code == 200 else {}
        check("cascade with empty order → 200, empty plan, suggestions key",
              r.status_code == 200 and j.get("lines") == []
              and j.get("feasibility") == [] and isinstance(j.get("suggestions"), list),
              f"got {r.status_code}")

        # Reorder sensitivity on live data: reversing the priority order must
        # never change total demand (pool math only shifts who gets it).
        if len(order) >= 2:
            r2 = await ac.post("/sme/plan/cascade", headers=H(admin_t),
                               json={"priority_order": list(reversed(order))})
            t1 = {t["Material_Code"]: t["Demand_Qty"] for t in body.get("totals", [])}
            t2 = {t["Material_Code"]: t["Demand_Qty"] for t in r2.json().get("totals", [])}
            check("reversed priority keeps per-material demand invariant",
                  r2.status_code == 200 and
                  all(abs(t1[k] - t2.get(k, -1)) < 1e-6 for k in t1), str(t2)[:200])

        # Phase S3: session exports rendered by the server oracle.
        r = await ac.post("/sme/plan/export", headers=H(admin_t),
                          json={"priority_order": order, "key": "session-full",
                                "format": "xlsx"})
        check("plan export (session-full xlsx) → 200 + spreadsheet",
              r.status_code == 200 and r.content[:2] == b"PK",
              f"got {r.status_code}")
        r = await ac.post("/sme/plan/export", headers=H(admin_t),
                          json={"priority_order": order, "key": "order-list",
                                "format": "csv"})
        proc = body.get("procurement", [])
        check("plan export (order-list csv) carries the oracle's shortages",
              r.status_code == 200 and
              (not proc or proc[0]["Material_Code"] in r.text),
              f"got {r.status_code}")
        r = await ac.post("/sme/plan/export", headers=H(worker_t),
                          json={"priority_order": [], "key": "session-full"})
        check("worker (lvl 0) → 403 on plan export", r.status_code == 403,
              f"got {r.status_code}")
        r = await ac.post("/sme/plan/export", headers=H(admin_t),
                          json={"priority_order": [], "key": "nope"})
        check("unknown plan-export key → 404", r.status_code == 404,
              f"got {r.status_code}")
        r = await ac.post("/sme/plan/export", headers=H(admin_t),
                          json={"priority_order": [], "key": "session-full",
                                "format": "yeet"})
        check("bad plan-export format → 400", r.status_code == 400,
              f"got {r.status_code}")

        # Phase S4: title override + system-code matrix export.
        r = await ac.post("/sme/plan/export", headers=H(hod_t),
                          json={"priority_order": [], "key": "order-list",
                                "format": "xlsx",
                                "title": "SME Location Report — TRAIN J"})
        check("hod plan export with title override → 200 + spreadsheet",
              r.status_code == 200 and r.content[:2] == b"PK",
              f"got {r.status_code}")
        r = await ac.get("/sme/export/system-code-report",
                         params={"format": "csv"}, headers=H(hod_t))
        check("system-code-report export → 200 with summary columns",
              r.status_code == 200 and "System_Code" in r.text
              and "Equipment_Count" in r.text, f"got {r.status_code}")

        # Phase S5: Execution Plan reads + Total Overview oracle export.
        r = await ac.get("/sme/production-log", headers=H(hod_t))
        check("production-log → 200 with items list (committed only)",
              r.status_code == 200 and isinstance(r.json().get("items"), list),
              f"got {r.status_code}")
        r = await ac.get("/sme/production-log", headers=H(worker_t))
        check("worker (lvl 0) → 403 on production-log", r.status_code == 403,
              f"got {r.status_code}")
        r = await ac.get("/sme/export/progress-list",
                         params={"format": "csv"}, headers=H(admin_t))
        check("progress-list export → 200 with plan-vs-done columns",
              r.status_code == 200 and "Completion_Pct" in r.text
              and "Remaining_SQM" in r.text, f"got {r.status_code}")
        r = await ac.post("/sme/plan/export", headers=H(admin_t),
                          json={"priority_order": order, "key": "overview",
                                "format": "csv"})
        n_rows = max(len(r.text.strip().splitlines()) - 1, 0) if r.status_code == 200 else -1
        n_lines_pairs = len({(x["Equipment_Tag_No"], x["Lining_System_Code"])
                             for x in body.get("lines", [])})
        check("overview export → one row per (tag, code) pair of the cascade",
              r.status_code == 200 and "Fulfillment_Pct" in r.text
              and n_rows >= n_lines_pairs,
              f"got {r.status_code}, rows {n_rows} vs pairs {n_lines_pairs}")

        # 2026-07-07: legacy-parity scoped downloads + client-rows renderer.
        some_tag = order[0] if order else None
        if some_tag:
            r = await ac.get("/sme/export/equipment-report",
                             params={"format": "xlsx", "tag": some_tag},
                             headers=H(admin_t))
            check("scoped equipment export (?tag=) → 200 + legacy filename stem",
                  r.status_code == 200 and r.content[:2] == b"PK"
                  and "equipment_" in r.headers.get("content-disposition", ""),
                  f"got {r.status_code}")
        r = await ac.get("/sme/export/equipment-report",
                         params={"format": "xlsx"}, headers=H(admin_t))
        check("equipment-report xlsx (all) → 200 multi-sheet workbook",
              r.status_code == 200 and r.content[:2] == b"PK"
              and "equipment_report_all_" in r.headers.get("content-disposition", ""),
              f"got {r.status_code}")
        r = await ac.get("/sme/export/system-code-report",
                         params={"format": "xlsx", "code": "5"}, headers=H(hod_t))
        check("scoped system-code export (?code=) → 200 + legacy filename stem",
              r.status_code == 200 and r.content[:2] == b"PK"
              and "system_code_5_" in r.headers.get("content-disposition", ""),
              f"got {r.status_code}")
        r = await ac.post("/sme/plan/export", headers=H(admin_t),
                          json={"priority_order": order, "key": "execution-plan",
                                "format": "xlsx",
                                "equipment_tag": some_tag or "x"})
        check("execution-plan export (per-tag order list) → 200 spreadsheet",
              r.status_code == 200 and r.content[:2] == b"PK",
              f"got {r.status_code}")
        r = await ac.post("/sme/plan/export", headers=H(admin_t),
                          json={"priority_order": [], "key": "execution-plan"})
        check("execution-plan export without equipment_tag → 400",
              r.status_code == 400, f"got {r.status_code}")
        r = await ac.post("/sme/export/rows", headers=H(admin_t),
                          json={"title": "Material Balance Report",
                                "columns": ["Code", "Qty"],
                                "rows": [["M-1", 2.5], ["M-2", 0]],
                                "format": "xlsx",
                                "filename": "dashboard_material_balance"})
        check("client-rows renderer → 200 + requested filename",
              r.status_code == 200 and r.content[:2] == b"PK"
              and "dashboard_material_balance.xlsx" in r.headers.get("content-disposition", ""),
              f"got {r.status_code}")
        r = await ac.post("/sme/export/rows", headers=H(admin_t),
                          json={"title": "t", "columns": ["a", "b"],
                                "rows": [["only-one"]], "format": "csv"})
        check("client-rows renderer rejects ragged rows → 400",
              r.status_code == 400, f"got {r.status_code}")
        r = await ac.post("/sme/export/rows", headers=H(worker_t),
                          json={"title": "t", "columns": ["a"], "rows": []})
        check("worker (lvl 0) → 403 on client-rows renderer",
              r.status_code == 403, f"got {r.status_code}")


async def test_sla_tracker():
    """T2 — admin SLA tracker: >24h aggregation, clear, and the URGENT nudge
    (exact template + audit). Uses one committed synthetic staged issue at
    CNCEC (30h old), fully cleaned up in `finally`."""
    import datetime as _dt

    from sqlalchemy import delete, insert

    pi = ledger._MD.tables["pending_issues"]
    dis = ledger._MD.tables["sla_dismissals"]
    notif = ledger._MD.tables["app_notifications"]

    async with SessionLocal() as s:
        rid = (await s.execute(insert(pi).values(
            Date="2026-07-06", SAP_Code="1001", Quantity=1.0,
            status="pending_hod", Site_ID="CNCEC", Remarks="svc-sla-test",
            Timestamp=_dt.datetime.now() - _dt.timedelta(hours=30),
        ).returning(pi.c["id"]))).scalar_one()
        await s.commit()
    try:
        transport = ASGITransport(app=app)
        async with AsyncClient(transport=transport, base_url="http://svc") as ac:
            # Own X-Real-IP: by suite H the shared client IP has burned through
            # the 10/min login cap (same isolation trick as the other suites).
            _ip = {"X-Real-IP": "203.0.113.52"}

            async def token(u, p):
                r = await ac.post("/auth/login", headers=_ip,
                                  json={"username": u, "password": p})
                return r.json().get("access_token")

            def H(t):
                return {"Authorization": f"Bearer {t}"}

            worker_t = await token("worker", "floor2026")
            admin_t = await token("admin", "admin2026")

            r = await ac.get("/admin/overdue-actions", headers=H(worker_t))
            check("worker (lvl 0) → 403 on overdue-actions", r.status_code == 403,
                  f"got {r.status_code}")
            r = await ac.get("/admin/overdue-actions", params={"hours": 0},
                             headers=H(admin_t))
            check("overdue-actions hours=0 → 422", r.status_code == 422,
                  f"got {r.status_code}")
            r = await ac.post("/admin/overdue-actions/not-a-kind/1/clear",
                              headers=H(admin_t))
            check("clear with an unknown kind → 404", r.status_code == 404,
                  f"got {r.status_code}")
            r = await ac.post("/admin/overdue-actions/hod-issue/99999999/notify",
                              headers=H(admin_t))
            check("notify on a non-pending ref → 404", r.status_code == 404,
                  f"got {r.status_code}")

            r = await ac.get("/admin/overdue-actions", headers=H(admin_t))
            j = r.json()
            item = next((x for x in j.get("items", [])
                         if x["kind"] == "hod-issue" and x["ref_id"] == str(rid)), None)
            check("30h-old staged issue surfaces in overdue-actions",
                  r.status_code == 200 and item is not None
                  and item["age_hours"] >= 29 and item["label"],
                  f"got {r.status_code} item={item}")
            check("responsible resolves the site's HOD",
                  item is not None and "hod" in item["responsible"],
                  f"resp={item and item['responsible']}")
            _items = j.get("items", [])
            check("items are age-sorted (oldest first)",
                  _items == sorted(_items, key=lambda x: -x["age_hours"]))

            r = await ac.post(f"/admin/overdue-actions/hod-issue/{rid}/notify",
                              headers=H(admin_t))
            check("notify → 200 + recipients include the HOD",
                  r.status_code == 200 and "hod" in r.json().get("recipients", []),
                  f"got {r.status_code} {r.text[:120]}")
            async with SessionLocal() as s:
                row = (await s.execute(select(
                    notif.c["body"], notif.c["severity"], notif.c["recipient_user"])
                    .where(notif.c["event_key"] == "sla_nudge",
                           notif.c["related_ref"] == str(rid),
                           notif.c["recipient_user"] == "hod"))).first()
            check("nudge uses the EXACT URGENT template + critical severity",
                  row is not None and row.severity == "critical"
                  and row.body.startswith(
                      f"URGENT — Dear hod, From: Admin. Subject: Action required "
                      f"on pending submission {rid}"),
                  f"row={row}")

            r = await ac.post(f"/admin/overdue-actions/hod-issue/{rid}/clear",
                              headers=H(admin_t))
            check("clear → 200", r.status_code == 200, f"got {r.status_code}")
            r = await ac.post(f"/admin/overdue-actions/hod-issue/{rid}/clear",
                              headers=H(admin_t))
            check("double clear → 409", r.status_code == 409, f"got {r.status_code}")
            r = await ac.get("/admin/overdue-actions", headers=H(admin_t))
            check("cleared item no longer surfaces",
                  all(not (x["kind"] == "hod-issue" and x["ref_id"] == str(rid))
                      for x in r.json().get("items", [])), r.text[:160])
    finally:
        async with SessionLocal() as s:  # full cleanup (audit rows stay, by design)
            await s.execute(delete(notif).where(
                notif.c["event_key"] == "sla_nudge",
                notif.c["related_ref"] == str(rid)))
            await s.execute(delete(dis).where(
                dis.c["kind"] == "hod-issue", dis.c["ref_id"] == str(rid)))
            await s.execute(delete(pi).where(pi.c["id"] == rid))
            await s.commit()


async def test_submission_intel():
    """T1 — Submission Intelligence: role guards, 404s, and the deterministic
    summary contract on synthetic staged-issue + cross-site rows (cleaned up).
    Ollama may be up or down here — `source` just has to be a valid value and
    the summary text non-empty (fallback is deterministic by construction)."""
    from sqlalchemy import delete, insert

    pi = ledger._MD.tables["pending_issues"]
    req = ledger._MD.tables["requests"]
    jobs = ledger._MD.tables["ai_jobs"]

    async with SessionLocal() as s:
        iid = (await s.execute(insert(pi).values(
            Date="2026-07-07", SAP_Code="1001", Quantity=2.0,
            status="pending_hod", Site_ID="CNCEC", Remarks="svc-t1-test",
        ).returning(pi.c["id"]))).scalar_one()
        rid = (await s.execute(insert(req).values(
            requesting_site="CNCEC", target_site="HQ", SAP_Code="1001",
            requested_qty=1.0, status="pending", requested_by="svc-t1",
        ).returning(req.c["id"]))).scalar_one()
        await s.commit()
    try:
        transport = ASGITransport(app=app)
        async with AsyncClient(transport=transport, base_url="http://svc") as ac:
            _ip = {"X-Real-IP": "203.0.113.61"}

            async def token(u, p):
                r = await ac.post("/auth/login", headers=_ip,
                                  json={"username": u, "password": p})
                return r.json().get("access_token")

            def H(t):
                return {"Authorization": f"Bearer {t}"}

            worker_t = await token("worker", "floor2026")
            hod_t = await token("hod", "hod2026")

            r = await ac.get("/ai/submission-summary",
                             params={"kind": "staged-issue", "ref_id": iid},
                             headers=H(worker_t))
            check("worker (lvl 0) → 403 on staged-issue summary",
                  r.status_code == 403, f"got {r.status_code}")
            r = await ac.get("/ai/submission-summary",
                             params={"kind": "nope", "ref_id": 1}, headers=H(hod_t))
            check("unknown submission kind → 404", r.status_code == 404,
                  f"got {r.status_code}")
            r = await ac.get("/ai/submission-summary",
                             params={"kind": "staged-issue", "ref_id": 99999999},
                             headers=H(hod_t))
            check("unknown staged-issue ref → 404", r.status_code == 404,
                  f"got {r.status_code}")

            r = await ac.get("/ai/submission-summary",
                             params={"kind": "staged-issue", "ref_id": iid},
                             headers=H(hod_t))
            j = r.json() if r.status_code == 200 else {}
            check("staged-issue summary → 200 with summary/tone/source/facts",
                  r.status_code == 200 and j.get("summary")
                  and j.get("tone") in ("success", "info", "warning", "error")
                  and j.get("source") in ("ai", "deterministic")
                  and j.get("facts", {}).get("kind") == "staged-issue",
                  f"got {r.status_code} {str(j)[:140]}")
            check("staged-issue facts carry the 30d stats block",
                  isinstance(j.get("facts", {}).get("stats_30d", {}).get("mean_issue_qty"),
                             (int, float)), str(j.get("facts", {}))[:140])

            r = await ac.get("/ai/submission-summary",
                             params={"kind": "xsite", "ref_id": rid}, headers=H(hod_t))
            j = r.json() if r.status_code == 200 else {}
            check("xsite summary → 200 with depletion facts",
                  r.status_code == 200 and j.get("summary")
                  and j.get("facts", {}).get("target_site") == "HQ"
                  and "days_cover_after" in j.get("facts", {}),
                  f"got {r.status_code} {str(j)[:140]}")
    finally:
        async with SessionLocal() as s:  # cleanup (incl. cached summaries)
            await s.execute(delete(jobs).where(
                jobs.c["kind"] == "submission_summary",
                jobs.c["payload_json"].in_([
                    f'{{"kind": "staged-issue", "ref": {iid}}}',
                    f'{{"kind": "xsite", "ref": {rid}}}'])))
            await s.execute(delete(pi).where(pi.c["id"] == iid))
            await s.execute(delete(req).where(req.c["id"] == rid))
            await s.commit()


async def test_bulk_entry():
    """Phase 1 — bulk issue staging + item snapshot. SK-gated; atomic (a bad or
    invalid row stages nothing); snapshot returns ledger-derived stock and a
    30-point trend. Synthetic staged rows cleaned up in finally."""
    from sqlalchemy import delete

    pi = ledger._MD.tables["pending_issues"]
    TAG = "svc-bulk-test"
    transport = ASGITransport(app=app)
    try:
        async with AsyncClient(transport=transport, base_url="http://svc") as ac:
            _ip = {"X-Real-IP": "203.0.113.71"}

            async def token(u, p):
                r = await ac.post("/auth/login", headers=_ip,
                                  json={"username": u, "password": p})
                return r.json().get("access_token")

            def H(t):
                return {"Authorization": f"Bearer {t}"}

            sk_t = await token("worker", "floor2026")
            hod_t = await token("hod", "hod2026")

            inv = (await ac.get("/inventory", params={"limit": 5}, headers=H(sk_t))).json()
            items = inv.get("items", [])
            check("bulk: SK sees inventory to pick from", len(items) > 0, str(inv)[:120])
            if not items:
                return
            sap = str(items[0]["SAP_Code"])
            site = items[0].get("Site_ID") or "CNCEC"

            r = await ac.post("/entry/bulk", headers=H(hod_t), json={
                "kind": "consumption",
                "rows": [{"Date": "2026-07-08", "SAP_Code": sap, "Quantity": 1, "Site_ID": site}]})
            check("bulk: non-SK (hod) → 403", r.status_code == 403, f"got {r.status_code}")

            r = await ac.post("/entry/bulk", headers=H(sk_t), json={
                "kind": "consumption",
                "rows": [{"Date": "2026-07-08", "SAP_Code": sap, "Quantity": 0, "Site_ID": site}]})
            check("bulk: invalid row (qty 0) → 422, nothing staged",
                  r.status_code == 422, f"got {r.status_code}")

            r = await ac.post("/entry/bulk", headers=H(sk_t), json={
                "kind": "consumption",
                "rows": [{"Date": "2026-07-08", "SAP_Code": "__nope__", "Quantity": 1, "Site_ID": site}]})
            check("bulk: unknown SAP → 404", r.status_code == 404, f"got {r.status_code}")

            rows = [{"Date": "2026-07-08", "SAP_Code": sap, "Quantity": 1.5, "Site_ID": site, "Remarks": TAG},
                    {"Date": "2026-07-08", "SAP_Code": sap, "Quantity": 2.5, "Site_ID": site, "Remarks": TAG}]
            r = await ac.post("/entry/bulk", headers=H(sk_t),
                              json={"kind": "consumption", "rows": rows})
            j = r.json() if r.status_code == 201 else {}
            check("bulk: 2 issue lines → 201 staged=2",
                  r.status_code == 201 and j.get("staged") == 2
                  and len(j.get("pending_ids", [])) == 2, f"got {r.status_code} {str(j)[:140]}")

            r = await ac.get(f"/entry/snapshot/{sap}", params={"site_id": site}, headers=H(sk_t))
            j = r.json() if r.status_code == 200 else {}
            check("snapshot → 200 with numeric current_stock",
                  r.status_code == 200 and isinstance(j.get("current_stock"), (int, float)),
                  f"got {r.status_code} {str(j)[:140]}")
            check("snapshot → 30-point trend", len(j.get("trend", [])) == 30,
                  str(len(j.get("trend", []))))

            r = await ac.get(f"/entry/snapshot/{sap}", params={"site_id": site}, headers=H(hod_t))
            check("snapshot: non-SK (hod) → 403", r.status_code == 403, f"got {r.status_code}")
    finally:
        async with SessionLocal() as s:
            await s.execute(delete(pi).where(pi.c["Remarks"] == TAG))
            await s.commit()


async def test_reschedule():
    """Phase 4 H7 — reschedule workflow: HOD raises → Logistics decides → the
    approved date is pushed onto the PO. Role-gated; dup-guarded; idempotent
    decide. Synthetic PO + request cleaned up in finally."""
    from sqlalchemy import delete, insert, select as _sel

    po = ledger._MD.tables["purchase_orders"]
    rr = ledger._MD.tables["po_reschedule_requests"]
    PO = "PO-SVC-RESCHED"
    async with SessionLocal() as s:
        await s.execute(delete(rr).where(rr.c["PO_Number"] == PO))
        await s.execute(delete(po).where(po.c["PO_Number"] == PO))
        await s.execute(insert(po).values(PO_Number=PO, Site_ID="CNCEC", status="open",
                                          Expected_Delivery="2026-08-01", created_by="svc"))
        await s.commit()
    try:
        transport = ASGITransport(app=app)
        async with AsyncClient(transport=transport, base_url="http://svc") as ac:
            _ip = {"X-Real-IP": "203.0.113.81"}

            async def token(u, p):
                r = await ac.post("/auth/login", headers=_ip, json={"username": u, "password": p})
                return r.json().get("access_token")

            def H(t):
                return {"Authorization": f"Bearer {t}"}

            worker_t = await token("worker", "floor2026")
            hod_t = await token("hod", "hod2026")
            admin_t = await token("admin", "admin2026")

            r = await ac.post("/hod/reschedule", headers=H(worker_t),
                              json={"po_number": PO, "requested_date": "2026-09-15", "reason": "x"})
            check("reschedule: worker (lvl0) → 403 on raise", r.status_code == 403, f"got {r.status_code}")

            r = await ac.post("/hod/reschedule", headers=H(hod_t),
                              json={"po_number": PO, "requested_date": "2026-09-15", "reason": "vendor delay"})
            j = r.json() if r.status_code == 200 else {}
            rid = j.get("id")
            check("reschedule: HOD raises → 200 with id", r.status_code == 200 and rid, f"got {r.status_code} {str(j)[:120]}")

            r = await ac.post("/hod/reschedule", headers=H(hod_t),
                              json={"po_number": PO, "requested_date": "2026-10-01", "reason": "again"})
            check("reschedule: duplicate pending → 409", r.status_code == 409, f"got {r.status_code}")

            r = await ac.get("/logistics/reschedules", params={"status": "pending"}, headers=H(admin_t))
            items = r.json().get("items", []) if r.status_code == 200 else []
            check("reschedule: logistics lists the pending request",
                  any(it.get("id") == rid for it in items), f"got {r.status_code} {len(items)} items")

            r = await ac.get("/logistics/reschedules", headers=H(worker_t))
            check("reschedule: worker → 403 on logistics list", r.status_code == 403, f"got {r.status_code}")

            r = await ac.post(f"/logistics/reschedules/{rid}/decide", headers=H(admin_t),
                              json={"action": "approve"})
            j = r.json() if r.status_code == 200 else {}
            check("reschedule: approve → 200 decided=approved new_date set",
                  r.status_code == 200 and j.get("decided") == "approved" and j.get("new_date") == "2026-09-15",
                  f"got {r.status_code} {str(j)[:120]}")

            async with SessionLocal() as s:
                nd = (await s.execute(_sel(po.c["Expected_Delivery"]).where(po.c["PO_Number"] == PO))).scalar_one()
            check("reschedule: PO Expected_Delivery pushed to the new date", nd == "2026-09-15", f"got {nd}")

            r = await ac.post(f"/logistics/reschedules/{rid}/decide", headers=H(admin_t),
                              json={"action": "approve"})
            check("reschedule: re-decide → 409 (already approved)", r.status_code == 409, f"got {r.status_code}")
    finally:
        async with SessionLocal() as s:
            await s.execute(delete(rr).where(rr.c["PO_Number"] == PO))
            await s.execute(delete(po).where(po.c["PO_Number"] == PO))
            await s.commit()


async def test_force_close():
    """Phase 4 H8 — force-close a PO (reason required) + 24h undo restores the
    prior state. Role-gated; idempotent. Synthetic PO + closure cleaned up."""
    from sqlalchemy import delete, insert, select as _sel

    po = ledger._MD.tables["purchase_orders"]
    fc = ledger._MD.tables["po_force_closures"]
    PO = "PO-SVC-FC"
    async with SessionLocal() as s:
        await s.execute(delete(fc).where(fc.c["target_ref"] == PO))
        await s.execute(delete(po).where(po.c["PO_Number"] == PO))
        await s.execute(insert(po).values(PO_Number=PO, Site_ID="CNCEC", status="open", created_by="svc"))
        await s.commit()
    try:
        transport = ASGITransport(app=app)
        async with AsyncClient(transport=transport, base_url="http://svc") as ac:
            _ip = {"X-Real-IP": "203.0.113.82"}

            async def token(u, p):
                r = await ac.post("/auth/login", headers=_ip, json={"username": u, "password": p})
                return r.json().get("access_token")

            def H(t):
                return {"Authorization": f"Bearer {t}"}

            worker_t = await token("worker", "floor2026")
            admin_t = await token("admin", "admin2026")

            r = await ac.post("/logistics/force-close", headers=H(worker_t),
                              json={"target_type": "po", "target_ref": PO, "reason": "x"})
            check("force-close: worker (lvl0) → 403", r.status_code == 403, f"got {r.status_code}")

            r = await ac.post("/logistics/force-close", headers=H(admin_t),
                              json={"target_type": "po", "target_ref": PO, "reason": ""})
            check("force-close: empty reason → 409", r.status_code == 409, f"got {r.status_code}")

            r = await ac.post("/logistics/force-close", headers=H(admin_t),
                              json={"target_type": "po", "target_ref": PO, "reason": "duplicate order"})
            j = r.json() if r.status_code == 201 else {}
            cid = j.get("id")
            check("force-close: PO → 201 closed", r.status_code == 201 and cid, f"got {r.status_code} {str(j)[:120]}")

            async with SessionLocal() as s:
                st = (await s.execute(_sel(po.c["status"]).where(po.c["PO_Number"] == PO))).scalar_one()
            check("force-close: PO status = force_closed", st == "force_closed", f"got {st}")

            r = await ac.post("/logistics/force-close", headers=H(admin_t),
                              json={"target_type": "po", "target_ref": PO, "reason": "again"})
            check("force-close: double-close → 409", r.status_code == 409, f"got {r.status_code}")

            r = await ac.get("/logistics/force-closures", headers=H(admin_t))
            items = r.json().get("items", []) if r.status_code == 200 else []
            mine = next((it for it in items if it.get("id") == cid), None)
            check("force-close: log lists it with numeric age_hours + not reverted",
                  mine is not None and isinstance(mine.get("age_hours"), (int, float))
                  and mine.get("reverted_at") is None, str(mine)[:140])

            r = await ac.post(f"/logistics/force-close/{cid}/undo", headers=H(admin_t))
            check("force-close: undo → 200 reverted", r.status_code == 200 and r.json().get("reverted"),
                  f"got {r.status_code}")

            async with SessionLocal() as s:
                st = (await s.execute(_sel(po.c["status"]).where(po.c["PO_Number"] == PO))).scalar_one()
            check("force-close: undo restored PO status = open", st == "open", f"got {st}")

            r = await ac.post(f"/logistics/force-close/{cid}/undo", headers=H(admin_t))
            check("force-close: re-undo → 409 (already undone)", r.status_code == 409, f"got {r.status_code}")
    finally:
        async with SessionLocal() as s:
            await s.execute(delete(fc).where(fc.c["target_ref"] == PO))
            await s.execute(delete(po).where(po.c["PO_Number"] == PO))
            await s.commit()


async def test_manual_po():
    """Phase 4 — manual PO creation (free-text lines, prices, unlisted PR) +
    vendor master round-trip. Role-gated; unique PO; total computed. Cleaned up."""
    from sqlalchemy import delete, select as _sel

    po = ledger._MD.tables["purchase_orders"]
    poi = ledger._MD.tables["po_items"]
    ven = ledger._MD.tables["vendors"]
    PO = "PO-SVC-MANUAL"
    VC = "V-SVC-TEST"
    async with SessionLocal() as s:
        await s.execute(delete(poi).where(poi.c["PO_Number"] == PO))
        await s.execute(delete(po).where(po.c["PO_Number"] == PO))
        await s.execute(delete(ven).where(ven.c["Vendor_Code"] == VC))
        await s.commit()
    try:
        transport = ASGITransport(app=app)
        async with AsyncClient(transport=transport, base_url="http://svc") as ac:
            _ip = {"X-Real-IP": "203.0.113.83"}

            async def token(u, p):
                r = await ac.post("/auth/login", headers=_ip, json={"username": u, "password": p})
                return r.json().get("access_token")

            def H(t):
                return {"Authorization": f"Bearer {t}"}

            worker_t = await token("worker", "floor2026")
            admin_t = await token("admin", "admin2026")

            body = {"po_number": PO, "site_id": "CNCEC", "pr_number": "PR-UNLISTED-X",
                    "vendor_code": VC, "vendor_name": "Svc Vendor",
                    "inco_terms": "FOB", "payment_terms": "Net 30",
                    "lines": [
                        {"Material_Code": "FREE-1", "Description": "free line 1", "Qty": 3, "Unit_Price": 10},
                        {"Description": "unlisted line 2", "Qty": 2, "Unit_Price": 25},
                    ]}

            r = await ac.post("/logistics/pos/manual", headers=H(worker_t), json=body)
            check("manual-po: worker (lvl0) → 403", r.status_code == 403, f"got {r.status_code}")

            r = await ac.post("/logistics/pos/manual", headers=H(admin_t),
                              json={**body, "lines": []})
            check("manual-po: empty lines → 422", r.status_code == 422, f"got {r.status_code}")

            r = await ac.post("/logistics/pos/manual", headers=H(admin_t), json=body)
            j = r.json() if r.status_code == 201 else {}
            check("manual-po: create → 201 total=80 lines=2",
                  r.status_code == 201 and j.get("lines") == 2 and j.get("total") == 80.0,
                  f"got {r.status_code} {str(j)[:120]}")

            r = await ac.post("/logistics/pos/manual", headers=H(admin_t), json=body)
            check("manual-po: duplicate PO → 409", r.status_code == 409, f"got {r.status_code}")

            async with SessionLocal() as s:
                n = (await s.execute(_sel(func.count()).select_from(poi).where(poi.c["PO_Number"] == PO))).scalar_one()
                tot = (await s.execute(_sel(po.c["Total_Amount"]).where(po.c["PO_Number"] == PO))).scalar_one()
            check("manual-po: 2 po_items persisted + header total 80", n == 2 and float(tot) == 80.0, f"n={n} tot={tot}")

            # vendor master round-trip (the picker's inline-add uses this path).
            r = await ac.post("/vendors", headers=H(admin_t), json={
                "Vendor_Code": VC, "Vendor_Name": "Svc Vendor",
                "Default_Inco_Terms": "FOB", "Default_Payment_Terms": "Net 30", "status": "active"})
            check("vendor: admin create → 2xx", r.status_code < 300, f"got {r.status_code} {r.text[:120]}")
            r = await ac.get("/vendors", params={"limit": 500}, headers=H(admin_t))
            items = r.json().get("items", []) if r.status_code == 200 else []
            v = next((it for it in items if it.get("Vendor_Code") == VC), None)
            check("vendor: appears in list with defaults",
                  v is not None and v.get("Default_Inco_Terms") == "FOB", str(v)[:120])
    finally:
        async with SessionLocal() as s:
            await s.execute(delete(poi).where(poi.c["PO_Number"] == PO))
            await s.execute(delete(po).where(po.c["PO_Number"] == PO))
            await s.execute(delete(ven).where(ven.c["Vendor_Code"] == VC))
            await s.commit()


async def test_ratelimit_ip():
    """Phase I-B — the rate-limiter resolves the client IP in priority order:
    CF-Connecting-IP (Cloudflare Tunnel) → X-Real-IP (nginx) → TCP peer. Without
    the CF header, every tunnelled tester would share one bucket."""
    from .ratelimit import _client_ip

    class _Stub:
        def __init__(self, headers, host="9.9.9.9"):
            self.headers = headers
            self.client = type("C", (), {"host": host})()

    check("ratelimit: CF-Connecting-IP wins over X-Real-IP",
          _client_ip(_Stub({"cf-connecting-ip": "1.1.1.1", "x-real-ip": "2.2.2.2"})) == "1.1.1.1")
    check("ratelimit: X-Real-IP used when no CF header",
          _client_ip(_Stub({"x-real-ip": "2.2.2.2"})) == "2.2.2.2")
    check("ratelimit: falls back to the TCP peer",
          _client_ip(_Stub({}, host="3.3.3.3")) == "3.3.3.3")


async def test_reporting_dashboard():
    """Phase 5 — PR-status report + dashboard metrics + admin system-overview.
    Read-only over the mirror; just shape + role-gate assertions (no fixtures)."""
    transport = ASGITransport(app=app)
    async with AsyncClient(transport=transport, base_url="http://svc") as ac:
        _ip = {"X-Real-IP": "203.0.113.84"}

        async def token(u, p):
            r = await ac.post("/auth/login", headers=_ip, json={"username": u, "password": p})
            return r.json().get("access_token")

        def H(t):
            return {"Authorization": f"Bearer {t}"}

        worker_t = await token("worker", "floor2026")
        admin_t = await token("admin", "admin2026")

        # PR-status report registered + downloadable.
        r = await ac.get("/reports", headers=H(admin_t))
        keys = [it.get("key") for it in r.json().get("reports", [])] if r.status_code == 200 else []
        check("reports: pr-status registered", "pr-status" in keys, str(keys)[:160])
        r = await ac.get("/reports/pr-status", params={"format": "csv"}, headers=H(admin_t))
        check("reports: pr-status downloads (csv, non-empty)",
              r.status_code == 200 and len(r.content) > 0, f"got {r.status_code} len={len(r.content)}")

        # Dashboard metrics — supervisor+; worker (lvl0) blocked.
        r = await ac.get("/dashboard/metrics", headers=H(worker_t))
        check("dashboard: worker (lvl0) → 403", r.status_code == 403, f"got {r.status_code}")
        r = await ac.get("/dashboard/metrics", headers=H(admin_t))
        j = r.json() if r.status_code == 200 else {}
        check("dashboard: metrics → 200 w/ valuation_total + 3 chart series",
              r.status_code == 200 and isinstance(j.get("valuation_total"), (int, float))
              and isinstance(j.get("stock_vs_min"), list) and isinstance(j.get("top_consumed"), list)
              and isinstance(j.get("burn_forecast"), list), f"got {r.status_code} {str(j)[:120]}")

        # Admin system-overview — admin only.
        r = await ac.get("/admin/system-overview", headers=H(worker_t))
        check("system-overview: worker → 403", r.status_code == 403, f"got {r.status_code}")
        r = await ac.get("/admin/system-overview", headers=H(admin_t))
        j = r.json() if r.status_code == 200 else {}
        check("system-overview: → 200 w/ db_size + txn total + valuation_by_site",
              r.status_code == 200 and j.get("db_size")
              and isinstance(j.get("transactions", {}).get("total"), int)
              and isinstance(j.get("valuation_by_site"), list), f"got {r.status_code} {str(j)[:120]}")


async def test_dn_approval():
    """Phase 6 — DN two-stage approval: WH submit → Logistics decide → HOD decide
    → ship gated to hod_approved. Role-gated; state-checked. Synthetic DN cleaned up."""
    from sqlalchemy import delete, insert, select as _sel

    dn = ledger._MD.tables["delivery_notes"]
    dni = ledger._MD.tables["dn_items"]
    DN = "DN-SVC-APPROVAL"
    async with SessionLocal() as s:
        await s.execute(delete(dni).where(dni.c["DN_Number"] == DN))
        await s.execute(delete(dn).where(dn.c["DN_Number"] == DN))
        await s.execute(insert(dn).values(DN_Number=DN, PO_Number="PO-SVC-DN", Warehouse_ID="HQ",
                                          Site_ID="CNCEC", status="draft", created_by="svc"))
        await s.execute(insert(dni).values(DN_Number=DN, po_item_id=1, Material_Code="M1", Qty=1.0, status="pending"))
        await s.commit()
    try:
        transport = ASGITransport(app=app)
        async with AsyncClient(transport=transport, base_url="http://svc") as ac:
            _ip = {"X-Real-IP": "203.0.113.85"}

            async def token(u, p):
                r = await ac.post("/auth/login", headers=_ip, json={"username": u, "password": p})
                return r.json().get("access_token")

            def H(t):
                return {"Authorization": f"Bearer {t}"}

            worker_t = await token("worker", "floor2026")
            hod_t = await token("hod", "hod2026")
            admin_t = await token("admin", "admin2026")

            async def dn_status():
                async with SessionLocal() as s:
                    return (await s.execute(_sel(dn.c["status"]).where(dn.c["DN_Number"] == DN))).scalar_one()

            r = await ac.post(f"/warehouse/dns/{DN}/submit", headers=H(admin_t))
            check("dn: submit → pending_logistics",
                  r.status_code == 200 and await dn_status() == "pending_logistics", f"got {r.status_code}")

            r = await ac.post(f"/warehouse/dns/{DN}/ship", headers=H(admin_t))
            check("dn: ship before approval → 409", r.status_code == 409, f"got {r.status_code}")

            r = await ac.post(f"/logistics/dns/{DN}/decide", headers=H(worker_t), json={"action": "approve"})
            check("dn: worker → 403 on logistics decide", r.status_code == 403, f"got {r.status_code}")

            r = await ac.post(f"/logistics/dns/{DN}/decide", headers=H(admin_t), json={"action": "approve"})
            check("dn: logistics approve → pending_hod",
                  r.status_code == 200 and await dn_status() == "pending_hod", f"got {r.status_code}")

            r = await ac.post(f"/hod/dns/{DN}/decide", headers=H(worker_t), json={"action": "approve"})
            check("dn: worker → 403 on HOD decide", r.status_code == 403, f"got {r.status_code}")

            r = await ac.post(f"/hod/dns/{DN}/decide", headers=H(hod_t), json={"action": "approve"})
            check("dn: HOD approve → hod_approved",
                  r.status_code == 200 and await dn_status() == "hod_approved", f"got {r.status_code}")

            r = await ac.post(f"/warehouse/dns/{DN}/ship", headers=H(admin_t))
            check("dn: ship after HOD approval → in_transit",
                  r.status_code == 200 and await dn_status() == "in_transit", f"got {r.status_code}")
    finally:
        async with SessionLocal() as s:
            await s.execute(delete(dni).where(dni.c["DN_Number"] == DN))
            await s.execute(delete(dn).where(dn.c["DN_Number"] == DN))
            await s.commit()


async def test_supervisor_parity():
    """Phase 6 — supervisor parity: Intent-vs-Actual JSON, live stock-check, and
    cancel-while-pending (own + not-own guards). Synthetic SMRs cleaned up."""
    from sqlalchemy import delete, insert, select as _sel

    smr = ledger._MD.tables["supervisor_material_requests"]

    def _mk(request_no, by, status="pending_sk"):
        return insert(smr).values(request_no=request_no, Site_ID="CNCEC", Worker_ID="W1",
                                  Worker_Name="Test Worker", Job_Tank_Place="T1",
                                  Old_PPE_Returned=1, requested_by=by, status=status
                                  ).returning(smr.c["id"])
    async with SessionLocal() as s:
        await s.execute(delete(smr).where(smr.c["request_no"].like("SMR-SVC-%")))
        mine = (await s.execute(_mk("SMR-SVC-MINE", "supervisor"))).scalar_one()
        other = (await s.execute(_mk("SMR-SVC-OTHER", "someone_else"))).scalar_one()
        await s.commit()
    try:
        transport = ASGITransport(app=app)
        async with AsyncClient(transport=transport, base_url="http://svc") as ac:
            _ip = {"X-Real-IP": "203.0.113.86"}

            async def token(u, p):
                r = await ac.post("/auth/login", headers=_ip, json={"username": u, "password": p})
                return r.json().get("access_token")

            def H(t):
                return {"Authorization": f"Bearer {t}"}

            sup_t = await token("supervisor", "super2026")
            worker_t = await token("worker", "floor2026")

            r = await ac.get("/requests/intent-vs-actual", headers=H(sup_t))
            j = r.json() if r.status_code == 200 else {}
            check("supervisor: intent-vs-actual → 200 w/ columns + rows",
                  r.status_code == 200 and isinstance(j.get("columns"), list)
                  and isinstance(j.get("rows"), list), f"got {r.status_code} {str(j)[:120]}")
            r = await ac.get("/requests/intent-vs-actual", headers=H(worker_t))
            check("supervisor: intent-vs-actual → 403 for store_keeper", r.status_code == 403, f"got {r.status_code}")

            r = await ac.get("/requests/stock/1001", headers=H(sup_t))
            j = r.json() if r.status_code == 200 else {}
            check("supervisor: stock-check → 200 w/ numeric current_stock",
                  r.status_code == 200 and isinstance(j.get("current_stock"), (int, float)),
                  f"got {r.status_code} {str(j)[:120]}")

            r = await ac.post(f"/requests/{other}/cancel", headers=H(sup_t))
            check("supervisor: cancel someone else's → 409", r.status_code == 409, f"got {r.status_code}")

            r = await ac.post(f"/requests/{mine}/cancel", headers=H(sup_t))
            check("supervisor: cancel own pending → 200", r.status_code == 200 and r.json().get("cancelled"),
                  f"got {r.status_code}")
            async with SessionLocal() as s:
                st = (await s.execute(_sel(smr.c["status"]).where(smr.c["id"] == mine))).scalar_one()
            check("supervisor: cancelled status persisted", st == "cancelled", f"got {st}")

            r = await ac.post(f"/requests/{mine}/cancel", headers=H(sup_t))
            check("supervisor: re-cancel → 409 (already cancelled)", r.status_code == 409, f"got {r.status_code}")
    finally:
        async with SessionLocal() as s:
            await s.execute(delete(smr).where(smr.c["request_no"].like("SMR-SVC-%")))
            await s.commit()


async def test_entry_guards():
    """Phase 6 — receipt guards: MTC gate for Rubber materials + pack→base UoM
    conversion. Synthetic inventory/conversion/MTC rows cleaned up in finally."""
    from sqlalchemy import delete, insert, select as _sel

    inv = ledger._MD.tables["inventory"]
    uom = ledger._MD.tables["uom_conversions"]
    mtc = ledger._MD.tables["mtc_documents"]
    pr = ledger._MD.tables["pending_receipts"]
    RUB, UOMSAP = "SVC-RUBBER", "SVC-UOM"

    async def _cleanup():
        async with SessionLocal() as s:
            await s.execute(delete(pr).where(pr.c["SAP_Code"].in_([RUB, UOMSAP])))
            await s.execute(delete(mtc).where(mtc.c["SAP_Code"] == RUB))
            await s.execute(delete(uom).where(uom.c["SAP_Code"] == UOMSAP))
            await s.execute(delete(inv).where(inv.c["SAP_Code"].in_([RUB, UOMSAP])))
            await s.commit()

    await _cleanup()
    async with SessionLocal() as s:
        await s.execute(insert(inv).values(SAP_Code=RUB, Equipment_Description="Rubber sheet",
                                           UOM="m2", Category="Surface Shields"))
        await s.execute(insert(inv).values(SAP_Code=UOMSAP, Equipment_Description="Solvent",
                                           UOM="L", Category="Others"))
        await s.execute(insert(uom).values(SAP_Code=UOMSAP, Pack_UOM="Drum", Factor=200.0))
        await s.commit()
    try:
        transport = ASGITransport(app=app)
        async with AsyncClient(transport=transport, base_url="http://svc") as ac:
            _ip = {"X-Real-IP": "203.0.113.87"}

            async def token(u, p):
                r = await ac.post("/auth/login", headers=_ip, json={"username": u, "password": p})
                return r.json().get("access_token")

            def H(t):
                return {"Authorization": f"Bearer {t}"}

            sk_t = await token("worker", "floor2026")

            r = await ac.get(f"/entry/receipt-meta/{RUB}", headers=H(sk_t))
            check("guards: receipt-meta flags rubber", r.status_code == 200 and r.json().get("is_rubber") is True,
                  f"got {r.status_code} {r.text[:100]}")
            r = await ac.get(f"/entry/receipt-meta/{UOMSAP}", headers=H(sk_t))
            convs = r.json().get("conversions", []) if r.status_code == 200 else []
            check("guards: receipt-meta lists pack conversion",
                  any(c.get("Pack_UOM") == "Drum" and float(c.get("Factor")) == 200.0 for c in convs), str(convs)[:120])

            # Rubber receipt without an MTC → blocked.
            base = {"Date": "2026-07-09", "SAP_Code": RUB, "Quantity": 5, "Site_ID": "CNCEC"}
            r = await ac.post("/entry/receipts", headers=H(sk_t), json=base)
            check("guards: rubber receipt without MTC → 422", r.status_code == 422, f"got {r.status_code}")

            # Upload an MTC, then the receipt is accepted + linked.
            up = await ac.post("/entry/mtc", headers=H(sk_t),
                               files={"file": ("mtc.pdf", b"%PDF-1.4 test", "application/pdf")},
                               data={"sap_code": RUB, "site_id": "CNCEC", "mtc_number": "MTC-1"})
            mtc_id = up.json().get("id") if up.status_code == 201 else None
            check("guards: MTC upload → 201 w/ id", up.status_code == 201 and mtc_id, f"got {up.status_code}")
            r = await ac.post("/entry/receipts", headers=H(sk_t), json={**base, "mtc_document_id": mtc_id})
            pid = r.json().get("pending_id") if r.status_code == 201 else None
            check("guards: rubber receipt WITH MTC → 201", r.status_code == 201 and pid, f"got {r.status_code} {r.text[:100]}")
            async with SessionLocal() as s:
                linked = (await s.execute(_sel(mtc.c["pending_receipt_id"]).where(mtc.c["id"] == mtc_id))).scalar_one()
            check("guards: MTC linked to the staged receipt", linked == pid, f"got {linked} want {pid}")

            # UoM conversion: 2 Drum × 200 = 400 L stored on the staged receipt.
            r = await ac.post("/entry/receipts", headers=H(sk_t),
                              json={"Date": "2026-07-09", "SAP_Code": UOMSAP, "Quantity": 2,
                                    "Site_ID": "CNCEC", "entry_uom": "Drum"})
            pid2 = r.json().get("pending_id") if r.status_code == 201 else None
            check("guards: UoM receipt → 201", r.status_code == 201 and pid2, f"got {r.status_code}")
            async with SessionLocal() as s:
                q = (await s.execute(_sel(pr.c["Quantity"]).where(pr.c["id"] == pid2))).scalar_one()
            check("guards: pack→base converted (2 Drum × 200 = 400)", float(q) == 400.0, f"got {q}")

            # Unknown pack UoM → 422.
            r = await ac.post("/entry/receipts", headers=H(sk_t),
                              json={"Date": "2026-07-09", "SAP_Code": UOMSAP, "Quantity": 1,
                                    "Site_ID": "CNCEC", "entry_uom": "Pallet"})
            check("guards: unknown pack UoM → 422", r.status_code == 422, f"got {r.status_code}")
    finally:
        await _cleanup()


async def test_vendor_returns():
    """Deferred-MED — logistics vendor returns: raise-to-vendor reopens the PO
    line, over-return is blocked, list + close (idempotent). Fixtures cleaned up."""
    from sqlalchemy import delete, insert, select as _sel

    po = ledger._MD.tables["purchase_orders"]
    poi = ledger._MD.tables["po_items"]
    ret = ledger._MD.tables["po_returns"]
    PO = "PO-SVC-VR"
    async with SessionLocal() as s:
        await s.execute(delete(ret).where(ret.c["PO_Number"] == PO))
        await s.execute(delete(poi).where(poi.c["PO_Number"] == PO))
        await s.execute(delete(po).where(po.c["PO_Number"] == PO))
        await s.execute(insert(po).values(PO_Number=PO, Site_ID="CNCEC", status="delivered", created_by="svc"))
        lid = (await s.execute(insert(poi).values(PO_Number=PO, line_no=1, Material_Code="M1",
               Qty=10.0, Delivered_Qty=10.0, Returned_Qty=0.0, line_status="delivered"
               ).returning(poi.c["id"]))).scalar_one()
        await s.commit()
    try:
        transport = ASGITransport(app=app)
        async with AsyncClient(transport=transport, base_url="http://svc") as ac:
            _ip = {"X-Real-IP": "203.0.113.88"}

            async def token(u, p):
                r = await ac.post("/auth/login", headers=_ip, json={"username": u, "password": p})
                return r.json().get("access_token")

            def H(t):
                return {"Authorization": f"Bearer {t}"}

            worker_t = await token("worker", "floor2026")
            admin_t = await token("admin", "admin2026")

            body = {"po_number": PO, "po_item_id": lid, "qty": 4, "reason": "defective", "expected_resupply": "2026-08-01"}
            r = await ac.post("/logistics/vendor-returns", headers=H(worker_t), json=body)
            check("vendor-return: worker (lvl0) → 403", r.status_code == 403, f"got {r.status_code}")

            r = await ac.post("/logistics/vendor-returns", headers=H(admin_t),
                              json={**body, "qty": 15})
            check("vendor-return: over-return → 409", r.status_code == 409, f"got {r.status_code}")

            r = await ac.post("/logistics/vendor-returns", headers=H(admin_t), json=body)
            j = r.json() if r.status_code == 201 else {}
            rid = j.get("id")
            check("vendor-return: raise → 201 reopened_line",
                  r.status_code == 201 and rid and j.get("reopened_line") is True, f"got {r.status_code} {str(j)[:120]}")

            async with SessionLocal() as s:
                lrow = (await s.execute(_sel(poi.c["Returned_Qty"], poi.c["line_status"]).where(poi.c["id"] == lid))).first()
                pst = (await s.execute(_sel(po.c["status"]).where(po.c["PO_Number"] == PO))).scalar_one()
            check("vendor-return: PO line reopened (Returned_Qty=4, line open, PO partial)",
                  float(lrow[0]) == 4.0 and lrow[1] == "open" and pst == "partially_delivered",
                  f"returned={lrow[0]} line={lrow[1]} po={pst}")

            r = await ac.get("/logistics/vendor-returns", params={"status": "open"}, headers=H(admin_t))
            items = r.json().get("items", []) if r.status_code == 200 else []
            check("vendor-return: appears in the open list", any(it.get("id") == rid for it in items), str(len(items)))

            r = await ac.post(f"/logistics/vendor-returns/{rid}/close", headers=H(admin_t), json={"notes": "resupplied"})
            check("vendor-return: close → 200", r.status_code == 200 and r.json().get("closed"), f"got {r.status_code}")
            r = await ac.post(f"/logistics/vendor-returns/{rid}/close", headers=H(admin_t), json={})
            check("vendor-return: re-close → 409", r.status_code == 409, f"got {r.status_code}")
    finally:
        async with SessionLocal() as s:
            await s.execute(delete(ret).where(ret.c["PO_Number"] == PO))
            await s.execute(delete(poi).where(poi.c["PO_Number"] == PO))
            await s.execute(delete(po).where(po.c["PO_Number"] == PO))
            await s.commit()


async def test_pr_management():
    """Deferred-MED — HOD draft-PR line-edit + PR rename (draft-only, dup-guard,
    role-gated). Synthetic PR rows cleaned up in finally."""
    from sqlalchemy import delete, insert, select as _sel

    prm = ledger._MD.tables["pr_master"]

    def _line(pr, qty=5.0, status="site_draft"):
        return insert(prm).values(PR_Number=pr, SAP_Code="1001", Requested_Qty=qty,
                                  Site_ID="CNCEC", status="open", logistics_status=status
                                  ).returning(prm.c["id"])
    async with SessionLocal() as s:
        await s.execute(delete(prm).where(prm.c["PR_Number"].like("PR-SVC-%")))
        lid = (await s.execute(_line("PR-SVC-EDIT"))).scalar_one()
        await s.execute(_line("PR-SVC-EDIT"))
        await s.execute(_line("PR-SVC-OTHER"))
        await s.commit()
    try:
        transport = ASGITransport(app=app)
        async with AsyncClient(transport=transport, base_url="http://svc") as ac:
            _ip = {"X-Real-IP": "203.0.113.89"}

            async def token(u, p):
                r = await ac.post("/auth/login", headers=_ip, json={"username": u, "password": p})
                return r.json().get("access_token")

            def H(t):
                return {"Authorization": f"Bearer {t}"}

            worker_t = await token("worker", "floor2026")
            admin_t = await token("admin", "admin2026")

            r = await ac.patch(f"/hod/prs/lines/{lid}", headers=H(worker_t), json={"fields": {"Requested_Qty": 12}})
            check("pr-edit: worker (lvl0) → 403", r.status_code == 403, f"got {r.status_code}")

            r = await ac.patch(f"/hod/prs/lines/{lid}", headers=H(admin_t),
                               json={"fields": {"Requested_Qty": 12, "Supplier": "ACME"}})
            check("pr-edit: draft line edit → 200", r.status_code == 200 and r.json().get("updated"), f"got {r.status_code}")
            async with SessionLocal() as s:
                q = (await s.execute(_sel(prm.c["Requested_Qty"]).where(prm.c["id"] == lid))).scalar_one()
            check("pr-edit: qty persisted (12)", float(q) == 12.0, f"got {q}")

            r = await ac.patch(f"/hod/prs/lines/{lid}", headers=H(admin_t), json={"fields": {"Requested_Qty": 0}})
            check("pr-edit: qty 0 → 409", r.status_code == 409, f"got {r.status_code}")

            r = await ac.get("/hod/prs/PR-SVC-EDIT/lines", params={"site_id": "CNCEC"}, headers=H(admin_t))
            check("pr-edit: HOD lines endpoint → 200 with 2 lines",
                  r.status_code == 200 and len(r.json().get("items", [])) == 2, f"got {r.status_code}")

            r = await ac.post("/hod/prs/PR-SVC-EDIT/rename", headers=H(admin_t),
                              json={"site_id": "CNCEC", "new_pr": "PR-SVC-RENAMED"})
            check("pr-rename: draft rename → 200 lines=2",
                  r.status_code == 200 and r.json().get("lines") == 2, f"got {r.status_code} {r.text[:120]}")
            async with SessionLocal() as s:
                n = (await s.execute(_sel(func.count()).select_from(prm).where(prm.c["PR_Number"] == "PR-SVC-RENAMED"))).scalar_one()
            check("pr-rename: rows carry the new number", n == 2, f"got {n}")

            r = await ac.post("/hod/prs/PR-SVC-RENAMED/rename", headers=H(admin_t),
                              json={"site_id": "CNCEC", "new_pr": "PR-SVC-OTHER"})
            check("pr-rename: collide with existing PR → 409", r.status_code == 409, f"got {r.status_code}")

            # A submitted PR line cannot be edited.
            async with SessionLocal() as s:
                await s.execute(prm.update().where(prm.c["PR_Number"] == "PR-SVC-RENAMED")
                                .values(logistics_status="submitted"))
                sid = (await s.execute(_sel(prm.c["id"]).where(prm.c["PR_Number"] == "PR-SVC-RENAMED").limit(1))).scalar_one()
                await s.commit()
            r = await ac.patch(f"/hod/prs/lines/{sid}", headers=H(admin_t), json={"fields": {"Requested_Qty": 3}})
            check("pr-edit: submitted line → 409 (draft-only)", r.status_code == 409, f"got {r.status_code}")
    finally:
        async with SessionLocal() as s:
            await s.execute(delete(prm).where(prm.c["PR_Number"].like("PR-SVC-%")))
            await s.commit()


async def test_lot_lifecycle():
    """Deferred-MED — admin lot lifecycle: quarantine → release → dispose (terminal),
    role-gated + validated. Synthetic lot cleaned up in finally."""
    from sqlalchemy import delete, insert, select as _sel

    lots = ledger._MD.tables["lots"]
    LOT = "LOT-SVC-LC"
    async with SessionLocal() as s:
        await s.execute(delete(lots).where(lots.c["Lot_Number"] == LOT))
        lid = (await s.execute(insert(lots).values(Lot_Number=LOT, SAP_Code="1001",
               Site_ID="CNCEC", Received_Date="2026-07-01", Status="open"
               ).returning(lots.c["id"]))).scalar_one()
        await s.commit()
    try:
        transport = ASGITransport(app=app)
        async with AsyncClient(transport=transport, base_url="http://svc") as ac:
            _ip = {"X-Real-IP": "203.0.113.90"}

            async def token(u, p):
                r = await ac.post("/auth/login", headers=_ip, json={"username": u, "password": p})
                return r.json().get("access_token")

            def H(t):
                return {"Authorization": f"Bearer {t}"}

            worker_t = await token("worker", "floor2026")
            admin_t = await token("admin", "admin2026")

            async def lot_status():
                async with SessionLocal() as s:
                    return (await s.execute(_sel(lots.c["Status"]).where(lots.c["id"] == lid))).scalar_one()

            r = await ac.post(f"/admin/lots/{lid}/status", headers=H(worker_t), json={"status": "quarantined"})
            check("lot: worker → 403", r.status_code == 403, f"got {r.status_code}")

            r = await ac.post(f"/admin/lots/{lid}/status", headers=H(admin_t), json={"status": "bogus"})
            check("lot: invalid status → 422", r.status_code == 422, f"got {r.status_code}")

            r = await ac.get("/admin/lots", params={"status": "open"}, headers=H(admin_t))
            check("lot: admin list includes the open lot",
                  any(it.get("id") == lid for it in r.json().get("items", [])), f"got {r.status_code}")

            r = await ac.post(f"/admin/lots/{lid}/status", headers=H(admin_t), json={"status": "quarantined", "reason": "damp"})
            check("lot: quarantine → 200", r.status_code == 200 and await lot_status() == "quarantined", f"got {r.status_code}")

            r = await ac.post(f"/admin/lots/{lid}/status", headers=H(admin_t), json={"status": "open"})
            check("lot: release → open", r.status_code == 200 and await lot_status() == "open", f"got {r.status_code}")

            r = await ac.post(f"/admin/lots/{lid}/status", headers=H(admin_t), json={"status": "disposed", "reason": "expired"})
            check("lot: dispose → 200", r.status_code == 200 and await lot_status() == "disposed", f"got {r.status_code}")

            r = await ac.post(f"/admin/lots/{lid}/status", headers=H(admin_t), json={"status": "open"})
            check("lot: change disposed → 409 (terminal)", r.status_code == 409, f"got {r.status_code}")
    finally:
        async with SessionLocal() as s:
            await s.execute(delete(lots).where(lots.c["Lot_Number"] == LOT))
            await s.commit()


async def test_whatsapp_outbox():
    """Phase 7 — native WhatsApp outbox. Meta HTTP is MOCKED (no live calls):
    xsite>5 + FEFO-override triggers enqueue+send, report delivery uploads+sends,
    the admin console lists + retries. Synthetic rows cleaned up in finally."""
    import os as _o
    import json as _j
    from sqlalchemy import delete, insert, select as _sel
    import backend.api.services.whatsapp as wamod

    ob = ledger._MD.tables["whatsapp_outbox"]
    async with SessionLocal() as s:
        base_id = (await s.execute(_sel(func.coalesce(func.max(ob.c["id"]), 0)))).scalar_one()

    saved = (wamod._post_message, wamod._upload_media)
    _ENVK = ("WHATSAPP_ESCALATION_TO", "WHATSAPP_PHONE_NUMBER_ID", "WHATSAPP_TOKEN")
    prev_env = {k: _o.environ.get(k) for k in _ENVK}
    sent_payloads: list = []

    async def ok_post(payload):
        sent_payloads.append(payload)
        return {"ok": True, "message_id": "wamid.TEST"}

    async def ok_upload(blob, filename, mime):
        return {"ok": True, "media_id": "media.TEST"}

    wamod._post_message, wamod._upload_media = ok_post, ok_upload
    _o.environ["WHATSAPP_ESCALATION_TO"] = "15550001111"   # fallback recipient
    # Enable WhatsApp so dispatch()-based triggers actually send (HTTP is mocked
    # above, so no live Meta call is made regardless of these dummy creds).
    _o.environ["WHATSAPP_PHONE_NUMBER_ID"] = "svc-test-pnid"
    _o.environ["WHATSAPP_TOKEN"] = "svc-test-token"
    try:
        transport = ASGITransport(app=app)
        async with AsyncClient(transport=transport, base_url="http://svc") as ac:
            _ip = {"X-Real-IP": "203.0.113.91"}

            async def token(u, p):
                r = await ac.post("/auth/login", headers=_ip, json={"username": u, "password": p})
                return r.json().get("access_token")

            def H(t):
                return {"Authorization": f"Bearer {t}"}

            worker_t = await token("worker", "floor2026")
            admin_t = await token("admin", "admin2026")

            async def rows_for(event_key):
                async with SessionLocal() as s:
                    return [dict(m) for m in (await s.execute(_sel(
                        ob.c["id"], ob.c["status"], ob.c["event_key"], ob.c["meta_message_id"]
                    ).where((ob.c["event_key"] == event_key) & (ob.c["id"] > base_id)))).mappings().all()]

            # Console list (role-gated).
            r = await ac.get("/admin/whatsapp", headers=H(worker_t))
            check("whatsapp: worker → 403 on console", r.status_code == 403, f"got {r.status_code}")
            r = await ac.get("/admin/whatsapp", headers=H(admin_t))
            check("whatsapp: admin console → 200 with counts",
                  r.status_code == 200 and "items" in r.json() and "counts" in r.json(), f"got {r.status_code}")

            # Trigger 1 — cross-site request > 5 units escalates.
            r = await ac.post("/xsite", headers=H(admin_t),
                              json={"requesting_site": "HQ", "target_site": "CNCEC", "SAP_Code": "1001", "requested_qty": 10})
            check("whatsapp: xsite>5 accepted", r.status_code == 201, f"got {r.status_code}")
            esc = await rows_for("xsite_escalation")
            check("whatsapp: xsite escalation enqueued + sent",
                  any(x["status"] == "sent" and x["meta_message_id"] == "wamid.TEST" for x in esc), str(esc)[:160])
            # Alert sends must be TEMPLATE messages (deliverable outside the
            # 24h customer-service window), carrying the alert in body params.
            def _params_text(p):
                comps = p.get("template", {}).get("components", [])
                return " ".join(str(pp.get("text", "")) for c in comps
                                for pp in c.get("parameters", [])).lower()
            xtpl = next((p for p in sent_payloads if p.get("type") == "template"
                         and "cross-site" in _params_text(p)), None)
            check("whatsapp: alert payload is a template w/ body param",
                  xtpl is not None and bool(xtpl["template"]["name"]), str(sent_payloads)[:200])
            check("whatsapp: escalation uses the gi_critical_alert template",
                  any(p.get("template", {}).get("name") == "gi_critical_alert"
                      for p in sent_payloads), str([p.get("template", {}).get("name")
                                                    for p in sent_payloads])[:200])

            # Dual-write invariant: a dispatch()-driven trigger writes BOTH an
            # in-app notification AND a WhatsApp outbox row for the same event.
            appn = ledger._MD.tables["app_notifications"]
            async with SessionLocal() as s:
                base_app = (await s.execute(_sel(func.coalesce(func.max(appn.c["id"]), 0)))).scalar_one()
            r = await ac.post("/xsite", headers=H(admin_t),
                              json={"requesting_site": "HQ", "target_site": "CNCEC",
                                    "SAP_Code": "1001", "requested_qty": 3})
            check("whatsapp: xsite<=5 accepted", r.status_code == 201, f"got {r.status_code}")
            async with SessionLocal() as s:
                in_app = (await s.execute(_sel(func.count()).select_from(appn).where(
                    (appn.c["id"] > base_app) & (appn.c["event_key"] == "cross_site_requested")))).scalar_one()
            wa_rows = await rows_for("cross_site_requested")
            check("whatsapp: dispatch writes in-app + outbox for one event",
                  in_app >= 1 and any(x["status"] == "sent" for x in wa_rows),
                  f"in_app={in_app} wa={str(wa_rows)[:120]}")

            # Trigger 2 — FEFO override on an issue alerts the HOD.
            inv = (await ac.get("/inventory", params={"limit": 3}, headers=H(worker_t))).json().get("items", [])
            if inv:
                sap = str(inv[0]["SAP_Code"]); site = inv[0].get("Site_ID") or "CNCEC"
                r = await ac.post("/entry/consumption", headers=H(worker_t),
                                  json={"Date": "2026-07-09", "SAP_Code": sap, "Quantity": 1,
                                        "Site_ID": site, "FEFO_Override": "yes"})
                check("whatsapp: FEFO-override issue accepted", r.status_code == 201, f"got {r.status_code}")
                fo = await rows_for("fefo_override")
                check("whatsapp: FEFO override alert enqueued + sent",
                      any(x["status"] == "sent" for x in fo), str(fo)[:160])

            # Trigger 3 — report delivery (upload media + send document).
            r = await ac.post("/reports/pr-status/whatsapp", headers=H(admin_t),
                              json={"to": "15550009999", "format": "csv"})
            j = r.json() if r.status_code == 200 else {}
            check("whatsapp: report delivery → 200 sent",
                  r.status_code == 200 and j.get("status") == "sent", f"got {r.status_code} {str(j)[:120]}")
            r = await ac.post("/reports/pr-status/whatsapp", headers=H(admin_t), json={"to": ""})
            check("whatsapp: report delivery without recipient → 422", r.status_code == 422, f"got {r.status_code}")

            # Retry path — synthetic failed rows.
            async with SessionLocal() as s:
                fid = (await s.execute(insert(ob).values(to_number="15550002222", message_type="text",
                       body="retry me", payload_json=_j.dumps({"messaging_product": "whatsapp",
                       "to": "15550002222", "type": "text", "text": {"body": "hi"}}),
                       status="failed", event_key="svc_retry", attempts=1, created_by="svc"
                       ).returning(ob.c["id"]))).scalar_one()
                nid = (await s.execute(insert(ob).values(to_number=None, message_type="text",
                       body="no recipient", payload_json="{}", status="failed",
                       event_key="svc_retry", attempts=1, created_by="svc").returning(ob.c["id"]))).scalar_one()
                await s.commit()

            r = await ac.post(f"/admin/whatsapp/{fid}/retry", headers=H(admin_t))
            check("whatsapp: retry failed → 200 sent", r.status_code == 200 and r.json().get("status") == "sent", f"got {r.status_code}")
            r = await ac.post(f"/admin/whatsapp/{fid}/retry", headers=H(admin_t))
            check("whatsapp: retry an already-sent → 409", r.status_code == 409, f"got {r.status_code}")
            r = await ac.post(f"/admin/whatsapp/{nid}/retry", headers=H(admin_t))
            check("whatsapp: retry with no recipient → 409", r.status_code == 409, f"got {r.status_code}")
    finally:
        wamod._post_message, wamod._upload_media = saved
        for _k, _v in prev_env.items():
            if _v is None:
                _o.environ.pop(_k, None)
            else:
                _o.environ[_k] = _v
        async with SessionLocal() as s:
            await s.execute(delete(ob).where(ob.c["id"] > base_id))
            await s.commit()


async def test_email_outbox():
    """Phase 7b — native SMTP outbox. The SMTP boundary is MOCKED (no live
    connections): MTC-missing + vendor-return triggers enqueue+send, the admin
    Email Console lists + retries. Synthetic rows cleaned up in finally."""
    import os as _o
    from sqlalchemy import delete, insert, select as _sel
    import backend.api.services.emailer as emod

    eb = ledger._MD.tables["email_outbox"]
    inv = ledger._MD.tables["inventory"]
    po = ledger._MD.tables["purchase_orders"]
    poi = ledger._MD.tables["po_items"]
    ret = ledger._MD.tables["po_returns"]
    RUB, PO = "SVC-EM-RUBBER", "PO-SVC-EMAIL"

    async with SessionLocal() as s:
        base_id = (await s.execute(_sel(func.coalesce(func.max(eb.c["id"]), 0)))).scalar_one()

    saved_send = emod._smtp_send
    prev_to = _o.environ.get("EMAIL_LOGISTICS_TO")
    sent_mails: list = []

    async def ok_send(to, subject, body, cc=None):
        sent_mails.append({"to": to, "subject": subject})
        return {"ok": True}

    emod._smtp_send = ok_send
    _o.environ["EMAIL_LOGISTICS_TO"] = "logistics@svc.test"

    async def _cleanup():
        async with SessionLocal() as s:
            await s.execute(delete(eb).where(eb.c["id"] > base_id))
            await s.execute(delete(ret).where(ret.c["PO_Number"] == PO))
            await s.execute(delete(poi).where(poi.c["PO_Number"] == PO))
            await s.execute(delete(po).where(po.c["PO_Number"] == PO))
            await s.execute(delete(inv).where(inv.c["SAP_Code"] == RUB))
            await s.commit()

    await _cleanup()
    async with SessionLocal() as s:
        await s.execute(insert(inv).values(SAP_Code=RUB, Equipment_Description="Rubber liner",
                                           UOM="m2", Category="Surface Shields"))
        await s.execute(insert(po).values(PO_Number=PO, Site_ID="CNCEC", status="delivered", created_by="svc"))
        lid = (await s.execute(insert(poi).values(PO_Number=PO, line_no=1, Material_Code="M1",
               Qty=10.0, Delivered_Qty=10.0, Returned_Qty=0.0, line_status="delivered"
               ).returning(poi.c["id"]))).scalar_one()
        await s.commit()
    try:
        transport = ASGITransport(app=app)
        async with AsyncClient(transport=transport, base_url="http://svc") as ac:
            _ip = {"X-Real-IP": "203.0.113.92"}

            async def token(u, p):
                r = await ac.post("/auth/login", headers=_ip, json={"username": u, "password": p})
                return r.json().get("access_token")

            def H(t):
                return {"Authorization": f"Bearer {t}"}

            worker_t = await token("worker", "floor2026")
            admin_t = await token("admin", "admin2026")

            async def rows_for(event_key):
                async with SessionLocal() as s:
                    return [dict(m) for m in (await s.execute(_sel(
                        eb.c["id"], eb.c["status"], eb.c["to_email"]
                    ).where((eb.c["event_key"] == event_key) & (eb.c["id"] > base_id)))).mappings().all()]

            r = await ac.get("/admin/email", headers=H(worker_t))
            check("email: worker → 403 on console", r.status_code == 403, f"got {r.status_code}")
            r = await ac.get("/admin/email", headers=H(admin_t))
            check("email: admin console → 200 with counts",
                  r.status_code == 200 and "items" in r.json() and "counts" in r.json(), f"got {r.status_code}")

            # Trigger 1 — MTC-missing: blocked rubber receipt still 422s AND emails logistics.
            r = await ac.post("/entry/receipts", headers=H(worker_t),
                              json={"Date": "2026-07-09", "SAP_Code": RUB, "Quantity": 2, "Site_ID": "CNCEC"})
            check("email: rubber receipt w/o MTC still → 422", r.status_code == 422, f"got {r.status_code}")
            mm = await rows_for("mtc_missing")
            check("email: MTC-missing alert enqueued + sent to logistics inbox",
                  any(x["status"] == "sent" and x["to_email"] == "logistics@svc.test" for x in mm), str(mm)[:160])

            # Trigger 2 — vendor return raised → logistics email draft.
            r = await ac.post("/logistics/vendor-returns", headers=H(admin_t),
                              json={"po_number": PO, "po_item_id": lid, "qty": 3, "reason": "damaged"})
            check("email: vendor return raised → 201", r.status_code == 201, f"got {r.status_code}")
            vr = await rows_for("vendor_return")
            check("email: vendor-return draft enqueued + sent",
                  any(x["status"] == "sent" for x in vr), str(vr)[:160])
            check("email: SMTP mock actually received the draft",
                  any("Vendor return raised" in m["subject"] for m in sent_mails), str(sent_mails)[:160])

            # Retry paths.
            async with SessionLocal() as s:
                fid = (await s.execute(insert(eb).values(to_email="x@svc.test", subject="retry me",
                       body="hello", status="failed", event_key="svc_retry", attempts=1,
                       created_by="svc").returning(eb.c["id"]))).scalar_one()
                nid = (await s.execute(insert(eb).values(to_email=None, subject="no rcpt",
                       body="x", status="failed", event_key="svc_retry", attempts=1,
                       created_by="svc").returning(eb.c["id"]))).scalar_one()
                await s.commit()
            r = await ac.post(f"/admin/email/{fid}/retry", headers=H(admin_t))
            check("email: retry failed → 200 sent", r.status_code == 200 and r.json().get("status") == "sent", f"got {r.status_code}")
            r = await ac.post(f"/admin/email/{fid}/retry", headers=H(admin_t))
            check("email: retry an already-sent → 409", r.status_code == 409, f"got {r.status_code}")
            r = await ac.post(f"/admin/email/{nid}/retry", headers=H(admin_t))
            check("email: retry with no recipient → 409", r.status_code == 409, f"got {r.status_code}")
    finally:
        emod._smtp_send = saved_send
        if prev_to is None:
            _o.environ.pop("EMAIL_LOGISTICS_TO", None)
        else:
            _o.environ["EMAIL_LOGISTICS_TO"] = prev_to
        await _cleanup()


async def test_phone_otp():
    """Phase 7c — self-service phone change via WhatsApp OTP. Meta HTTP is MOCKED
    and _gen_otp is monkeypatched to a fixed code (no live send, deterministic
    verify). Covers request → wrong code → correct code → number saved, the
    single-active-code rule, and the admin no-OTP override. Cleans up in finally."""
    import os as _o
    from sqlalchemy import delete, select as _sel, update as _upd
    import backend.api.services.whatsapp as wamod
    import backend.api.auth as authmod

    ob = ledger._MD.tables["whatsapp_outbox"]
    otp = ledger._MD.tables["phone_otp"]
    users = ledger._MD.tables["users"]

    async with SessionLocal() as s:
        base_ob = (await s.execute(_sel(func.coalesce(func.max(ob.c["id"]), 0)))).scalar_one()
        prev_phone = (await s.execute(_sel(users.c["Phone_Number"])
                      .where(users.c["username"] == "worker"))).scalar_one_or_none()
        # Start from the FIRST-TIME state (no number on file) so the bootstrap
        # path — code to the NEW number — is what the first request exercises.
        await s.execute(_upd(users).where(users.c["username"] == "worker")
                        .values(Phone_Number=None))
        await s.commit()

    saved_post = wamod._post_message
    saved_gen = authmod._gen_otp
    _ENVK = ("WHATSAPP_PHONE_NUMBER_ID", "WHATSAPP_TOKEN")
    prev_env = {k: _o.environ.get(k) for k in _ENVK}
    otp_sends: list = []

    async def ok_post(payload):
        otp_sends.append(payload)
        return {"ok": True, "message_id": "wamid.OTP"}

    wamod._post_message = ok_post
    _o.environ["WHATSAPP_PHONE_NUMBER_ID"] = "svc-test-pnid"
    _o.environ["WHATSAPP_TOKEN"] = "svc-test-token"
    authmod._gen_otp = lambda: "654321"
    try:
        transport = ASGITransport(app=app)
        async with AsyncClient(transport=transport, base_url="http://svc") as ac:
            _ip = {"X-Real-IP": "203.0.113.93"}

            async def token(u, p):
                r = await ac.post("/auth/login", headers=_ip, json={"username": u, "password": p})
                return r.json().get("access_token")

            def H(t):
                return {"Authorization": f"Bearer {t}"}

            worker_t = await token("worker", "floor2026")
            admin_t = await token("admin", "admin2026")

            r = await ac.get("/auth/phone", headers=H(worker_t))
            check("otp: GET /auth/phone → 200", r.status_code == 200 and "phone_number" in r.json(),
                  f"got {r.status_code}")

            # A malformed number is rejected before any code is generated.
            r = await ac.post("/auth/phone/request-otp", headers={**H(worker_t), **_ip},
                              json={"new_number": "12"})
            check("otp: bad number → 422", r.status_code == 422, f"got {r.status_code}")

            # Request a code for a new number (dashes/spaces normalized away;
            # stored canonical = strict E.164 WITH the leading '+').
            r = await ac.post("/auth/phone/request-otp", headers={**H(worker_t), **_ip},
                              json={"new_number": "+1 555-123-4567"})
            check("otp: request-otp → 200 sent (first-time setup skips the old-number stage)",
                  r.status_code == 200 and r.json().get("sent") is True
                  and r.json().get("sent_to") == "new"
                  and r.json().get("stage") == "new",
                  f"got {r.status_code} {str(r.json())[:120]}")
            async with SessionLocal() as s:
                pend = (await s.execute(_sel(func.count()).select_from(otp).where(
                    (otp.c["username"] == "worker") & (otp.c["new_number"] == "+15551234567")
                    & otp.c["consumed_at"].is_(None)))).scalar_one()
            check("otp: one active code row for the new number (+E.164)", pend == 1, f"pend={pend}")
            # Meta payload `to` must be digits-only even though storage keeps '+'.
            check("otp: Meta payload strips the '+' (digits-only to)",
                  any(p.get("to") == "15551234567" for p in otp_sends), str(otp_sends)[:160])
            # The code must NOT appear in the outbox preview (redacted).
            async with SessionLocal() as s:
                rows = [dict(m) for m in (await s.execute(_sel(ob.c["body"], ob.c["status"])
                        .where((ob.c["event_key"] == "otp_verification") & (ob.c["id"] > base_ob)))).mappings().all()]
            check("otp: code sent via WhatsApp, preview redacted",
                  any(r["status"] == "sent" for r in rows) and all("654321" not in (r["body"] or "") for r in rows),
                  str(rows)[:160])

            # A second request supersedes the first (still exactly one active).
            r = await ac.post("/auth/phone/request-otp", headers={**H(worker_t), **_ip},
                              json={"new_number": "15551234567"})
            async with SessionLocal() as s:
                active = (await s.execute(_sel(func.count()).select_from(otp).where(
                    (otp.c["username"] == "worker") & otp.c["consumed_at"].is_(None)))).scalar_one()
            check("otp: re-request supersedes prior code (one active)", active == 1, f"active={active}")

            # Wrong code is rejected; the number is unchanged.
            r = await ac.post("/auth/phone/verify-otp", headers={**H(worker_t), **_ip},
                              json={"new_number": "15551234567", "code": "000000"})
            check("otp: wrong code → 400", r.status_code == 400, f"got {r.status_code}")

            # Correct code saves the new number (canonical +E.164).
            r = await ac.post("/auth/phone/verify-otp", headers={**H(worker_t), **_ip},
                              json={"new_number": "15551234567", "code": "654321"})
            check("otp: correct code → updated", r.status_code == 200 and r.json().get("phone_number") == "+15551234567",
                  f"got {r.status_code} {str(r.json())[:120]}")
            r = await ac.get("/auth/phone", headers=H(worker_t))
            check("otp: phone now reflects the verified number",
                  r.json().get("phone_number") == "+15551234567", str(r.json())[:120])

            # The code is single-use — verifying again fails.
            r = await ac.post("/auth/phone/verify-otp", headers={**H(worker_t), **_ip},
                              json={"new_number": "15551234567", "code": "654321"})
            check("otp: reused code → 404 (already consumed)", r.status_code == 404, f"got {r.status_code}")

            # ── Dual-OTP change flow: with a number ON FILE, code 1 goes to
            # the OLD registered number (stage='old'); verifying it does NOT
            # save anything — it dispatches code 2 to the NEW number
            # (stage='new'), and only THAT verification commits the change.
            otp_sends.clear()
            r = await ac.post("/auth/phone/request-otp", headers={**H(worker_t), **_ip},
                              json={"new_number": "+1 555 999 8888"})
            check("otp: change request → stage 'old', code dispatched to the OLD number only",
                  r.status_code == 200 and r.json().get("sent_to") == "current"
                  and r.json().get("stage") == "old"
                  and any(p.get("to") == "15551234567" for p in otp_sends)
                  and not any(p.get("to") == "15559998888" for p in otp_sends),
                  f"got {r.status_code} {str(r.json())[:100]} sends={str(otp_sends)[:120]}")
            r = await ac.post("/auth/phone/verify-otp", headers={**H(worker_t), **_ip},
                              json={"new_number": "+15559998888", "code": "654321"})
            j = r.json()
            check("otp: old-device code authorizes → NOT saved yet, code 2 → NEW number",
                  r.status_code == 200 and j.get("updated") is False
                  and j.get("stage") == "new" and j.get("sent") is True
                  and any(p.get("to") == "15559998888" for p in otp_sends),
                  f"got {r.status_code} {str(j)[:140]} sends={str(otp_sends)[:140]}")
            r = await ac.get("/auth/phone", headers=H(worker_t))
            check("otp: number UNCHANGED until the new device verifies",
                  r.json().get("phone_number") == "+15551234567", str(r.json())[:120])
            async with SessionLocal() as s:
                st = (await s.execute(_sel(otp.c["stage"]).where(
                    (otp.c["username"] == "worker") & otp.c["consumed_at"].is_(None))
                    .order_by(otp.c["id"].desc()).limit(1))).scalar_one()
            check("otp: active code row is stage='new'", st == "new", f"stage={st}")
            r = await ac.post("/auth/phone/verify-otp", headers={**H(worker_t), **_ip},
                              json={"new_number": "+15559998888", "code": "654321"})
            check("otp: new-device code commits the new number",
                  r.status_code == 200 and r.json().get("updated") is True
                  and r.json().get("phone_number") == "+15559998888",
                  f"got {r.status_code} {str(r.json())[:120]}")

            # ── Meta sandbox restriction (#131030) is handled gracefully ─────
            # The Graph error parser maps 131030 to a user-facing message while
            # keeping the raw Meta detail for the outbox row.
            import json as _json
            _sandbox_body = _json.dumps({"error": {
                "message": "(#131030) Recipient phone number not in allowed list",
                "code": 131030, "type": "OAuthException",
                "error_data": {"messaging_product": "whatsapp",
                               "details": "Recipient phone number not in allowed list: "
                                          "Add recipient phone number to recipient list "
                                          "and try again."}}})
            ge = wamod._graph_error(400, _sandbox_body)
            check("otp: _graph_error maps #131030 to a friendly message",
                  ge["ok"] is False and ge.get("code") == 131030
                  and "not whitelisted in the Meta Developer Sandbox" in ge["error"]
                  and "400 (#131030)" in ge["error"], str(ge)[:200])

            async def sandbox_post(payload):
                return wamod._graph_error(400, _sandbox_body)

            wamod._post_message = sandbox_post
            try:
                r = await ac.post("/auth/phone/request-otp", headers={**H(worker_t), **_ip},
                                  json={"new_number": "+1555777333"})
                j = r.json()
                check("otp: sandbox rejection → 200 (no generic 4xx/5xx), sent=false",
                      r.status_code == 200 and j.get("sent") is False,
                      f"got {r.status_code} {str(j)[:120]}")
                check("otp: sandbox rejection surfaces the descriptive message",
                      "not whitelisted in the Meta Developer Sandbox" in (j.get("error") or ""),
                      str(j)[:200])
                async with SessionLocal() as s:
                    fr = (await s.execute(_sel(ob.c["status"], ob.c["error"]).where(
                        (ob.c["id"] > base_ob) & (ob.c["status"] == "failed"))
                        .order_by(ob.c["id"].desc()).limit(1))).first()
                check("otp: sandbox rejection recorded FAILED in outbox with detail",
                      fr is not None and "Sandbox" in (fr[1] or "") and "131030" in (fr[1] or ""),
                      str(fr)[:200])
            finally:
                wamod._post_message = ok_post

            # Admin override sets any user's number directly — no OTP; input is
            # normalized to the same canonical +E.164.
            r = await ac.patch("/admin/users/worker", headers=H(admin_t),
                               json={"phone_number": "1555 999-0000"})
            check("otp: admin override → 200", r.status_code == 200, f"got {r.status_code}")
            r = await ac.get("/auth/phone", headers=H(worker_t))
            check("otp: admin override applied without OTP (+E.164)",
                  r.json().get("phone_number") == "+15559990000", str(r.json())[:120])
            r = await ac.patch("/admin/users/worker", headers=H(admin_t),
                               json={"phone_number": "12"})
            check("otp: admin override rejects a malformed number → 422",
                  r.status_code == 422, f"got {r.status_code}")
    finally:
        wamod._post_message = saved_post
        authmod._gen_otp = saved_gen
        for _k, _v in prev_env.items():
            if _v is None:
                _o.environ.pop(_k, None)
            else:
                _o.environ[_k] = _v
        async with SessionLocal() as s:
            await s.execute(delete(otp).where(otp.c["username"] == "worker"))
            await s.execute(delete(ob).where(ob.c["id"] > base_ob))
            await s.execute(_upd(users).where(users.c["username"] == "worker")
                            .values(Phone_Number=prev_phone))
            await s.commit()


async def test_returnables_notify():
    """Phase 1 UAT fixes — tool loans: (a) expected_return_time is stored as the
    LOCAL wall-clock the SK picked (no UTC shift), (b) loan/return/overdue all
    notify — in-app for the site SKs + WhatsApp direct to the borrower. Meta
    HTTP mocked; synthetic rows cleaned up in finally."""
    import datetime as _dtm
    import os as _o
    from sqlalchemy import delete, insert as _ins, select as _sel
    import backend.api.services.whatsapp as wamod

    ob = ledger._MD.tables["whatsapp_outbox"]
    rt = ledger._MD.tables["returnable_items"]
    appn = ledger._MD.tables["app_notifications"]
    BORROWER = "+966569233053"

    async with SessionLocal() as s:
        base_ob = (await s.execute(_sel(func.coalesce(func.max(ob.c["id"]), 0)))).scalar_one()
        base_rt = (await s.execute(_sel(func.coalesce(func.max(rt.c["id"]), 0)))).scalar_one()

    saved_post = wamod._post_message
    _ENVK = ("WHATSAPP_PHONE_NUMBER_ID", "WHATSAPP_TOKEN", "WHATSAPP_ESCALATION_TO")
    prev_env = {k: _o.environ.get(k) for k in _ENVK}
    sent: list = []

    async def ok_post(payload):
        sent.append(payload)
        return {"ok": True, "message_id": "wamid.LOAN"}

    wamod._post_message = ok_post
    _o.environ["WHATSAPP_PHONE_NUMBER_ID"] = "svc-test-pnid"
    _o.environ["WHATSAPP_TOKEN"] = "svc-test-token"
    _o.environ["WHATSAPP_ESCALATION_TO"] = "+15550001111"
    try:
        transport = ASGITransport(app=app)
        async with AsyncClient(transport=transport, base_url="http://svc") as ac:
            _ip = {"X-Real-IP": "203.0.113.94"}
            r = await ac.post("/auth/login", headers=_ip,
                              json={"username": "worker", "password": "floor2026"})
            H = {"Authorization": f"Bearer {r.json()['access_token']}"}

            async def wa_rows(event_key):
                async with SessionLocal() as s:
                    return [dict(m) for m in (await s.execute(_sel(
                        ob.c["to_number"], ob.c["status"], ob.c["payload_json"]
                    ).where((ob.c["event_key"] == event_key) & (ob.c["id"] > base_ob)))).mappings().all()]

            # (a) Local naive datetime survives verbatim — no UTC shift.
            r = await ac.post("/entry/returnables", headers=H,
                              json={"material_name": "SVC Torque Wrench", "borrower_name": "Svc Borrower",
                                    "borrower_phone": BORROWER, "qty": 1,
                                    "expected_return_time": "2026-07-15T21:59:00", "site_id": "CNCEC"})
            check("loan: created → 201", r.status_code == 201, f"got {r.status_code} {r.text[:120]}")
            rid = r.json().get("id")
            async with SessionLocal() as s:
                due = (await s.execute(_sel(rt.c["expected_return_time"])
                       .where(rt.c["id"] == rid))).scalar_one()
            check("loan: naive local due-time stored VERBATIM (no UTC shift)",
                  due.strftime("%H:%M") == "21:59", f"stored {due}")

            # A tz-aware (Z) input converts to LOCAL wall-clock, not raw-stripped.
            aware = _dtm.datetime(2026, 7, 16, 12, 0, tzinfo=_dtm.timezone.utc)
            expect_local = aware.astimezone().replace(tzinfo=None)
            r = await ac.post("/entry/returnables", headers=H,
                              json={"material_name": "SVC Zulu Tool", "borrower_name": "Svc Borrower",
                                    "qty": 1, "expected_return_time": "2026-07-16T12:00:00Z",
                                    "site_id": "CNCEC"})
            rid2 = r.json().get("id")
            async with SessionLocal() as s:
                due2 = (await s.execute(_sel(rt.c["expected_return_time"])
                        .where(rt.c["id"] == rid2))).scalar_one()
            check("loan: Z-suffixed input converts UTC → local before storing",
                  due2 == expect_local, f"stored {due2}, expected {expect_local}")

            # (b) Loan → borrower WhatsApp (digits-only to) + in-app for SKs.
            lw = await wa_rows("loan_created")
            check("loan: borrower WhatsApp sent (canonical + stored, digits sent)",
                  any(x["status"] == "sent" and x["to_number"] == BORROWER
                      and '"to": "966569233053"' in (x["payload_json"] or "").replace("'", '"')
                      or (x["status"] == "sent" and "966569233053" in (x["payload_json"] or ""))
                      for x in lw), str(lw)[:200])
            async with SessionLocal() as s:
                n_created = (await s.execute(_sel(func.count()).select_from(appn).where(
                    (appn.c["event_key"] == "loan_created")
                    & (appn.c["related_ref"] == str(rid))))).scalar_one()
            check("loan: in-app row for the site store keepers", n_created >= 1, f"n={n_created}")

            # Return → borrower confirmation + in-app.
            r = await ac.post(f"/entry/returnables/{rid}/return", headers=H)
            check("loan: mark returned → 200", r.status_code == 200, f"got {r.status_code}")
            rw = await wa_rows("loan_returned")
            check("loan: return confirmation WhatsApp sent",
                  any(x["status"] == "sent" for x in rw), str(rw)[:160])

            # Overdue sweep — insert an already-late loan, list triggers alerts.
            async with SessionLocal() as s:
                late = (await s.execute(_ins(rt).values(
                    material_name="SVC Late Tool", borrower_name="Svc Late",
                    borrower_phone=BORROWER, qty=1, status="borrowed", Site_ID="CNCEC",
                    expected_return_time=_dtm.datetime(2020, 1, 1, 8, 0),
                    whatsapp_alert_sent=0).returning(rt.c["id"]))).scalar_one()
                await s.commit()
            r = await ac.get("/entry/returnables", headers=H)
            check("loan: overdue sweep runs → 200", r.status_code == 200, f"got {r.status_code}")
            ow = await wa_rows("returnable_overdue")
            check("loan: overdue chases the borrower on WhatsApp",
                  any(x["status"] == "sent" and x["to_number"] == BORROWER for x in ow), str(ow)[:200])
            async with SessionLocal() as s:
                flag = (await s.execute(_sel(rt.c["whatsapp_alert_sent"])
                        .where(rt.c["id"] == late))).scalar_one()
            check("loan: overdue alert deduped (flag set)", flag == 1, f"flag={flag}")
    finally:
        wamod._post_message = saved_post
        for _k, _v in prev_env.items():
            if _v is None:
                _o.environ.pop(_k, None)
            else:
                _o.environ[_k] = _v
        async with SessionLocal() as s:
            await s.execute(delete(rt).where(rt.c["id"] > base_rt))
            await s.execute(delete(ob).where(ob.c["id"] > base_ob))
            await s.commit()


async def test_search_filters():
    """UAT Phase 2 — global search & filtering + the PR browse entity:
    `q` free-text on generic read entities and the derived stock views,
    `category` on stock views, /meta/categories, and /purchase-requests."""
    transport = ASGITransport(app=app)
    async with AsyncClient(transport=transport, base_url="http://svc") as ac:
        _ip = {"X-Real-IP": "203.0.113.95"}

        async def token(u, p):
            r = await ac.post("/auth/login", headers=_ip, json={"username": u, "password": p})
            return r.json().get("access_token")

        def H(t):
            return {"Authorization": f"Bearer {t}"}

        admin_t = await token("admin", "admin2026")
        worker_t = await token("worker", "floor2026")

        # /meta/categories — distinct, non-empty inventory categories.
        r = await ac.get("/meta/categories", headers=H(admin_t))
        cats = r.json().get("categories", [])
        check("search: /meta/categories → 200 + list",
              r.status_code == 200 and isinstance(cats, list) and len(cats) >= 1,
              f"got {r.status_code} {str(cats)[:100]}")

        # Generic entity q: matching description narrows, nonsense → 0.
        r = await ac.get("/inventory", headers=H(admin_t), params={"q": "WATER STORAGE"})
        hit = r.json().get("total", 0)
        r2 = await ac.get("/inventory", headers=H(admin_t), params={"q": "zzz-no-such-material"})
        miss = r2.json().get("total", -1)
        check("search: /inventory?q narrows (hit ≥ 1, nonsense = 0)",
              hit >= 1 and miss == 0, f"hit={hit} miss={miss}")

        # q also hits the SAP code itself.
        r = await ac.get("/inventory", headers=H(admin_t), params={"q": "1001"})
        check("search: /inventory?q matches the SAP code", r.json().get("total", 0) >= 1,
              str(r.json())[:100])

        # Derived stock views: q by description via the inventory join.
        r = await ac.get("/stock/live", headers=H(admin_t), params={"q": "WATER STORAGE"})
        check("search: /stock/live?q via inventory join", r.status_code == 200
              and r.json().get("total", 0) >= 1, f"got {r.status_code} {str(r.json())[:80]}")
        r = await ac.get("/stock/by-site", headers=H(admin_t), params={"q": "zzz-nope"})
        check("search: /stock/by-site nonsense q → 0 rows", r.json().get("total", -1) == 0,
              str(r.json())[:80])
        if cats:
            r = await ac.get("/stock/by-site", headers=H(admin_t), params={"category": cats[0]})
            check("search: /stock/by-site?category filters → 200",
                  r.status_code == 200 and r.json().get("total", -1) >= 0, f"got {r.status_code}")

        # Site scoping still binds under q (worker pinned to CNCEC).
        r = await ac.get("/consumption", headers=H(worker_t), params={"q": "1001", "limit": 5})
        items = r.json().get("items", [])
        check("search: scoped user q respects site pinning",
              r.status_code == 200 and all((i.get("Site_ID") or "CNCEC") == "CNCEC" for i in items),
              f"got {r.status_code} {str(items)[:100]}")

        # New PR browse entity (PO-page standard): hod+ read, list shape.
        r = await ac.get("/purchase-requests", headers=H(admin_t), params={"limit": 5})
        check("search: /purchase-requests browse → 200 + paged shape",
              r.status_code == 200 and {"total", "items"} <= set(r.json().keys()),
              f"got {r.status_code}")
        r = await ac.get("/purchase-requests", headers=H(admin_t), params={"q": "zzz-no-such-pr"})
        check("search: /purchase-requests?q → 0 on nonsense", r.json().get("total", -1) == 0,
              str(r.json())[:80])


async def test_notification_qa():
    """UAT Phase 4 — automated notification QA: fire EVERY WhatsApp pathway with
    the target test number +966569233053 and prove each one resolves recipients
    and lands a SENT row in whatsapp_outbox (Meta HTTP mocked; no live sends).
    Pathways not re-fired here are covered by sister suites: loans/overdue (Y),
    OTP (X), MTC-missing email (W). All synthetic data is cleaned up in finally."""
    import os as _o
    from sqlalchemy import delete, insert as _ins, select as _sel, update as _upd
    import backend.api.services.whatsapp as wamod
    from .services import procurement as _proc
    from .services import supervisor as _sup
    from .services import warehouse as _wh

    TARGET = "+966569233053"
    USERS = ("worker", "hod", "admin", "supervisor")
    ob = ledger._MD.tables["whatsapp_outbox"]
    users = ledger._MD.tables["users"]
    appn = ledger._MD.tables["app_notifications"]
    po_t = ledger._MD.tables["purchase_orders"]
    poi_t = ledger._MD.tables["po_items"]
    dn_t = ledger._MD.tables["delivery_notes"]
    dni_t = ledger._MD.tables["dn_items"]
    lots_tt = ledger._MD.tables["lots"]
    cons_t = ledger._MD.tables["consumption"]
    pend_i = ledger._MD.tables["pending_issues"]
    mtc_t = ledger._MD.tables["mtc_documents"]
    bugs_tt = ledger._MD.tables["bug_reports"]
    req_t = ledger._MD.tables["requests"]
    prm_t = ledger._MD.tables["pr_master"]
    smr_tt = ledger._MD.tables["supervisor_material_requests"]
    smri_t = ledger._MD.tables["supervisor_material_request_items"]
    resch_t = ledger._MD.tables["po_reschedule_requests"]
    pret_t = ledger._MD.tables["po_returns"]
    pofc_t = ledger._MD.tables["po_force_closures"]
    poa_t = ledger._MD.tables["po_assignments"]
    wh_t = ledger._MD.tables["warehouses"]
    PO, DN, LOT, WH = "PO-SVC-QA", "DN-SVC-QA", "SVC-QA-LOT", "WH-SVC-QA"

    async with SessionLocal() as s:
        base_ob = (await s.execute(_sel(func.coalesce(func.max(ob.c["id"]), 0)))).scalar_one()
        base_app = (await s.execute(_sel(func.coalesce(func.max(appn.c["id"]), 0)))).scalar_one()
        base_cons = (await s.execute(_sel(func.coalesce(func.max(cons_t.c["id"]), 0)))).scalar_one()
        base_pend = (await s.execute(_sel(func.coalesce(func.max(pend_i.c["id"]), 0)))).scalar_one()
        base_mtc = (await s.execute(_sel(func.coalesce(func.max(mtc_t.c["id"]), 0)))).scalar_one()
        prev_phones = {u: p for u, p in (await s.execute(
            _sel(users.c["username"], users.c["Phone_Number"])
            .where(users.c["username"].in_(USERS)))).all()}
        # Force the target number on every test role (the QA requirement).
        await s.execute(_upd(users).where(users.c["username"].in_(USERS))
                        .values(Phone_Number=TARGET))
        await s.commit()

    saved = (wamod._post_message, wamod._upload_media)
    _ENVK = ("WHATSAPP_PHONE_NUMBER_ID", "WHATSAPP_TOKEN", "WHATSAPP_ESCALATION_TO")
    prev_env = {k: _o.environ.get(k) for k in _ENVK}
    sent_payloads: list = []

    async def ok_post(payload):
        sent_payloads.append(payload)
        return {"ok": True, "message_id": "wamid.QA"}

    async def ok_upload(blob, filename, mime):
        return {"ok": True, "media_id": "media.QA"}

    wamod._post_message, wamod._upload_media = ok_post, ok_upload
    _o.environ["WHATSAPP_PHONE_NUMBER_ID"] = "svc-test-pnid"
    _o.environ["WHATSAPP_TOKEN"] = "svc-test-token"
    _o.environ["WHATSAPP_ESCALATION_TO"] = TARGET  # catch-all → target number

    async def _cleanup():
        async with SessionLocal() as s:
            await s.execute(delete(pret_t).where(pret_t.c["PO_Number"] == PO))
            await s.execute(delete(resch_t).where(resch_t.c["PO_Number"] == PO))
            await s.execute(delete(pofc_t).where(pofc_t.c["PO_Number"] == PO))
            await s.execute(delete(poa_t).where(poa_t.c["PO_Number"] == PO))
            await s.execute(delete(dni_t).where(dni_t.c["DN_Number"] == DN))
            await s.execute(delete(dn_t).where(dn_t.c["DN_Number"] == DN))
            await s.execute(delete(poi_t).where(poi_t.c["PO_Number"] == PO))
            await s.execute(delete(po_t).where(po_t.c["PO_Number"] == PO))
            await s.execute(delete(wh_t).where(wh_t.c["Warehouse_ID"] == WH))
            await s.execute(delete(lots_tt).where(lots_tt.c["Lot_Number"] == LOT))
            await s.execute(delete(cons_t).where(cons_t.c["id"] > base_cons))
            await s.execute(delete(pend_i).where(pend_i.c["id"] > base_pend))
            await s.execute(delete(mtc_t).where(mtc_t.c["id"] > base_mtc))
            await s.execute(delete(bugs_tt).where(bugs_tt.c["description"] == "svc notification qa"))
            await s.execute(delete(req_t).where(req_t.c["notes"] == "svc-qa"))
            await s.execute(delete(ob).where(ob.c["id"] > base_ob))
            await s.commit()

    await _cleanup()
    async with SessionLocal() as s:  # synthetic PO + DN + lot fixtures
        await s.execute(_ins(po_t).values(PO_Number=PO, Site_ID="CNCEC", status="open",
                                          Expected_Delivery="2026-07-01", created_by="svc"))
        lid = (await s.execute(_ins(poi_t).values(
            PO_Number=PO, line_no=1, Material_Code="M1", Qty=10.0, Delivered_Qty=5.0,
            Returned_Qty=0.0, line_status="open").returning(poi_t.c["id"]))).scalar_one()
        await s.execute(_ins(dn_t).values(DN_Number=DN, PO_Number=PO, Warehouse_ID="HQ",
                                          Site_ID="CNCEC", status="draft", created_by="svc"))
        await s.execute(_ins(dni_t).values(DN_Number=DN, po_item_id=1, Material_Code="M1",
                                           Qty=1.0, status="pending"))
        await s.execute(_ins(wh_t).values(Warehouse_ID=WH, Name="Svc QA Warehouse",
                                          status="active"))
        lot_id = (await s.execute(_ins(lots_tt).values(
            Lot_Number=LOT, SAP_Code="1001", Site_ID="CNCEC", Received_Date="2026-07-01",
            Status="open").returning(lots_tt.c["id"]))).scalar_one()
        await s.commit()

    smr_no = pr_no = None
    try:
        transport = ASGITransport(app=app)
        async with AsyncClient(transport=transport, base_url="http://svc") as ac:
            _ip = {"X-Real-IP": "203.0.113.96"}

            async def token(u, p):
                r = await ac.post("/auth/login", headers=_ip, json={"username": u, "password": p})
                return r.json().get("access_token")

            def H(t):
                return {"Authorization": f"Bearer {t}"}

            worker_t = await token("worker", "floor2026")
            admin_t = await token("admin", "admin2026")

            # 1-3 · Entry pathway: stage → FEFO override → HOD approval feedback.
            r = await ac.post("/entry/consumption", headers=H(worker_t),
                              json={"Date": "2026-07-10", "SAP_Code": "1001",
                                    "Quantity": 1, "Site_ID": "CNCEC"})
            pid = r.json().get("pending_id")
            await ac.post("/entry/consumption", headers=H(worker_t),
                          json={"Date": "2026-07-10", "SAP_Code": "1001", "Quantity": 1,
                                "Site_ID": "CNCEC", "FEFO_Override": "yes"})
            await ac.post(f"/hod/pending/issues/{pid}/approve", headers=H(admin_t))

            # 4 · MTC upload → logistics.
            await ac.post("/entry/mtc", headers=H(worker_t),
                          data={"sap_code": "1001", "site_id": "CNCEC", "mtc_number": "MTC-QA"},
                          files={"file": ("mtc.pdf", b"%PDF-1.4 svc qa", "application/pdf")})

            # 5-10 · Procurement pathway (services fire the same dispatch()).
            async with SessionLocal() as s:
                res = await _proc.create_pr(s, username="hod", site_id="CNCEC",
                                            lines=[{"SAP_Code": "1001", "Requested_Qty": 2}])
                pr_no = res.get("pr_number")
                await _proc.submit_pr(s, username="hod", pr_number=pr_no, site_id="CNCEC")
                await _proc.assign_po(s, username="admin", po_number=PO,
                                      warehouse_id=WH, expected_delivery="2026-07-20", notes="")
                rr = await _proc.raise_reschedule(s, username="hod", role="hod", po_number=PO,
                                                  requested_date="2026-07-25", reason="svc qa")
                await _proc.decide_reschedule(s, username="admin", req_id=rr["id"],
                                              action="approve", decision_notes="")
                await _proc.raise_vendor_return(s, username="admin", po_number=PO,
                                                po_item_id=lid, qty=1, reason="svc qa damaged")
                await _proc.force_close(s, username="admin", target_type="po",
                                        target_ref=PO, reason="svc qa close")
                await s.commit()

            # 11-14 · DN multi-stage machine.
            async with SessionLocal() as s:
                await _wh.submit_dn(s, username="admin", dn_number=DN)
                await _wh.decide_dn_logistics(s, username="admin", dn_number=DN, action="approve")
                await _wh.decide_dn_hod(s, username="hod", dn_number=DN, action="approve")
                await _wh.ship_dn(s, username="admin", dn_number=DN)
                await s.commit()

            # 15-16 · Supervisor request → SK approval.
            async with SessionLocal() as s:
                res = await _sup.create_smr(s, supervisor="supervisor", site_id="CNCEC",
                                            worker_id="30001", job_tank_place="svc qa",
                                            old_ppe_returned=1, no_return_reason=None,
                                            items=[{"SAP_Code": "1001", "Requested_Qty": 1}])
                smr_no = res.get("request_no")
                await _sup.approve_smr(s, sk_username="worker", request_id=res["request_id"])
                await s.commit()

            # 17-19 · Cross-site request (>5 escalates) + decision.
            r = await ac.post("/xsite", headers=H(admin_t),
                              json={"requesting_site": "HQ", "target_site": "CNCEC",
                                    "SAP_Code": "1001", "requested_qty": 9, "notes": "svc-qa"})
            rid = r.json().get("id")
            await ac.post(f"/xsite/{rid}/decide", headers=H(admin_t), json={"action": "approve"})

            # 20 · Lot lifecycle.
            await ac.post(f"/admin/lots/{lot_id}/status", headers=H(admin_t),
                          json={"status": "quarantined", "reason": "svc qa"})

            # 21 · Feedback status update back to the reporter.
            r = await ac.post("/feedback", headers=H(worker_t),
                              json={"type": "bug", "page": "/qa", "description": "svc notification qa"})
            fid = r.json().get("id")
            await ac.patch(f"/admin/feedback/{fid}", headers=H(admin_t),
                           json={"status": "resolved", "admin_response": "svc qa done"})

            # 22 · Report document delivery straight to the target number.
            await ac.post("/reports/pr-status/whatsapp", headers=H(admin_t),
                          json={"to": TARGET, "format": "csv"})

        EXPECTED = ["entry_staged", "fefo_override", "entry_approved", "mtc_uploaded",
                    "pr_submitted_to_logistics", "po_assigned_to_warehouse",
                    "reschedule_raised", "reschedule_decided", "vendor_return_raised",
                    "force_close", "dn_pending_logistics", "dn_pending_hod",
                    "dn_hod_approved", "dn_shipped", "smr_created", "smr_approved",
                    "cross_site_requested", "xsite_escalation", "cross_site_decided",
                    "lot_quarantined", "feedback_updated", "report_delivery"]
        async with SessionLocal() as s:
            rows = [dict(m) for m in (await s.execute(_sel(
                ob.c["event_key"], ob.c["status"], ob.c["to_number"], ob.c["error"]
            ).where(ob.c["id"] > base_ob))).mappings().all()]
            in_app = {r[0] for r in (await s.execute(_sel(appn.c["event_key"])
                      .where(appn.c["id"] > base_app))).all()}
        by_event: dict = {}
        for r in rows:
            by_event.setdefault(r["event_key"], []).append(r)
        for ev in EXPECTED:
            ok = any(x["status"] == "sent" for x in by_event.get(ev, []))
            check(f"qa: {ev} → whatsapp_outbox SENT", ok,
                  str(by_event.get(ev, "NO ROWS"))[:140])
        bad = [r for r in rows if r["status"] != "sent"]
        check("qa: ZERO failed/unresolved rows across the whole sweep",
              not bad, str(bad)[:200])
        check("qa: payloads went to the target number (digits-only to)",
              any(p.get("to") == "966569233053" for p in sent_payloads),
              str([p.get("to") for p in sent_payloads[:10]]))
        DISPATCHED = [e for e in EXPECTED if e not in ("xsite_escalation", "report_delivery")]
        missing_app = [e for e in DISPATCHED if e not in in_app]
        check("qa: every dispatch pathway also wrote its in-app twin",
              not missing_app, f"missing in-app: {missing_app}")
    finally:
        wamod._post_message, wamod._upload_media = saved
        for _k, _v in prev_env.items():
            if _v is None:
                _o.environ.pop(_k, None)
            else:
                _o.environ[_k] = _v
        async with SessionLocal() as s:
            for u, p in prev_phones.items():
                await s.execute(_upd(users).where(users.c["username"] == u)
                                .values(Phone_Number=p))
            if smr_no:
                sid = (await s.execute(_sel(smr_tt.c["id"]).where(
                    smr_tt.c["request_no"] == smr_no))).scalar_one_or_none()
                if sid is not None:
                    await s.execute(delete(smri_t).where(smri_t.c["request_id"] == sid))
                    await s.execute(delete(smr_tt).where(smr_tt.c["id"] == sid))
            if pr_no:
                await s.execute(delete(prm_t).where(prm_t.c["PR_Number"] == pr_no))
            # The sweep's in-app rows must go too: deleting the PR/SMR fixtures
            # frees their sequence numbers, so a later run would re-mint the
            # same refs and collide with suite A's exact-count assertions.
            await s.execute(delete(appn).where(appn.c["id"] > base_app))
            await s.commit()
        await _cleanup()


async def test_webhook_and_digest():
    """Phase 6 — inbound WhatsApp webhook + dynamic delivery preference.

    Meta HTTP is MOCKED (payloads captured). Covers: the GET verification
    handshake, X-Hub-Signature-256 enforcement, unknown-sender drop, the
    site-scoped STOCK command, RESET PASSWORD (temp credential actually logs
    in; sessions revoked), evening staging vs urgent, the critical-alert
    bypass, the digest compiler formatting, the batch aggregator send +
    processed marking, and the X-Delivery-Preference header end-to-end.
    All fixtures restored / rows deleted in finally."""
    import hashlib as _hl
    import hmac as _hm
    import json as _json
    import os as _o
    import re as _re

    from sqlalchemy import delete, select as _sel, update as _upd
    import backend.api.services.whatsapp as wamod
    from .services.notifications import _compile_digest, dispatch as _dispatch, send_evening_digests

    ob = ledger._MD.tables["whatsapp_outbox"]
    ps = ledger._MD.tables["pending_summary_notifications"]
    users = ledger._MD.tables["users"]
    appn = ledger._MD.tables["app_notifications"]
    sess_t = ledger._MD.tables["auth_sessions"]
    req_t = ledger._MD.tables["requests"]

    WK_PHONE = "+15551230001"
    VERIFY_TOK = "svc-verify-token"
    APP_SECRET = "svc-app-secret"

    async with SessionLocal() as s:
        base_ob = (await s.execute(_sel(func.coalesce(func.max(ob.c["id"]), 0)))).scalar_one()
        base_ps = (await s.execute(_sel(func.coalesce(func.max(ps.c["id"]), 0)))).scalar_one()
        base_app = (await s.execute(_sel(func.coalesce(func.max(appn.c["id"]), 0)))).scalar_one()
        prev = {u: (await s.execute(_sel(users.c["Phone_Number"])
                    .where(users.c["username"] == u))).scalar_one_or_none()
                for u in ("worker", "admin")}
        prev_hash = (await s.execute(_sel(users.c["password_hash"])
                     .where(users.c["username"] == "worker"))).scalar_one()
        wk_site = ((await s.execute(_sel(users.c["Site_ID"])
                    .where(users.c["username"] == "worker"))).scalar_one_or_none() or "").strip()
        await s.execute(_upd(users).where(users.c["username"] == "worker")
                        .values(Phone_Number=WK_PHONE))
        await s.execute(_upd(users).where(users.c["username"] == "admin")
                        .values(Phone_Number="+15559990002"))
        await s.commit()
        # A SAP with activity at the worker's site (for a deterministic STOCK hit).
        from .stock import SQL_SITE_STOCK as _SQL_SITE
        from sqlalchemy import text as _sqt
        srow = (await s.execute(_sqt(
            f'SELECT * FROM ({_SQL_SITE}) sub WHERE sub."Site_ID" = :site LIMIT 1'
        ), {"site": wk_site})).mappings().first()
        stock_sap = (srow or {}).get("SAP_Code")

    saved_post = wamod._post_message
    _ENVK = ("WHATSAPP_PHONE_NUMBER_ID", "WHATSAPP_TOKEN",
             "WHATSAPP_WEBHOOK_VERIFY_TOKEN", "WHATSAPP_APP_SECRET")
    prev_env = {k: _o.environ.get(k) for k in _ENVK}
    payloads: list = []

    async def ok_post(payload):
        payloads.append(payload)
        return {"ok": True, "message_id": f"wamid.AB{len(payloads)}"}

    wamod._post_message = ok_post
    _o.environ["WHATSAPP_PHONE_NUMBER_ID"] = "svc-test-pnid"
    _o.environ["WHATSAPP_TOKEN"] = "svc-test-token"
    _o.environ["WHATSAPP_WEBHOOK_VERIFY_TOKEN"] = VERIFY_TOK
    _o.environ["WHATSAPP_APP_SECRET"] = APP_SECRET

    def sign(raw: bytes) -> str:
        return "sha256=" + _hm.new(APP_SECRET.encode(), raw, _hl.sha256).hexdigest()

    def inbound(sender_digits: str, text_body: str) -> bytes:
        return _json.dumps({"entry": [{"changes": [{"value": {"messages": [
            {"type": "text", "from": sender_digits, "text": {"body": text_body}}
        ]}}]}]}).encode()

    xsite_rid = None
    try:
        transport = ASGITransport(app=app)
        async with AsyncClient(transport=transport, base_url="http://svc") as ac:
            _ip = {"X-Real-IP": "203.0.113.97"}

            async def token(u, p):
                r = await ac.post("/auth/login", headers=_ip, json={"username": u, "password": p})
                return r.json().get("access_token")

            # 1) Meta verification handshake.
            r = await ac.get("/whatsapp/webhook", params={
                "hub.mode": "subscribe", "hub.verify_token": VERIFY_TOK,
                "hub.challenge": "42challenge"})
            check("wh: handshake echoes hub.challenge", r.status_code == 200
                  and r.text == "42challenge", f"got {r.status_code} {r.text[:60]}")
            r = await ac.get("/api/v1/whatsapp/webhook", params={
                "hub.mode": "subscribe", "hub.verify_token": "wrong",
                "hub.challenge": "x"})
            check("wh: wrong verify_token → 403 (both mounts live)",
                  r.status_code == 403, f"got {r.status_code}")

            # 2) Signature enforcement.
            raw = inbound("15551230001", "HELP")
            r = await ac.post("/whatsapp/webhook", content=raw,
                              headers={"X-Hub-Signature-256": "sha256=deadbeef",
                                       "Content-Type": "application/json"})
            check("wh: invalid X-Hub-Signature-256 → 403", r.status_code == 403,
                  f"got {r.status_code}")

            # 3) Unknown sender: logged + silently dropped (no reply row).
            raw = inbound("19998887777", "STOCK 123")
            r = await ac.post("/whatsapp/webhook", content=raw,
                              headers={"X-Hub-Signature-256": sign(raw),
                                       "Content-Type": "application/json"})
            async with SessionLocal() as s:
                n_ob = (await s.execute(_sel(func.count()).select_from(ob)
                        .where(ob.c["id"] > base_ob))).scalar_one()
            check("wh: unregistered number dropped (ack 200, no reply, no send)",
                  r.status_code == 200 and r.json().get("handled") == 0 and n_ob == 0,
                  f"got {r.status_code} {str(r.json())[:80]} outbox+{n_ob}")

            # 4) STOCK command from a verified store keeper — site-scoped.
            q_sap = stock_sap or "SVC-NO-SUCH-SAP"
            raw = inbound("15551230001", f"STOCK {q_sap}")
            r = await ac.post("/whatsapp/webhook", content=raw,
                              headers={"X-Hub-Signature-256": sign(raw),
                                       "Content-Type": "application/json"})
            reply = (payloads[-1] if payloads else {})
            body_txt = ((reply.get("text") or {}).get("body") or "")
            ok_stock = (reply.get("type") == "text" and reply.get("to") == "15551230001"
                        and ((stock_sap and "Current stock" in body_txt
                              and str(stock_sap) in body_txt)
                             or (not stock_sap and "not found" in body_txt)))
            check("wh: STOCK → session-text reply with the site balance",
                  r.json().get("handled") == 1 and ok_stock,
                  f"{str(reply)[:160]}")
            async with SessionLocal() as s:
                wrow = (await s.execute(_sel(ob.c["status"], ob.c["event_key"], ob.c["message_type"])
                        .where(ob.c["id"] > base_ob).order_by(ob.c["id"].desc()).limit(1))).first()
            check("wh: reply recorded in whatsapp_outbox as sent/text",
                  wrow is not None and wrow[0] == "sent" and wrow[1] == "webhook_reply"
                  and wrow[2] == "text", str(wrow))

            # 5) Unknown command → help text.
            raw = inbound("15551230001", "what can you do?")
            await ac.post("/whatsapp/webhook", content=raw,
                          headers={"X-Hub-Signature-256": sign(raw),
                                   "Content-Type": "application/json"})
            check("wh: unknown command → command help",
                  "GI Hub WhatsApp commands" in ((payloads[-1].get("text") or {}).get("body") or ""),
                  str(payloads[-1])[:120])

            # 6) RESET PASSWORD → single-use temp credential that really logs in.
            raw = inbound("15551230001", "RESET PASSWORD")
            r = await ac.post("/whatsapp/webhook", content=raw,
                              headers={"X-Hub-Signature-256": sign(raw),
                                       "Content-Type": "application/json"})
            m = _re.search(r"temporary password: (\S+)",
                           ((payloads[-1].get("text") or {}).get("body") or ""))
            async with SessionLocal() as s:
                n_sess = (await s.execute(_sel(func.count()).select_from(sess_t)
                          .where(sess_t.c["username"] == "worker"))).scalar_one()
            check("wh: RESET PASSWORD sends a temp credential + revokes sessions",
                  r.json().get("handled") == 1 and m is not None and n_sess == 0,
                  f"handled={str(r.json())[:60]} temp={'yes' if m else 'no'} sessions={n_sess}")
            rl = await ac.post("/auth/login", headers=_ip,
                               json={"username": "worker", "password": m.group(1) if m else "x"})
            check("wh: the temp password actually authenticates",
                  rl.status_code == 200 and bool(rl.json().get("access_token")),
                  f"got {rl.status_code} {rl.text[:80]}")
            async with SessionLocal() as s:  # restore the real credential at once
                await s.execute(_upd(users).where(users.c["username"] == "worker")
                                .values(password_hash=prev_hash))
                await s.execute(delete(sess_t).where(sess_t.c["username"] == "worker"))
                await s.commit()

            # 7) Delivery preference: evening stages, critical bypasses.
            ob_before = len(payloads)
            async with SessionLocal() as s:
                await _dispatch(s, event_key="p6_evening_a", title="PR SVC-P6 submitted",
                                body="awaiting HOD review", recipient_user="worker",
                                wa_template="action_required", delivery="evening")
                await _dispatch(s, event_key="p6_evening_b", title="DN SVC-P6 approved",
                                body="", recipient_user="worker",
                                wa_template="status_update", delivery="evening")
                await _dispatch(s, event_key="p6_critical", title="FEFO override at HQ",
                                body="lot skipped", recipient_user="worker", severity="critical",
                                wa_template="critical_alert", delivery="evening")
                await s.commit()
                staged = [dict(m) for m in (await s.execute(
                    _sel(ps.c["id"], ps.c["event_key"], ps.c["processed_at"])
                    .where(ps.c["id"] > base_ps))).mappings().all()]
            check("digest: evening events staged, no immediate send",
                  len([x for x in staged if x["event_key"].startswith("p6_evening")]) == 2
                  and len(payloads) == ob_before + 1,  # ONLY the critical went out
                  f"staged={str(staged)[:120]} sends+{len(payloads) - ob_before}")
            check("digest: critical alert bypasses the evening queue",
                  not any(x["event_key"] == "p6_critical" for x in staged)
                  and (payloads[-1].get("template") or {}).get("name") == "gi_critical_alert",
                  str(payloads[-1])[:140])

            # 8) The compiler: clean bullets, hard cap with explicit remainder.
            d = _compile_digest([{"title": "PR-1 submitted", "body": "3 items"},
                                 {"title": "DN-2 approved", "body": ""}])
            check("digest: compiler formats •-bullets with title — body",
                  d == "• PR-1 submitted — 3 items  • DN-2 approved", d)
            many = [{"title": f"event {i} with a reasonably long tail", "body": "x" * 40}
                    for i in range(60)]
            dm = _compile_digest(many)
            check("digest: compiler caps under the Meta 1024 limit with (+N more)",
                  len(dm) <= 1024 and "more)" in dm, f"len={len(dm)} tail={dm[-30:]}")

            # 9) The batch aggregator: ONE message per recipient, rows marked.
            async with SessionLocal() as s:
                res = await send_evening_digests(s)
                await s.commit()
                left = [dict(m) for m in (await s.execute(
                    _sel(ps.c["id"], ps.c["event_key"], ps.c["processed_at"], ps.c["digest_outbox_id"])
                    .where(ps.c["id"] > base_ps))).mappings().all()]
            dig = next((p for p in reversed(payloads)
                        if (p.get("template") or {}).get("name") == "gi_evening_summary"), None)
            dig_params = [prm["text"] for comp in ((dig or {}).get("template") or {}).get("components", [])
                          for prm in comp.get("parameters", [])]
            check("digest: aggregator sends ONE gi_evening_summary per recipient",
                  res.get("sent") == 1 and res.get("recipients") == 1 and dig is not None
                  and any("PR SVC-P6 submitted" in v and "DN SVC-P6 approved" in v
                          for v in dig_params),
                  f"res={res} params={str(dig_params)[:160]}")
            check("digest: staged rows marked processed + linked to the outbox row",
                  all(x["processed_at"] is not None and x["digest_outbox_id"] for x in left),
                  str(left)[:160])
            async with SessionLocal() as s:
                res2 = await send_evening_digests(s)
                await s.commit()
            check("digest: second run is a no-op (nothing pending)",
                  res2.get("recipients") == 0, str(res2))

            # 10) X-Delivery-Preference header end-to-end (middleware → dispatch).
            hod_t = await token("hod", "hod2026")
            r = await ac.post("/xsite", headers={"Authorization": f"Bearer {hod_t}",
                                                 "X-Delivery-Preference": "evening", **_ip},
                              json={"target_site": "SVC-P6-TARGET", "SAP_Code": "SVC-P6-SAP",
                                    "requested_qty": 1})
            xsite_rid = (r.json() or {}).get("id")
            async with SessionLocal() as s:
                hdr_staged = (await s.execute(_sel(func.count()).select_from(ps).where(
                    (ps.c["id"] > base_ps) & (ps.c["event_key"] == "cross_site_requested")
                    & ps.c["processed_at"].is_(None)))).scalar_one()
                hdr_sent = (await s.execute(_sel(func.count()).select_from(ob).where(
                    (ob.c["id"] > base_ob) & (ob.c["event_key"] == "cross_site_requested")))).scalar_one()
            check("digest: X-Delivery-Preference header stages instead of sending",
                  r.status_code == 201 and hdr_staged >= 1 and hdr_sent == 0,
                  f"got {r.status_code} staged={hdr_staged} sent={hdr_sent}")
    finally:
        wamod._post_message = saved_post
        for _k, _v in prev_env.items():
            if _v is None:
                _o.environ.pop(_k, None)
            else:
                _o.environ[_k] = _v
        async with SessionLocal() as s:
            await s.execute(_upd(users).where(users.c["username"] == "worker")
                            .values(Phone_Number=prev["worker"], password_hash=prev_hash))
            await s.execute(_upd(users).where(users.c["username"] == "admin")
                            .values(Phone_Number=prev["admin"]))
            if xsite_rid:
                await s.execute(delete(req_t).where(req_t.c["id"] == xsite_rid))
            await s.execute(delete(ps).where(ps.c["id"] > base_ps))
            await s.execute(delete(ob).where(ob.c["id"] > base_ob))
            await s.execute(delete(appn).where(appn.c["id"] > base_app))
            await s.execute(delete(sess_t).where(sess_t.c["username"] == "worker"))
            await s.commit()


async def test_executive_summary():
    """HOD Executive Summary — endpoint shape, exact site-scoped KPI math,
    role guard, date validation, Excel bytes, and the SME capacity rollup
    (unit-tested against a synthetic model). Read-only — nothing to clean."""
    from sqlalchemy import text as _sqt
    from . import exec_summary as esmod

    DF, DT = "2026-01-01", "2026-12-31"
    transport = ASGITransport(app=app)
    async with AsyncClient(transport=transport, base_url="http://svc") as ac:
        _ip = {"X-Real-IP": "203.0.113.99"}

        async def token(u, p):
            r = await ac.post("/auth/login", headers=_ip, json={"username": u, "password": p})
            return r.json().get("access_token")

        hod_t = await token("hod", "hod2026")
        H = {"Authorization": f"Bearer {hod_t}"}

        r = await ac.get("/hod/executive-summary", headers=H,
                         params={"date_from": DF, "date_to": DT})
        d = r.json() if r.status_code == 200 else {}
        sections = ("kpis", "receipts_detail", "consumption_detail", "returns_detail",
                    "sqm_detail", "manpower", "pr_status", "po_status", "delivery_plan",
                    "actions", "sqm_capacity", "cross_site")
        check("exec: GET → 200 with every section present",
              r.status_code == 200 and all(k in d for k in sections),
              f"got {r.status_code} missing={[k for k in sections if k not in d]}")

        # HOD is pinned to their own site; the receipts KPI must equal the
        # database truth for that site + range exactly.
        async with SessionLocal() as s:
            hod_site = (await s.execute(_sqt(
                "SELECT COALESCE(\"Site_ID\",'HQ') FROM users WHERE username='hod'"))).scalar_one()
            truth = (await s.execute(_sqt(
                'SELECT COUNT(*), COALESCE(SUM("Quantity"),0) FROM receipts '
                'WHERE substring("Date" FROM 1 FOR 10) BETWEEN :a AND :b '
                "AND COALESCE(\"Site_ID\",'HQ') = :st"),
                {"a": DF, "b": DT, "st": hod_site})).first()
        check("exec: site pinned to the HOD's site + exact receipts KPI",
              d.get("site_id") == hod_site
              and d["kpis"]["receipts"]["count"] == truth[0]
              and abs(d["kpis"]["receipts"]["qty"] - float(truth[1])) < 1e-6,
              f"site={d.get('site_id')}/{hod_site} kpi={d['kpis']['receipts']} truth={tuple(truth)}")

        # Manpower invariant: present + absent = active headcount.
        mp = d["kpis"]["manpower"]
        check("exec: manpower present+absent == active headcount",
              mp["present"] + mp["absent"] == mp["active_total"], str(mp))

        # Date validation.
        r = await ac.get("/hod/executive-summary", headers=H,
                         params={"date_from": "2026-1-1"})
        r2 = await ac.get("/hod/executive-summary", headers=H,
                          params={"date_from": "2026-02-02", "date_to": "2026-02-01"})
        check("exec: malformed date / inverted range → 422",
              r.status_code == 422 and r2.status_code == 422,
              f"got {r.status_code}/{r2.status_code}")

        # Role guard: store keeper locked out (hod+admin exact lock).
        wk_t = await token("worker", "floor2026")
        r = await ac.get("/hod/executive-summary",
                         headers={"Authorization": f"Bearer {wk_t}"})
        check("exec: store keeper → 403", r.status_code == 403, f"got {r.status_code}")

        # Admin: unrestricted (all sites) or pinned by ?site_id=.
        adm_t = await token("admin", "admin2026")
        HA = {"Authorization": f"Bearer {adm_t}"}
        ra = await ac.get("/hod/executive-summary", headers=HA,
                          params={"date_from": DF, "date_to": DT})
        rb = await ac.get("/hod/executive-summary", headers=HA,
                          params={"date_from": DF, "date_to": DT, "site_id": hod_site})
        check("exec: admin sees all sites by default, one site on request",
              ra.status_code == 200 and ra.json().get("site_id") is None
              and rb.status_code == 200 and rb.json().get("site_id") == hod_site,
              f"{ra.status_code}/{ra.json().get('site_id')} · {rb.status_code}/{rb.json().get('site_id')}")

        # Excel export: valid xlsx bytes + attachment filename.
        r = await ac.get("/hod/executive-summary/export.xlsx", headers=H,
                         params={"date_from": DF, "date_to": DT})
        cd = r.headers.get("content-disposition", "")
        check("exec: Excel export → valid workbook + filename",
              r.status_code == 200 and r.content[:2] == b"PK"
              and "executive_summary_" in cd and cd.endswith('.xlsx"'),
              f"got {r.status_code} {len(r.content)}b cd={cd[:80]}")

        # PDF export: server-rendered (fpdf2) — valid PDF magic, >1 page for a
        # full-year window, attachment filename. (Replaces the print-the-page
        # approach the HOD rejected in UAT.)
        r = await ac.get("/hod/executive-summary/export.pdf", headers=H,
                         params={"date_from": DF, "date_to": DT})
        cd = r.headers.get("content-disposition", "")
        pages = r.content.count(b"/Type /Page") - r.content.count(b"/Type /Pages")
        check("exec: PDF export → valid paginated PDF + filename",
              r.status_code == 200 and r.content[:5] == b"%PDF-" and pages >= 1
              and r.headers.get("content-type") == "application/pdf"
              and "executive_summary_" in cd and cd.endswith('.pdf"'),
              f"got {r.status_code} {len(r.content)}b pages={pages} cd={cd[:80]}")

    # SME capacity rollup — synthetic model, strict-bottleneck expectation:
    # unit remaining 100 SQM, two materials at 50% and 100% coverage → the
    # 50% material caps achievable at 50 SQM.
    model = {"units": {("TK-1", "1"): {"remaining": 100.0, "short_name": "CBL30"}},
             "tag_meta": {"TK-1": {"Name": "Tank 1"}}}
    lines = [
        {"Equipment_Tag_No": "TK-1", "Lining_System_Code": "1", "Material_Code": "M-A",
         "Material_Name": "Mat A", "Demand_Qty": 200.0, "Allocated_Qty": 100.0,
         "Shortfall_Qty": 100.0},
        {"Equipment_Tag_No": "TK-1", "Lining_System_Code": "1", "Material_Code": "M-B",
         "Material_Name": "Mat B", "Demand_Qty": 50.0, "Allocated_Qty": 50.0,
         "Shortfall_Qty": 0.0},
    ]
    eq, sy = esmod._capacity_from_lines(model, lines)
    check("exec: capacity math = strict bottleneck (50% mat → 50/100 SQM)",
          len(eq) == 1 and eq[0]["Achievable_SQM"] == 50.0 and eq[0]["Coverage_Pct"] == 50.0
          and "M-A" in eq[0]["Bottleneck"]
          and len(sy) == 1 and sy[0]["System_Code"] == "1" and sy[0]["Achievable_SQM"] == 50.0,
          f"eq={eq} sy={sy}")




async def test_lining_coverage():
    """Phase 8-1 — predictive lining analytics (/analytics/lining-coverage).
    Read-only; the SME engine input pool must be the LIVE ledger stock."""
    from sqlalchemy import text as _sqt

    transport = ASGITransport(app=app)
    async with AsyncClient(transport=transport, base_url="http://svc") as ac:
        _ip = {"X-Real-IP": "203.0.113.105"}

        async def token(u, p):
            r = await ac.post("/auth/login", headers=_ip, json={"username": u, "password": p})
            return r.json().get("access_token")

        def H(t):
            return {"Authorization": f"Bearer {t}", **_ip}

        worker_t = await token("worker", "floor2026")
        sup_t = await token("supervisor", "super2026")
        hod_t = await token("hod", "hod2026")

        r = await ac.get("/analytics/lining-coverage", headers=H(worker_t))
        check("lining: store keeper → 403", r.status_code == 403, f"{r.status_code}")
        r = await ac.get("/analytics/lining-coverage", headers=H(sup_t))
        check("lining: supervisor → 403", r.status_code == 403, f"{r.status_code}")

        r = await ac.get("/analytics/lining-coverage", headers=H(hod_t))
        d = r.json() if r.status_code == 200 else {}
        check("lining: HOD → 200 with families/per_system/materials",
              r.status_code == 200 and all(k in d for k in
                                           ("families", "per_system", "materials", "source")),
              f"{r.status_code} {str(d)[:160]}")

        async with SessionLocal() as ses:
            hod_site = (await ses.execute(_sqt(
                "SELECT COALESCE(\"Site_ID\",'HQ') FROM users WHERE username='hod'"))).scalar_one()
        check("lining: scoped HOD is pinned to their own site",
              d.get("site") == hod_site, f"site={d.get('site')} vs {hod_site}")

        fams = {f["family"] for f in d.get("families", [])}
        check("lining: RL and BL families present with sane math",
              {"RL", "BL"} <= fams and all(
                  f["achievable_sqm"] <= f["remaining_sqm"] + 1e-6
                  and 0 <= f["coverage_pct"] <= 100 for f in d["families"]),
              f"fams={fams}")

        # per_system sorted worst-coverage-first; each system carries a family tag
        cov = [s_["Coverage_Pct"] for s_ in d.get("per_system", [])]
        check("lining: systems sorted worst-coverage-first",
              cov == sorted(cov), f"{cov[:8]}")

        # the availability pool really is the LIVE ledger: pick a material the
        # engine consumed that exists in the live master and compare its
        # live_stock to the ledger truth for the HOD's site
        live_rows = [m for m in d.get("materials", []) if m["stock_source"] == "live"]
        check("lining: live-ledger materials present in the pool",
              len(live_rows) >= 1 and d["source"]["live"] >= 1,
              f"live={d.get('source')}")
        if live_rows:
            probe = live_rows[0]
            async with SessionLocal() as ses:
                truth = (await ses.execute(_sqt('''
                    SELECT COALESCE(SUM(x.q), 0) FROM (
                      SELECT (SELECT COALESCE(SUM(r."Quantity"),0) FROM receipts r
                              WHERE TRIM(r."SAP_Code") = TRIM(i."SAP_Code")
                                AND COALESCE(r."Site_ID",'HQ') = :st)
                           - (SELECT COALESCE(SUM(c."Quantity"),0) FROM consumption c
                              WHERE TRIM(c."SAP_Code") = TRIM(i."SAP_Code")
                                AND COALESCE(c."Site_ID",'HQ') = :st)
                           - (SELECT COALESCE(SUM(rt."Quantity"),0) FROM returns rt
                              WHERE TRIM(rt."SAP_Code") = TRIM(i."SAP_Code")
                                AND COALESCE(rt."Site_ID",'HQ') = :st) AS q
                      FROM inventory i
                      WHERE TRIM(i."Material_Code") = :mc
                        AND COALESCE(i."Site_ID",'HQ') = :st) x'''),
                    {"mc": probe["material_code"], "st": hod_site})).scalar_one()
            check("lining: live_stock equals the ledger truth",
                  abs(float(probe["live_stock"]) - float(truth)) < 1e-6,
                  f'{probe["material_code"]}: api={probe["live_stock"]} db={truth}')

        # materials sorted biggest-shortfall-first
        shorts = [m["shortfall_qty"] for m in d.get("materials", [])]
        check("lining: materials sorted by shortfall desc",
              shorts == sorted(shorts, reverse=True), f"{shorts[:8]}")


async def test_data_query():
    """Phase C — chat-with-your-data (/ai/query). The template lane must be
    deterministic, site-pinned for scoped roles, and independent of Ollama;
    the NL lane must only ever serve unscoped roles. Read-only."""
    from datetime import date, timedelta
    from sqlalchemy import text as _sqt
    from .ai import analytics as _an

    transport = ASGITransport(app=app)
    async with AsyncClient(transport=transport, base_url="http://svc") as ac:
        _ip = {"X-Real-IP": "203.0.113.104"}

        async def token(u, p):
            r = await ac.post("/auth/login", headers=_ip, json={"username": u, "password": p})
            return r.json().get("access_token")

        hod_t = await token("hod", "hod2026")
        worker_t = await token("worker", "floor2026")
        admin_t = await token("admin", "admin2026")

        def H(t):
            return {"Authorization": f"Bearer {t}", **_ip}

        # role gate: store keeper (level 0) is refused
        r = await ac.post("/ai/query", headers=H(worker_t), json={"question": "stock"})
        check("query: store keeper → 403", r.status_code == 403, f"{r.status_code}")
        r = await ac.get("/ai/query/examples", headers=H(worker_t))
        check("query: examples gated the same way", r.status_code == 403, f"{r.status_code}")
        r = await ac.get("/ai/query/examples", headers=H(hod_t))
        check("query: HOD gets example questions",
              r.status_code == 200 and len(r.json().get("examples", [])) >= 3, f"{r.status_code}")

        async with SessionLocal() as ses:
            hod_site = (await ses.execute(_sqt(
                "SELECT COALESCE(\"Site_ID\",'HQ') FROM users WHERE username='hod'"))).scalar_one()

        # template lane: returns, last week, HOD → every row pinned to own site
        r = await ac.post("/ai/query", headers=H(hod_t),
                          json={"question": "Show me all material returns from last week"})
        d = r.json() if r.status_code == 200 else {}
        site_col = d.get("columns", []).index("site") if "site" in d.get("columns", []) else -1
        check("query: HOD returns/last-week → template lane, ok",
              r.status_code == 200 and d.get("ok") is True and d.get("mode") == "template"
              and d.get("intent") == "returns" and ":site" in d.get("sql", ""),
              f"{r.status_code} {str(d)[:200]}")
        check("query: every returned row is the HOD's own site",
              site_col >= 0 and all(row[site_col] == hod_site for row in d.get("rows", [])),
              f"site_col={site_col} rows={len(d.get('rows', []))}")

        # naming ANOTHER site must not widen the scope for a scoped role
        other = "HQ" if hod_site != "HQ" else "CNCEC"
        r = await ac.post("/ai/query", headers=H(hod_t),
                          json={"question": f"returns in {other} last week"})
        d = r.json()
        rows = d.get("rows", [])
        sc = d["columns"].index("site") if "site" in d.get("columns", []) else -1
        check("query: scoped user naming another site stays pinned to their own",
              d.get("ok") is True and (sc < 0 or all(row[sc] == hod_site for row in rows)),
              f"{str(d)[:160]}")

        # count/total → metric card, and the numbers equal DB truth exactly
        r = await ac.post("/ai/query", headers=H(hod_t),
                          json={"question": "How many issues in the last 30 days?"})
        d = r.json()
        dfrom = (date.today() - timedelta(days=30)).isoformat()
        async with SessionLocal() as ses:
            truth = (await ses.execute(_sqt(
                'SELECT COUNT(*), COALESCE(SUM("Quantity"),0) FROM consumption '
                'WHERE "Date" >= :a AND COALESCE("Site_ID",\'HQ\') = :st'),
                {"a": dfrom, "st": hod_site})).first()
        m = d.get("metric") or {}
        check("query: count question → metric equals DB truth",
              d.get("ok") is True and m.get("entries") == truth[0]
              and abs(float(m.get("value", -1)) - float(truth[1])) < 1e-6,
              f"metric={m} truth={tuple(truth)}")

        # unmatched question, SCOPED role → friendly refusal + examples (never NL)
        r = await ac.post("/ai/query", headers=H(hod_t),
                          json={"question": "please summarise everything interesting"})
        d = r.json()
        check("query: unmatched + scoped → ok:false with examples, not the NL lane",
              r.status_code == 200 and d.get("ok") is False and d.get("examples"),
              f"{str(d)[:160]}")

        # unmatched question, UNSCOPED role → the NL lane is invoked (stubbed)
        async def _fake_nl(question):
            return {"ok": True, "sql": "SELECT 1", "columns": ["one"], "rows": [[1]],
                    "message": "stub", "question": question}
        orig = _an.run_nl_query
        _an.run_nl_query = _fake_nl
        try:
            r = await ac.post("/ai/query", headers=H(admin_t),
                              json={"question": "please summarise everything interesting"})
            d = r.json()
            check("query: unmatched + admin → NL fallback lane",
                  r.status_code == 200 and d.get("mode") == "nl" and d.get("ok") is True,
                  f"{str(d)[:160]}")
        finally:
            _an.run_nl_query = orig

        # admin naming a site narrows the template query to it
        r = await ac.post("/ai/query", headers=H(admin_t),
                          json={"question": f"receipts in {hod_site} last 90 days"})
        d = r.json()
        check("query: admin + site mention → template narrowed to that site",
              d.get("ok") is True and d.get("mode") == "template"
              and f"site {hod_site}" in d.get("message", ""), f"{str(d)[:160]}")

        # empty question → 422
        r = await ac.post("/ai/query", headers=H(admin_t), json={"question": "   "})
        check("query: blank question → 422", r.status_code == 422, f"{r.status_code}")



async def test_strict_limits():
    """Phase 8-2 — SMS-toll-fraud + HMAC-probing hardening. The strict rules
    are relaxed under GI_DOTENV=0 (so every other suite runs freely); this
    suite forces them on via GI_FORCE_STRICT_LIMITS and verifies:
      · OTP: 3/hour per source IP  · 3/hour per target phone (across IPs)
      · webhook: 5 invalid HMAC signatures in 10 min ⇒ 15-min IP ban (429)
      · Retry-After present on every 429."""
    import hashlib as _hl
    import hmac as _hm
    import os as _os

    transport = ASGITransport(app=app)
    async with AsyncClient(transport=transport, base_url="http://svc") as ac:
        _ip = {"X-Real-IP": "203.0.113.106"}

        async def token(u, p):
            r = await ac.post("/auth/login", headers=_ip, json={"username": u, "password": p})
            return r.json().get("access_token")

        worker_t = await token("worker", "floor2026")

        def H(t, ip):
            return {"Authorization": f"Bearer {t}", "X-Real-IP": ip}

        _os.environ["GI_FORCE_STRICT_LIMITS"] = "1"
        try:
            # ── OTP per-IP: 4th request from ONE IP inside the hour → 429 ────
            ip_a = "198.51.100.31"
            codes = []
            for i in range(4):
                r = await ac.post("/auth/phone/request-otp", headers=H(worker_t, ip_a),
                                  json={"new_number": f"+96650000010{i}"})
                codes.append(r.status_code)
            # hermetic env ⇒ WhatsApp off ⇒ allowed calls surface as 503, but
            # the 4th must be the limiter's 429 (quota burns on attempts)
            check("limits: OTP per-IP — 3 attempts pass the gate, 4th → 429",
                  codes[:3] == [503, 503, 503] and codes[3] == 429, f"{codes}")
            r = await ac.post("/auth/phone/request-otp", headers=H(worker_t, ip_a),
                              json={"new_number": "+966500000109"})
            check("limits: 429 carries Retry-After",
                  r.status_code == 429 and int(r.headers.get("retry-after", 0)) > 0,
                  f'{r.status_code} ra={r.headers.get("retry-after")}')

            # ── OTP per-phone: same number from ROTATING IPs still capped ────
            phone = "+966500000777"
            codes = []
            for i in range(4):
                r = await ac.post("/auth/phone/request-otp",
                                  headers=H(worker_t, f"198.51.100.{40 + i}"),
                                  json={"new_number": phone})
                codes.append(r.status_code)
            check("limits: OTP per-phone — rotating IPs, 4th for the number → 429",
                  codes[:3] == [503, 503, 503] and codes[3] == 429
                  and "number" in r.json().get("detail", ""), f"{codes}")

            # ── webhook HMAC penalty box ──────────────────────────────────────
            _os.environ["WHATSAPP_APP_SECRET"] = "af-test-secret"
            try:
                bad = {"X-Hub-Signature-256": "sha256=" + "0" * 64, "X-Real-IP": "198.51.100.60"}
                sigs = []
                for _ in range(5):
                    r = await ac.post("/whatsapp/webhook", headers=bad, content=b"{}")
                    sigs.append(r.status_code)
                check("limits: invalid HMACs are 403 until the ban trips",
                      sigs == [403] * 5, f"{sigs}")
                # 6th request from the SAME IP — even with a VALID signature — 429
                good_sig = "sha256=" + _hm.new(b"af-test-secret", b"{}", _hl.sha256).hexdigest()
                r = await ac.post("/whatsapp/webhook",
                                  headers={"X-Hub-Signature-256": good_sig,
                                           "X-Real-IP": "198.51.100.60"}, content=b"{}")
                check("limits: banned IP → 429 with Retry-After (even valid sig)",
                      r.status_code == 429 and int(r.headers.get("retry-after", 0)) > 0,
                      f'{r.status_code} ra={r.headers.get("retry-after")}')
                # an unrelated IP with a valid signature is untouched
                r = await ac.post("/whatsapp/webhook",
                                  headers={"X-Hub-Signature-256": good_sig,
                                           "X-Real-IP": "198.51.100.61"}, content=b"{}")
                check("limits: other IPs unaffected by the ban",
                      r.status_code == 200, f"{r.status_code}")
            finally:
                _os.environ.pop("WHATSAPP_APP_SECRET", None)
        finally:
            _os.environ.pop("GI_FORCE_STRICT_LIMITS", None)



async def test_weekly_report():
    """Phase 8-3 — automated weekly executive PDF: Friday-17:00 schedule math,
    render + tokenized storage, secure expiring download, admin run-now, and
    the bell/WhatsApp dispatch fan-out to admins + HODs. Cleans up after."""
    import datetime as _dtm
    from sqlalchemy import text as _sqt

    from .weekly_report import next_friday_1700

    # schedule math (Friday = weekday 4)
    wed = _dtm.datetime(2026, 7, 8, 9, 0)       # Wednesday
    fri_after = next_friday_1700(wed)
    check("weekly: Wednesday → this Friday 17:00",
          fri_after == _dtm.datetime(2026, 7, 10, 17, 0), f"{fri_after}")
    fri_eve = _dtm.datetime(2026, 7, 10, 18, 0)  # Friday after 17:00
    nxt = next_friday_1700(fri_eve)
    check("weekly: Friday 18:00 → NEXT Friday 17:00",
          nxt == _dtm.datetime(2026, 7, 17, 17, 0), f"{nxt}")
    fri_noon = _dtm.datetime(2026, 7, 10, 12, 0)
    check("weekly: Friday noon → same day 17:00",
          next_friday_1700(fri_noon) == _dtm.datetime(2026, 7, 10, 17, 0),
          f"{next_friday_1700(fri_noon)}")

    transport = ASGITransport(app=app)
    async with AsyncClient(transport=transport, base_url="http://svc") as ac:
        _ip = {"X-Real-IP": "203.0.113.107"}

        async def token(u, p):
            r = await ac.post("/auth/login", headers=_ip, json={"username": u, "password": p})
            return r.json().get("access_token")

        def H(t):
            return {"Authorization": f"Bearer {t}", **_ip}

        worker_t = await token("worker", "floor2026")
        admin_t = await token("admin", "admin2026")

        r = await ac.post("/admin/reports/weekly-exec/run", headers=H(worker_t))
        check("weekly: run-now is admin-only (SK → 403)", r.status_code == 403, f"{r.status_code}")

        r = await ac.post("/admin/reports/weekly-exec/run", headers=H(admin_t))
        d = r.json() if r.status_code == 200 else {}
        check("weekly: admin run-now → ≥1 PDF, ≥2 recipients (admins + HODs)",
              r.status_code == 200 and d.get("reports", 0) >= 1
              and d.get("recipients", 0) >= 2 and "ALL" in d.get("links", {}),
              f"{r.status_code} {str(d)[:200]}")

        # the WhatsApp-style link really serves a PDF, no auth required
        url = d["links"]["ALL"]
        tok = url.rsplit("/", 1)[-1]
        r = await ac.get(f"/reports/weekly-exec/{tok}")
        check("weekly: tokenized link → 200 application/pdf with %PDF magic",
              r.status_code == 200
              and r.headers.get("content-type", "").startswith("application/pdf")
              and r.content[:5] == b"%PDF-", f"{r.status_code}")

        r = await ac.get("/reports/weekly-exec/not-a-real-token")
        check("weekly: unknown token → 404", r.status_code == 404, f"{r.status_code}")

        # force-expire the artifact → the same link dies with 410
        async with SessionLocal() as ses:
            await ses.execute(_sqt(
                "UPDATE generated_reports SET expires_at = CURRENT_TIMESTAMP - INTERVAL '1 hour' "
                "WHERE kind = 'weekly_exec'"))
            await ses.commit()
        r = await ac.get(f"/reports/weekly-exec/{tok}")
        check("weekly: expired link → 410", r.status_code == 410, f"{r.status_code}")

        # every admin/hod got the in-app bell row with the download link
        async with SessionLocal() as ses:
            n_bell = (await ses.execute(_sqt(
                "SELECT COUNT(*) FROM app_notifications "
                "WHERE event_key = 'weekly_exec_report' "
                "AND body LIKE '%/reports/weekly-exec/%'"))).scalar_one()
            n_recip = (await ses.execute(_sqt(
                "SELECT COUNT(*) FROM users WHERE role IN ('admin','hod')"))).scalar_one()
        check("weekly: one bell row per admin/HOD recipient, carrying the link",
              n_bell >= n_recip, f"bell={n_bell} recipients={n_recip}")

        # cleanup — this suite COMMITS real rows
        async with SessionLocal() as ses:
            await ses.execute(_sqt("DELETE FROM generated_reports WHERE kind='weekly_exec'"))
            await ses.execute(_sqt(
                "DELETE FROM app_notifications WHERE event_key='weekly_exec_report'"))
            await ses.commit()



async def test_entry_documents():
    """Parity A1/A2/A4 — the legacy entry-document system, return gates and
    WBS enforcement, tested with the master switch ON (the suite flips
    require_entry_documents to '1' and restores '0' afterwards)."""
    import io as _io
    from sqlalchemy import text as _sqt

    async def _set_gate(v: str):
        async with SessionLocal() as s_:
            await s_.execute(_sqt(
                "INSERT INTO app_settings (key, value) VALUES ('require_entry_documents', :v) "
                "ON CONFLICT (key) DO UPDATE SET value = :v"), {"v": v})
            await s_.commit()

    transport = ASGITransport(app=app)
    async with AsyncClient(transport=transport, base_url="http://svc") as ac:
        _ip = {"X-Real-IP": "203.0.113.108"}

        async def token(u, p):
            r = await ac.post("/auth/login", headers=_ip, json={"username": u, "password": p})
            return r.json().get("access_token")

        def H(t):
            return {"Authorization": f"Bearer {t}", **_ip}

        worker_t = await token("worker", "floor2026")
        hod_t = await token("hod", "hod2026")
        sup_t = await token("supervisor", "super2026")

        async with SessionLocal() as ses:
            site = (await ses.execute(_sqt(
                "SELECT COALESCE(\"Site_ID\",'HQ') FROM users WHERE username='worker'"))).scalar_one()

        await _set_gate("1")
        try:
            # gate: no document → receipt refused
            r = await ac.post("/entry/receipts", headers=H(worker_t), json={
                "Date": "2026-07-13", "SAP_Code": "1001", "Quantity": 1, "Site_ID": site})
            check("docs: receipt without document → 422",
                  r.status_code == 422 and "document" in str(r.json().get("detail", "")),
                  f"{r.status_code} {str(r.json())[:120]}")

            # upload (SK-only) then the same receipt passes and links
            r = await ac.post("/entry/attachments", headers=H(sup_t),
                              files={"file": ("n.pdf", b"%PDF-1.4 note", "application/pdf")},
                              data={"doc_type": "receipt", "site_id": site})
            check("docs: upload is store-keeper-gated (supervisor → 403)",
                  r.status_code == 403, f"{r.status_code}")
            r = await ac.post("/entry/attachments", headers=H(worker_t),
                              files={"file": ("note.pdf", b"%PDF-1.4 handwritten", "application/pdf")},
                              data={"doc_type": "receipt", "site_id": site, "doc_number": "DN-777"})
            check("docs: SK upload → 201 with doc number",
                  r.status_code == 201 and r.json().get("doc_number") == "DN-777",
                  f"{r.status_code} {str(r.json())[:120]}")
            aid = r.json()["id"]

            r = await ac.post("/entry/receipts", headers=H(worker_t), json={
                "Date": "2026-07-13", "SAP_Code": "1001", "Quantity": 1, "Site_ID": site,
                "Supplier": "AH-DOCS", "attachment_ids": [aid]})
            check("docs: receipt WITH document → 201", r.status_code == 201, f"{r.status_code} {r.text[:120]}")
            pid = r.json().get("pending_id")
            async with SessionLocal() as ses:
                linked = (await ses.execute(_sqt(
                    "SELECT entry_table, entry_date FROM entry_attachments WHERE id=:i"),
                    {"i": aid})).first()
            check("docs: attachment linked to the staged batch",
                  linked and linked.entry_table == "pending_receipts"
                  and str(linked.entry_date)[:10] == "2026-07-13", f"{tuple(linked) if linked else None}")

            # wrong doc_type refused; someone else's doc refused
            r = await ac.post("/entry/attachments", headers=H(worker_t),
                              files={"file": ("x.pdf", b"%PDF-1.4", "application/pdf")},
                              data={"doc_type": "consumption", "site_id": site})
            cons_aid = r.json()["id"]
            r = await ac.post("/entry/receipts", headers=H(worker_t), json={
                "Date": "2026-07-13", "SAP_Code": "1001", "Quantity": 1, "Site_ID": site,
                "attachment_ids": [cons_aid]})
            check("docs: doc_type mismatch → 422", r.status_code == 422, f"{r.status_code}")

            # library: HOD browses + downloads; SK sees own via mine=1
            r = await ac.get("/entry/attachments", headers=H(hod_t),
                             params={"doc_type": "receipt", "doc_number": "DN-777"})
            items = r.json().get("items", [])
            check("docs: HOD library lists the upload",
                  r.status_code == 200 and any(x["id"] == aid for x in items),
                  f"{r.status_code} n={len(items)}")
            r = await ac.get(f"/entry/attachments/{aid}/download", headers=H(hod_t))
            check("docs: HOD download streams the original bytes",
                  r.status_code == 200 and r.content.startswith(b"%PDF-1.4 handwritten"),
                  f"{r.status_code}")
            r = await ac.get("/entry/attachments", headers=H(worker_t))
            check("docs: full library is level ≥2 (SK → 403)", r.status_code == 403, f"{r.status_code}")
            r = await ac.get("/entry/attachments", headers=H(worker_t), params={"mine": "1"})
            check("docs: SK can list own uploads (mine=1)",
                  r.status_code == 200 and any(x["id"] == aid for x in r.json()["items"]),
                  f"{r.status_code}")

            # linked docs can't be deleted
            r = await ac.delete(f"/entry/attachments/{aid}", headers=H(worker_t))
            check("docs: linked attachment delete → 409", r.status_code == 409, f"{r.status_code}")

            # ── A2: return gates ─────────────────────────────────────────────
            r = await ac.post("/entry/attachments", headers=H(worker_t),
                              files={"file": ("ret.jpg", b"\xff\xd8jpegret", "image/jpeg")},
                              data={"doc_type": "return", "site_id": site})
            ret_aid = r.json()["id"]
            r = await ac.post("/entry/returns", headers=H(worker_t), json={
                "Date": "2026-07-13", "SAP_Code": "1001", "Quantity": 1, "Site_ID": site,
                "Reason": "AH-no-dn", "attachment_ids": [ret_aid]})
            check("return: missing Return DN No. → 422",
                  r.status_code == 422 and "DN" in str(r.json().get("detail", "")), f"{r.status_code}")

            # source-receipt discipline: qty capped + 30-day window
            async with SessionLocal() as ses:
                src_ok = (await ses.execute(_sqt(
                    'INSERT INTO receipts ("Date","SAP_Code","Quantity","Site_ID","DN_No") '
                    "VALUES (CURRENT_DATE::text, '1001', 5, :s, 'AH-SRC-DN') RETURNING id"),
                    {"s": site})).scalar_one()
                src_old = (await ses.execute(_sqt(
                    'INSERT INTO receipts ("Date","SAP_Code","Quantity","Site_ID","DN_No") '
                    "VALUES ((CURRENT_DATE - INTERVAL '90 days')::date::text, '1001', 5, :s, 'AH-OLD-DN') "
                    "RETURNING id"), {"s": site})).scalar_one()
                await ses.commit()

            base = {"Date": "2026-07-13", "SAP_Code": "1001", "Site_ID": site,
                    "Reason": "AH-ret", "Return_DN_No": "AH-RDN-1"}
            r = await ac.post("/entry/returns", headers=H(worker_t), json={
                **base, "Quantity": 9, "source_receipt_id": src_ok,
                "attachment_ids": [ret_aid]})
            check("return: qty above the source receipt → 422",
                  r.status_code == 422 and "exceeds" in str(r.json().get("detail", "")), f"{r.status_code}")
            r = await ac.post("/entry/returns", headers=H(worker_t), json={
                **base, "Quantity": 1, "source_receipt_id": src_old,
                "attachment_ids": [ret_aid]})
            check("return: >30-day-old receipt without justification → 422",
                  r.status_code == 422 and "older" in str(r.json().get("detail", "")), f"{r.status_code}")
            r = await ac.post("/entry/returns", headers=H(worker_t), json={
                **base, "Quantity": 1, "source_receipt_id": src_old,
                "override_reason": "vendor recall approved by PM",
                "attachment_ids": [ret_aid]})
            check("return: override justification → 201 + flagged for HOD",
                  r.status_code == 201, f"{r.status_code} {r.text[:120]}")
            ret_pid = r.json().get("pending_id")
            async with SessionLocal() as ses:
                flag = (await ses.execute(_sqt(
                    "SELECT override_required, override_reason, received_dn_no "
                    "FROM pending_returns WHERE id=:i"), {"i": ret_pid})).first()
            check("return: pending row carries override flag + source provenance",
                  flag and flag.override_required == 1 and "recall" in flag.override_reason
                  and flag.received_dn_no == "AH-OLD-DN", f"{tuple(flag) if flag else None}")

            # ── A4: WBS gate (bites only once the site has active WBS) ──────
            r = await ac.post("/hod/site-config/wbs", headers=H(worker_t),
                              json={"WBS_Number": "WBS-AH-1"})
            check("wbs: config is HOD-gated (SK → 403)", r.status_code == 403, f"{r.status_code}")
            r = await ac.post("/hod/site-config/wbs", headers=H(hod_t),
                              json={"WBS_Number": "WBS-AH-1", "Description": "AH test"})
            check("wbs: HOD adds a WBS → 201", r.status_code == 201, f"{r.status_code}")
            wid = r.json()["id"]
            # (WBS is checked before the document gate, so no attachment needed here)
            r = await ac.post("/entry/receipts", headers=H(worker_t), json={
                "Date": "2026-07-13", "SAP_Code": "1001", "Quantity": 1, "Site_ID": site})
            check("wbs: entry without WBS once configured → 422",
                  r.status_code == 422 and "WBS" in str(r.json().get("detail", "")), f"{r.status_code} {r.text[:120]}")
            r = await ac.post("/entry/attachments", headers=H(worker_t),
                              files={"file": ("w.pdf", b"%PDF-1.4 w", "application/pdf")},
                              data={"doc_type": "receipt", "site_id": site})
            aid2 = r.json()["id"]
            r = await ac.post("/entry/receipts", headers=H(worker_t), json={
                "Date": "2026-07-13", "SAP_Code": "1001", "Quantity": 1, "Site_ID": site,
                "wbs": "WBS-AH-1", "Bin_Location": "R2-S4", "attachment_ids": [aid2]})
            check("wbs: entry with an active WBS (+bin) → 201", r.status_code == 201, f"{r.status_code} {r.text[:160]}")
            pid2 = r.json().get("pending_id")
            async with SessionLocal() as ses:
                row = (await ses.execute(_sqt(
                    'SELECT wbs, "Bin_Location" FROM pending_receipts WHERE id=:i'),
                    {"i": pid2})).first()
            check("wbs: staged row stores wbs + bin",
                  row and row.wbs == "WBS-AH-1" and row.Bin_Location == "R2-S4",
                  f"{tuple(row) if row else None}")
            r = await ac.patch(f"/hod/site-config/wbs/{wid}", headers=H(hod_t),
                               params={"status": "closed"})
            check("wbs: HOD closes the WBS (gate lifts)", r.status_code == 200, f"{r.status_code}")

            # ── A3: MTC category detection (legacy Surface Shields rule) ────
            r = await ac.get("/entry/receipt-meta/1001", headers=H(worker_t))
            check("mtc: ordinary category is NOT rubber-gated",
                  r.status_code == 200 and r.json().get("is_rubber") is False, f"{r.text[:100]}")

            # cleanup: staged rows + uploads + wbs + source receipts
            async with SessionLocal() as ses:
                await ses.execute(_sqt("DELETE FROM pending_receipts WHERE id IN (:a,:b)"),
                                  {"a": pid or -1, "b": pid2 or -1})
                await ses.execute(_sqt("DELETE FROM pending_returns WHERE id=:i"), {"i": ret_pid or -1})
                await ses.execute(_sqt(
                    "DELETE FROM entry_attachments WHERE uploaded_by='worker' AND id >= :i"), {"i": aid})
                await ses.execute(_sqt("DELETE FROM wbs_master WHERE \"WBS_Number\" LIKE 'WBS-AH-%'"))
                await ses.execute(_sqt("DELETE FROM receipts WHERE \"DN_No\" IN ('AH-SRC-DN','AH-OLD-DN')"))
                await ses.execute(_sqt(
                    "DELETE FROM email_outbox WHERE event_key IN ('return_approved')")) \
                    if False else None
                await ses.commit()
        finally:
            await _set_gate("0")


async def test_sme_master_crud():
    """Suite AI — SME Phase S6 (cutover day): Master Data CRUD.

    The write endpoints COMMIT real rows (unique SVC6- prefixes), so this
    suite cleans up after itself. Audit rows are left in place (never delete
    system_audit_log; assertions are delta-counted)."""
    from sqlalchemy import text as _sqt

    async def _scalar(sql: str, **params):
        async with SessionLocal() as s:
            return (await s.execute(_sqt(sql), params)).scalar()

    transport = ASGITransport(app=app)
    async with AsyncClient(transport=transport, base_url="http://svc") as ac:
        # Own X-Real-IP: by this suite the shared client IP has long burned
        # through the 10/min login budget.
        _ip = {"X-Real-IP": "203.0.113.86"}

        async def token(u, p):
            r = await ac.post("/auth/login", json={"username": u, "password": p},
                              headers=_ip)
            return r.json().get("access_token")

        def H(t):
            return {"Authorization": f"Bearer {t}"}

        worker_t = await token("worker", "floor2026")       # store_keeper (0)
        sup_t = await token("supervisor", "super2026")      # supervisor (1)
        hod_t = await token("hod", "hod2026")               # hod @ CNCEC (2)
        admin_t = await token("admin", "admin2026")         # admin (4)

        # ── exact-lock {hod, admin} ─────────────────────────────────────────
        r = await ac.get("/sme/master/equipment", headers=H(worker_t))
        check("s6: worker → 403 on master reads", r.status_code == 403, f"got {r.status_code}")
        r = await ac.post("/sme/master/recipes", headers=H(sup_t),
                          json={"Lining_System_Code": "9901", "Material_Code": "X"})
        check("s6: supervisor → 403 on master writes", r.status_code == 403, f"got {r.status_code}")

        # ── equipment: create pairs with a progress seed ────────────────────
        audit_before = await _scalar(
            "SELECT COUNT(*) FROM system_audit_log WHERE action_type='SME_CREATE_EQUIPMENT'")
        r = await ac.post("/sme/master/equipment", headers=H(hod_t), json={
            "Equipment_Tag_No": "SVC6-TANK-1", "Lining_System_Code": "9901",
            "Surface_Area_SQM": 50, "Name": "S6 test tank", "Location": "SVC6-LOC"})
        check("s6: hod creates equipment → 201", r.status_code == 201, f"{r.status_code} {r.text[:120]}")
        eq_id = r.json().get("id")
        prog_orig = await _scalar(
            "SELECT \"Original_SQM\" FROM sme_sqm_progress WHERE \"Site_ID\"='CNCEC' "
            "AND \"Equipment_Tag_No\"='SVC6-TANK-1' AND \"Lining_System_Code\"='9901'")
        check("s6: create seeds the SQM-progress row (Original=50)",
              prog_orig is not None and abs(float(prog_orig) - 50) < 1e-9, f"got {prog_orig}")
        audit_after = await _scalar(
            "SELECT COUNT(*) FROM system_audit_log WHERE action_type='SME_CREATE_EQUIPMENT'")
        check("s6: create writes an SME_CREATE_EQUIPMENT audit",
              audit_after == audit_before + 1, f"{audit_before} → {audit_after}")

        r = await ac.post("/sme/master/equipment", headers=H(hod_t), json={
            "Equipment_Tag_No": "SVC6-TANK-1", "Lining_System_Code": "9901",
            "Surface_Area_SQM": 10})
        check("s6: duplicate (site, tag, code) → 409", r.status_code == 409, f"got {r.status_code}")
        r = await ac.post("/sme/master/equipment", headers=H(hod_t), json={
            "Equipment_Tag_No": "SVC6-TANK-X", "Lining_System_Code": "9901",
            "Surface_Area_SQM": 10, "site_id": "HQ"})
        check("s6: hod naming a foreign site → 403", r.status_code == 403, f"got {r.status_code}")
        r = await ac.post("/sme/master/equipment", headers=H(admin_t), json={
            "Equipment_Tag_No": "SVC6-TANK-X", "Lining_System_Code": "9901",
            "Surface_Area_SQM": 10})
        check("s6: admin without site_id → 422", r.status_code == 422, f"got {r.status_code}")

        # the new tag flows into the prediction model snapshot (CNCEC scope)
        r = await ac.get("/sme/model-snapshot", headers=H(hod_t))
        tags = {e["Equipment_Tag_No"] for e in r.json().get("equipment", [])}
        check("s6: model snapshot picks up the new equipment",
              r.status_code == 200 and "SVC6-TANK-1" in tags, f"got {r.status_code}")

        # update: legacy cell-edit semantics — progress is NOT cascaded
        r = await ac.patch(f"/sme/master/equipment/{eq_id}", headers=H(hod_t),
                           json={"Surface_Area_SQM": 60, "Name": "S6 renamed"})
        check("s6: hod edits equipment → 200", r.status_code == 200, f"got {r.status_code}")
        new_name = await _scalar(
            'SELECT "Name" FROM sme_equipment WHERE id=:i', i=eq_id)
        prog_after = await _scalar(
            "SELECT \"Original_SQM\" FROM sme_sqm_progress WHERE \"Site_ID\"='CNCEC' "
            "AND \"Equipment_Tag_No\"='SVC6-TANK-1' AND \"Lining_System_Code\"='9901'")
        check("s6: edit lands on the row; progress untouched (legacy semantics)",
              new_name == "S6 renamed" and abs(float(prog_after) - 50) < 1e-9,
              f"name={new_name} orig={prog_after}")

        # ── progress: None-preserving upsert ────────────────────────────────
        r = await ac.put("/sme/master/progress", headers=H(hod_t), json={
            "Equipment_Tag_No": "SVC6-TANK-1", "Lining_System_Code": "9901",
            "Done_SQM": 5})
        prog_row = None
        if r.status_code == 200:
            async with SessionLocal() as s:
                prog_row = (await s.execute(_sqt(
                    "SELECT \"Original_SQM\", \"Done_SQM\" FROM sme_sqm_progress "
                    "WHERE \"Site_ID\"='CNCEC' AND \"Equipment_Tag_No\"='SVC6-TANK-1' "
                    "AND \"Lining_System_Code\"='9901'"))).first()
        check("s6: progress upsert with only Done_SQM preserves Original_SQM",
              r.status_code == 200 and prog_row is not None
              and abs(float(prog_row[0]) - 50) < 1e-9 and abs(float(prog_row[1]) - 5) < 1e-9,
              f"{r.status_code} row={tuple(prog_row) if prog_row else None}")
        r = await ac.put("/sme/master/progress", headers=H(hod_t), json={
            "Equipment_Tag_No": "SVC6-TANK-1", "Lining_System_Code": "9901"})
        check("s6: progress upsert with no values → 422", r.status_code == 422, f"got {r.status_code}")

        # ── settings: add / dup / in-use guard / delete ─────────────────────
        r = await ac.post("/sme/master/settings/locations", headers=H(hod_t),
                          json={"value": "SVC6-LOC"})
        check("s6: add location → 201", r.status_code == 201, f"got {r.status_code}")
        r = await ac.post("/sme/master/settings/locations", headers=H(hod_t),
                          json={"value": "SVC6-LOC"})
        check("s6: duplicate location → 409", r.status_code == 409, f"got {r.status_code}")
        r = await ac.get("/sme/master/settings", headers=H(hod_t))
        check("s6: settings list carries the new value",
              r.status_code == 200 and "SVC6-LOC" in r.json().get("locations", []),
              f"{r.status_code} {r.text[:120]}")
        r = await ac.delete("/sme/master/settings/locations", headers=H(hod_t),
                            params={"value": "SVC6-LOC"})
        check("s6: delete refused while equipment uses the location (409)",
              r.status_code == 409, f"got {r.status_code}")
        r = await ac.delete("/sme/master/settings/nope", headers=H(hod_t),
                            params={"value": "x"})
        check("s6: unknown setting kind → 404", r.status_code == 404, f"got {r.status_code}")

        # ── equipment delete cascades progress; location then removable ─────
        r = await ac.delete(f"/sme/master/equipment/{eq_id}", headers=H(hod_t))
        prog_count = await _scalar(
            "SELECT COUNT(*) FROM sme_sqm_progress WHERE \"Equipment_Tag_No\"='SVC6-TANK-1'")
        check("s6: delete cascades the SQM-progress row",
              r.status_code == 200 and prog_count == 0,
              f"{r.status_code} progress={prog_count}")
        r = await ac.delete("/sme/master/settings/locations", headers=H(hod_t),
                            params={"value": "SVC6-LOC"})
        check("s6: location removable once unused", r.status_code == 200, f"got {r.status_code}")
        r = await ac.delete(f"/sme/master/equipment/{eq_id}", headers=H(hod_t))
        check("s6: deleting a gone row → 404", r.status_code == 404, f"got {r.status_code}")

        # ── recipes (global) ────────────────────────────────────────────────
        r = await ac.post("/sme/master/recipes", headers=H(hod_t), json={
            "Lining_System_Code": "9901", "Material_Code": "SVC6-MAT-1",
            "Material_Name": "S6 resin", "UOM": "KG", "For_1_SQM": 2.5})
        check("s6: create recipe → 201", r.status_code == 201, f"{r.status_code} {r.text[:120]}")
        rec_id = r.json().get("id")
        r = await ac.post("/sme/master/recipes", headers=H(hod_t), json={
            "Lining_System_Code": "9901", "Material_Code": "SVC6-MAT-1"})
        check("s6: duplicate (code, material) recipe → 409", r.status_code == 409, f"got {r.status_code}")
        r = await ac.patch(f"/sme/master/recipes/{rec_id}", headers=H(hod_t),
                           json={"For_1_SQM": 3.25})
        got = await _scalar('SELECT "For_1_SQM" FROM sme_recipe WHERE id=:i', i=rec_id)
        check("s6: recipe patch lands", r.status_code == 200 and abs(float(got) - 3.25) < 1e-9,
              f"{r.status_code} got={got}")
        r = await ac.delete(f"/sme/master/recipes/{rec_id}", headers=H(hod_t))
        check("s6: recipe delete → 200", r.status_code == 200, f"got {r.status_code}")
        r = await ac.patch("/sme/master/recipes/999999999", headers=H(hod_t),
                           json={"UOM": "L"})
        check("s6: patch on a missing recipe → 404", r.status_code == 404, f"got {r.status_code}")

        # ── materials: seed-only upsert (Canon Rule 2: ERP inventory untouched)
        inv_before = await _scalar("SELECT COUNT(*) FROM inventory")
        r = await ac.post("/sme/master/materials", headers=H(hod_t), json={
            "Material_Code": "SVC6-MAT-1", "Material_Name": "S6 resin",
            "UOM": "KG", "Initial_Available_Qty": 100})
        check("s6: material seed create → 201", r.status_code == 201, f"{r.status_code} {r.text[:120]}")
        r = await ac.post("/sme/master/materials", headers=H(hod_t), json={
            "Material_Code": "SVC6-MAT-1", "Material_Name": "S6 resin v2",
            "UOM": "KG", "Initial_Available_Qty": 130})
        n_seed = await _scalar(
            "SELECT COUNT(*) FROM sme_inventory_seed WHERE \"Material_Code\"='SVC6-MAT-1'")
        qty = await _scalar(
            "SELECT \"Initial_Available_Qty\" FROM sme_inventory_seed "
            "WHERE \"Material_Code\"='SVC6-MAT-1'")
        check("s6: re-POST is an upsert (1 row, re-baselined to 130)",
              r.status_code == 201 and n_seed == 1 and abs(float(qty) - 130) < 1e-9,
              f"{r.status_code} rows={n_seed} qty={qty}")
        r = await ac.get("/sme/master/materials", headers=H(hod_t))
        mat = next((m for m in r.json().get("items", [])
                    if m["material_code"] == "SVC6-MAT-1"), None)
        check("s6: materials grid derives availability for the seed",
              mat is not None and abs(float(mat["available_qty"]) - 130) < 1e-9,
              f"row={mat}")
        r = await ac.patch("/sme/master/materials/SVC6-MAT-1", headers=H(hod_t),
                           json={"Vendor": "S6 Vendor"})
        check("s6: material patch → 200", r.status_code == 200, f"got {r.status_code}")
        r = await ac.patch("/sme/master/materials/SVC6-NOPE", headers=H(hod_t),
                           json={"Vendor": "x"})
        check("s6: patch on a missing material → 404", r.status_code == 404, f"got {r.status_code}")
        r = await ac.delete("/sme/master/materials/SVC6-MAT-1", headers=H(hod_t))
        inv_after = await _scalar("SELECT COUNT(*) FROM inventory")
        check("s6: material delete → 200; ERP inventory untouched throughout",
              r.status_code == 200 and inv_before == inv_after,
              f"{r.status_code} inv {inv_before}→{inv_after}")

        # cleanup — remove every SVC6- artifact this suite committed
        # (audit rows stay: system_audit_log is append-only by contract).
        async with SessionLocal() as s:
            await s.execute(_sqt(
                "DELETE FROM sme_sqm_progress WHERE \"Equipment_Tag_No\" LIKE 'SVC6-%'"))
            await s.execute(_sqt(
                "DELETE FROM sme_equipment WHERE \"Equipment_Tag_No\" LIKE 'SVC6-%'"))
            await s.execute(_sqt(
                "DELETE FROM sme_recipe WHERE \"Material_Code\" LIKE 'SVC6-%'"))
            await s.execute(_sqt(
                "DELETE FROM sme_inventory_seed WHERE \"Material_Code\" LIKE 'SVC6-%'"))
            await s.execute(_sqt(
                "DELETE FROM system_settings WHERE value LIKE 'SVC6-%' "
                "AND category IN ('sme_location','sme_equipment_type')"))
            await s.commit()


def _xlsx(sheets: dict[str, list[list]]) -> bytes:
    """Build a small workbook in memory: {sheet_name: [row, row, …]}."""
    import io as _io

    import openpyxl as _px
    wb = _px.Workbook()
    wb.remove(wb.active)
    for name, rows in sheets.items():
        ws = wb.create_sheet(title=name)
        for row in rows:
            ws.append(row)
    buf = _io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


async def test_bulk_import():
    """Suite AJ — Bulk Excel Import (dry-run/commit, upsert-only, reconcile).

    Hermetic: every workbook is built in memory with unique SVCJ- keys and
    all committed rows are cleaned up at the end (audit rows stay)."""
    from sqlalchemy import text as _sqt

    async def _scalar(sql: str, **params):
        async with SessionLocal() as s:
            return (await s.execute(_sqt(sql), params)).scalar()

    _INV_HDR = ["Sl. No.", "SAP CODE", "Material Code", "Equipment Description",
                "UOM", "Category", "Opening Stock", "Receipt", "Consumption",
                "Return", "Current Stock", "Minimum Qty"]
    _RCT_HDR = ["Date ", "SAP CODE", "Material Code", "Equipment Description",
                "UOM", "Qty.", "Serial No.", "PR#", "WBS#", "Location",
                "Vehicle No.", "Driver Name", "DN. No.", "Pallet No.",
                "Mob. From", "Prepared by"]
    _TITLE = ["CNCEC PROJECT Equipement and"]

    transport = ASGITransport(app=app)
    async with AsyncClient(transport=transport, base_url="http://svc") as ac:
        _ip = {"X-Real-IP": "203.0.113.87"}

        async def token(u, p):
            r = await ac.post("/auth/login", json={"username": u, "password": p},
                              headers=_ip)
            return r.json().get("access_token")

        def H(t):
            return {"Authorization": f"Bearer {t}"}

        def up(data, name="t.xlsx"):
            return {"file": (name, data,
                             "application/vnd.openxmlformats-officedocument"
                             ".spreadsheetml.sheet")}

        worker_t = await token("worker", "floor2026")
        hod_t = await token("hod", "hod2026")
        admin_t = await token("admin", "admin2026")

        # ── guards ──────────────────────────────────────────────────────────
        wb = _xlsx({"Sheet1": [["Lining_System_Code", "Material_Code", "For_1_SQM"]]})
        r = await ac.post("/import/sme-recipes", headers=H(worker_t), files=up(wb))
        check("aj: worker → 403 on SME import", r.status_code == 403, f"got {r.status_code}")
        r = await ac.post("/import/inventory", headers=H(hod_t), files=up(wb))
        check("aj: hod → 403 on inventory import (admin-only)",
              r.status_code == 403, f"got {r.status_code}")
        r = await ac.post("/import/nope", headers=H(hod_t), files=up(wb))
        check("aj: unknown kind → 404", r.status_code == 404, f"got {r.status_code}")
        r = await ac.post("/import/sme-recipes", headers=H(hod_t),
                          files={"file": ("t.xlsx", b"not a zip", "application/octet-stream")})
        check("aj: non-xlsx payload → 422", r.status_code == 422, f"got {r.status_code}")

        # ── inventory master: insert + canonicalised category ───────────────
        mat_1001 = await _scalar(
            "SELECT \"Material_Code\" FROM inventory WHERE \"SAP_Code\"='1001'")
        inv_wb = _xlsx({"Inventory": [
            _TITLE, _INV_HDR,
            ["901", "SVCJ-1", "GI-SVCJ-1", "SVCJ test item", "EA",
             "Surface Shield", 3, 0, 0, 0, 3, 1],
            ["902", "SVCJ-2", mat_1001, "conflict row", "EA", "Office",
             0, 0, 0, 0, 0, 0],
        ]})
        r = await ac.post("/import/inventory", headers=H(admin_t), files=up(inv_wb),
                          params={"site_id": "CNCEC"})
        j = r.json() if r.status_code == 200 else {}
        check("aj: inventory dry-run plans 2 inserts; foreign Material_Code "
              "dropped with a warning (owner not re-mapped)",
              r.status_code == 200 and j.get("summary", {}).get("inserts") == 2
              and not j.get("committed")
              and any("stays with SAP 1001" in w for w in j.get("warnings", [])),
              f"{r.status_code} {j.get('summary')} warn={j.get('warnings')}")
        n = await _scalar("SELECT COUNT(*) FROM inventory WHERE \"SAP_Code\" LIKE 'SVCJ-%'")
        check("aj: dry-run wrote nothing", n == 0, f"got {n}")
        r = await ac.post("/import/inventory", headers=H(admin_t), files=up(inv_wb),
                          params={"site_id": "CNCEC", "commit": "true"})
        cat = await _scalar(
            "SELECT \"Category\" FROM inventory WHERE \"SAP_Code\"='SVCJ-1'")
        check("aj: commit inserts + canonicalises 'Surface Shield' → 'Surface Shields'"
              " (the MTC gate keeps matching)",
              r.status_code == 200 and cat == "Surface Shields", f"{r.status_code} cat={cat}")
        r = await ac.get("/entry/receipt-meta/SVCJ-1", headers=H(worker_t))
        check("aj: imported Surface-Shields item IS rubber/MTC-gated",
              r.status_code == 200 and r.json().get("is_rubber") is True, f"{r.text[:100]}")
        inv_wb2 = _xlsx({"Inventory": [
            _TITLE, _INV_HDR,
            ["901", "SVCJ-1", "GI-SVCJ-1", "SVCJ test item", "EA",
             "Surface Shield", 3, 0, 0, 0, 3, 7],
        ]})
        r = await ac.post("/import/inventory", headers=H(admin_t), files=up(inv_wb2),
                          params={"site_id": "CNCEC", "commit": "true"})
        mq = await _scalar(
            "SELECT \"Minimum_Qty\" FROM inventory WHERE \"SAP_Code\"='SVCJ-1'")
        check("aj: re-import updates changed fields only (Minimum_Qty 1→7)",
              r.status_code == 200 and abs(float(mq) - 7) < 1e-9, f"{r.status_code} mq={mq}")

        # ── ledger reconcile: insert → exact-skip → qty correction ──────────
        led_wb = _xlsx({"Receipt Log": [
            _TITLE, _RCT_HDR,
            ["2026-07-01 00:00:00", "SVCJ-1", "GI-SVCJ-1", "SVCJ test item",
             "EA", 5, None, None, None, None, None, None, "SVCJ-DN1", None, None, "svc"],
            ["2026-07-02 00:00:00", "SVCJ-NOPE", None, None, "EA", 1,
             None, None, None, None, None, None, "SVCJ-DN2", None, None, "svc"],
            ["2026-07-03 00:00:00", "SVCJ-1", None, None, "EA", 0,
             None, None, None, None, None, None, "SVCJ-DN3", None, None, "svc"],
        ]})
        r = await ac.post("/import/ledger", headers=H(admin_t), files=up(led_wb),
                          params={"site_id": "CNCEC", "commit": "true"})
        j = r.json() if r.status_code == 200 else {}
        sec = j.get("summary", {}).get("receipts", {})
        check("aj: ledger commit — 1 insert, 1 orphan-SAP reject, 1 zero-skip",
              r.status_code == 200 and sec.get("inserts") == 1
              and sec.get("zero_skipped") == 1 and len(j.get("rejects", [])) == 1,
              f"{r.status_code} {sec} rejects={len(j.get('rejects', []))}")
        r = await ac.post("/import/ledger", headers=H(admin_t), files=up(led_wb),
                          params={"site_id": "CNCEC", "commit": "true"})
        sec = (r.json().get("summary", {}) or {}).get("receipts", {})
        check("aj: re-import is idempotent (exact multiset match, 0 inserts)",
              sec.get("inserts") == 0 and sec.get("matched") == 1, f"{sec}")
        led_wb2 = _xlsx({"Receipt Log": [
            _TITLE, _RCT_HDR,
            ["2026-07-01 00:00:00", "SVCJ-1", None, None, "EA", 3,
             None, None, None, None, None, None, "SVCJ-DN1", None, None, "svc"],
        ]})
        r = await ac.post("/import/ledger", headers=H(admin_t), files=up(led_wb2),
                          params={"site_id": "CNCEC", "commit": "true"})
        sec = (r.json().get("summary", {}) or {}).get("receipts", {})
        qty = await _scalar(
            "SELECT \"Quantity\" FROM receipts WHERE \"DN_No\"='SVCJ-DN1'")
        check("aj: workbook qty change lands as a CORRECTION (5→3), not a dup row",
              sec.get("corrections") == 1 and abs(float(qty) - 3) < 1e-9,
              f"{sec} qty={qty}")
        n = await _scalar("SELECT COUNT(*) FROM receipts WHERE \"DN_No\" LIKE 'SVCJ-%'")
        check("aj: exactly one SVCJ ledger row exists after all passes", n == 1, f"got {n}")

        # ── 2026-07-14 workbook restructure: columns resolve by NAME ────────
        # Reordered headers + operator-added columns (Current Location /
        # Audit …) must map identically and surface an "ignored" warning.
        inv_wb3 = _xlsx({"Inventory": [
            _TITLE,
            ["Category", "SAP CODE", "Equipment Description", "UOM",
             "Current Location", "Opening Stock", "Material Code",
             "Minimum Qty", "Audit 13/06/26", "Sl. No."],
            ["Surface Shield", "SVCJ-1", "SVCJ test item", "EA", "Rack A",
             3, "GI-SVCJ-1", 7, "ok", "901"],
        ]})
        r = await ac.post("/import/inventory", headers=H(admin_t), files=up(inv_wb3),
                          params={"site_id": "CNCEC"})
        j = r.json() if r.status_code == 200 else {}
        check("aj: reordered inventory headers map by NAME (row unchanged) + "
              "unknown columns warned",
              r.status_code == 200 and j.get("summary", {}).get("unchanged") == 1
              and j.get("summary", {}).get("updates") == 0
              and any("Current Location" in w and "Audit 13/06/26" in w
                      for w in j.get("warnings", [])),
              f"{r.status_code} {j.get('summary')} warn={j.get('warnings')}")
        led_wb3 = _xlsx({
            "Receipt Log": [
                _TITLE,
                ["Qty.", "Date ", "DN. No.", "SAP CODE", "Remarks ",
                 "DN. Copy", "Mystery Col"],
                [4, "2026-07-04 00:00:00", "SVCJ-DN9", "SVCJ-1", "note-9",
                 "Yes", "zzz"],
            ],
            "Consumption Log": [
                _TITLE,
                ["Date ", "SAP CODE", "Qty.", "Tank No.", "Cons. Paper No."],
                ["2026-07-05 00:00:00", "SVCJ-1", 2, "SVCJ-TK1", "CP-9"],
            ]})
        r = await ac.post("/import/ledger", headers=H(admin_t), files=up(led_wb3),
                          params={"site_id": "CNCEC"})
        j = r.json() if r.status_code == 200 else {}
        rins = (j.get("preview", {}).get("receipts", {}) or {}).get("inserts", [])
        check("aj: reordered Receipt Log captures Remarks + DN_Copy; truly "
              "unknown column warned",
              r.status_code == 200 and len(rins) == 1
              and rins[0].get("Remarks") == "note-9"
              and rins[0].get("DN_Copy") == "Yes"
              and any("Mystery Col" in w for w in j.get("warnings", [])),
              f"{r.status_code} ins={rins} warn={j.get('warnings')}")
        csec = j.get("summary", {}).get("consumption", {})
        check("aj: 'Cons. Paper No.' is spec-ignored (no DB home, no warning "
              "noise); reordered Consumption Log still plans the insert",
              csec.get("inserts") == 1
              and not any("Cons. Paper No." in w for w in j.get("warnings", [])),
              f"{csec} warn={j.get('warnings')}")

        # ── sme-equipment: area aggregation + Name identity + Done preserved ─
        _EQ_HDR = ["Sl. #", "Project", "WBS #", "Sub_Location", "Location",
                   "Type", "Substrate", "Equipment_Tag_No.", "Name", "Drawing #",
                   "Design", "Dia / L", "Ht. /W", "Equipment Total SQM",
                   "Remaraks", "Lining_System_Code", "Lining_System_Short_Name",
                   "Lining_Type", "Lining_System", "Material Spec.",
                   "Lining_Area/location", "Surface_Area_SQM"]

        def eq_rows(sqm_a, sqm_b):
            return [_EQ_HDR,
                    ["1", "SVCJ", None, None, "BROWN FIELD", "ME", "TANK",
                     "SVCJ-T1", "svc tank", None, None, None, None, 25, None,
                     "9905", "SVCJ30", None, None, None, "Bottom", sqm_a],
                    ["2", "SVCJ", None, None, "Brown Field ", "ME", "TANK",
                     "SVCJ-T1", "svc tank", None, None, None, None, 25, None,
                     "9905", "SVCJ30", None, None, None, "Shell", sqm_b],
                    ["3", "SVCJ", None, None, "TRAIN J", "CIVIL", "AREA",
                     None, "SVCJ Area", None, None, None, None, 5, None,
                     "9906", None, None, None, None, None, 5],
                    ["4", "SVCJ", None, None, "TRAIN J", "ME", "TANK",
                     "SVCJ-T2", "tbc tank", None, None, None, None, 1, None,
                     "To_Be_Confirmed_LSC", None, None, None, None, None, 1]]

        eq_wb = _xlsx({"Data Input": eq_rows(10, 15)})
        r = await ac.post("/import/sme-equipment", headers=H(hod_t), files=up(eq_wb),
                          params={"commit": "true"})
        j = r.json() if r.status_code == 200 else {}
        sqm = await _scalar(
            "SELECT \"Surface_Area_SQM\" FROM sme_equipment "
            "WHERE \"Equipment_Tag_No\"='SVCJ-T1' AND \"Lining_System_Code\"='9905'")
        loc = await _scalar(
            "SELECT \"Location\" FROM sme_equipment WHERE \"Equipment_Tag_No\"='SVCJ-T1'")
        check("aj: equipment area rows aggregate (10+15=25) + Location canonicalised",
              r.status_code == 200 and sqm is not None and abs(float(sqm) - 25) < 1e-9
              and loc == "Brown Field", f"{r.status_code} sqm={sqm} loc={loc}")
        name_row = await _scalar(
            "SELECT COUNT(*) FROM sme_equipment WHERE \"Equipment_Tag_No\"='SVCJ Area'")
        tbc = await _scalar(
            "SELECT COUNT(*) FROM sme_equipment WHERE \"Equipment_Tag_No\"='SVCJ-T2'")
        check("aj: Name-identity area imported; non-numeric code row skipped",
              name_row == 1 and tbc == 0, f"area={name_row} tbc={tbc}")
        prog = await _scalar(
            "SELECT \"Original_SQM\" FROM sme_sqm_progress "
            "WHERE \"Equipment_Tag_No\"='SVCJ-T1' AND \"Lining_System_Code\"='9905'")
        check("aj: import seeds the SQM-progress baseline",
              prog is not None and abs(float(prog) - 25) < 1e-9, f"got {prog}")
        async with SessionLocal() as s:  # simulate real progress before re-import
            await s.execute(_sqt(
                "UPDATE sme_sqm_progress SET \"Done_SQM\"=4 "
                "WHERE \"Equipment_Tag_No\"='SVCJ-T1' AND \"Lining_System_Code\"='9905'"))
            await s.commit()
        eq_wb2 = _xlsx({"Data Input": eq_rows(12, 18)})
        r = await ac.post("/import/sme-equipment", headers=H(hod_t), files=up(eq_wb2),
                          params={"commit": "true"})
        row = None
        async with SessionLocal() as s:
            row = (await s.execute(_sqt(
                "SELECT \"Original_SQM\", \"Done_SQM\" FROM sme_sqm_progress "
                "WHERE \"Equipment_Tag_No\"='SVCJ-T1' AND \"Lining_System_Code\"='9905'"
            ))).first()
        check("aj: re-import re-baselines Original_SQM (30) and PRESERVES Done_SQM (4)",
              r.status_code == 200 and row is not None
              and abs(float(row[0]) - 30) < 1e-9 and abs(float(row[1]) - 4) < 1e-9,
              f"{r.status_code} row={tuple(row) if row else None}")

        # ── sme-recipes ──────────────────────────────────────────────────────
        rec_hdr = ["Sl. #", "Lining_System_Code", "Substrate", "Lining_System",
                   "System Key's", "Lining_Thicknes", "Lining_System_Short_Name",
                   "Lining_Type", "Material_Code", "Material_Description",
                   "Material_Name", "For_1_SQM", "UOM", "PACKAGE SIZE"]
        rec_wb = _xlsx({"LINING SYSTEM MATERIAL CONSM": [
            rec_hdr,
            ["1", "9905", "Steel", "SVCJ Lining", "SVCJ", "30 mm", "SVCJ30",
             "Brick", "SVCJ-MAT-1", "Primer", "SVCJ Primer", 2.5, "KG", "25"],
            ["2", "9905", "Steel", "SVCJ Lining", "SVCJ", "30 mm", "SVCJ30",
             "Brick", "SVCJ-MAT-1", "dup", "dup", 9, "KG", "25"],
        ]})
        r = await ac.post("/import/sme-recipes", headers=H(hod_t), files=up(rec_wb),
                          params={"commit": "true"})
        j = r.json() if r.status_code == 200 else {}
        got = await _scalar(
            "SELECT \"For_1_SQM\" FROM sme_recipe WHERE \"Material_Code\"='SVCJ-MAT-1'")
        check("aj: recipe import inserts once; in-file repeat skipped first-wins",
              r.status_code == 200 and j["summary"]["inserts"] == 1
              and abs(float(got) - 2.5) < 1e-9
              and any("first occurrence wins" in w for w in j.get("warnings", [])),
              f"{r.status_code} {j.get('summary')} warn={j.get('warnings')} got={got}")
        n2 = await _scalar(
            "SELECT COUNT(*) FROM sme_recipe WHERE \"Material_Code\" LIKE 'SVCJ-%'")
        check("aj: comma-split + first-wins leaves exactly one SVCJ recipe row",
              n2 == 1, f"got {n2}")

        # ── sme-materials: PO lines aggregate on Material_Code ──────────────
        mat_hdr = ["Item", "Vendor/supplying plant", "Purchasing Document",
                   "Document Date", "Material_Code", "Material_Name", "Nature",
                   "UOM", "Available_Qty", "Ordered_Qty"]
        mat_wb = _xlsx({"Materials": [
            mat_hdr,
            ["1", "SVCJ Vendor", "4700000001", "2026-01-05", "SVCJ-MAT-1",
             "SVCJ Primer", "Liquid", "KG", 100, 40],
            ["2", "SVCJ Vendor", "4700000002", "2026-02-05", "SVCJ-MAT-1",
             "SVCJ Primer", "Liquid", "KG", 30, None],
        ]})
        r = await ac.post("/import/sme-materials", headers=H(hod_t), files=up(mat_wb),
                          params={"commit": "true"})
        row = None
        async with SessionLocal() as s:
            row = (await s.execute(_sqt(
                "SELECT \"Initial_Available_Qty\", \"Initial_Ordered_Qty\", "
                "\"Document_Date\" FROM sme_inventory_seed "
                "WHERE \"Material_Code\"='SVCJ-MAT-1'"))).first()
        check("aj: materials PO lines aggregate (130/40) + latest Document_Date wins",
              r.status_code == 200 and row is not None
              and abs(float(row[0]) - 130) < 1e-9 and abs(float(row[1]) - 40) < 1e-9
              and str(row[2]) == "2026-02-05", f"{r.status_code} row={tuple(row) if row else None}")

        # ── SAP-aware SME files (2026-07-18 layout): (code, mat, SAP) identity ─
        rec_hdr2 = rec_hdr[:9] + ["SAP_Code"] + rec_hdr[9:]
        rec_wb2 = _xlsx({"LINING SYSTEM MATERIAL CONSM": [
            rec_hdr2,
            ["1", "9906", "Steel", "SVCJ PU", "SVCJ", "3 mm", "SVCJPU3",
             "PU", "SVCJ-MAT-2", "SVCJ-S1", "Comp-A", "SVCJ PU", 0.5, "KG", None],
            ["2", "9906", "Steel", "SVCJ PU", "SVCJ", "3 mm", "SVCJPU3",
             "PU", "SVCJ-MAT-2", "SVCJ-S1-1", "Comp-B", "SVCJ PU", 0.7, "KG", None],
            ["3", "9906", "Steel", "SVCJ PU", "SVCJ", "3 mm", "SVCJPU3",
             "PU", "SVCJ-MAT-3", "SVCJ-S2", "Primer", "SVCJ Primer", 0.2, "KG", None],
            ["4", "9906", "Steel", "SVCJ PU", "SVCJ", "3 mm", "SVCJPU3",
             "PU", "SVCJ-MAT-3", "SVCJ-S2", "Primer", "SVCJ Primer", 0.3, "KG", None],
        ]})
        r = await ac.post("/import/sme-recipes", headers=H(hod_t), files=up(rec_wb2),
                          params={"commit": "true"})
        j = r.json() if r.status_code == 200 else {}
        async with SessionLocal() as s:
            comp = (await s.execute(_sqt(
                "SELECT \"SAP_Code\", \"For_1_SQM\" FROM sme_recipe "
                "WHERE \"Material_Code\"='SVCJ-MAT-2' ORDER BY \"SAP_Code\""))).all()
            coat = (await s.execute(_sqt(
                "SELECT \"For_1_SQM\" FROM sme_recipe "
                "WHERE \"Material_Code\"='SVCJ-MAT-3'"))).scalar()
        check("aj: SAP-aware recipe file keeps variant-SAP component lines apart",
              r.status_code == 200 and j.get("summary", {}).get("inserts") == 3
              and [t[0] for t in comp] == ["SVCJ-S1", "SVCJ-S1-1"],
              f"{r.status_code} {j.get('summary')} comp={comp}")
        check("aj: repeated (code, mat, SAP) coat lines merge — For_1_SQM sums",
              coat is not None and abs(float(coat) - 0.5) < 1e-9
              and any("coat line" in w for w in j.get("warnings", [])),
              f"coat={coat} warn={j.get('warnings')}")
        mat_hdr2 = mat_hdr[:5] + ["SAP_Code"] + mat_hdr[5:]
        mat_wb2 = _xlsx({"Materials": [
            mat_hdr2,
            ["1", "SVCJ Vendor", "4700000003", "2026-03-01", "SVCJ-MAT-2",
             "SVCJ-S1", "SVCJ PU", "Liquid", "KG", 10, 5],
            ["2", "SVCJ Vendor", "4700000003", "2026-03-01", "SVCJ-MAT-2",
             "SVCJ-S1-1", "SVCJ PU", "Liquid", "KG", 4, None],
        ]})
        r = await ac.post("/import/sme-materials", headers=H(hod_t), files=up(mat_wb2),
                          params={"commit": "true"})
        sap_join = await _scalar(
            "SELECT \"SAP_Code\" FROM sme_inventory_seed "
            "WHERE \"Material_Code\"='SVCJ-MAT-2'")
        check("aj: materials seed records the distinct variant SAP list",
              r.status_code == 200 and sap_join == "SVCJ-S1, SVCJ-S1-1",
              f"{r.status_code} sap={sap_join}")

        audits = await _scalar(
            "SELECT COUNT(DISTINCT action_type) FROM system_audit_log "
            "WHERE action_type LIKE 'BULK_IMPORT_%'")
        check("aj: every committed kind wrote its audit action", audits >= 4, f"got {audits}")

        # cleanup — every SVCJ- artifact (audit rows stay, append-only contract)
        async with SessionLocal() as s:
            await s.execute(_sqt("DELETE FROM receipts WHERE \"DN_No\" LIKE 'SVCJ-%'"))
            await s.execute(_sqt("DELETE FROM inventory WHERE \"SAP_Code\" LIKE 'SVCJ-%'"))
            await s.execute(_sqt(
                "DELETE FROM sme_sqm_progress WHERE \"Equipment_Tag_No\" LIKE 'SVCJ%'"))
            await s.execute(_sqt(
                "DELETE FROM sme_equipment WHERE \"Equipment_Tag_No\" LIKE 'SVCJ%'"))
            await s.execute(_sqt(
                "DELETE FROM sme_recipe WHERE \"Material_Code\" LIKE 'SVCJ-%'"))
            await s.execute(_sqt(
                "DELETE FROM sme_inventory_seed WHERE \"Material_Code\" LIKE 'SVCJ-%'"))
            await s.commit()


async def test_ocr_doc_assist():
    """Suite AK — parity C3: OCR doc assist on the entry forms
    (POST /ai/jobs/from-attachment) + ask-data SCHEMA_HINT pins."""
    import io as _io

    from sqlalchemy import text as _sqt

    from .ai import analytics as _ana
    from .ai import client as aic

    # ── prompt pins: the NL lane must know the post-injection schema/terms ──
    for needle in ('"DN_No"', '"Opening_Stock"', "Surface Shields",
                   "ILIKE 'surface shield%'", '"WBS"', '"Reason"'):
        check(f"ak: SCHEMA_HINT teaches {needle}", needle in _ana.SCHEMA_HINT)

    def tiny_jpeg() -> bytes:
        from PIL import Image
        buf = _io.BytesIO()
        Image.new("RGB", (32, 32), (200, 180, 40)).save(buf, format="JPEG")
        return buf.getvalue()

    transport = ASGITransport(app=app)
    async with AsyncClient(transport=transport, base_url="http://svc") as ac:
        _ip = {"X-Real-IP": "203.0.113.88"}

        async def token(u, p):
            r = await ac.post("/auth/login", json={"username": u, "password": p},
                              headers=_ip)
            return r.json().get("access_token")

        def H(t):
            return {"Authorization": f"Bearer {t}"}

        worker_t = await token("worker", "floor2026")
        hod_t = await token("hod", "hod2026")

        async def poll(jid):
            import asyncio as _aio
            for _ in range(80):
                r = await ac.get(f"/ai/jobs/{jid}", headers=H(worker_t))
                j = r.json()
                if j["status"] in ("done", "error"):
                    return j
                await _aio.sleep(0.05)
            return {"status": "timeout"}

        # attach a photo the normal entry-form way
        r = await ac.post("/entry/attachments", headers=H(worker_t),
                          files={"file": ("dn.jpg", tiny_jpeg(), "image/jpeg")},
                          data={"doc_type": "receipt", "site_id": "CNCEC"})
        check("ak: photo attachment uploads", r.status_code in (200, 201), f"{r.status_code} {r.text[:100]}")
        aid = r.json()["id"]

        r = await ac.post("/ai/jobs/from-attachment", headers=H(hod_t),
                          json={"attachment_id": aid, "kind": "ocr_delivery_note"})
        check("ak: hod → 403 (SK exact-lock, like the legacy Daily Issue Log)",
              r.status_code == 403, f"got {r.status_code}")
        r = await ac.post("/ai/jobs/from-attachment", headers=H(worker_t),
                          json={"attachment_id": 999999999, "kind": "ocr_delivery_note"})
        check("ak: unknown attachment → 404", r.status_code == 404, f"got {r.status_code}")
        r = await ac.post("/ai/jobs/from-attachment", headers=H(worker_t),
                          json={"attachment_id": aid, "kind": "nope"})
        check("ak: unknown kind → 422", r.status_code == 422, f"got {r.status_code}")

        real_gen, real_health = aic.generate, aic.health
        try:
            async def up_health():
                return True

            async def fake_models():
                return [aic.MODEL_VISION, aic.MODEL_CHAT, aic.MODEL_CODER]

            async def fake_vision_dn(model, prompt, **kw):
                import json as _json
                return _json.dumps({
                    "header": {"DN_No": "15733", "Date": "2026-07-13",
                               "Mob_From": "GI - ABU HADRIYAH",
                               "Driver_Name": "Imran", "Vehicle_No": "3909",
                               "Prepared_by": "H", "Mob_To": "CNCEC"},
                    "items": []})
            real_list = aic.list_models
            aic.health, aic.generate, aic.list_models = up_health, fake_vision_dn, fake_models
            r = await ac.post("/ai/jobs/from-attachment", headers=H(worker_t),
                              json={"attachment_id": aid, "kind": "ocr_delivery_note"})
            check("ak: from-attachment job accepted (202)",
                  r.status_code == 202, f"{r.status_code} {r.text[:120]}")
            j = await poll(r.json().get("job_id"))
            check("ak: stored attachment bytes → DN header extracted",
                  j.get("status") == "done"
                  and j.get("result", {}).get("header", {}).get("DN_No") == "15733",
                  str(j)[:160])
        finally:
            aic.generate, aic.health, aic.list_models = real_gen, real_health, real_list

        # a PDF attachment can't be OCR'd — friendly 422, not a dead job
        r = await ac.post("/entry/attachments", headers=H(worker_t),
                          files={"file": ("note.pdf", b"%PDF-1.4 fake", "application/pdf")},
                          data={"doc_type": "receipt", "site_id": "CNCEC"})
        pdf_aid = r.json().get("id")
        r = await ac.post("/ai/jobs/from-attachment", headers=H(worker_t),
                          json={"attachment_id": pdf_aid, "kind": "ocr_delivery_note"})
        check("ak: PDF attachment → friendly 422 (photograph instead)",
              r.status_code == 422 and "photograph" in r.text.lower(),
              f"{r.status_code} {r.text[:120]}")

        # cleanup — test attachments + jobs
        async with SessionLocal() as s:
            await s.execute(_sqt(
                "DELETE FROM entry_attachments WHERE id IN (:a, :b)"),
                {"a": aid, "b": pdf_aid or -1})
            await s.execute(_sqt(
                "DELETE FROM ai_jobs WHERE actor='worker' AND kind='ocr_delivery_note' "
                "AND created_at > NOW() - INTERVAL '10 minutes'"))
            await s.commit()


async def test_qr_returnables_parity():
    """Suite AL — QR + returnables parity: single-badge PNG, label-quantity
    repeats, and the Smart-Scan adoption audit on loan creation."""
    from sqlalchemy import text as _sqt

    async def _scalar(sql: str, **params):
        async with SessionLocal() as s:
            return (await s.execute(_sqt(sql), params)).scalar()

    transport = ASGITransport(app=app)
    async with AsyncClient(transport=transport, base_url="http://svc") as ac:
        _ip = {"X-Real-IP": "203.0.113.89"}

        async def token(u, p):
            r = await ac.post("/auth/login", json={"username": u, "password": p},
                              headers=_ip)
            return r.json().get("access_token")

        def H(t):
            return {"Authorization": f"Bearer {t}"}

        worker_t = await token("worker", "floor2026")
        hod_t = await token("hod", "hod2026")

        emp_id = await _scalar(
            "SELECT \"ID_Number\" FROM employees WHERE status='active' "
            "AND COALESCE(\"Site_ID\",'')='CNCEC' LIMIT 1")
        r = await ac.get(f"/documents/employee-badge/{emp_id}", headers=H(hod_t))
        check("al: single employee badge downloads as a PNG",
              r.status_code == 200 and r.content[:8] == b"\x89PNG\r\n\x1a\n",
              f"{r.status_code} {r.content[:8]!r}")
        r = await ac.get("/documents/employee-badge/NOPE-999", headers=H(hod_t))
        check("al: unknown employee badge → 404", r.status_code == 404, f"got {r.status_code}")
        r = await ac.get(f"/documents/employee-badge/{emp_id}", headers=H(worker_t))
        check("al: store keeper (level 0) → 403 on badge PNG",
              r.status_code == 403, f"got {r.status_code}")

        sap = await _scalar("SELECT \"SAP_Code\" FROM inventory "
                            "WHERE \"Site_ID\"='CNCEC' LIMIT 1")
        r = await ac.get("/documents/qr-labels", headers=H(hod_t),
                         params={"sap_codes": ",".join([sap] * 3)})
        check("al: label sheet accepts repeated codes (copies) and is a PDF",
              r.status_code == 200 and r.content[:5] == b"%PDF-",
              f"{r.status_code} {r.content[:5]!r}")

        # loan with Smart-Scan provenance → cv_* audit columns land
        r = await ac.post("/entry/returnables", headers=H(worker_t), json={
            "material_name": "SVCL torque wrench", "borrower_name": "SVCL Borrower",
            "expected_return_time": "2027-01-01T10:00:00", "qty": 1,
            "cv_employee_id": str(emp_id), "cv_tool_class": "torque wrench",
            "cv_confidence": 0.9})
        rid = r.json().get("id")
        row = None
        async with SessionLocal() as s:
            row = (await s.execute(_sqt(
                "SELECT cv_detected, cv_employee_id, cv_tool_class, cv_confidence "
                "FROM returnable_items WHERE id=:i"), {"i": rid or -1})).first()
        check("al: loan stores the Smart-Scan adoption audit (cv_* columns)",
              r.status_code in (200, 201) and row is not None and row[0] == 1
              and row[1] == str(emp_id) and row[2] == "torque wrench"
              and abs(float(row[3]) - 0.9) < 1e-9,
              f"{r.status_code} row={tuple(row) if row else None}")
        r = await ac.post("/entry/returnables", headers=H(worker_t), json={
            "material_name": "SVCL manual loan", "borrower_name": "SVCL Borrower",
            "expected_return_time": "2027-01-01T10:00:00", "qty": 1})
        rid2 = r.json().get("id")
        cvd = await _scalar("SELECT cv_detected FROM returnable_items WHERE id=:i",
                            i=rid2 or -1)
        check("al: manual loan stays cv_detected=0", cvd == 0, f"got {cvd}")

        async with SessionLocal() as s:  # cleanup (notifications keyed by ref stay)
            await s.execute(_sqt(
                "DELETE FROM returnable_items WHERE material_name LIKE 'SVCL %'"))
            await s.execute(_sqt(
                "DELETE FROM app_notifications WHERE related_table='returnable_items' "
                "AND related_ref IN (:a, :b)"), {"a": str(rid), "b": str(rid2)})
            await s.commit()


async def test_ai_phase2_upgrades():
    """Suite AM — handwritten-form OCR spec stages + ask-data deep filters.

    The deterministic stages (docs/features/handwritten-ocr) are pinned as
    pure-function checks; the endpoint and the query-router category/keyword
    filters run hermetically with SVCM- fixtures."""
    from sqlalchemy import text as _sqt

    from .ai import handwritten as hw

    async def _scalar(sql: str, **params):
        async with SessionLocal() as s:
            return (await s.execute(_sqt(sql), params)).scalar()

    from datetime import date as _date
    _today = _date(2026, 7, 18)

    # ── pure spec stages ────────────────────────────────────────────────────
    d1, _ = hw.parse_form_date("13/07/26", _today)
    d2, _ = hw.parse_form_date("7.6.26", _today)
    d3, _ = hw.parse_form_date("l8/07/26", _today)   # OCR l → 1
    _, f4 = hw.parse_form_date("31/02/26", _today)   # impossible date
    _, f5 = hw.parse_form_date("13/07/20", _today)   # outside year window
    check("am: form dates — 3 formats, digit fixes, validity gates",
          d1 == "2026-07-13" and d2 == "2026-06-07" and d3 == "2026-07-18"
          and f4 == "CRIT_DATE_UNPARSEABLE" and f5 == "CRIT_DATE_UNPARSEABLE",
          f"{d1} {d2} {d3} {f4} {f5}")
    c1, n1 = hw.apply_corrections("Yloues blasting large")
    c2, _ = hw.apply_corrections("Mask")
    c3, _ = hw.apply_corrections("Face Mask")  # regex is ^Mask$ — no touch
    check("am: corrections — substring_ci + anchored regex_ci",
          c1 == "Gloves blasting large" and n1 and c2 == "Dust Mask"
          and c3 == "Face Mask", f"{c1!r} {c2!r} {c3!r}")
    q1, fl1 = hw.parse_qty("2+3", True)
    q2, fl2 = hw.parse_qty("", True)
    q3, fl3 = hw.parse_qty("0", True)
    q4, fl4 = hw.parse_qty("~4", True)
    check("am: qty rules — additive sums, blank defaults to 1, zero rejects",
          q1 == 5 and "WARN_ADDITIVE_QTY" in fl1
          and q2 == 1 and "WARN_QTY_DEFAULTED" in fl2
          and "CRIT_QTY_ZERO_OR_NEGATIVE" in fl3
          and q4 == 4 and "WARN_QTY_APPROXIMATE" in fl4,
          f"{q1}{fl1} {q2}{fl2} {q3}{fl3} {q4}{fl4}")
    r_sub = {"sap_code": "1107", "flags": []}
    hw.validate_stock(r_sub, {"1107": 0.0, "1097": 3.0})
    r_blk = {"sap_code": "9998", "flags": []}
    hw.validate_stock(r_blk, {"9998": 0.0})
    check("am: substitution R2 rewrites SAP with reason; R3 blocks",
          r_sub["sap_code"] == "1097" and "WARN_SUBSTITUTED" in r_sub["flags"]
          and r_sub.get("substituted_from") == "1107"
          and "CRIT_ZERO_STOCK_NO_SUBSTITUTE" in r_blk["flags"],
          f"{r_sub} {r_blk}")

    # ── endpoint: batch process + TSV ───────────────────────────────────────
    transport = ASGITransport(app=app)
    async with AsyncClient(transport=transport, base_url="http://svc") as ac:
        _ip = {"X-Real-IP": "203.0.113.90"}

        async def token(u, p):
            r = await ac.post("/auth/login", json={"username": u, "password": p},
                              headers=_ip)
            return r.json().get("access_token")

        def H(t):
            return {"Authorization": f"Bearer {t}"}

        worker_t = await token("worker", "floor2026")
        hod_t = await token("hod", "hod2026")
        admin_t = await token("admin", "admin2026")

        async with SessionLocal() as s:
            await s.execute(_sqt(
                'INSERT INTO inventory ("SAP_Code", "Equipment_Description", '
                '"Category", "UOM", "Site_ID") VALUES '
                "('SVCM-1', 'GLOVES SVCMALPHA', 'Safety', 'PAIR', 'CNCEC'), "
                "('SVCM-CAT', 'SVCM Category Item', 'Surface Shields', 'EA', 'CNCEC')"))
            await s.execute(_sqt(
                'INSERT INTO receipts ("Date", "SAP_Code", "Quantity", "Site_ID") '
                "VALUES ('2026-07-01 00:00:00', 'SVCM-1', 8, 'CNCEC')"))
            await s.execute(_sqt(
                'INSERT INTO sme_recipe ("Lining_System_Code", "Material_Code", '
                '"SAP_Code", "Material_Name", "For_1_SQM") VALUES '
                "('9907', 'SVCM-MAT-9', 'SVCM-CAT', 'SVCM Furanoid Syrup', 1.5)"))
            await s.commit()

        batch = {"forms": [{"form_id": "form_svcm", "date_text": "l8/07/26",
                            "rows": [
            {"sno": 1, "issued_to": "Imran", "tank_no": "TK-1",
             "material_text": "Yloues SVCMALPHA", "qty_text": "", "work_type": "blasting"},
            {"sno": 2, "issued_to": '"', "tank_no": '"',
             "material_text": '"', "qty_text": "2+3", "work_type": '"'},
            {"sno": 3, "issued_to": "Ali", "tank_no": "",
             "material_text": "SVCM UNKNOWNIUM", "qty_text": "1", "work_type": ""},
            {"sno": 4, "issued_to": "X", "tank_no": "",
             "material_text": "cancelled thing", "qty_text": "9",
             "struck_through": True},
        ]}]}
        r = await ac.post("/ai/ocr/handwritten-process", headers=H(hod_t), json=batch)
        check("am: handwritten-process is SK exact-locked (hod → 403)",
              r.status_code == 403, f"got {r.status_code}")
        r = await ac.post("/ai/ocr/handwritten-process", headers=H(worker_t), json=batch)
        j = r.json() if r.status_code == 200 else {}
        rows = j.get("rows", [])
        summ = j.get("summary", {})
        by_no = {x["source_row_no"]: x for x in rows}
        check("am: batch processes — auto-match, ditto, defaults, struck count",
              r.status_code == 200 and summ.get("total_rows") == 3
              and summ.get("struck_through_excluded") == 1
              and by_no.get(1, {}).get("sap_code") == "SVCM-1"
              and by_no.get(1, {}).get("date_iso") == "2026-07-18"
              and "WARN_QTY_DEFAULTED" in by_no.get(1, {}).get("flags", [])
              and by_no.get(2, {}).get("sap_code") == "SVCM-1"
              and by_no.get(2, {}).get("qty") == 5
              and "WARN_ADDITIVE_QTY" in by_no.get(2, {}).get("flags", [])
              and "INFO_MATCH_UNCERTAIN" in by_no.get(3, {}).get("flags", []),
              f"{r.status_code} summ={summ} rows={[(x.get('source_row_no'), x.get('sap_code'), x.get('qty'), x.get('flags')) for x in rows]}")
        check("am: batch simulation — 8 − 1 − 5 crosses the low-stock line",
              "WARN_LOW_STOCK_CROSSED" in by_no.get(2, {}).get("flags", [])
              and by_no.get(2, {}).get("stock_after") == 2,
              f"{by_no.get(2, {}).get('flags')} after={by_no.get(2, {}).get('stock_after')}")
        tsv_lines = (j.get("tsv") or "").split("\n")
        cols0 = tsv_lines[0].split("\t") if tsv_lines and tsv_lines[0] else []
        check("am: TSV export — 17 positional columns, spec slots filled",
              len(tsv_lines) == 3 and len(cols0) == 17
              and cols0[0] == "2026-07-18" and cols0[3] == "GLOVES SVCMALPHA"
              and cols0[5] == "1" and cols0[9] == "blasting"
              and cols0[10] == "TK-1" and cols0[15] == "Imran",
              f"lines={len(tsv_lines)} cols={len(cols0)} {cols0[:6]}")

        # ── ask-data deep filters ───────────────────────────────────────────
        r = await ac.post("/ai/query", headers=H(admin_t),
                          json={"question": "Current stock for surface shield category items"})
        j = r.json() if r.status_code == 200 else {}
        check("am: ask-data — category question filters with bound ILIKE",
              r.status_code == 200 and j.get("mode") == "template"
              and 'ILIKE :cat' in j.get("sql", "")
              and "category ≈ Surface Shields" in j.get("message", "")
              and j.get("rows"), f"{r.status_code} {j.get('message')} sql={j.get('sql', '')[:120]}")
        r = await ac.post("/ai/query", headers=H(admin_t),
                          json={"question": "current stock of furanoid materials"})
        j = r.json() if r.status_code == 200 else {}
        saps = {row[0] for row in j.get("rows", [])}
        check("am: ask-data — material family resolves via the SME SAP join",
              r.status_code == 200 and "sme_recipe" in j.get("sql", "")
              and "SVCM-CAT" in saps,
              f"{r.status_code} saps={sorted(saps)[:5]} sql={j.get('sql', '')[:120]}")

        async with SessionLocal() as s:  # cleanup
            await s.execute(_sqt("DELETE FROM receipts WHERE \"SAP_Code\" LIKE 'SVCM-%'"))
            await s.execute(_sqt("DELETE FROM sme_recipe WHERE \"Material_Code\" LIKE 'SVCM-%'"))
            await s.execute(_sqt("DELETE FROM inventory WHERE \"SAP_Code\" LIKE 'SVCM-%'"))
            await s.commit()


async def test_sme_sk_upgrades():
    """Suite AN — SK Surface-Shields workflow helper + SME Smart Calculator."""
    from sqlalchemy import text as _sqt

    async def _scalar(sql: str, **params):
        async with SessionLocal() as s:
            return (await s.execute(_sqt(sql), params)).scalar()

    transport = ASGITransport(app=app)
    async with AsyncClient(transport=transport, base_url="http://svc") as ac:
        _ip = {"X-Real-IP": "203.0.113.91"}

        async def token(u, p):
            r = await ac.post("/auth/login", json={"username": u, "password": p},
                              headers=_ip)
            return r.json().get("access_token")

        def H(t):
            return {"Authorization": f"Bearer {t}"}

        worker_t = await token("worker", "floor2026")
        hod_t = await token("hod", "hod2026")

        async with SessionLocal() as s:
            await s.execute(_sqt(
                'INSERT INTO inventory ("SAP_Code", "Equipment_Description", '
                '"Category", "UOM", "Site_ID") VALUES '
                "('SVCN-1', 'SVCN Primer', 'Surface Shields', 'KG', 'CNCEC'), "
                "('SVCN-2', 'SVCN Mortar', 'Surface Shields', 'KG', 'CNCEC')"))
            await s.execute(_sqt(
                'INSERT INTO receipts ("Date", "SAP_Code", "Quantity", "Site_ID") '
                "VALUES ('2026-07-01 00:00:00', 'SVCN-1', 150, 'CNCEC')"))
            await s.execute(_sqt(
                'INSERT INTO sme_recipe ("Lining_System_Code", "Material_Code", '
                '"SAP_Code", "Material_Name", "Material_Description", "UOM", '
                '"For_1_SQM", "Package_Size", "Lining_System_Name", "Substrate") VALUES '
                "('9908', 'SVCN-M1', 'SVCN-1', 'SVCN Primer', 'Primer', 'KG', "
                " 2.5, '25', 'SVCN30', 'Steel'), "
                "('9908', 'SVCN-M2', 'SVCN-2', 'SVCN Mortar', 'Mortar', 'KG', "
                " 0.4, NULL, 'SVCN30', 'Steel')"))
            await s.execute(_sqt(
                'INSERT INTO sme_equipment ("Site_ID", "Equipment_Tag_No", '
                '"Lining_System_Code", "Surface_Area_SQM") VALUES '
                "('CNCEC', 'SVCN-T1', '9908', 60)"))
            await s.execute(_sqt(
                'INSERT INTO sme_sqm_progress ("Site_ID", "Equipment_Tag_No", '
                '"Lining_System_Code", "Original_SQM", "Done_SQM") VALUES '
                "('CNCEC', 'SVCN-T1', '9908', 60, 24)"))
            await s.commit()

        r = await ac.get("/entry/lining-systems", headers=H(worker_t))
        j = r.json() if r.status_code == 200 else {}
        sys9908 = next((s for s in j.get("systems", []) if s["code"] == "9908"), None)
        check("an: SK reads lining systems — recipe SAPs + Done/Pending SQM",
              r.status_code == 200 and sys9908 is not None
              and sys9908["saps"] == ["SVCN-1", "SVCN-2"]
              and sys9908["sqm"]["done_sqm"] == 24
              and sys9908["sqm"]["pending_sqm"] == 36
              and sys9908["short_name"] == "SVCN30",
              f"{r.status_code} {sys9908}")
        check("an: sap_index reverses SAP → system codes",
              (j.get("sap_index", {}).get("SVCN-1") or []) == ["9908"],
              f"{j.get('sap_index', {}).get('SVCN-1')}")

        r = await ac.get("/sme/calculator", headers=H(worker_t),
                         params={"code": "9908", "sqm": 40})
        check("an: calculator is HOD+ (SK level 0 → 403)",
              r.status_code == 403, f"got {r.status_code}")
        r = await ac.get("/sme/calculator", headers=H(hod_t),
                         params={"code": "9908", "sqm": 40})
        j = r.json() if r.status_code == 200 else {}
        lines = j.get("lines", [])
        l1 = next((x for x in lines if x["sap_code"] == "SVCN-1"), {})
        l2 = next((x for x in lines if x["sap_code"] == "SVCN-2"), {})
        check("an: calculator — demand math, pack counts, live stock coverage",
              r.status_code == 200 and len(lines) == 2
              and l1.get("required_qty") == 100 and l1.get("packages_needed") == 4
              and l1.get("available_stock") == 150 and not l1.get("shortfall_qty")
              and l2.get("required_qty") == 16 and l2.get("shortfall_qty") == 16
              and j.get("totals", {}).get("shortfall_lines") == 1
              and "2.5 KG/SQM × 40 SQM = 100 KG" in l1.get("explanation", ""),
              f"{r.status_code} l1={l1} l2={l2}")
        r = await ac.get("/sme/calculator", headers=H(hod_t),
                         params={"code": "nope-999", "sqm": 40})
        r2 = await ac.get("/sme/calculator", headers=H(hod_t),
                          params={"code": "9908", "sqm": 0})
        check("an: calculator guards — unknown code 404, non-positive sqm 422",
              r.status_code == 404 and r2.status_code == 422,
              f"{r.status_code}/{r2.status_code}")

        # ── report column scoping (2026-07-18 polish) ───────────────────────
        r = await ac.get("/reports/stock", headers=H(hod_t),
                         params={"format": "csv"})
        head = r.text.splitlines()[0] if r.status_code == 200 and r.text else ""
        check("an: Current Stock report is stock-scoped (no ledger totals)",
              r.status_code == 200 and "Current_Stock" in head
              and "Total_Received" not in head and "Total_Consumed" not in head,
              f"{r.status_code} head={head[:120]}")
        r = await ac.get("/reports/consumption", headers=H(hod_t),
                         params={"format": "csv", "site_id": "CNCEC"})
        head = r.text.splitlines()[0] if r.status_code == 200 and r.text else ""
        check("an: Consumption report carries the Material description",
              r.status_code == 200 and "Material" in head
              and "Total_Consumed" in head,
              f"{r.status_code} head={head[:120]}")

        async with SessionLocal() as s:  # cleanup
            await s.execute(_sqt("DELETE FROM receipts WHERE \"SAP_Code\" LIKE 'SVCN-%'"))
            await s.execute(_sqt("DELETE FROM inventory WHERE \"SAP_Code\" LIKE 'SVCN-%'"))
            await s.execute(_sqt("DELETE FROM sme_recipe WHERE \"Material_Code\" LIKE 'SVCN-%'"))
            await s.execute(_sqt(
                "DELETE FROM sme_sqm_progress WHERE \"Equipment_Tag_No\" LIKE 'SVCN%'"))
            await s.execute(_sqt(
                "DELETE FROM sme_equipment WHERE \"Equipment_Tag_No\" LIKE 'SVCN%'"))
            await s.commit()


async def _relax_entry_gates() -> None:
    """Parity A1 — require_entry_documents defaults ON in production. Switch
    it OFF for the functional suites (they submit entries without documents);
    suite AH flips it back on to test the gate itself."""
    from sqlalchemy import text as _sqt
    async with SessionLocal() as s:
        await s.execute(_sqt(
            "INSERT INTO app_settings (key, value) VALUES ('require_entry_documents','0') "
            "ON CONFLICT (key) DO UPDATE SET value='0'"))
        await s.commit()


async def main() -> int:
    await _relax_entry_gates()
    print("Service-level invariants (rolled back) + auth/role guards:\n")
    print(" A. service invariants")
    await test_create_and_submit_pr()
    await test_smr_create_and_approve()
    await test_receipt_ledger()
    await test_submitter_resolution()
    await test_notification_visibility()
    test_config_jwt()
    print("\n B. auth/role guards")
    await test_auth_guards()
    print("\n C. site scoping (multi-site isolation)")
    await test_site_scoping()
    print("\n D. token refresh (rotation + revocation)")
    await test_token_refresh()
    print("\n E. man-hours portal (Phase 10)")
    await test_manhours()
    print("\n F. intelligence layer (AI-0/AI-1, Ollama mocked)")
    await test_ai_layer()
    print("\n G. SME plan layer (Phase S1 — engine port + parity oracle)")
    await test_sme_plan_layer()
    print("\n H. admin SLA tracker (T2 — overdue actions + nudges)")
    await test_sla_tracker()
    print("\n I. submission intelligence (T1 — reviewer summaries)")
    await test_submission_intel()
    print("\n J. bulk entry + item snapshot (Phase 1)")
    await test_bulk_entry()
    print("\n K. reschedule workflow (Phase 4 H7)")
    await test_reschedule()
    print("\n L. force-close + 24h undo (Phase 4 H8)")
    await test_force_close()
    print("\n M. manual PO + vendor master (Phase 4)")
    await test_manual_po()
    print("\n N. rate-limiter client-IP resolution (Phase I-B)")
    await test_ratelimit_ip()
    print("\n O. reporting + dashboard parity (Phase 5)")
    await test_reporting_dashboard()
    print("\n P. DN two-stage approval (Phase 6)")
    await test_dn_approval()
    print("\n Q. supervisor parity (Phase 6)")
    await test_supervisor_parity()
    print("\n R. receipt entry guards — MTC + UoM (Phase 6)")
    await test_entry_guards()
    print("\n S. logistics vendor-returns (deferred MED)")
    await test_vendor_returns()
    print("\n T. HOD PR line-edit + rename (deferred MED)")
    await test_pr_management()
    print("\n U. admin lot lifecycle (deferred MED)")
    await test_lot_lifecycle()
    print("\n V. WhatsApp outbox + triggers (Phase 7, Meta mocked)")
    await test_whatsapp_outbox()
    print("\n W. email outbox + triggers (Phase 7b, SMTP mocked)")
    await test_email_outbox()
    print("\n X. phone-change OTP + admin override (Phase 7c, Meta mocked)")
    await test_phone_otp()
    print("\n Y. tool-loan notifications + timezone (UAT Phase 1, Meta mocked)")
    await test_returnables_notify()
    print("\n Z. search & filtering + PR browse (UAT Phase 2)")
    await test_search_filters()
    print("\n AA. notification QA — every pathway → whatsapp_outbox (UAT Phase 4)")
    await test_notification_qa()
    print("\n AB. inbound webhook + dynamic delivery (Phase 6)")
    await test_webhook_and_digest()
    print("\n AC. HOD executive summary")
    await test_executive_summary()
    print("\n AD. chat-with-your-data query router (Phase C)")
    await test_data_query()
    print("\n AE. predictive lining-coverage analytics (Phase 8-1)")
    await test_lining_coverage()
    print("\n AF. strict abuse limits — OTP toll fraud + webhook HMAC ban (Phase 8-2)")
    await test_strict_limits()
    print("\n AG. automated weekly executive PDF (Phase 8-3)")
    await test_weekly_report()
    print("\n AH. entry documents + return gates + WBS (parity A1/A2/A3/A4)")
    await test_entry_documents()
    print("\n AI. SME S6 master-data CRUD (cutover day)")
    await test_sme_master_crud()
    print("\n AJ. bulk Excel import (dry-run/commit, reconcile)")
    await test_bulk_import()
    print("\n AK. OCR doc assist (C3) + ask-data prompt pins")
    await test_ocr_doc_assist()
    print("\n AL. QR + returnables parity (badge PNG, label copies, cv audit)")
    await test_qr_returnables_parity()
    print("\n AM. Handwritten-OCR spec stages + ask-data deep filters")
    await test_ai_phase2_upgrades()
    print("\n AN. Surface-Shields issue workflow + Smart Calculator")
    await test_sme_sk_upgrades()
    await engine.dispose()

    print(f"\n== SERVICE TESTS: {'✅ PASS' if not FAILED else '❌ FAIL'} "
          f"({len(PASSED)} passed, {len(FAILED)} failed) ==")
    if FAILED:
        print("   failed:", ", ".join(FAILED))
    return 1 if FAILED else 0


if __name__ == "__main__":
    sys.exit(asyncio.run(main()))
