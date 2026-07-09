"""
backend/api/reports.py — downloadable reports (Excel / PDF / CSV).

The old app's exportable reports are the biggest capability the new stack
lacked. Each report is a query over the live data; the same rows render to any
of three formats. Reports are a management view → gated at role level ≥ 2
(hod / logistics / admin).

  GET /reports                      → list of available reports + their filters
  GET /reports/{key}?format=xlsx    → the file download (xlsx | pdf | csv)
"""
from __future__ import annotations

import csv
import datetime as _dt
import io
from decimal import Decimal
from typing import Optional

from fastapi import APIRouter, Body, Depends, HTTPException, Query
from fastapi.responses import StreamingResponse
from pydantic import BaseModel
from sqlalchemy import text
from sqlalchemy.ext.asyncio import AsyncSession

from .auth import require_level, resolve_site_param, site_scope
from .db import get_session
from .services import whatsapp as wa
from .stock import SQL_EXPIRING, SQL_SITE_STOCK

router = APIRouter(prefix="/reports", tags=["reports"],
                   dependencies=[Depends(require_level(2))])


# --- data (each returns title, columns, rows) --------------------------------
async def _run(session: AsyncSession, sql: str, params: dict):
    res = await session.execute(text(sql), params)
    columns = list(res.keys())
    rows = [list(r) for r in res.all()]
    return columns, rows


def _cutoff(days: int) -> str:
    days = max(1, min(days, 3650))
    return (_dt.date.today() - _dt.timedelta(days=days)).isoformat()


async def rep_stock(session, *, site_id=None, **_):
    sql = f'SELECT * FROM ({SQL_SITE_STOCK}) s'
    params: dict = {}
    if site_id:
        sql += ' WHERE s."Site_ID" = :site'
        params["site"] = site_id
    sql += ' ORDER BY s."SAP_Code", s."Site_ID"'
    cols, rows = await _run(session, sql, params)
    return "Current Stock by Site", cols, rows


async def rep_expiring(session, *, within_days=30, site_id=None, **_):
    where = 'e."Days_Until_Expiry" <= :w'
    params: dict = {"w": max(0, min(within_days, 3650))}
    if site_id:
        where += ' AND e."Site_ID" = :site'
        params["site"] = site_id
    sql = f'SELECT * FROM ({SQL_EXPIRING}) e WHERE {where} ORDER BY e."Days_Until_Expiry"'
    cols, rows = await _run(session, sql, params)
    return f"Expiring Stock (≤ {within_days} days)", cols, rows


async def rep_consumption(session, *, site_id=None, days=30, **_):
    where = '"Date" >= :cutoff'
    params = {"cutoff": _cutoff(days)}
    if site_id:
        where += ' AND COALESCE("Site_ID", \'HQ\') = :site'
        params["site"] = site_id
    sql = f'''
        SELECT TRIM("SAP_Code") AS "SAP_Code", COALESCE("Site_ID",'HQ') AS "Site_ID",
               ROUND(CAST(SUM("Quantity") AS NUMERIC), 3) AS "Total_Consumed",
               COUNT(*) AS "Transactions", MAX("Date") AS "Last_Issue"
        FROM consumption WHERE {where}
        GROUP BY TRIM("SAP_Code"), COALESCE("Site_ID",'HQ')
        ORDER BY "Total_Consumed" DESC'''
    cols, rows = await _run(session, sql, params)
    return f"Consumption (last {days} days)", cols, rows


async def rep_receipts(session, *, site_id=None, days=30, **_):
    where = '"Date" >= :cutoff'
    params = {"cutoff": _cutoff(days)}
    if site_id:
        where += ' AND COALESCE("Site_ID", \'HQ\') = :site'
        params["site"] = site_id
    sql = f'''
        SELECT "Date", TRIM("SAP_Code") AS "SAP_Code", "Quantity", "Supplier",
               COALESCE("Site_ID",'HQ') AS "Site_ID", "PR_Number", "Lot_Number", "Expiry_Date"
        FROM receipts WHERE {where}
        ORDER BY "Date" DESC, "SAP_Code" LIMIT 5000'''
    cols, rows = await _run(session, sql, params)
    return f"Goods Receipts (last {days} days)", cols, rows


async def rep_purchase_orders(session, *, status=None, site_id=None, **_):
    conds, params = [], {}
    if status:
        conds.append("status = :status")
        params["status"] = status
    if site_id:
        conds.append('"Site_ID" = :site')
        params["site"] = site_id
    where = ("WHERE " + " AND ".join(conds)) if conds else ""
    sql = f'''
        SELECT "PO_Number", "PR_Number", "Site_ID", "Vendor_Name", "PO_Date",
               "Expected_Delivery", status, created_by, created_at
        FROM purchase_orders {where}
        ORDER BY "PO_Number" DESC LIMIT 5000'''
    cols, rows = await _run(session, sql, params)
    return "Purchase Orders", cols, rows


async def rep_pr_status(session, *, status=None, site_id=None, **_):
    # One row per PR (lines rolled up) with its workflow / logistics status —
    # the legacy "PR Status" report (distinct from the PO-status report above).
    conds, params = [], {}
    if site_id:
        conds.append('"Site_ID" = :site')
        params["site"] = site_id
    if status:
        conds.append("COALESCE(logistics_status, 'site_draft') = :status")
        params["status"] = status
    where = ("WHERE " + " AND ".join(conds)) if conds else ""
    sql = f'''
        SELECT "PR_Number", MAX("Site_ID") AS "Site_ID", COUNT(*) AS "Lines",
               ROUND(CAST(SUM("Requested_Qty") AS numeric), 3) AS "Total_Qty",
               MAX(COALESCE(logistics_status, 'site_draft')) AS "Logistics_Status",
               MAX(status) AS "Status",
               MAX(submitted_to_logistics_at) AS "Submitted_At",
               MAX(submitted_to_logistics_by) AS "Submitted_By",
               MIN(created_at) AS "Created_At"
        FROM pr_master {where}
        GROUP BY "PR_Number"
        ORDER BY "PR_Number" DESC LIMIT 5000'''
    cols, rows = await _run(session, sql, params)
    return "PR Status", cols, rows


async def rep_inventory(session, *, site_id=None, **_):
    where, params = "", {}
    if site_id:
        where = 'WHERE "Site_ID" = :site'
        params["site"] = site_id
    sql = f'''
        SELECT "SAP_Code", "Equipment_Description", "Material_Code", "Category",
               "UOM", "Minimum_Qty", "Unit_Cost", "Opening_Stock", "Site_ID", "Expiry_Date"
        FROM inventory {where} ORDER BY "SAP_Code" LIMIT 10000'''
    cols, rows = await _run(session, sql, params)
    return "Inventory Master", cols, rows


# --- Phase-5 parity reports ---------------------------------------------------
async def rep_daily_consumption(session, *, site_id=None, date_from=None, date_to=None, **_):
    f = date_from or _dt.date.today().isoformat()
    t = date_to or f
    where = '"Date" >= :f AND "Date" <= :t'
    params: dict = {"f": f, "t": t}
    if site_id:
        where += ' AND COALESCE("Site_ID", \'HQ\') = :site'
        params["site"] = site_id
    sql = f'''
        SELECT "Date", TRIM("SAP_Code") AS "SAP_Code", "Quantity", "Work_Type",
               "Issued_By", "Issued_To", "Tank_No", COALESCE("Site_ID",'HQ') AS "Site_ID",
               "Lot_Number", "Remarks"
        FROM consumption WHERE {where}
        ORDER BY "Date", "SAP_Code" LIMIT 5000'''
    cols, rows = await _run(session, sql, params)
    return f"Daily Consumption ({f} → {t})", cols, rows


async def rep_monthly_summary(session, *, site_id=None, month=None, **_):
    m = month or _dt.date.today().strftime("%Y-%m")
    try:
        start = _dt.date.fromisoformat(f"{m}-01")
    except ValueError:
        start = _dt.date.today().replace(day=1)
        m = start.strftime("%Y-%m")
    nxt = (start + _dt.timedelta(days=32)).replace(day=1)
    site_w = ' AND COALESCE("Site_ID", \'HQ\') = :site' if site_id else ""
    params: dict = {"start": start.isoformat(), "nxt": nxt.isoformat()}
    if site_id:
        params["site"] = site_id
    sql = f'''
        WITH r AS (SELECT TRIM("SAP_Code") sap,
                          SUM(CASE WHEN "Date" < :start THEN "Quantity" ELSE 0 END) pre,
                          SUM(CASE WHEN "Date" >= :start AND "Date" < :nxt THEN "Quantity" ELSE 0 END) cur
                   FROM receipts WHERE 1=1{site_w} GROUP BY TRIM("SAP_Code")),
             c AS (SELECT TRIM("SAP_Code") sap,
                          SUM(CASE WHEN "Date" < :start THEN "Quantity" ELSE 0 END) pre,
                          SUM(CASE WHEN "Date" >= :start AND "Date" < :nxt THEN "Quantity" ELSE 0 END) cur
                   FROM consumption WHERE 1=1{site_w} GROUP BY TRIM("SAP_Code")),
             x AS (SELECT TRIM("SAP_Code") sap,
                          SUM(CASE WHEN "Date" < :start THEN "Quantity" ELSE 0 END) pre,
                          SUM(CASE WHEN "Date" >= :start AND "Date" < :nxt THEN "Quantity" ELSE 0 END) cur
                   FROM returns WHERE 1=1{site_w} GROUP BY TRIM("SAP_Code"))
        SELECT i."SAP_Code", i."Equipment_Description", i."UOM",
               ROUND(CAST(COALESCE(r.pre,0)-COALESCE(c.pre,0)-COALESCE(x.pre,0) AS NUMERIC),3) AS "Opening",
               COALESCE(r.cur,0) AS "Received", COALESCE(c.cur,0) AS "Issued",
               COALESCE(x.cur,0) AS "Returned",
               ROUND(CAST(COALESCE(r.pre,0)-COALESCE(c.pre,0)-COALESCE(x.pre,0)
                     + COALESCE(r.cur,0)-COALESCE(c.cur,0)-COALESCE(x.cur,0) AS NUMERIC),3) AS "Closing"
        FROM inventory i
        LEFT JOIN r ON r.sap = TRIM(i."SAP_Code")
        LEFT JOIN c ON c.sap = TRIM(i."SAP_Code")
        LEFT JOIN x ON x.sap = TRIM(i."SAP_Code")
        WHERE COALESCE(r.cur,0)+COALESCE(c.cur,0)+COALESCE(x.cur,0)
              +ABS(COALESCE(r.pre,0)-COALESCE(c.pre,0)-COALESCE(x.pre,0)) > 0
        ORDER BY i."SAP_Code"'''
    cols, rows = await _run(session, sql, params)
    return f"Monthly Summary ({m})", cols, rows


async def rep_wbs(session, *, site_id=None, days=90, **_):
    where = '"Date" >= :cutoff'
    params: dict = {"cutoff": _cutoff(days)}
    if site_id:
        where += ' AND COALESCE("Site_ID", \'HQ\') = :site'
        params["site"] = site_id
    sql = f'''
        SELECT COALESCE(NULLIF(TRIM(wbs), ''), '(no WBS)') AS "WBS",
               TRIM("SAP_Code") AS "SAP_Code",
               ROUND(CAST(SUM("Quantity") AS NUMERIC),3) AS "Consumed",
               COUNT(*) AS "Transactions", MAX("Date") AS "Last_Issue"
        FROM consumption WHERE {where}
        GROUP BY COALESCE(NULLIF(TRIM(wbs), ''), '(no WBS)'), TRIM("SAP_Code")
        ORDER BY "WBS", "Consumed" DESC'''
    cols, rows = await _run(session, sql, params)
    return f"WBS Report (last {days} days)", cols, rows


async def rep_low_stock(session, *, site_id=None, **_):
    where = 's."Minimum_Qty" > 0 AND s."Current_Stock" < s."Minimum_Qty"'
    params: dict = {"cutoff": _cutoff(30)}
    if site_id:
        where += ' AND s."Site_ID" = :site'
        params["site"] = site_id
    sql = f'''
        SELECT s."SAP_Code", s."Site_ID", s."Equipment_Description", s."UOM",
               s."Minimum_Qty", s."Current_Stock",
               s."Minimum_Qty" - s."Current_Stock" AS "Shortage",
               ROUND(CAST(COALESCE(b.daily_avg,0) AS NUMERIC),3) AS "Daily_Burn",
               ROUND(CAST((s."Minimum_Qty" - s."Current_Stock")
                     + COALESCE(b.daily_avg,0)*30 AS NUMERIC),3) AS "Suggested_Reorder"
        FROM ({SQL_SITE_STOCK}) s
        LEFT JOIN (SELECT TRIM("SAP_Code") sap, COALESCE("Site_ID",'HQ') site,
                          SUM("Quantity")/30.0 daily_avg
                   FROM consumption WHERE "Date" >= :cutoff
                   GROUP BY TRIM("SAP_Code"), COALESCE("Site_ID",'HQ')) b
          ON b.sap = s."SAP_Code" AND b.site = s."Site_ID"
        WHERE {where} ORDER BY "Shortage" DESC'''
    cols, rows = await _run(session, sql, params)
    return "Low Stock Alert", cols, rows


async def rep_burn_rate(session, *, site_id=None, days=30, **_):
    days = max(1, min(int(days or 30), 365))
    where = '"Date" >= :cutoff'
    params: dict = {"cutoff": _cutoff(days), "days": days}
    if site_id:
        where += ' AND COALESCE("Site_ID", \'HQ\') = :site'
        params["site"] = site_id
    sql = f'''
        SELECT TRIM("SAP_Code") AS "SAP_Code",
               ROUND(CAST(SUM("Quantity") AS NUMERIC),3) AS "Consumed",
               ROUND(CAST(SUM("Quantity")/:days AS NUMERIC),3) AS "Daily_Avg"
        FROM consumption WHERE {where}
        GROUP BY TRIM("SAP_Code") ORDER BY "Consumed" DESC LIMIT 500'''
    cols, rows = await _run(session, sql, params)
    return f"Burn Rate ({days} days)", cols, rows


async def rep_valuation(session, *, site_id=None, **_):
    where, params = "1=1", {}
    if site_id:
        where = 's."Site_ID" = :site'
        params["site"] = site_id
    sql = f'''
        SELECT s."SAP_Code", s."Site_ID", s."Equipment_Description", s."UOM",
               s."Current_Stock", COALESCE(i."Unit_Cost",0) AS "Unit_Cost",
               ROUND(CAST(s."Current_Stock"*COALESCE(i."Unit_Cost",0) AS NUMERIC),2) AS "Stock_Value_SAR"
        FROM ({SQL_SITE_STOCK}) s
        LEFT JOIN inventory i ON TRIM(i."SAP_Code") = s."SAP_Code"
        WHERE {where} ORDER BY "Stock_Value_SAR" DESC'''
    cols, rows = await _run(session, sql, params)
    return "Inventory Valuation (standard cost)", cols, rows


async def rep_fefo(session, *, site_id=None, days=90, **_):
    where = '"Date" >= :cutoff'
    params: dict = {"cutoff": _cutoff(days)}
    if site_id:
        where += ' AND COALESCE("Site_ID", \'HQ\') = :site'
        params["site"] = site_id
    sql = f'''
        SELECT "Date", TRIM("SAP_Code") AS "SAP_Code", "Quantity", "Lot_Number",
               CASE WHEN COALESCE("FEFO_Override",'') <> '' THEN 'OVERRIDE' ELSE 'FEFO' END AS "Pick",
               "FEFO_Override" AS "Override_Reason", "Issued_By",
               COALESCE("Site_ID",'HQ') AS "Site_ID"
        FROM consumption WHERE {where} AND COALESCE("Lot_Number",'') <> ''
        ORDER BY CASE WHEN COALESCE("FEFO_Override",'') <> '' THEN 0 ELSE 1 END, "Date" DESC
        LIMIT 5000'''
    cols, rows = await _run(session, sql, params)
    return f"FEFO Compliance (last {days} days)", cols, rows


async def rep_audit(session, *, days=30, **_):
    sql = '''
        SELECT id, timestamp, username, action_type, target_table, details
        FROM system_audit_log
        WHERE timestamp >= CAST(:cutoff AS timestamp)
        ORDER BY id DESC LIMIT 5000'''
    # asyncpg types the param as timestamp — it must be a datetime, not a str.
    cutoff = _dt.datetime.fromisoformat(_cutoff(days))
    cols, rows = await _run(session, sql, {"cutoff": cutoff})
    return f"Full Audit Log (last {days} days)", cols, rows


async def rep_warehouse_throughput(session, **_):
    sql = '''
        SELECT COALESCE("Warehouse_ID",'—') AS "Warehouse", status,
               COALESCE(rl_bl_family,'—') AS "Family", COUNT(*) AS "DNs"
        FROM delivery_notes
        GROUP BY COALESCE("Warehouse_ID",'—'), status, COALESCE(rl_bl_family,'—')
        ORDER BY "Warehouse", status'''
    cols, rows = await _run(session, sql, {})
    return "Warehouse Throughput (DN counts)", cols, rows


async def rep_force_closures(session, *, site_id=None, **_):
    where, params = "1=1", {}
    if site_id:
        where = '"Site_ID" = :site'
        params["site"] = site_id
    sql = f'''
        SELECT id, target_type, target_ref, "Site_ID", "PR_Number", "PO_Number",
               reason, closed_by, closed_at, notes, reverted_at, reverted_by
        FROM po_force_closures WHERE {where}
        ORDER BY id DESC LIMIT 5000'''
    cols, rows = await _run(session, sql, params)
    return "Force-Closures", cols, rows


async def rep_intent_vs_actual(session, *, site_id=None, days=90, **_):
    where = "h.status = 'approved' AND CAST(h.requested_at AS date) >= CAST(:cutoff AS date)"
    # asyncpg types the param as date — pass a date object, not a str.
    params: dict = {"cutoff": _dt.date.fromisoformat(_cutoff(days))}
    if site_id:
        where += ' AND h."Site_ID" = :site'
        params["site"] = site_id
    sql = f'''
        SELECT h.request_no AS "Request", h."Site_ID", h."Worker_Name",
               i."SAP_Code", i."Requested_Qty" AS "Approved_Qty",
               COALESCE(c.qty, 0) AS "Consumed_Qty",
               ROUND(CAST(COALESCE(c.qty,0) - i."Requested_Qty" AS NUMERIC),3) AS "Variance",
               CASE WHEN i."Requested_Qty" > 0
                    THEN ROUND(CAST(100.0*(COALESCE(c.qty,0)-i."Requested_Qty")/i."Requested_Qty" AS NUMERIC),1)
               END AS "Variance_Pct"
        FROM supervisor_material_request_items i
        JOIN supervisor_material_requests h ON h.id = i.request_id
        LEFT JOIN (SELECT "Source_Ref", SUM("Quantity") qty FROM consumption
                   WHERE "Source_Ref" LIKE 'SMR:%' GROUP BY "Source_Ref") c
          ON c."Source_Ref" = 'SMR:' || h.request_no || ':' || i.id
        WHERE {where}
        ORDER BY h.request_no, i."SAP_Code"'''
    cols, rows = await _run(session, sql, params)
    return f"Supervisor Intent vs Actual (last {days} days)", cols, rows


REPORTS = {
    "stock":           {"fn": rep_stock,           "label": "Current Stock",   "filters": ["site_id"],
                        "desc": "Current stock per item and site (received − consumed − returned)."},
    "expiring":        {"fn": rep_expiring,        "label": "Expiring Stock",  "filters": ["site_id", "within_days"],
                        "desc": "Lots by days-to-expiry, flagged expired / short-dated."},
    "consumption":     {"fn": rep_consumption,     "label": "Consumption",     "filters": ["site_id", "days"],
                        "desc": "Consumption per item over a period."},
    "receipts":        {"fn": rep_receipts,        "label": "Goods Receipts",  "filters": ["site_id", "days"],
                        "desc": "Goods receipts over a period."},
    "purchase-orders": {"fn": rep_purchase_orders, "label": "Purchase Orders", "filters": ["site_id", "status"],
                        "desc": "Purchase orders with status."},
    "pr-status":       {"fn": rep_pr_status,       "label": "PR Status",       "filters": ["site_id", "status"],
                        "desc": "Purchase requests grouped by PR number with workflow / logistics status."},
    "inventory":       {"fn": rep_inventory,       "label": "Inventory Master","filters": ["site_id"],
                        "desc": "The full inventory master list."},
    "daily-consumption": {"fn": rep_daily_consumption, "label": "Daily Consumption", "filters": ["site_id", "date_from", "date_to"],
                        "desc": "Every issue in a date range, with work type / lot / issuer."},
    "monthly-summary": {"fn": rep_monthly_summary,  "label": "Monthly Summary",  "filters": ["site_id", "month"],
                        "desc": "Per-SAP opening / received / issued / returned / closing for a month."},
    "wbs":             {"fn": rep_wbs,              "label": "WBS Report",       "filters": ["site_id", "days"],
                        "desc": "Consumption grouped by WBS number."},
    "low-stock":       {"fn": rep_low_stock,        "label": "Low Stock Alert",  "filters": ["site_id"],
                        "desc": "Items below minimum, with burn rate + suggested reorder."},
    "burn-rate":       {"fn": rep_burn_rate,        "label": "Burn Rate",        "filters": ["site_id", "days"],
                        "desc": "Consumption per material with daily average."},
    "valuation":       {"fn": rep_valuation,        "label": "Inventory Valuation", "filters": ["site_id"],
                        "desc": "Current stock × standard unit cost (SAR)."},
    "fefo":            {"fn": rep_fefo,             "label": "FEFO Compliance",  "filters": ["site_id", "days"],
                        "desc": "Lot-tagged issues, overrides listed first."},
    "audit":           {"fn": rep_audit,            "label": "Full Audit Log",   "filters": ["days"],
                        "desc": "System audit trail (admin/logistics only).", "global_only": True},
    "warehouse-throughput": {"fn": rep_warehouse_throughput, "label": "Warehouse Throughput", "filters": [],
                        "desc": "DN counts by warehouse, status and RL/BL family."},
    "force-closures":  {"fn": rep_force_closures,   "label": "Force-Closures",   "filters": ["site_id"],
                        "desc": "Audit of force-closed PRs / POs / lines."},
    "intent-vs-actual": {"fn": rep_intent_vs_actual, "label": "Intent vs Actual", "filters": ["site_id", "days"],
                        "desc": "Approved supervisor requests vs actual consumption + variance."},
}


# --- renderers (all take title, columns, rows, username) ---------------------
def _xl_val(v):
    if v is None:
        return ""
    if isinstance(v, Decimal):
        return float(v)
    if isinstance(v, (int, float, str)):
        return v
    return str(v)


def _latin(s: str) -> str:
    return s.encode("latin-1", "ignore").decode("latin-1")


def to_csv(title, columns, rows, username) -> bytes:
    buf = io.StringIO()
    w = csv.writer(buf)
    w.writerow(columns)
    w.writerows(rows)
    return buf.getvalue().encode("utf-8-sig")  # BOM so Excel opens UTF-8 cleanly


def to_xlsx(title, columns, rows, username) -> bytes:
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.utils import get_column_letter
    wb = Workbook()
    ws = wb.active
    ws.title = (title[:31] or "Report")
    ws.append(columns)
    fill = PatternFill("solid", fgColor="0A192F")
    font = Font(bold=True, color="FFFFFF")
    for c in ws[1]:
        c.fill = fill
        c.font = font
        c.alignment = Alignment(horizontal="center")
    for r in rows:
        ws.append([_xl_val(v) for v in r])
    for i, col in enumerate(columns, 1):
        widths = [len(str(col))] + [len(str(r[i - 1])) for r in rows[:200]]
        ws.column_dimensions[get_column_letter(i)].width = min(max(max(widths) + 2, 10), 45)
    ws.freeze_panes = "A2"
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def to_xlsx_sheets(sheets, username) -> bytes:
    """Multi-sheet workbook — legacy SME parity (one sheet per report section).
    `sheets` is a list of (sheet_title, columns, rows); styling matches
    to_xlsx(). Sheet titles are trimmed to Excel's 31-char limit and deduped."""
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.utils import get_column_letter
    wb = Workbook()
    wb.remove(wb.active)
    if not sheets:
        sheets = [("Report", [], [])]
    seen: set[str] = set()
    for sheet_title, columns, rows in sheets:
        name = str(sheet_title or "Sheet")
        for ch in "[]:*?/\\":
            name = name.replace(ch, "-")
        name = name[:31] or "Sheet"
        base, n = name, 2
        while name in seen:
            suffix = f" ({n})"
            name, n = base[:31 - len(suffix)] + suffix, n + 1
        seen.add(name)
        ws = wb.create_sheet(title=name)
        ws.append(list(columns))
        fill = PatternFill("solid", fgColor="0A192F")
        font = Font(bold=True, color="FFFFFF")
        for c in ws[1]:
            c.fill = fill
            c.font = font
            c.alignment = Alignment(horizontal="center")
        for r in rows:
            ws.append([_xl_val(v) for v in r])
        for i, col in enumerate(columns, 1):
            widths = [len(str(col))] + [len(str(r[i - 1])) for r in rows[:200]]
            ws.column_dimensions[get_column_letter(i)].width = min(max(max(widths) + 2, 10), 45)
        ws.freeze_panes = "A2"
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def to_pdf_sheets(title, sheets, username) -> bytes:
    """Sectioned PDF — one header + table per sheet (legacy _pdf_from_sheets)."""
    from fpdf import FPDF
    pdf = FPDF(orientation="L", unit="mm", format="A4")
    pdf.add_page()
    pdf.set_font("helvetica", "B", 14)
    pdf.cell(0, 10, _latin(str(title).upper()), new_x="LMARGIN", new_y="NEXT")
    pdf.set_font("helvetica", "", 9)
    stamp = _dt.datetime.now().strftime("%Y-%m-%d %H:%M")
    pdf.cell(0, 6, _latin(f"Generated by {username}  ·  {stamp}"),
             new_x="LMARGIN", new_y="NEXT")
    for sheet_title, columns, rows in sheets:
        pdf.ln(4)
        if pdf.get_y() > 170:
            pdf.add_page()
        pdf.set_font("helvetica", "B", 11)
        pdf.set_text_color(0, 0, 0)
        pdf.cell(0, 8, _latin(str(sheet_title)), new_x="LMARGIN", new_y="NEXT")
        if not rows:
            pdf.set_font("helvetica", "", 8)
            pdf.cell(0, 6, "No data.", new_x="LMARGIN", new_y="NEXT")
            continue
        col_w = pdf.epw / max(len(columns), 1)
        pdf.set_font("helvetica", "B", 8)
        pdf.set_fill_color(10, 25, 47)
        pdf.set_text_color(255, 255, 255)
        for c in columns:
            pdf.cell(col_w, 7, _latin(str(c))[:18], border=1, fill=True, align="C")
        pdf.ln(7)
        pdf.set_font("helvetica", "", 7)
        pdf.set_text_color(0, 0, 0)
        for row in rows:
            if pdf.get_y() > 190:
                pdf.add_page()
            for v in row:
                pdf.cell(col_w, 6, _latin("" if v is None else str(v))[:24], border=1, align="C")
            pdf.ln(6)
    return bytes(pdf.output())


def to_pdf(title, columns, rows, username) -> bytes:
    from fpdf import FPDF
    pdf = FPDF(orientation="L", unit="mm", format="A4")
    pdf.add_page()
    pdf.set_font("helvetica", "B", 14)
    pdf.cell(0, 10, _latin(title.upper()), new_x="LMARGIN", new_y="NEXT")
    pdf.set_font("helvetica", "", 9)
    stamp = _dt.datetime.now().strftime("%Y-%m-%d %H:%M")
    pdf.cell(0, 6, _latin(f"Generated by {username}  ·  {stamp}  ·  {len(rows)} rows"),
             new_x="LMARGIN", new_y="NEXT")
    pdf.ln(3)
    if not rows:
        pdf.cell(0, 10, "No data for this report.", new_x="LMARGIN", new_y="NEXT")
        return bytes(pdf.output())
    col_w = pdf.epw / max(len(columns), 1)
    pdf.set_font("helvetica", "B", 8)
    pdf.set_fill_color(10, 25, 47)
    pdf.set_text_color(255, 255, 255)
    for c in columns:
        pdf.cell(col_w, 7, _latin(str(c))[:18], border=1, fill=True, align="C")
    pdf.ln(7)
    pdf.set_font("helvetica", "", 7)
    pdf.set_text_color(0, 0, 0)
    for row in rows:
        if pdf.get_y() > 190:
            pdf.add_page()
        for v in row:
            pdf.cell(col_w, 6, _latin("" if v is None else str(v))[:24], border=1, align="C")
        pdf.ln(6)
    return bytes(pdf.output())


_FORMATS = {
    "xlsx": (to_xlsx, "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"),
    "pdf":  (to_pdf,  "application/pdf"),
    "csv":  (to_csv,  "text/csv"),
}


@router.get("", summary="Available reports + their filters")
async def list_reports(user: dict = Depends(require_level(2))):
    scoped = site_scope(user) is not None
    return {"reports": [
        {"key": k, "label": v["label"], "description": v["desc"], "filters": v["filters"]}
        for k, v in REPORTS.items()
        if not (scoped and v.get("global_only"))
    ]}


async def render_report(session, key: str, *, format: str, user: dict,
                        site_id: Optional[str] = None, days: int = 30,
                        within_days: int = 30, status: Optional[str] = None,
                        date_from: Optional[str] = None, date_to: Optional[str] = None,
                        month: Optional[str] = None) -> tuple[bytes, str, str]:
    """Shared pipeline for the download endpoint, the archive, and the
    scheduler: validates, applies site scoping, renders → (bytes, filename,
    media type)."""
    if key not in REPORTS:
        raise HTTPException(404, f"unknown report {key!r}")
    fmt = (format or "xlsx").lower()
    if fmt not in _FORMATS:
        raise HTTPException(400, f"format must be one of {sorted(_FORMATS)}")
    # Site scoping: every report is forced to the caller's own site below
    # logistics level; a scoped user with no site gets a clear 403.
    site_id = resolve_site_param(user, site_id)
    if site_scope(user) is not None:
        if site_id == "":
            raise HTTPException(403, "no site is assigned to your account — reports unavailable")
        if REPORTS[key].get("global_only"):
            raise HTTPException(403, "this report is restricted to logistics/admin")
    title, columns, rows = await REPORTS[key]["fn"](
        session, site_id=site_id, days=days, within_days=within_days, status=status,
        date_from=date_from, date_to=date_to, month=month)
    render, media = _FORMATS[fmt]
    data = render(title, columns, rows, user["username"])
    fname = f"{key}-{_dt.date.today().isoformat()}.{fmt}"
    return data, fname, media


@router.get("/{key}", summary="Download a report (xlsx | pdf | csv)")
async def download_report(key: str, format: str = Query("xlsx"),
                          site_id: Optional[str] = None, days: int = 30,
                          within_days: int = 30, status: Optional[str] = None,
                          date_from: Optional[str] = None, date_to: Optional[str] = None,
                          month: Optional[str] = None,
                          user: dict = Depends(require_level(2)),
                          session: AsyncSession = Depends(get_session)):
    data, fname, media = await render_report(
        session, key, format=format, user=user, site_id=site_id, days=days,
        within_days=within_days, status=status, date_from=date_from,
        date_to=date_to, month=month)
    return StreamingResponse(io.BytesIO(data), media_type=media,
                             headers={"Content-Disposition": f'attachment; filename="{fname}"'})


class ReportWhatsAppIn(BaseModel):
    to: str
    format: str = "pdf"
    site_id: Optional[str] = None
    days: int = 30
    status: Optional[str] = None


@router.post("/{key}/whatsapp", summary="Send a report to a WhatsApp number (as a document)")
async def whatsapp_report(key: str, body: ReportWhatsAppIn = Body(...),
                          user: dict = Depends(require_level(2)),
                          session: AsyncSession = Depends(get_session)):
    to = (body.to or "").strip()
    if not to:
        raise HTTPException(422, "a recipient WhatsApp number is required")
    data, fname, media = await render_report(
        session, key, format=body.format, user=user, site_id=body.site_id,
        days=body.days, status=body.status)
    res = await wa.send_document(session, to=to, blob=data, filename=fname, mime=media,
                                 caption=f"{key} report", event_key="report_delivery",
                                 created_by=user["username"])
    await session.commit()
    return res
