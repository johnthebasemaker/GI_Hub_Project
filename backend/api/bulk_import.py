"""
backend/api/bulk_import.py — Bulk Excel Import (operator "update from Excel").

Four structured workbooook kinds, each with a DRY-RUN (default) that returns a
full plan and a COMMIT mode that applies it in one transaction + audits:

  inventory      CNCEC_Inventory.xlsx sheet "Inventory" → `inventory` master
                 upsert on SAP_Code. Aggregate columns (Receipt/Consumption/
                 Return/Current Stock) are IGNORED — stock is ledger-derived.
                 Category values are canonicalised (e.g. the workbook's
                 "Surface Shield" → the DB's "Surface Shields") so the MTC
                 hard-block keeps matching `mtc_required_category`.
  ledger         The same workbook's "Receipt Log" / "Consumption Log" /
                 "Return Log" sheets → append-only ledger backfill with a
                 three-tier reconcile per row key (day, SAP, qty, ref):
                 exact multiset match → skip; same (day, SAP, ref) with a
                 different qty → UPDATE (workbook corrections, e.g. a
                 mis-entry zeroed out); otherwise INSERT. Rows whose SAP is
                 missing from `inventory` are rejected (soft-FK protection).
                 DB rows absent from the workbook are only REPORTED — this
                 importer never deletes ledger history.
  sme-equipment  Equipment.xlsx "Data Input" → sme_equipment upsert on
                 (Site_ID, tag, code) + sme_sqm_progress re-seed that
                 PRESERVES Done_SQM (ports legacy sme_bootstrap
                 _clean_equipment: Name-identity tag backfill, short-name →
                 code backfill, non-numeric code skip, per-(tag,code) area
                 aggregation with SQM summing, Location canonicalisation).
  sme-recipes    For_1_SQM.xlsx → sme_recipe upsert on (code, material).
  sme-materials  Materials_DetailsAvailable_Qty.xlsx → sme_inventory_seed
                 upsert on Material_Code (ports _clean_inventory_seed: one
                 row per code, quantities summed across PO lines).

Roles: SME kinds are the Master-Data exact-lock {hod, admin}; `inventory` and
`ledger` are admin-only. HOD site pinning follows sme_master._write_site.
"""
from __future__ import annotations

import io
from collections import Counter
from datetime import date, datetime
from typing import Any, Optional

from fastapi import APIRouter, Depends, File, HTTPException, Query, UploadFile
from sqlalchemy import delete, func, insert, select, update
from sqlalchemy.dialects.postgresql import insert as pg_insert
from sqlalchemy.ext.asyncio import AsyncSession

from .auth import require_level, require_roles
from .db import get_session
from .services.ledger import _MD, write_audit
from .sme_master import _upsert_progress, _write_site

router = APIRouter(prefix="/import", tags=["Bulk Excel import"])

inventory_t = _MD.tables["inventory"]
receipts_t = _MD.tables["receipts"]
consumption_t = _MD.tables["consumption"]
returns_t = _MD.tables["returns"]
equipment_t = _MD.tables["sme_equipment"]
recipe_t = _MD.tables["sme_recipe"]
seed_t = _MD.tables["sme_inventory_seed"]

MAX_XLSX_BYTES = 8 * 1024 * 1024

# Workbook category spellings → the DB's canonical values. "Surface Shields"
# drives the MTC hard-block (app_settings.mtc_required_category) — an
# unnormalised singular would silently disarm the gate for imported rows.
CATEGORY_CANON = {"surface shield": "Surface Shields",
                  "surface shields": "Surface Shields",
                  "r/l cons": "R/L Consumables"}

_LOCATION_CANON = {c.lower(): c for c in ("Brown Field", "TRAIN J", "TRAIN K")}


def _s(v: Any) -> Optional[str]:
    """Cell → stripped string or None ('', 'nan', 'None', 'N/A' → None)."""
    if v is None:
        return None
    s = str(v).strip()
    if s in ("", "nan", "None", "N/A", "n/a", "NA"):
        return None
    if s.endswith(".0") and s[:-2].isdigit():  # Excel float-ified codes
        s = s[:-2]
    return s


def _f(v: Any) -> Optional[float]:
    try:
        f = float(v)
        return f if f == f else None
    except (TypeError, ValueError):
        return None


def _iso(v: Any) -> Optional[str]:
    """Excel date/datetime/text → the ledger's 'YYYY-MM-DD HH:MM:SS' text."""
    if isinstance(v, datetime):
        return v.strftime("%Y-%m-%d %H:%M:%S")
    if isinstance(v, date):
        return v.strftime("%Y-%m-%d 00:00:00")
    s = _s(v)
    if not s:
        return None
    for fmt in ("%Y-%m-%d %H:%M:%S", "%Y-%m-%d", "%d/%m/%Y", "%d-%m-%Y"):
        try:
            return datetime.strptime(s, fmt).strftime("%Y-%m-%d %H:%M:%S")
        except ValueError:
            continue
    return None


def _sheet_rows(data: bytes, want: str | None, header_probe: tuple[str, ...],
                required: bool = True) -> tuple[list[str], list[tuple]]:
    """Load one worksheet and find its header row (workbooks carry a title
    banner above the real header). Returns (headers, data_rows)."""
    import openpyxl
    try:
        wb = openpyxl.load_workbook(io.BytesIO(data), read_only=True, data_only=True)
    except Exception:
        raise HTTPException(422, "not a readable .xlsx workbook")
    try:
        ws = None
        if want is not None:
            for cand in wb.sheetnames:
                if cand.strip().lower() == want.strip().lower():
                    ws = wb[cand]
                    break
            if ws is None:
                if required:
                    raise HTTPException(422, f"worksheet {want!r} not found "
                                             f"(has: {wb.sheetnames})")
                return [], []
        else:
            ws = wb.worksheets[0]
        rows = list(ws.iter_rows(values_only=True))
    finally:
        wb.close()
    probe = {p.lower() for p in header_probe}
    for i, row in enumerate(rows[:5]):
        cells = {str(c).strip().lower() for c in row if c is not None}
        if probe <= cells:
            headers = [str(c).strip() if c is not None else "" for c in row]
            return headers, rows[i + 1:]
    raise HTTPException(422, f"header row not found (need columns {sorted(probe)})")


def _col(headers: list[str], *names: str) -> Optional[int]:
    lower = [h.lower() for h in headers]
    for n in names:
        if n.lower() in lower:
            return lower.index(n.lower())
    return None


# ─── inventory master ─────────────────────────────────────────────────────────
async def plan_inventory(session: AsyncSession, data: bytes, site_id: str) -> dict:
    headers, rows = _sheet_rows(data, "Inventory",
                                ("sap code", "category", "opening stock"),
                                required=False)
    if not headers:  # single-sheet master file fallback
        headers, rows = _sheet_rows(data, None, ("sap code", "category"))
    colspec = {
        "sl": ("Sl. No.", "Sl_No", "Sl. #"), "sap": ("SAP CODE", "SAP_Code"),
        "mat": ("Material Code", "Material_Code"),
        "desc": ("Equipment Description", "Equipment_Description"),
        "uom": ("UOM",), "cat": ("Category",),
        "open": ("Opening Stock", "Opening_Stock"),
        "min": ("Minimum Qty", "Minimum_Qty"),
    }
    ix = {k: _col(headers, *names) for k, names in colspec.items()}
    if ix["sap"] is None:
        raise HTTPException(422, "SAP CODE column missing")
    # Columns are resolved by NAME (order-independent). Anything the workbook
    # carries beyond the mapped set + the ledger-derived aggregates is ignored
    # — but LOUDLY, so a restructured sheet never loses data silently.
    _known = {n.lower() for names in colspec.values() for n in names}
    _known |= {"receipt", "consumption", "return", "current stock"}
    extra_cols = [h for h in headers if h and h.lower() not in _known]

    existing = {r["SAP_Code"]: dict(r) for r in
                (await session.execute(select(inventory_t))).mappings().all()}
    mat_owner = {r["Material_Code"]: r["SAP_Code"] for r in existing.values()
                 if r.get("Material_Code")}

    # First pass — collect every row so Material_Code conflicts can be judged
    # against the WHOLE file (a code may legitimately change owners when its
    # current owner is re-mapped in the same workbook).
    parsed: list[dict] = []
    rejects, warnings = [], []
    if extra_cols:
        warnings.append("ignored unmapped column(s): " + ", ".join(extra_cols))
    normalised_cats = Counter()
    seen_saps = set()
    for n, row in enumerate(rows, start=1):
        sap = _s(row[ix["sap"]]) if ix["sap"] < len(row) else None
        if not sap:
            continue
        if sap in seen_saps:
            rejects.append({"row": n, "sap": sap, "reason": "duplicate SAP in file"})
            continue
        seen_saps.add(sap)

        def cell(key):
            i = ix[key]
            return row[i] if i is not None and i < len(row) else None

        cat_raw = _s(cell("cat"))
        cat = CATEGORY_CANON.get(cat_raw.lower(), cat_raw) if cat_raw else None
        if cat_raw and cat != cat_raw:
            normalised_cats[f"{cat_raw} → {cat}"] += 1
        parsed.append({"row": n, "sap": sap, "mat": _s(cell("mat")),
                       "fields": {"Sl_No": _s(cell("sl")),
                                  "Equipment_Description": _s(cell("desc")),
                                  "UOM": _s(cell("uom")), "Category": cat,
                                  "Opening_Stock": _f(cell("open")),
                                  "Minimum_Qty": _f(cell("min"))}})

    # Material_Code resolution (unique across inventory):
    #   in-file duplicate        → first row keeps it, later rows import codeless
    #   owner re-mapped in file  → release + reassign (the workbook is truth)
    #   owner NOT in the file    → keep the owner's code; import row codeless
    file_mat: dict[str, str] = {}
    for p in parsed:
        if p["mat"] and p["mat"] not in file_mat:
            file_mat[p["mat"]] = p["sap"]
    file_saps = {p["sap"] for p in parsed}
    file_mat_of_sap = {p["sap"]: p["mat"] for p in parsed}
    releases: list[dict] = []
    for p in parsed:
        mat = p["mat"]
        if not mat:
            continue
        if file_mat.get(mat) != p["sap"]:
            warnings.append(f"SAP {p['sap']}: Material_Code {mat} already used by "
                            f"SAP {file_mat[mat]} in this file — imported without it")
            p["mat"] = None
            continue
        owner = mat_owner.get(mat)
        if owner and owner != p["sap"]:
            if owner in file_saps and file_mat_of_sap.get(owner) != mat:
                releases.append({"sap": owner, "mat": mat})  # re-mapped → release
            else:
                warnings.append(f"SAP {p['sap']}: Material_Code {mat} stays with "
                                f"SAP {owner} (not re-mapped here) — imported "
                                f"without it")
                p["mat"] = None

    inserts, updates, unchanged = [], [], 0
    for p in parsed:
        fields = {k: v for k, v in p["fields"].items() if v is not None}
        if p["mat"] is not None:
            fields["Material_Code"] = p["mat"]
        cur = existing.get(p["sap"])
        if cur is None:
            inserts.append({"SAP_Code": p["sap"], "Site_ID": site_id, **fields})
        else:
            diff = {k: v for k, v in fields.items() if cur.get(k) != v}
            if diff:
                if "Opening_Stock" in diff:
                    warnings.append(f"SAP {p['sap']}: Opening_Stock "
                                    f"{cur.get('Opening_Stock')} → {diff['Opening_Stock']}")
                updates.append({"SAP_Code": p["sap"], "diff": diff})
            else:
                unchanged += 1
    if normalised_cats:
        warnings.append("category canonicalised: " +
                        ", ".join(f"{k} ×{v}" for k, v in normalised_cats.items()))
    return {"inserts": inserts, "updates": updates, "unchanged": unchanged,
            "rejects": rejects, "warnings": warnings, "releases": releases}


async def apply_inventory(session: AsyncSession, plan: dict, username: str) -> None:
    # Free re-mapped Material_Codes FIRST so the unique constraint never trips
    # mid-plan (their new values arrive with the updates below).
    for rel in plan.get("releases", []):
        await session.execute(update(inventory_t)
                              .where(inventory_t.c["SAP_Code"] == rel["sap"],
                                     inventory_t.c["Material_Code"] == rel["mat"])
                              .values(Material_Code=None))
    for row in plan["inserts"]:
        await session.execute(insert(inventory_t).values(**row))
    for u in plan["updates"]:
        await session.execute(update(inventory_t)
                              .where(inventory_t.c["SAP_Code"] == u["SAP_Code"])
                              .values(**u["diff"]))
    await write_audit(session, username, "BULK_IMPORT_INVENTORY", "inventory",
                      f"+{len(plan['inserts'])} ~{len(plan['updates'])} "
                      f"rejected={len(plan['rejects'])}")


# ─── ledger backfill (receipts / consumption / returns) ───────────────────────
# Every sheet's columns resolve by NAME (order-independent). "ignore" lists
# workbook columns that deliberately have no DB home — anything else that is
# unmapped raises a warning so a restructured sheet never drops data silently.
# ("Material Code" / "Equipment Description" / "UOM" repeat the inventory
# master on every log sheet and are always ignored.)
_LEDGER_ALWAYS_IGNORED = ("date", "sap code", "sap_code", "material code",
                          "equipment description", "uom")
_LEDGER_SHEETS = {
    "receipts": {
        "sheet": "Receipt Log", "table": receipts_t,
        "cols": {"Quantity": ("Qty.",), "Serial_No": ("Serial No.",),
                 "PR_Number": ("PR#",), "WBS": ("WBS#",),
                 "Location": ("Location",), "Vehicle_No": ("Vehicle No.",),
                 "Driver_Name": ("Driver Name",), "DN_No": ("DN. No.",),
                 "Pallet_No": ("Pallet No.",), "Mob_From": ("Mob. From",),
                 "Mob_To": ("Mob. To",), "Prepared_by": ("Prepared by",),
                 "Received_by": ("Received by",), "DN_Copy": ("DN. Copy",),
                 "Remarks": ("Remarks",)},
        "ref": "DN_No", "ignore": (),
    },
    "consumption": {
        "sheet": "Consumption Log", "table": consumption_t,
        "cols": {"Quantity": ("Qty.",), "Serial_No": ("Serial No.",),
                 "PR_Number": ("PR#",), "Work_Type": ("Work Type",),
                 "Tank_No": ("Tank No.",), "WBS": ("WBS#",),
                 "Approved By": ("Approved By",), "Issued_To": ("Received by",),
                 "Issued_By": ("Prepared by",), "Remarks": ("Remarks",)},
        # `consumption` has no Pallet_No / paper-number columns — 2026-07-14
        # workbook restructure adds both to the sheet; ignored by design.
        "ref": "Tank_No", "ignore": ("cons. paper no.", "pallet no."),
    },
    "returns": {
        "sheet": "Return Log", "table": returns_t,
        "cols": {"Quantity": ("Qty.",), "Reason": ("Reason",),
                 "Remarks": ("Remarks",)},
        # the Return Log reuses the Receipt Log template; `returns` is a
        # narrow table (Date/SAP/Qty/Reason/Remarks) so the rest has no home
        "ref": "Reason",
        "ignore": ("serial no.", "pr#", "wbs#", "location", "vehicle no.",
                   "driver name", "dn. no.", "pallet no.", "mob. from",
                   "mob. to", "prepared by", "received by", "dn. copy"),
    },
}


def _day(v) -> str:
    return str(v or "")[:10]


async def plan_ledger(session: AsyncSession, data: bytes, site_id: str,
                      extra_saps: set[str] | None = None) -> dict:
    known_saps = {r[0] for r in
                  (await session.execute(select(inventory_t.c["SAP_Code"]))).all()}
    known_saps |= extra_saps or set()  # dry-run chained after an inventory plan
    out = {"sections": {}, "rejects": [], "warnings": []}
    for kind, spec in _LEDGER_SHEETS.items():
        headers, rows = _sheet_rows(data, spec["sheet"], ("sap code", "qty."),
                                    required=False)
        section = {"inserts": [], "corrections": [], "matched": 0,
                   "zero_skipped": 0, "db_only": 0}
        out["sections"][kind] = section
        if not headers:
            out["warnings"].append(f"{spec['sheet']}: sheet missing — skipped")
            continue
        sap_i = _col(headers, "SAP CODE", "SAP_Code")
        date_i = _col(headers, "Date", "Date ")
        colmap = {field: _col(headers, *names)
                  for field, names in spec["cols"].items()}
        _known = set(_LEDGER_ALWAYS_IGNORED) | set(spec["ignore"])
        _known |= {n.lower() for names in spec["cols"].values() for n in names}
        extra_cols = [h for h in headers if h and h.lower() not in _known]
        if extra_cols:
            out["warnings"].append(f"{spec['sheet']}: ignored unmapped "
                                   f"column(s): {', '.join(extra_cols)}")

        file_rows = []
        for n, row in enumerate(rows, start=1):
            sap = _s(row[sap_i]) if sap_i is not None and sap_i < len(row) else None
            if not sap:
                continue
            d = _iso(row[date_i]) if date_i is not None and date_i < len(row) else None
            qty = _f(row[colmap["Quantity"]]) if colmap["Quantity"] is not None else None
            if d is None or qty is None:
                out["rejects"].append({"sheet": spec["sheet"], "row": n, "sap": sap,
                                       "reason": "missing/unparseable Date or Qty"})
                continue
            if sap not in known_saps:
                out["rejects"].append({"sheet": spec["sheet"], "row": n, "sap": sap,
                                       "reason": "SAP not in inventory master"})
                continue
            vals = {"Date": d, "SAP_Code": sap, "Quantity": qty, "Site_ID": site_id}
            for field, i in colmap.items():
                if field == "Quantity" or i is None or i >= len(row):
                    continue
                v = _s(row[i])
                if v is not None:
                    vals[field] = v
            file_rows.append(vals)

        table, ref = spec["table"], spec["ref"]
        db_rows = [dict(m) for m in (await session.execute(
            select(table).where(table.c["Site_ID"] == site_id))).mappings().all()]

        def key(r):  # exact multiset identity
            return (_day(r.get("Date")), r.get("SAP_Code"),
                    round(float(r.get("Quantity") or 0), 4),
                    _s(r.get(ref)) or "")

        def refkey(r):  # correction identity: same day+sap+ref, any qty
            return (_day(r.get("Date")), r.get("SAP_Code"), _s(r.get(ref)) or "")

        db_exact = Counter(key(r) for r in db_rows)
        # tier 1 — exact matches consume DB copies
        remaining = []
        for fr in file_rows:
            k = key(fr)
            if db_exact.get(k, 0) > 0:
                db_exact[k] -= 1
                section["matched"] += 1
            else:
                remaining.append(fr)
        # unmatched DB copies, grouped for the correction tier
        db_left: dict[tuple, list[dict]] = {}
        for r in db_rows:
            k = key(r)
            if db_exact.get(k, 0) > 0:
                db_exact[k] -= 1
                db_left.setdefault(refkey(r), []).append(r)
        # tier 2 — qty corrections (workbook is truth for the same day+sap+ref)
        inserts = []
        for fr in remaining:
            cands = db_left.get(refkey(fr)) or []
            if cands:
                target = cands.pop(0)
                section["corrections"].append(
                    {"id": target["id"], "sap": fr["SAP_Code"],
                     "date": _day(fr["Date"]),
                     "qty_from": target["Quantity"], "qty_to": fr["Quantity"]})
            elif float(fr["Quantity"]) == 0.0:
                section["zero_skipped"] += 1  # zero-qty history line, no DB twin
            else:
                inserts.append(fr)
        section["inserts"] = inserts
        section["db_only"] = sum(len(v) for v in db_left.values())
        if section["db_only"]:
            out["warnings"].append(
                f"{spec['sheet']}: {section['db_only']} DB row(s) have no workbook "
                f"counterpart — left untouched (this importer never deletes)")
    return out


async def apply_ledger(session: AsyncSession, plan: dict, username: str) -> None:
    for kind, spec in _LEDGER_SHEETS.items():
        section = plan["sections"].get(kind) or {}
        table = spec["table"]
        for row in section.get("inserts", []):
            await session.execute(insert(table).values(**row))
        for c in section.get("corrections", []):
            await session.execute(update(table).where(table.c["id"] == c["id"])
                                  .values(Quantity=c["qty_to"]))
        if section.get("inserts") or section.get("corrections"):
            await write_audit(session, username, "BULK_IMPORT_LEDGER",
                              table.name,
                              f"+{len(section['inserts'])} rows, "
                              f"{len(section['corrections'])} qty corrections")


# ─── SME masters ──────────────────────────────────────────────────────────────
async def plan_sme_equipment(session: AsyncSession, data: bytes, site_id: str) -> dict:
    headers, rows = _sheet_rows(data, "Data Input",
                                ("equipment_tag_no.", "lining_system_code"),
                                required=False)
    if not headers:
        headers, rows = _sheet_rows(data, None,
                                    ("equipment_tag_no.", "lining_system_code"))
    field_names = {  # workbook header → sme_equipment column
        "Sl. #": "Sl_No", "Project": "Project", "WBS #": "WBS_No",
        "IO#": "IO_No", "Sub_Location": "Sub_Location", "Location": "Location",
        "Type": "Type", "Substrate": "Substrate", "Name": "Name",
        "Drawing #": "Drawing_No", "Design": "Design", "Dia / L": "Dia_L",
        "Ht. /W": "Ht_W", "Equipment Total SQM": "Equipment_Total_SQM",
        "Remaraks": "Remaraks",
        "Lining_System_Short_Name": "Lining_System_Short_Name",
        "Lining_Type": "Lining_Type", "Lining_System": "Lining_System",
        "Material Spec.": "Material_Spec",
        "Lining_Area/location": "Lining_Area_Location",
    }
    ix = {col: _col(headers, hdr) for hdr, col in field_names.items()}
    tag_i = _col(headers, "Equipment_Tag_No.", "Equipment_Tag_No")
    code_i = _col(headers, "Lining_System_Code")
    sqm_i = _col(headers, "Surface_Area_SQM")
    if tag_i is None or code_i is None or sqm_i is None:
        raise HTTPException(422, "Equipment sheet needs Equipment_Tag_No., "
                                 "Lining_System_Code and Surface_Area_SQM")

    # short-name → code backfill map (recipes already in the DB)
    sn_map = {}
    for r in (await session.execute(
            select(recipe_t.c["Lining_System_Name"],
                   recipe_t.c["Lining_System_Code"]))).all():
        if r[0] and r[0].strip() and r[0].strip() not in sn_map:
            sn_map[r[0].strip()] = str(r[1]).strip()

    agg: dict[tuple[str, str], dict] = {}
    warnings, skipped_nonnum, backfilled_tags = [], 0, 0
    for row in rows:
        def cell(i):
            return row[i] if i is not None and i < len(row) else None
        tag = _s(cell(tag_i))
        name = _s(cell(ix["Name"]))
        if not tag and name:
            tag, backfilled_tags = name, backfilled_tags + 1  # Name IS the identity
        code = _s(cell(code_i))
        short = _s(cell(ix["Lining_System_Short_Name"]))
        if not code and short and short in sn_map:
            code = sn_map[short]
        sqm = _f(cell(sqm_i))
        if not tag or not code:
            continue
        try:
            code = str(int(float(code)))
        except (TypeError, ValueError):
            skipped_nonnum += 1  # e.g. To_Be_Confirmed_LSC placeholders
            continue
        if sqm is None or sqm <= 0:
            continue
        a = agg.setdefault((tag, code), {"Surface_Area_SQM": 0.0, "_areas": []})
        a["Surface_Area_SQM"] += sqm
        area = _s(cell(ix["Lining_Area_Location"]))
        if area and area not in a["_areas"]:
            a["_areas"].append(area)
        for col, i in ix.items():
            if col == "Lining_Area_Location":
                continue
            v = cell(i)
            if col == "Equipment_Total_SQM":
                v = _f(v)
            else:
                v = _s(v)
                if col == "Location" and v:
                    v = _LOCATION_CANON.get(v.lower(), v)
            if v is not None and col not in a:
                a[col] = v
    if skipped_nonnum:
        warnings.append(f"skipped {skipped_nonnum} row(s) with non-numeric "
                        f"Lining_System_Code")
    if backfilled_tags:
        warnings.append(f"backfilled Equipment_Tag_No from Name for "
                        f"{backfilled_tags} area row(s)")

    existing = {(r["Equipment_Tag_No"], r["Lining_System_Code"]): dict(r)
                for r in (await session.execute(
                    select(equipment_t).where(equipment_t.c["Site_ID"] == site_id)
                )).mappings().all()}
    inserts, updates, unchanged = [], [], 0
    for (tag, code), a in agg.items():
        areas = ", ".join(a.pop("_areas", [])) or None
        fields = {k: v for k, v in a.items() if v is not None}
        if areas:
            fields["Lining_Area_Location"] = areas
        fields["Surface_Area_SQM"] = round(float(fields["Surface_Area_SQM"]), 4)
        cur = existing.get((tag, code))
        if cur is None:
            inserts.append({"Site_ID": site_id, "Equipment_Tag_No": tag,
                            "Lining_System_Code": code, **fields})
        else:
            diff = {k: v for k, v in fields.items() if cur.get(k) != v}
            if diff:
                updates.append({"id": cur["id"], "tag": tag, "code": code,
                                "diff": diff,
                                "sqm": fields["Surface_Area_SQM"]})
            else:
                unchanged += 1
    not_in_file = [f"{t}/{c}" for (t, c) in existing if (t, c) not in agg]
    if not_in_file:
        warnings.append(f"{len(not_in_file)} DB equipment row(s) not in the file "
                        f"— left untouched (delete via Master Data if intended)")
    return {"inserts": inserts, "updates": updates, "unchanged": unchanged,
            "rejects": [], "warnings": warnings, "site_id": site_id}


async def apply_sme_equipment(session: AsyncSession, plan: dict, username: str) -> None:
    site = plan["site_id"]
    for row in plan["inserts"]:
        await session.execute(insert(equipment_t).values(**row))
        await _upsert_progress(session, site, row["Equipment_Tag_No"],
                               row["Lining_System_Code"],
                               original_sqm=row["Surface_Area_SQM"], done_sqm=None)
    for u in plan["updates"]:
        await session.execute(update(equipment_t)
                              .where(equipment_t.c["id"] == u["id"])
                              .values(**u["diff"]))
        # re-seed the baseline, PRESERVING Done_SQM (legacy bootstrap contract)
        await _upsert_progress(session, site, u["tag"], u["code"],
                               original_sqm=u["sqm"], done_sqm=None)
    await write_audit(session, username, "BULK_IMPORT_SME_EQUIPMENT",
                      "sme_equipment",
                      f"{site}: +{len(plan['inserts'])} ~{len(plan['updates'])}")


async def plan_sme_recipes(session: AsyncSession, data: bytes) -> dict:
    headers, rows = _sheet_rows(data, None, ("lining_system_code", "material_code"))
    cols = {"Lining_System_Name": ("Lining_System_Short_Name",),
            "Substrate": ("Substrate",), "System_Keys": ("System Key's", "System_Keys"),
            "Lining_Thickness": ("Lining_Thicknes", "Lining_Thickness"),
            "Lining_System": ("Lining_System",), "Lining_Type": ("Lining_Type",),
            "Material_Description": ("Material_Description",),
            "Material_Name": ("Material_Name",), "UOM": ("UOM",),
            "Package_Size": ("PACKAGE SIZE", "Package_Size"),
            "Sl_No": ("Sl. #",)}
    ix = {field: _col(headers, *names) for field, names in cols.items()}
    code_i = _col(headers, "Lining_System_Code")
    mat_i = _col(headers, "Material_Code")
    sqm_i = _col(headers, "For_1_SQM")
    sap_i = _col(headers, "SAP_Code", "SAP CODE")
    sap_aware = sap_i is not None  # 2026-07-18 workbook layout
    if code_i is None or mat_i is None or sqm_i is None:
        raise HTTPException(422, "recipe sheet needs Lining_System_Code, "
                                 "Material_Code and For_1_SQM")
    existing = {(str(r["Lining_System_Code"]).strip(), r["Material_Code"],
                 _s(r.get("SAP_Code")) or ""): dict(r)
                for r in (await session.execute(select(recipe_t))).mappings().all()}
    rejects: list[dict] = []
    # Line identity is (code, material, SAP). PU systems carry Comp-A/B/C/D
    # lines that share one Material_Code and differ only by variant SAP; a
    # repeat of the SAME identity in a SAP-aware file is a deliberate coat
    # line — For_1_SQM sums (e.g. CONDL2 primer + body coat). Legacy files
    # (no SAP column) keep the historical first-occurrence-wins dedupe.
    agg: dict[tuple, dict] = {}
    dup_skips, coat_merges = 0, 0
    for n, row in enumerate(rows, start=1):
        code, mat_cell = _s(row[code_i]), _s(row[mat_i])
        if not code or not mat_cell:
            continue
        try:
            code = str(int(float(code)))
        except ValueError:
            rejects.append({"row": n, "reason": f"non-numeric code {code!r}"})
            continue
        sap = _s(row[sap_i]) if sap_aware and sap_i < len(row) else None
        qty = _f(row[sqm_i]) or 0.0
        fields = {}
        for field, i in ix.items():
            v = _s(row[i]) if i is not None and i < len(row) else None
            if v is not None:
                fields[field] = v
        # a comma-separated Material_Code cell is one line per material
        for mat in (m.strip() for m in mat_cell.split(",")):
            if not mat:
                continue
            key = (code, mat, sap or "")
            cur = agg.get(key)
            if cur is not None:
                if sap_aware:
                    cur["For_1_SQM"] += qty
                    coat_merges += 1
                else:
                    dup_skips += 1
                continue
            agg[key] = {"For_1_SQM": qty, **fields,
                        **({"SAP_Code": sap} if sap else {})}

    inserts, updates, unchanged = [], [], 0
    for (code, mat, sap), fields in agg.items():
        cur = existing.get((code, mat, sap))
        if cur is None:
            inserts.append({"Lining_System_Code": code, "Material_Code": mat,
                            **fields})
        else:
            diff = {k: v for k, v in fields.items() if cur.get(k) != v}
            if diff:
                updates.append({"id": cur["id"], "diff": diff})
            else:
                unchanged += 1
    warnings = []
    if dup_skips:
        warnings.append(f"{dup_skips} repeated (code, material) line(s) skipped "
                        f"— first occurrence wins (legacy bootstrap rule)")
    if coat_merges:
        warnings.append(f"{coat_merges} repeated (code, material, SAP) coat "
                        f"line(s) merged — For_1_SQM summed")
    return {"inserts": inserts, "updates": updates, "unchanged": unchanged,
            "rejects": rejects, "warnings": warnings}


async def apply_sme_recipes(session: AsyncSession, plan: dict, username: str) -> None:
    for row in plan["inserts"]:
        await session.execute(insert(recipe_t).values(**row))
    for u in plan["updates"]:
        await session.execute(update(recipe_t).where(recipe_t.c["id"] == u["id"])
                              .values(**u["diff"]))
    await write_audit(session, username, "BULK_IMPORT_SME_RECIPES", "sme_recipe",
                      f"+{len(plan['inserts'])} ~{len(plan['updates'])}")


async def plan_sme_materials(session: AsyncSession, data: bytes) -> dict:
    headers, rows = _sheet_rows(data, None, ("material_code", "material_name"))
    ix = {"Item": _col(headers, "Item"),
          "Vendor": _col(headers, "Vendor/supplying plant", "Vendor"),
          "Purchasing_Document": _col(headers, "Purchasing Document"),
          "Document_Date": _col(headers, "Document Date"),
          "Material_Name": _col(headers, "Material_Name"),
          "Nature": _col(headers, "Nature"), "UOM": _col(headers, "UOM"),
          "sap": _col(headers, "SAP_Code", "SAP CODE"),
          "avail": _col(headers, "Available_Qty"),
          "ordered": _col(headers, "Ordered_Qty")}
    mat_i = _col(headers, "Material_Code")
    if mat_i is None:
        raise HTTPException(422, "Material_Code column missing")
    agg: dict[str, dict] = {}
    for row in rows:
        mat = _s(row[mat_i]) if mat_i < len(row) else None
        if not mat:
            continue

        def cell(key):
            i = ix[key]
            return row[i] if i is not None and i < len(row) else None
        a = agg.setdefault(mat, {"Initial_Available_Qty": 0.0,
                                 "Initial_Ordered_Qty": 0.0})
        a["Initial_Available_Qty"] += _f(cell("avail")) or 0.0
        a["Initial_Ordered_Qty"] += _f(cell("ordered")) or 0.0
        dd = cell("Document_Date")
        dd = (_iso(dd) or "")[:10] or None
        if dd and dd > (a.get("Document_Date") or ""):
            a["Document_Date"] = dd  # most recent PO date wins
        sap = _s(cell("sap"))
        if sap is not None:  # one material can span variant SAPs (1041-1 …)
            saps = a.setdefault("_saps", [])
            if sap not in saps:
                saps.append(sap)
        for field in ("Item", "Vendor", "Purchasing_Document",
                      "Material_Name", "Nature", "UOM"):
            v = _s(cell(field))
            if v is not None and field not in a:
                a[field] = v
    existing = {r["Material_Code"]: dict(r) for r in
                (await session.execute(select(seed_t))).mappings().all()}
    inserts, updates, unchanged = [], [], 0
    for mat, a in agg.items():
        saps = a.pop("_saps", None)
        if saps:
            a["SAP_Code"] = ", ".join(saps)
        a["Initial_Available_Qty"] = round(a["Initial_Available_Qty"], 4)
        a["Initial_Ordered_Qty"] = round(a["Initial_Ordered_Qty"], 4)
        cur = existing.get(mat)
        if cur is None:
            inserts.append({"Material_Code": mat, **a})
        else:
            diff = {k: v for k, v in a.items() if cur.get(k) != v}
            if diff:
                updates.append({"Material_Code": mat, "diff": diff})
            else:
                unchanged += 1
    return {"inserts": inserts, "updates": updates, "unchanged": unchanged,
            "rejects": [], "warnings": []}


async def apply_sme_materials(session: AsyncSession, plan: dict, username: str) -> None:
    for row in plan["inserts"]:
        stmt = pg_insert(seed_t).values(**row, updated_at=func.now())
        stmt = stmt.on_conflict_do_update(
            index_elements=["Material_Code"],
            set_={**{k: stmt.excluded[k] for k in row if k != "Material_Code"},
                  "updated_at": func.now()})
        await session.execute(stmt)
    for u in plan["updates"]:
        await session.execute(update(seed_t)
                              .where(seed_t.c["Material_Code"] == u["Material_Code"])
                              .values(**u["diff"], updated_at=func.now()))
    await write_audit(session, username, "BULK_IMPORT_SME_MATERIALS",
                      "sme_inventory_seed",
                      f"+{len(plan['inserts'])} ~{len(plan['updates'])}")


# ─── endpoints ────────────────────────────────────────────────────────────────
def _summary(plan: dict) -> dict:
    if "sections" in plan:  # ledger
        return {k: {"inserts": len(s["inserts"]),
                    "corrections": len(s["corrections"]),
                    "matched": s["matched"], "zero_skipped": s["zero_skipped"],
                    "db_only": s["db_only"]}
                for k, s in plan["sections"].items()}
    return {"inserts": len(plan["inserts"]), "updates": len(plan["updates"]),
            "unchanged": plan["unchanged"], "rejects": len(plan["rejects"])}


async def _read_upload(file: UploadFile) -> bytes:
    data = await file.read()
    if len(data) > MAX_XLSX_BYTES:
        raise HTTPException(413, "workbook exceeds the 8 MB import cap")
    if not data[:4] == b"PK\x03\x04":
        raise HTTPException(422, "expected an .xlsx upload")
    return data


_SME_KINDS = {
    "sme-equipment": (plan_sme_equipment, apply_sme_equipment, True),
    "sme-recipes": (plan_sme_recipes, apply_sme_recipes, False),
    "sme-materials": (plan_sme_materials, apply_sme_materials, False),
}


@router.post("/{kind}", summary="Dry-run (default) or commit a bulk Excel import")
async def bulk_import(kind: str,
                      file: UploadFile = File(...),
                      commit: bool = Query(False),
                      site_id: Optional[str] = Query(None),
                      user: dict = Depends(require_roles("hod")),
                      session: AsyncSession = Depends(get_session)):
    data = await _read_upload(file)
    if kind in ("inventory", "ledger"):
        if user["level"] < 4:
            raise HTTPException(403, "inventory/ledger import is admin-only")
        site = (site_id or "CNCEC").strip()
        plan = await (plan_inventory(session, data, site) if kind == "inventory"
                      else plan_ledger(session, data, site))
        if commit:
            await (apply_inventory if kind == "inventory"
                   else apply_ledger)(session, plan, user["username"])
            await session.commit()
    elif kind in _SME_KINDS:
        plan_fn, apply_fn, scoped = _SME_KINDS[kind]
        if scoped:
            site = _write_site(user, site_id)
            plan = await plan_fn(session, data, site)
        else:
            plan = await plan_fn(session, data)
        if commit:
            await apply_fn(session, plan, user["username"])
            await session.commit()
    else:
        raise HTTPException(404, f"unknown import kind {kind!r} (use one of "
                                 f"inventory, ledger, sme-equipment, "
                                 f"sme-recipes, sme-materials)")
    resp = {"kind": kind, "committed": bool(commit), "summary": _summary(plan),
            "warnings": plan.get("warnings", []),
            "rejects": plan.get("rejects", [])[:200]}
    if not commit:  # preview payload for the UI (trimmed)
        if "sections" in plan:
            resp["preview"] = {k: {"inserts": s["inserts"][:20],
                                   "corrections": s["corrections"][:20]}
                               for k, s in plan["sections"].items()}
        else:
            resp["preview"] = {"inserts": plan["inserts"][:20],
                               "updates": plan["updates"][:20]}
    return resp
