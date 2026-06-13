"""
mailer.py — General Industries Lightning Hub
============================================
Excel report generation engine + SMTP email delivery.

Pure-Python API (testable without Streamlit):
  build_daily_report(conn, report_date)  → bytes  (.xlsx)
  build_monthly_report(conn)             → bytes  (.xlsx)
  build_low_stock_report(conn)           → bytes  (.xlsx)
  parse_recipients(raw)                  → list[str]
  get_default_recipients()               → str
  _build_email_message(...)              → MIMEMultipart
  send_email(recipients, subject, ...)   → (bool, str)
  send_eod_report(recipients, conn)      → (bool, str)
"""

import io
import os
import platform
import datetime

import pandas as pd
import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from dotenv import load_dotenv

from database import get_connection, load_live_inventory, get_low_stock_items

# Load .env on import
load_dotenv()

# Windows-only Outlook COM bindings — exposed at module level so tests can
# patch them, and so we don't pay the import cost on every send.
try:
    import win32com.client as win32  # type: ignore
    import pythoncom  # type: ignore
except ImportError:
    # Non-Windows runtime: provide attribute stubs so unittest.mock.patch can
    # target `mailer.win32.Dispatch` / `mailer.pythoncom.CoInitialize`. The
    # actual Outlook code path is gated on `platform.system() == "Windows"`.
    from types import SimpleNamespace
    win32 = SimpleNamespace(Dispatch=None)
    pythoncom = SimpleNamespace(CoInitialize=None, CoUninitialize=None)

# ---------------------------------------------------------------------------
# SMTP CONFIGURATION  (all sourced from .env)
# ---------------------------------------------------------------------------

DEFAULT_RECIPIENTS = os.getenv("DEFAULT_RECIPIENTS", "")


# ---------------------------------------------------------------------------
# EXCEL STYLE TOKENS  — GI Corporate Theme
# ---------------------------------------------------------------------------
# Header: GI Deep Blue (#003366) fill, white bold text
_HEADER_FILL  = PatternFill(start_color="003366", end_color="003366", fill_type="solid")
_HEADER_FONT  = Font(bold=True, color="FFFFFF", name="Calibri", size=11)
_HEADER_ALIGN = Alignment(horizontal="center", vertical="center", wrap_text=True)

# Gold accent (totals rows)
_GOLD_FILL   = PatternFill(start_color="D4AF37", end_color="D4AF37", fill_type="solid")
_GOLD_FONT   = Font(bold=True, color="001F40", name="Calibri", size=11)

# Alternating row colour
_ALT_FILL    = PatternFill(start_color="E8EEF4", end_color="E8EEF4", fill_type="solid")

# Low-stock semantic fills
_RED_FILL    = PatternFill(start_color="FFDEDE", end_color="FFDEDE", fill_type="solid")
_AMBER_FILL  = PatternFill(start_color="FFF3CD", end_color="FFF3CD", fill_type="solid")

_BODY_FONT   = Font(name="Calibri", size=10)
_BOLD_FONT   = Font(bold=True, name="Calibri", size=10)

_THIN_BORDER = Border(
    left=Side(style="thin", color="CCCCCC"),
    right=Side(style="thin", color="CCCCCC"),
    top=Side(style="thin", color="CCCCCC"),
    bottom=Side(style="thin", color="CCCCCC"),
)


# ---------------------------------------------------------------------------
# PRIVATE STYLING HELPER
# ---------------------------------------------------------------------------
def _apply_gi_style(
    ws,
    df: pd.DataFrame,
    low_stock_col: str = None,
    min_qty_col: str = None,
) -> None:
    """
    Writes `df` into `ws` (starting at row 1) with GI corporate styling:
      - Row 1  : Blue header
      - Even rows: light blue alternating fill
      - Columns auto-sized (capped at 45 chars)
      - Row 1 frozen
      - Optional: red/amber row fill when low_stock_col value is < min_qty_col value
    """
    # Write data
    for r_idx, row in enumerate(
        [df.columns.tolist()] + df.values.tolist(), start=1
    ):
        for c_idx, value in enumerate(row, start=1):
            cell = ws.cell(row=r_idx, column=c_idx, value=value)

            if r_idx == 1:
                # Header
                cell.fill  = _HEADER_FILL
                cell.font  = _HEADER_FONT
                cell.alignment = _HEADER_ALIGN
            else:
                # Body
                cell.font = _BODY_FONT
                if r_idx % 2 == 0:
                    cell.fill = _ALT_FILL

                # Low-stock colour override
                if low_stock_col and min_qty_col:
                    cols = df.columns.tolist()
                    if low_stock_col in cols and min_qty_col in cols:
                        stock_val = df.iloc[r_idx - 2][low_stock_col]
                        min_val   = df.iloc[r_idx - 2][min_qty_col]
                        try:
                            if float(stock_val) <= 0:
                                cell.fill = _RED_FILL
                            elif float(stock_val) < float(min_val):
                                cell.fill = _AMBER_FILL
                        except (TypeError, ValueError):
                            pass

            cell.border = _THIN_BORDER

    # Auto-size columns
    for col_cells in ws.columns:
        max_len = max(
            (len(str(c.value)) if c.value is not None else 0) for c in col_cells
        )
        ws.column_dimensions[col_cells[0].column_letter].width = min(max_len + 4, 45)

    # Freeze header row
    ws.freeze_panes = "A2"


def _empty_xlsx(message: str = "No data available.") -> bytes:
    """Returns a minimal valid .xlsx with a single message cell."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws["A1"] = message
    ws["A1"].font = _BOLD_FONT
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def _df_to_xlsx_bytes(wb: openpyxl.Workbook) -> bytes:
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


# ---------------------------------------------------------------------------
# REPORT BUILDERS
# ---------------------------------------------------------------------------
def build_daily_report(
    conn=None,
    report_date: datetime.date = None,
) -> bytes:
    """
    Daily Issue Log Report (.xlsx).
    Combines:
      - Committed consumption records for `report_date` (default: today)
      - All currently pending (staged) items
    Returns raw xlsx bytes.
    """
    _owns = conn is None
    if _owns:
        conn = get_connection()

    if report_date is None:
        report_date = datetime.date.today()

    date_str = str(report_date)

    try:
        committed = pd.read_sql(
            "SELECT * FROM consumption WHERE Date = ?", conn, params=(date_str,)
        )
        pending = pd.read_sql("SELECT * FROM pending_issues", conn)

        # Pull item descriptions
        try:
            inv = pd.read_sql(
                "SELECT SAP_Code, Equipment_Description, UOM FROM inventory", conn
            )
        except Exception:
            inv = pd.DataFrame(columns=["SAP_Code", "Equipment_Description", "UOM"])
    finally:
        if _owns:
            conn.close()

    # ── Committed sheet ────────────────────────────────────────────────────
    wb = openpyxl.Workbook()

    # Sheet 1: Committed
    ws_committed = wb.active
    ws_committed.title = "Committed Today"

    if committed.empty:
        ws_committed["A1"] = f"No committed consumption records for {date_str}."
        ws_committed["A1"].font = _BOLD_FONT
    else:
        df_c = pd.merge(committed, inv, on="SAP_Code", how="left")
        # Reorder: descriptive columns first
        priority = ["Date", "SAP_Code", "Equipment_Description", "UOM",
                    "Quantity", "Work_Type", "Issued_By", "Issued_To",
                    "Tank_No", "Serial_No", "PR_Number", "Remarks"]
        ordered_c = [c for c in priority if c in df_c.columns] + \
                    [c for c in df_c.columns if c not in priority]
        _apply_gi_style(ws_committed, df_c[ordered_c])

    # Sheet 2: Pending
    ws_pending = wb.create_sheet("Pending (Unconfirmed)")
    if pending.empty:
        ws_pending["A1"] = "No items currently in the staging queue."
        ws_pending["A1"].font = _BOLD_FONT
    else:
        df_p = pd.merge(pending, inv, on="SAP_Code", how="left")
        drop_cols = [c for c in ["id", "Timestamp"] if c in df_p.columns]
        df_p = df_p.drop(columns=drop_cols)
        _apply_gi_style(ws_pending, df_p)

    # Summary tab
    ws_summary = wb.create_sheet("Summary")
    summary_data = {
        "Report Date":     [date_str],
        "Committed Items": [len(committed)],
        "Pending Items":   [len(pending)],
        "Total Qty Committed": [
            pd.to_numeric(committed.get("Quantity", pd.Series([0])), errors="coerce").sum()
            if not committed.empty else 0
        ],
        "Generated At": [datetime.datetime.now().strftime("%Y-%m-%d %H:%M")],
    }
    _apply_gi_style(ws_summary, pd.DataFrame(summary_data))

    return _df_to_xlsx_bytes(wb)


def build_monthly_report(conn=None) -> bytes:
    """
    Monthly Consumption Pivot Report (.xlsx).
    Pivots the consumption table: rows = SAP_Code, columns = YYYY-MM months.
    Adds Grand Total column sorted descending.
    Returns raw xlsx bytes.
    """
    _owns = conn is None
    if _owns:
        conn = get_connection()

    try:
        df = pd.read_sql("SELECT * FROM consumption", conn)
        try:
            inv = pd.read_sql(
                "SELECT SAP_Code, Equipment_Description, UOM FROM inventory", conn
            )
        except Exception:
            inv = pd.DataFrame(columns=["SAP_Code", "Equipment_Description", "UOM"])
    finally:
        if _owns:
            conn.close()

    if df.empty:
        return _empty_xlsx("No consumption data available for monthly report.")

    # Detect quantity column
    qty_col = next(
        (c for c in df.columns if "qty" in c.lower() or "quantity" in c.lower()), None
    )
    if not qty_col:
        return _empty_xlsx("No Quantity column found in consumption table.")

    df[qty_col] = pd.to_numeric(df[qty_col], errors="coerce").fillna(0)
    df["Date"]  = pd.to_datetime(df["Date"], errors="coerce")
    df["Month"] = df["Date"].dt.strftime("%Y-%m")
    df = df.dropna(subset=["Month"])

    pivot = df.pivot_table(
        index="SAP_Code",
        columns="Month",
        values=qty_col,
        aggfunc="sum",
        fill_value=0,
    ).reset_index()

    pivot.columns.name = None
    month_cols = [c for c in pivot.columns if c != "SAP_Code"]
    pivot["Grand Total"] = pivot[month_cols].sum(axis=1)
    pivot = pivot.sort_values("Grand Total", ascending=False).reset_index(drop=True)

    # Merge description
    pivot = pd.merge(pivot, inv[["SAP_Code", "Equipment_Description", "UOM"]],
                     on="SAP_Code", how="left")
    front = ["SAP_Code", "Equipment_Description", "UOM"]
    ordered = front + [c for c in pivot.columns if c not in front]
    pivot = pivot[ordered]

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Monthly Consumption"

    _apply_gi_style(ws, pivot)

    # Gold fill for Grand Total column
    gt_col_idx = pivot.columns.tolist().index("Grand Total") + 1
    for row_idx in range(2, len(pivot) + 2):
        cell = ws.cell(row=row_idx, column=gt_col_idx)
        cell.fill = _GOLD_FILL
        cell.font = _GOLD_FONT

    return _df_to_xlsx_bytes(wb)


def build_low_stock_report(conn=None) -> bytes:
    """
    Low-Stock Warning Report (.xlsx).
    Lists all items where Current_Stock < Minimum_Qty.
    Red fill  = Current_Stock <= 0 (critical/empty)
    Amber fill = 0 < Current_Stock < Minimum_Qty (low)
    Returns raw xlsx bytes.
    """
    _owns = conn is None
    if _owns:
        conn = get_connection()

    low_df = get_low_stock_items(conn)

    if _owns:
        conn.close()

    if low_df.empty:
        return _empty_xlsx("✅ All stock levels are adequate. No low-stock items.")

    col_order = [c for c in [
        "SAP_Code", "Equipment_Description", "UOM",
        "Current_Stock", "Minimum_Qty", "Shortage",
        "Total_Received", "Total_Consumed", "Total_Returned",
    ] if c in low_df.columns]

    display = low_df[col_order].copy()
    display = display.sort_values("Shortage", ascending=False).reset_index(drop=True)

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Low-Stock Warning"

    _apply_gi_style(
        ws, display,
        low_stock_col="Current_Stock",
        min_qty_col="Minimum_Qty",
    )

    return _df_to_xlsx_bytes(wb)


# ---------------------------------------------------------------------------
# RECIPIENT HELPERS
# ---------------------------------------------------------------------------
def parse_recipients(raw: str) -> list[str]:
    """
    Parses a comma or newline-separated string of email addresses.
    Strips whitespace, filters empty strings.
    """
    import re
    parts = re.split(r"[,\n;]+", raw or "")
    return [p.strip() for p in parts if p.strip()]


def get_default_recipients() -> str:
    """Returns DEFAULT_RECIPIENTS from .env as a raw comma-separated string."""
    return DEFAULT_RECIPIENTS


# ---------------------------------------------------------------------------
# EMAIL ENGINE
# ---------------------------------------------------------------------------
def send_eod_report(
    recipients: list[str],
    conn=None,
    report_date: datetime.date = None,
) -> tuple[bool, str]:
    """
    Builds the Daily Report and opens a local email draft.
    - Windows: uses the Outlook Desktop App (COM automation)
    - Mac/Linux: opens the system default mail app via a mailto: link
    """
    if not recipients:
        return False, "No recipients specified."

    if report_date is None:
        report_date = datetime.date.today()

    date_str = report_date.strftime("%d %b %Y")
    filename = f"GI_EOD_Report_{report_date.strftime('%Y%m%d')}.xlsx"
    temp_filepath = os.path.abspath(filename)

    # 1. Build and save the Excel report (shared by both paths)
    xlsx_bytes = build_daily_report(conn=conn, report_date=report_date)
    with open(temp_filepath, "wb") as f:
        f.write(xlsx_bytes)

    # ── Windows path: Outlook COM automation ────────────────────────────────
    if platform.system() == "Windows":
        try:
            pythoncom.CoInitialize()

            outlook = win32.Dispatch('outlook.application')
            mail = outlook.CreateItem(0)  # 0 = Mail Item
            mail.To = "; ".join(recipients)  # Outlook uses semicolons
            mail.Subject = f"GI Inventory — EOD Report {date_str}"
            mail.HTMLBody = f"""
            <html><body style="font-family: Calibri, Arial, sans-serif; color: #222;">
              <div style="background:#003366; padding:20px; border-radius:8px 8px 0 0;">
                <h2 style="color:#D4AF37; margin:0;">⚡ General Industries</h2>
                <p style="color:#ccc; margin:4px 0 0 0;">End-of-Day Inventory Report — {date_str}</p>
              </div>
              <div style="padding:20px; background:#f8f9fa; border:1px solid #ddd; border-radius:0 0 8px 8px;">
                <p>Dear Management,</p>
                <p>Please find attached the <strong>End-of-Day Inventory Issue Report</strong>
                   for <strong>{date_str}</strong>.</p>
                <p>The report includes:</p>
                <ul>
                  <li>All material issues committed to the Master Log today</li>
                  <li>Any items still in the pending staging queue</li>
                  <li>Daily summary statistics</li>
                </ul>
                <p style="color:#666; font-size:12px; margin-top:30px; border-top:1px solid #ddd; padding-top:10px;">
                  This is an automated message from the GI Lightning Hub v2.0.<br>
                  Do not reply to this email.
                </p>
              </div>
            </body></html>
            """
            mail.Attachments.Add(temp_filepath)
            mail.Display(True)

            if os.path.exists(temp_filepath):
                os.remove(temp_filepath)

            return True, "Draft opened in Outlook! Please review and click Send."

        except Exception as e:
            if os.path.exists(temp_filepath):
                os.remove(temp_filepath)
            return False, f"Outlook Desktop App error: {str(e)}"

        finally:
            try:
                pythoncom.CoUninitialize()
            except Exception:
                pass

    # ── Mac / Linux path: default mail app via mailto: ──────────────────────
    else:
        try:
            import subprocess
            import urllib.parse

            subject = urllib.parse.quote(f"GI Inventory — EOD Report {date_str}")
            body_text = (
                f"Dear Management,\n\n"
                f"Please find attached the End-of-Day Inventory Issue Report for {date_str}.\n\n"
                f"The report includes:\n"
                f"  - All material issues committed to the Master Log today\n"
                f"  - Any items still in the pending staging queue\n"
                f"  - Daily summary statistics\n\n"
                f"Attachment saved at:\n{temp_filepath}\n\n"
                f"This is an automated message from the GI Lightning Hub v2.0."
            )
            to_field = urllib.parse.quote(", ".join(recipients))
            body_field = urllib.parse.quote(body_text)
            mailto_url = f"mailto:{to_field}?subject={subject}&body={body_field}"

            # Open the system default mail app
            subprocess.Popen(["open", mailto_url])

            return (
                True,
                f"Draft opened in your default Mail app! "
                f"Please attach the report manually from:\n{temp_filepath}"
            )

        except Exception as e:
            return False, f"Failed to open Mail app: {str(e)}"


def send_pr_logistics_alert(pr_number: str, sap_code: str, mat_name: str, site_id: str, qty_received: float, balance_info: dict) -> tuple[bool, str]:
    """
    Module 6: Sends an immediate notification email to the Logistics department
    whenever materials are received against an open Purchase Request.
    """
    import smtplib
    from email.mime.multipart import MIMEMultipart
    from email.mime.text import MIMEText
    import os
    from dotenv import load_dotenv
    
    load_dotenv()
    
    # Load configuration settings from environment
    sender = os.getenv("SMTP_USER", "hub-alerts@generalindustries.net")
    password = os.getenv("SMTP_PASS", "")
    smtp_server = os.getenv("SMTP_SERVER", "smtp.office365.com")
    try:
        smtp_port = int(os.getenv("SMTP_PORT", "587"))
    except ValueError:
        smtp_port = 587
        
    # Logistics department email destination
    logistics_recipient = os.getenv("LOGISTICS_EMAIL", "logistics@generalindustries.net")

    msg = MIMEMultipart("alternative")
    msg["Subject"] = f"🔔 PR Balance Update: PR {pr_number} ({site_id})"
    msg["From"] = sender
    msg["To"] = logistics_recipient

    # Calculate status representation
    status_tag = "🔴 PARTIAL DELIVERY" if balance_info["balance"] > 0 else "🟢 FULLY FULFILLED & CLOSED"

    # Branded HTML email template matching company guidelines
    html_content = f"""
    <html>
    <body style="font-family: 'Segoe UI', Tahoma, Geneva, Verdana, sans-serif; color: #333; line-height: 1.6;">
        <div style="background-color: #0A192F; padding: 15px; text-align: center; border-radius: 6px 6px 0 0;">
            <h2 style="color: #F5A623; margin: 0; font-size: 20px;">General Industries Hub — Logistics Advisory</h2>
        </div>
        <div style="padding: 20px; border: 1px solid #E0E0E0; border-top: none; border-radius: 0 0 6px 6px; background-color: #FAFAFA;">
            <p>Dear Logistics Team,</p>
            <p>An incoming shipment has been logged at <strong>{site_id}</strong> against Purchase Request <strong>{pr_number}</strong>.</p>
            
            <div style="background: #FFF; border-left: 4px solid #F5A623; padding: 12px; margin: 15px 0; border-radius: 4px;">
                <table style="width: 100%; border-collapse: collapse;">
                    <tr><td style="padding: 4px 0; font-weight: bold; width: 40%;">Material Code / SAP:</td><td>[{sap_code}] {mat_name}</td></tr>
                    <tr><td style="padding: 4px 0; font-weight: bold;">Quantity Just Received:</td><td style="color: #0A192F; font-weight: bold;">{qty_received}</td></tr>
                    <tr><td style="padding: 4px 0; font-weight: bold;">Delivery Site:</td><td>{site_id}</td></tr>
                </table>
            </div>

            <h3 style="color: #0A192F; border-bottom: 1px solid #DDD; padding-bottom: 5px;">PR Summary Balance</h3>
            <table style="width: 100%; text-align: left; border-collapse: collapse; margin-top: 10px;">
                <thead>
                    <tr style="background: #0A192F; color: #FFF;">
                        <th style="padding: 8px;">Metric</th>
                        <th style="padding: 8px;">Quantity</th>
                    </tr>
                </thead>
                <tbody>
                    <tr><td style="padding: 8px; border-bottom: 1px solid #EEE;">Total Requested in PR:</td><td style="padding: 8px; border-bottom: 1px solid #EEE;">{balance_info['requested']}</td></tr>
                    <tr><td style="padding: 8px; border-bottom: 1px solid #EEE;">Total Accumulated Received:</td><td style="padding: 8px; border-bottom: 1px solid #EEE; color: green; font-weight: bold;">{balance_info['received']}</td></tr>
                    <tr style="background: #F1F1F1;"><td style="padding: 8px; font-weight: bold;">Remaining Balance Owed:</td><td style="padding: 8px; font-weight: bold; color: {'#C0392B' if balance_info['balance'] > 0 else '#27AE60'};">{balance_info['balance']}</td></tr>
                </tbody>
            </table>

            <div style="margin-top: 20px; padding: 10px; background: #EEE; border-radius: 4px; text-align: center; font-weight: bold; color: #333;">
                Status: {status_tag}
            </div>

            <p style="font-size: 11px; color: #777; margin-top: 25px; border-top: 1px solid #E5E5E5; padding-top: 10px;">
                This is an automated operational alert generated by the General Industries HUB. Please do not reply directly to this email.
            </p>
        </div>
    </body>
    </html>
    """
    msg.attach(MIMEText(html_content, "html"))

    if not password:
        # If credentials aren't provided in .env yet, fall back to simulation log so the app doesn't crash
        print(f"✉️ [SIMULATION] Logistics Email Triggered for PR {pr_number}. Balance remaining: {balance_info['balance']}")
        return True, f"✅ Receipt saved. Logistics Email simulated successfully (No password in .env)!"

    try:
        server = smtplib.SMTP(smtp_server, smtp_port)
        server.starttls()
        server.login(sender, password)
        server.sendmail(sender, [logistics_recipient], msg.as_string())
        server.close()
        return True, "✅ Receipt saved and balance statement emailed to Logistics!"
    except Exception as e:
        return False, f"⚠️ Receipt saved, but automated email failed to transmit: {str(e)}"



def draft_logistics_email_via_outlook(pr_number: str, site_id: str, pr_df) -> tuple[bool, str]:
    """
    Module 6: Opens a local email draft with a well-formatted HTML table 
    showing the pending balance for a specific Purchase Request.
    - Windows: uses the Outlook Desktop App (COM automation)
    - Mac/Linux: opens the system default mail app via a mailto: link
    """
    import os
    from dotenv import load_dotenv
    load_dotenv()
    logistics_recipient = os.getenv("LOGISTICS_EMAIL", "logistics@generalindustries.net")

    # Build the HTML Table Rows dynamically from the Database data
    table_rows = ""
    for _, row in pr_df.iterrows():
        mat_code = row.get("Material_Code", "N/A")
        name = row.get("Material_Name", "Unknown Material")
        req = float(row.get("Requested_Qty", 0))
        pend = float(row.get("Pending_Qty", 0))
        rec = req - pend
        status_color = "#27AE60" if pend <= 0 else "#C0392B"
        table_rows += f"""
        <tr>
            <td style="padding: 8px; border: 1px solid #ddd;">{mat_code}</td>
            <td style="padding: 8px; border: 1px solid #ddd;">{name}</td>
            <td style="padding: 8px; border: 1px solid #ddd; text-align: center;">{req}</td>
            <td style="padding: 8px; border: 1px solid #ddd; text-align: center;">{rec}</td>
            <td style="padding: 8px; border: 1px solid #ddd; text-align: center; color: {status_color}; font-weight: bold;">{max(0, pend)}</td>
        </tr>
        """

    html_body = f"""
    <html>
    <body style="font-family: 'Segoe UI', Arial, sans-serif; color: #333; line-height: 1.6;">
        <h2 style="color: #0A192F;">General Industries Hub — Logistics Advisory</h2>
        <p>Dear Logistics Team,</p>
        <p>Please find the current pending delivery status for <strong>PR {pr_number}</strong> requested by <strong>{site_id}</strong>.</p>

        <table style="width: 100%; border-collapse: collapse; margin-top: 15px; font-size: 14px;">
            <thead>
                <tr style="background-color: #0A192F; color: #FFF;">
                    <th style="padding: 10px; border: 1px solid #0A192F; text-align: left;">Material Code</th>
                    <th style="padding: 10px; border: 1px solid #0A192F; text-align: left;">Description</th>
                    <th style="padding: 10px; border: 1px solid #0A192F;">Requested</th>
                    <th style="padding: 10px; border: 1px solid #0A192F;">Received</th>
                    <th style="padding: 10px; border: 1px solid #0A192F;">Pending Balance</th>
                </tr>
            </thead>
            <tbody>
                {table_rows}
            </tbody>
        </table>
        
        <p style="margin-top: 20px;">Kindly expedite the pending materials to ensure uninterrupted site operations.</p>
        <p>Best Regards,<br><strong>{site_id} Hub Management</strong></p>
    </body>
    </html>
    """

    # ── Windows path: Outlook COM automation ────────────────────────────────
    if platform.system() == "Windows":
        try:
            pythoncom.CoInitialize()

            outlook = win32.Dispatch('outlook.application')
            mail = outlook.CreateItem(0)
            mail.To = logistics_recipient
            mail.Subject = f"🔔 Pending Delivery Advisory: PR {pr_number} ({site_id})"
            mail.HTMLBody = html_body
            mail.Display(True)

            return True, f"Outlook draft opened for PR {pr_number}!"

        except ImportError:
            return False, "pywin32 library not found. Please run: pip install pywin32"
        except Exception as e:
            return False, f"Failed to open Outlook Desktop App: {str(e)}"
        finally:
            try:
                pythoncom.CoUninitialize()
            except Exception:
                pass

    # ── Mac / Linux path: default mail app via mailto: ──────────────────────
    # On Mac we first try to drive Mail.app via AppleScript so the body
    # arrives as HTML (matching the Windows/Outlook treatment). The
    # mailto:// fallback ships a monospaced fixed-width text table that
    # renders cleanly in every email client.
    else:
        try:
            import subprocess
            import urllib.parse

            subject_raw = f"🔔 Pending Delivery Advisory: PR {pr_number} ({site_id})"

            # ── Try Mail.app via AppleScript (HTML body) ─────────────────
            if platform.system() == "Darwin":
                try:
                    # AppleScript needs single-quoted strings escaped.
                    safe_html = (
                        html_body.replace("\\", "\\\\")
                                 .replace('"', '\\"')
                    )
                    osa = (
                        'tell application "Mail"\n'
                        f'  set newMessage to make new outgoing message with properties '
                        f'{{visible:true, subject:"{subject_raw}", '
                        f'content:""}}\n'
                        f'  tell newMessage to make new to recipient at end of to recipients '
                        f'with properties {{address:"{logistics_recipient}"}}\n'
                        '  set html content of newMessage to "' + safe_html + '"\n'
                        'end tell'
                    )
                    res = subprocess.run(
                        ["osascript", "-e", osa],
                        capture_output=True, text=True, timeout=10,
                    )
                    if res.returncode == 0:
                        return True, (
                            f"Mail.app draft opened with formatted table "
                            f"for PR {pr_number}."
                        )
                    # else → fall through to mailto: with a pretty plaintext table
                except Exception:
                    pass

            # ── Fallback: mailto: with a fixed-width plain-text table ────
            # Build a clean column-aligned table that renders well even in
            # plain-text only email clients.
            COL_MC, COL_DESC, COL_REQ, COL_REC, COL_PND = 14, 38, 10, 10, 10
            sep = (
                "+-" + "-" * COL_MC + "-+-" + "-" * COL_DESC + "-+-"
                + "-" * COL_REQ + "-+-" + "-" * COL_REC + "-+-"
                + "-" * COL_PND + "-+"
            )
            header = (
                "| " + "Material Code".ljust(COL_MC)
                + " | " + "Description".ljust(COL_DESC)
                + " | " + "Requested".rjust(COL_REQ)
                + " | " + "Received".rjust(COL_REC)
                + " | " + "Pending".rjust(COL_PND) + " |"
            )

            data_rows = []
            for _, row in pr_df.iterrows():
                mat_code = str(row.get("Material_Code", "N/A"))[:COL_MC]
                name = str(row.get("Material_Name", "Unknown"))[:COL_DESC]
                req = float(row.get("Requested_Qty", 0))
                pend = float(row.get("Pending_Qty", 0))
                rec = req - pend
                data_rows.append(
                    "| " + mat_code.ljust(COL_MC)
                    + " | " + name.ljust(COL_DESC)
                    + " | " + f"{req:.1f}".rjust(COL_REQ)
                    + " | " + f"{rec:.1f}".rjust(COL_REC)
                    + " | " + f"{max(0, pend):.1f}".rjust(COL_PND) + " |"
                )

            lines = [
                "Dear Logistics Team,",
                "",
                f"Please find below the pending delivery status for "
                f"PR {pr_number} requested by {site_id}.",
                "",
                sep,
                header,
                sep,
                *data_rows,
                sep,
                "",
                "Kindly expedite the pending materials to ensure uninterrupted "
                "site operations.",
                "",
                "Best Regards,",
                f"{site_id} Hub Management",
            ]

            subject = urllib.parse.quote(subject_raw)
            body_field = urllib.parse.quote("\n".join(lines))
            to_field = urllib.parse.quote(logistics_recipient)
            mailto_url = f"mailto:{to_field}?subject={subject}&body={body_field}"
            subprocess.Popen(["open", mailto_url])

            return True, f"Draft opened in your default Mail app for PR {pr_number}!"

        except Exception as e:
            return False, f"Failed to open Mail app: {str(e)}"


def draft_return_logistics_email(site_id: str, return_row: dict) -> tuple[bool, str]:
    """
    Drafts a logistics-notification email after HOD approves a return.
    `return_row` keys: SAP_Code, Equipment_Description, Material_Code,
    Quantity, Return_Reason, Return_DN_No, received_date, received_dn_no,
    received_qty, PR_Number, Lot_Number.
    """
    import os
    try:
        from dotenv import load_dotenv
        load_dotenv()
    except Exception:
        pass
    logistics_recipient = os.environ.get("LOGISTICS_EMAIL", "logistics@generalindustries.net")

    subject_raw = (
        f"↩️ Return Approved — [{return_row.get('SAP_Code','')}] "
        f"{return_row.get('Equipment_Description','')[:40]} ({site_id})"
    )
    html_body = (
        "<html><body style='font-family:Segoe UI, Arial, sans-serif;color:#333;'>"
        "<h2 style='color:#0A192F;'>General Industries Hub — Return Notification</h2>"
        f"<p>Dear Logistics Team,</p>"
        f"<p>The following material has been <b>returned by {site_id}</b> "
        "after HOD approval. Please collect / disposition as required.</p>"
        "<table style='width:100%;border-collapse:collapse;font-size:14px;'>"
        "<tbody>"
        f"<tr><th style='text-align:left;padding:8px;background:#F4F6F8;'>SAP Code</th>"
        f"<td style='padding:8px;border:1px solid #ddd;'>{return_row.get('SAP_Code','')}</td></tr>"
        f"<tr><th style='text-align:left;padding:8px;background:#F4F6F8;'>Material Code</th>"
        f"<td style='padding:8px;border:1px solid #ddd;'>{return_row.get('Material_Code','—') or '—'}</td></tr>"
        f"<tr><th style='text-align:left;padding:8px;background:#F4F6F8;'>Description</th>"
        f"<td style='padding:8px;border:1px solid #ddd;'>{return_row.get('Equipment_Description','')}</td></tr>"
        f"<tr><th style='text-align:left;padding:8px;background:#F4F6F8;'>Return Qty</th>"
        f"<td style='padding:8px;border:1px solid #ddd;'><b>{return_row.get('Quantity','')}</b></td></tr>"
        f"<tr><th style='text-align:left;padding:8px;background:#F4F6F8;'>Reason</th>"
        f"<td style='padding:8px;border:1px solid #ddd;'>{return_row.get('Return_Reason','')}</td></tr>"
        f"<tr><th style='text-align:left;padding:8px;background:#F4F6F8;'>Return DN No.</th>"
        f"<td style='padding:8px;border:1px solid #ddd;'>{return_row.get('Return_DN_No','')}</td></tr>"
        f"<tr><th style='text-align:left;padding:8px;background:#F4F6F8;'>Original Receipt</th>"
        f"<td style='padding:8px;border:1px solid #ddd;'>"
        f"{return_row.get('received_date','')} · DN {return_row.get('received_dn_no','—') or '—'}"
        f" · PR {return_row.get('PR_Number','—') or '—'}"
        f" · Lot {return_row.get('Lot_Number','—') or '—'}"
        f" · Received Qty {return_row.get('received_qty','')}"
        f"</td></tr>"
        "</tbody></table>"
        f"<p style='margin-top:18px;'>Best Regards,<br>"
        f"<b>{site_id} Hub Management</b></p>"
        "</body></html>"
    )

    if platform.system() == "Windows":
        try:
            pythoncom.CoInitialize()
            outlook = win32.Dispatch('outlook.application')
            mail = outlook.CreateItem(0)
            mail.To = logistics_recipient
            mail.Subject = subject_raw
            mail.HTMLBody = html_body
            mail.Display(True)
            return True, "Outlook draft opened for return."
        except Exception as e:
            return False, f"Outlook draft failed: {e}"
        finally:
            try: pythoncom.CoUninitialize()
            except Exception: pass

    try:
        import subprocess, urllib.parse
        text_body = (
            "Dear Logistics Team,\n\n"
            f"The following material has been returned by {site_id} after HOD approval:\n\n"
            f"SAP Code      : {return_row.get('SAP_Code','')}\n"
            f"Material Code : {return_row.get('Material_Code','—')}\n"
            f"Description   : {return_row.get('Equipment_Description','')}\n"
            f"Return Qty    : {return_row.get('Quantity','')}\n"
            f"Reason        : {return_row.get('Return_Reason','')}\n"
            f"Return DN     : {return_row.get('Return_DN_No','')}\n"
            f"Original DN   : {return_row.get('received_dn_no','—')}  "
            f"({return_row.get('received_date','')})\n"
            f"Original PR   : {return_row.get('PR_Number','—')}\n"
            f"Lot Number    : {return_row.get('Lot_Number','—')}\n\n"
            f"Best Regards,\n{site_id} Hub Management\n"
        )
        url = (
            f"mailto:{urllib.parse.quote(logistics_recipient)}"
            f"?subject={urllib.parse.quote(subject_raw)}"
            f"&body={urllib.parse.quote(text_body)}"
        )
        opener = "open" if platform.system() == "Darwin" else "xdg-open"
        subprocess.Popen([opener, url])
        return True, "Mail draft opened for return."
    except Exception as e:
        return False, f"Mail draft failed: {e}"


def draft_rubber_mtc_email(site_id: str, mtc_df) -> tuple[bool, str]:
    """
    Drafts an email to logistics listing rubber materials received without
    an MTC. mtc_df rows expected to have: SAP_Code, Equipment_Description,
    Lot_Number, Quantity.
    """
    import os
    try:
        from dotenv import load_dotenv
        load_dotenv()
    except Exception:
        pass
    logistics_recipient = os.environ.get("LOGISTICS_EMAIL", "logistics@generalindustries.net")

    subject_raw = f"⚠️ Rubber materials received without MTC — {site_id}"
    table_rows = ""
    for _, row in mtc_df.iterrows():
        table_rows += (
            "<tr>"
            f"<td style='padding:8px;border:1px solid #ddd;'>{row.get('SAP_Code','')}</td>"
            f"<td style='padding:8px;border:1px solid #ddd;'>{row.get('Equipment_Description','')}</td>"
            f"<td style='padding:8px;border:1px solid #ddd;'>{row.get('Lot_Number','—') or '—'}</td>"
            f"<td style='padding:8px;border:1px solid #ddd;text-align:center;'>{row.get('Quantity','')}</td>"
            "</tr>"
        )
    html_body = (
        "<html><body style='font-family:Segoe UI, Arial, sans-serif;color:#333;'>"
        "<h2 style='color:#0A192F;'>General Industries Hub — MTC Document Request</h2>"
        "<p>Dear Logistics Team,</p>"
        f"<p>The following rubber materials were received at <b>{site_id}</b> "
        "without an accompanying MTC. Please forward the certificates at the earliest:</p>"
        "<table style='width:100%;border-collapse:collapse;font-size:14px;'>"
        "<thead><tr style='background:#0A192F;color:#fff;'>"
        "<th style='padding:10px;border:1px solid #0A192F;text-align:left;'>SAP Code</th>"
        "<th style='padding:10px;border:1px solid #0A192F;text-align:left;'>Description</th>"
        "<th style='padding:10px;border:1px solid #0A192F;text-align:left;'>Lot Number</th>"
        "<th style='padding:10px;border:1px solid #0A192F;'>Quantity</th>"
        "</tr></thead>"
        f"<tbody>{table_rows}</tbody></table>"
        f"<p style='margin-top:20px;'>Thank you,<br><b>{site_id} Hub Management</b></p>"
        "</body></html>"
    )

    if platform.system() == "Windows":
        try:
            pythoncom.CoInitialize()
            outlook = win32.Dispatch('outlook.application')
            mail = outlook.CreateItem(0)
            mail.To = logistics_recipient
            mail.Subject = subject_raw
            mail.HTMLBody = html_body
            mail.Display(True)
            return True, "Outlook draft opened for missing MTC items."
        except Exception as e:
            return False, f"Outlook draft failed: {e}"
        finally:
            try: pythoncom.CoUninitialize()
            except Exception: pass

    try:
        import subprocess, urllib.parse
        COL_SAP, COL_DESC, COL_LOT, COL_QTY = 12, 36, 14, 8
        sep = ("+-" + "-"*COL_SAP + "-+-" + "-"*COL_DESC + "-+-"
               + "-"*COL_LOT + "-+-" + "-"*COL_QTY + "-+")
        header = ("| " + "SAP".ljust(COL_SAP)
                  + " | " + "Description".ljust(COL_DESC)
                  + " | " + "Lot".ljust(COL_LOT)
                  + " | " + "Qty".rjust(COL_QTY) + " |")
        body_rows = []
        for _, r in mtc_df.iterrows():
            body_rows.append(
                "| " + str(r.get("SAP_Code",""))[:COL_SAP].ljust(COL_SAP)
                + " | " + str(r.get("Equipment_Description",""))[:COL_DESC].ljust(COL_DESC)
                + " | " + str(r.get("Lot_Number","") or "—")[:COL_LOT].ljust(COL_LOT)
                + " | " + f"{float(r.get('Quantity',0) or 0):.1f}".rjust(COL_QTY) + " |"
            )
        lines = [
            "Dear Logistics Team,", "",
            f"The following rubber materials were received at {site_id} without an MTC. "
            "Please forward the certificates:", "",
            sep, header, sep, *body_rows, sep, "",
            "Thank you,", f"{site_id} Hub Management",
        ]
        url = (
            f"mailto:{urllib.parse.quote(logistics_recipient)}"
            f"?subject={urllib.parse.quote(subject_raw)}"
            f"&body={urllib.parse.quote(chr(10).join(lines))}"
        )
        opener = "open" if platform.system() == "Darwin" else "xdg-open"
        subprocess.Popen([opener, url])
        return True, "Mail draft opened for missing MTC items."
    except Exception as e:
        return False, f"Mail draft failed: {e}"